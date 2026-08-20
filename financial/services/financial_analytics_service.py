# financial/services/financial_analytics_service.py
"""
خدمة التحليلات والمؤشرات المالية المتقدمة - Enterprise Financial Analytics & CFO Command Center (v3.0)
توفر مركز قيادة مالي واستراتيجي متكامل لمتخذي القرار المالي والإدارة العليا،
تشمل المحاور الخمسة للتحليل المالي المؤسسي:
1. نسب السيولة ورأس المال العامل (Liquidity Ratios)
2. نسب الربحية ومعدلات العائد المحولة سنوياً (Profitability & Annualized Returns)
3. نسب الملاءة والهيكل التمويلي والرفع المالي (Solvency & Financial Leverage)
4. نسب النشاط والكفاءة ودورة التحول النقدي (Activity & Cash Conversion Cycle)
5. نموذج دوبونت الثلاثي الاستراتيجي (DuPont 3-Step Analysis)
بالإضافة إلى:
- مؤشر ألتمان للسلامة والتعثر المالي (Altman Z-Score for Private Firms)
- بطاقة التقييم والصحة المالية الشاملة (Executive Health Scorecard)
- محرك المقارنة الزمنية بالفترات السابقة مع احتساب فروق التغير
- استعلامات SQL فائقة السرعة O(1) باستخدام TruncMonth وتصدير Excel المعتمد.
"""

import logging
from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Any, Union
from io import BytesIO

from django.db.models import Sum, Q, F, Count
from django.db.models.functions import Coalesce, TruncMonth
from django.utils import timezone
from django.utils.translation import gettext as _

from financial.models.chart_of_accounts import ChartOfAccounts
from financial.models.journal_entry import JournalEntryLine, JournalEntry
from financial.models.cost_center import CostCenter
from financial.services.exchange_rate_service import ExchangeRateService
from financial.services.income_statement_service import IncomeStatementService

logger = logging.getLogger(__name__)


class FinancialAnalyticsService:
    """
    خدمة التحليلات والمؤشرات المالية المؤسسية
    """

    @classmethod
    def get_complete_analytics(
        cls,
        date_from: Optional[Union[date, str]] = None,
        date_to: Optional[Union[date, str]] = None,
        comp_date_from: Optional[Union[date, str]] = None,
        comp_date_to: Optional[Union[date, str]] = None,
        cost_center_id: Optional[Union[int, str]] = None,
        include_unposted: bool = False,
    ) -> Dict[str, Any]:
        """
        توليد كافة التحليلات والمؤشرات المالية التنفيذية للفترة الحالية وفترة المقارنة.
        """
        try:
            # 1. ضبط التواريخ والفترة الحالية
            today = timezone.now().date()
            if isinstance(date_to, str) and date_to.strip():
                try:
                    date_to = datetime.strptime(date_to.strip(), "%Y-%m-%d").date()
                except ValueError:
                    date_to = today
            elif not isinstance(date_to, date):
                date_to = today

            if isinstance(date_from, str) and date_from.strip():
                try:
                    date_from = datetime.strptime(date_from.strip(), "%Y-%m-%d").date()
                except ValueError:
                    date_from = date(date_to.year, 1, 1)
            elif not isinstance(date_from, date):
                date_from = date(date_to.year, 1, 1)

            days_in_period = max(1, (date_to - date_from).days + 1)

            # 2. ضبط تواريخ فترة المقارنة (تلقائياً إذا لم تحدد)
            if isinstance(comp_date_to, str) and comp_date_to.strip():
                try:
                    comp_date_to = datetime.strptime(comp_date_to.strip(), "%Y-%m-%d").date()
                except ValueError:
                    comp_date_to = None
            if isinstance(comp_date_from, str) and comp_date_from.strip():
                try:
                    comp_date_from = datetime.strptime(comp_date_from.strip(), "%Y-%m-%d").date()
                except ValueError:
                    comp_date_from = None

            if not comp_date_from or not comp_date_to:
                # التحديد الذكي التلقائي للفترة السابقة المماثلة
                comp_date_to = date_from - timedelta(days=1)
                comp_date_from = comp_date_to - timedelta(days=days_in_period - 1)

            comp_days = max(1, (comp_date_to - comp_date_from).days + 1)
            has_comparison = True

            # 3. تحديد العملة الوظيفية
            functional_currency = ExchangeRateService.get_functional_currency()
            currency_code = functional_currency.code if functional_currency else "EGP"
            currency_symbol = functional_currency.symbol or currency_code if functional_currency else "ج.م"
            status_list = ["posted", "draft"] if include_unposted else ["posted"]

            # 4. احتساب مؤشرات الفترة الحالية
            curr_metrics = cls._calculate_period_full_metrics(
                date_from=date_from,
                date_to=date_to,
                days_in_period=days_in_period,
                cost_center_id=cost_center_id,
                status_list=status_list,
                include_unposted=include_unposted
            )

            # 5. احتساب مؤشرات فترة المقارنة
            comp_metrics = cls._calculate_period_full_metrics(
                date_from=comp_date_from,
                date_to=comp_date_to,
                days_in_period=comp_days,
                cost_center_id=cost_center_id,
                status_list=status_list,
                include_unposted=include_unposted
            )

            # 6. دمج مقارنات التغير (Deltas & Variances)
            merged_analytics = cls._merge_comparative_deltas(curr_metrics, comp_metrics)

            # 7. الرسوم البيانية: اتجاهات 12 شهراً وتوزيع المصروفات
            monthly_trends = cls._get_twelve_months_trends(date_to=date_to, status_list=status_list)
            expense_distribution = cls._get_expense_distribution(
                date_from=date_from,
                date_to=date_to,
                cost_center_id=cost_center_id,
                status_list=status_list
            )

            # معلومات فلتر مركز التكلفة
            cost_center_obj = None
            if cost_center_id:
                try:
                    cost_center_obj = CostCenter.objects.filter(id=cost_center_id).first()
                except Exception:
                    cost_center_obj = None

            return {
                "date_from": date_from,
                "date_to": date_to,
                "days_in_period": days_in_period,
                "comp_date_from": comp_date_from,
                "comp_date_to": comp_date_to,
                "comp_days": comp_days,
                "has_comparison": has_comparison,
                "currency_code": currency_code,
                "currency_symbol": currency_symbol,
                "cost_center_id": cost_center_id,
                "cost_center_name": f"{cost_center_obj.code} - {cost_center_obj.name}" if cost_center_obj else "",
                "include_unposted": include_unposted,
                "basic_metrics": merged_analytics["basic_metrics"],
                "liquidity": merged_analytics["liquidity"],
                "profitability": merged_analytics["profitability"],
                "solvency": merged_analytics["solvency"],
                "activity": merged_analytics["activity"],
                "dupont": merged_analytics["dupont"],
                "altman_z": merged_analytics["altman_z"],
                "health_scorecard": merged_analytics["health_scorecard"],
                "monthly_trends": monthly_trends,
                "expense_distribution": expense_distribution,
            }

        except Exception as e:
            logger.error(f"Error generating complete Financial Analytics: {str(e)}", exc_info=True)
            raise

    @classmethod
    def _calculate_period_full_metrics(
        cls,
        date_from: date,
        date_to: date,
        days_in_period: int,
        cost_center_id: Optional[Union[int, str]],
        status_list: List[str],
        include_unposted: bool,
    ) -> Dict[str, Any]:
        """
        حساب كافة المؤشرات المالية التفصيلية لفترة زمنية محددة.
        """
        prev_day = date_from - timedelta(days=1)

        # 1. إجمالي الأصول (1)
        assets_qs = ChartOfAccounts.objects.filter(code__startswith="1", is_active=True, is_leaf=True)
        beg_assets_agg = JournalEntryLine.objects.filter(
            account__in=assets_qs, journal_entry__status__in=status_list, journal_entry__date__lte=prev_day
        ).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        beginning_assets = max(Decimal("0"), beg_assets_agg["d"] - beg_assets_agg["c"])

        end_assets_agg = JournalEntryLine.objects.filter(
            account__in=assets_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to
        ).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        ending_assets = max(Decimal("0"), end_assets_agg["d"] - end_assets_agg["c"])
        avg_assets = (beginning_assets + ending_assets) / 2 if (beginning_assets + ending_assets) > 0 else ending_assets

        # 2. إجمالي الخصوم (2)
        liab_qs = ChartOfAccounts.objects.filter(code__startswith="2", is_active=True, is_leaf=True)
        end_liab_agg = JournalEntryLine.objects.filter(
            account__in=liab_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to
        ).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        ending_liabilities = max(Decimal("0"), end_liab_agg["c"] - end_liab_agg["d"])

        # 3. إجمالي حقوق الملكية (3)
        equity_qs = ChartOfAccounts.objects.filter(code__startswith="3", is_active=True, is_leaf=True)
        beg_eq_agg = JournalEntryLine.objects.filter(
            account__in=equity_qs, journal_entry__status__in=status_list, journal_entry__date__lte=prev_day
        ).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        beginning_equity = beg_eq_agg["c"] - beg_eq_agg["d"]

        end_eq_agg = JournalEntryLine.objects.filter(
            account__in=equity_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to
        ).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        ending_equity = end_eq_agg["c"] - end_eq_agg["d"]
        avg_equity = (beginning_equity + ending_equity) / 2 if (beginning_equity + ending_equity) != 0 else ending_equity

        # 4. أرقام قائمة الدخل المعتمدة
        inc_data = IncomeStatementService.generate_income_statement(
            date_from=date_from,
            date_to=date_to,
            cost_center_id=cost_center_id,
            include_unposted=include_unposted,
            hide_zero_balances=False
        )
        net_sales = inc_data.get("operating_revenues", {}).get("total", Decimal("0"))
        net_cogs = inc_data.get("cogs", {}).get("total", Decimal("0"))
        gross_profit = inc_data.get("gross_profit", Decimal("0"))
        operating_profit = inc_data.get("operating_profit", Decimal("0"))
        net_income = inc_data.get("net_income", Decimal("0"))
        total_expenses = max(Decimal("0"), net_sales - net_income)

        # 5. حساب الأرباح المرحلة والاحتياطيات التراكمية لنموذج ألتمان
        retained_qs = ChartOfAccounts.objects.filter(
            Q(code__startswith="321") | Q(code__startswith="322"),
            is_active=True, is_leaf=True
        )
        retained_agg = JournalEntryLine.objects.filter(
            account__in=retained_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to
        ).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        retained_earnings_total = (retained_agg["c"] - retained_agg["d"]) + net_income

        # 6. حساب المحاور التفصيلية
        liquidity = cls._calculate_liquidity_ratios(date_to, status_list)
        profitability = cls._calculate_profitability_ratios(
            net_sales=net_sales,
            net_cogs=net_cogs,
            gross_profit=gross_profit,
            operating_profit=operating_profit,
            net_income=net_income,
            days_in_period=days_in_period,
            avg_assets=avg_assets,
            avg_equity=avg_equity,
            ending_equity=ending_equity,
        )
        solvency = cls._calculate_solvency_ratios(ending_assets, ending_liabilities, ending_equity)
        activity = cls._calculate_activity_ratios(
            date_from=date_from,
            date_to=date_to,
            days_in_period=days_in_period,
            net_sales=net_sales,
            net_cogs=net_cogs,
            avg_assets=avg_assets,
            status_list=status_list
        )
        dupont = cls._calculate_dupont_model(
            net_margin=profitability["net_margin"],
            asset_turnover=activity["asset_turnover"],
            leverage=solvency["equity_multiplier"],
            is_negative_equity=profitability["is_negative_equity"]
        )
        altman_z = cls._calculate_altman_z_score(
            working_capital=liquidity["net_working_capital"],
            retained_earnings=retained_earnings_total,
            ebit=operating_profit,
            total_equity=ending_equity,
            total_liabilities=ending_liabilities,
            net_sales=net_sales,
            total_assets=ending_assets,
        )
        health_scorecard = cls._calculate_health_score(liquidity, profitability, solvency, activity)

        return {
            "basic_metrics": {
                "monthly_income": net_sales,
                "monthly_expenses": total_expenses,
                "net_profit": net_income,
                "profit_margin": profitability["net_margin"],
                "current_ratio": liquidity["current_ratio"],
                "annualized_roe": profitability["annualized_roe"],
            },
            "liquidity": liquidity,
            "profitability": profitability,
            "solvency": solvency,
            "activity": activity,
            "dupont": dupont,
            "altman_z": altman_z,
            "health_scorecard": health_scorecard,
        }

    @classmethod
    def _calculate_liquidity_ratios(cls, date_to: date, status_list: List[str]) -> Dict[str, Any]:
        """
        1. حساب نسب السيولة ورأس المال العامل مع المعالجة الصفرية النظيفة
        """
        cur_assets_qs = ChartOfAccounts.objects.filter(code__startswith="11", is_active=True, is_leaf=True)
        cur_assets_agg = JournalEntryLine.objects.filter(
            account__in=cur_assets_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to
        ).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        current_assets = max(Decimal("0"), cur_assets_agg["d"] - cur_assets_agg["c"])

        cash_qs = ChartOfAccounts.objects.filter(code__startswith="111", is_active=True, is_leaf=True)
        cash_agg = JournalEntryLine.objects.filter(
            account__in=cash_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to
        ).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        cash_equivalents = max(Decimal("0"), cash_agg["d"] - cash_agg["c"])

        inv_qs = ChartOfAccounts.objects.filter(code__startswith="113", is_active=True, is_leaf=True)
        inv_agg = JournalEntryLine.objects.filter(
            account__in=inv_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to
        ).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        inventory = max(Decimal("0"), inv_agg["d"] - inv_agg["c"])

        cur_liab_qs = ChartOfAccounts.objects.filter(code__startswith="21", is_active=True, is_leaf=True)
        cur_liab_agg = JournalEntryLine.objects.filter(
            account__in=cur_liab_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to
        ).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        current_liabilities = max(Decimal("0"), cur_liab_agg["c"] - cur_liab_agg["d"])

        net_working_capital = current_assets - current_liabilities

        if current_liabilities > 0:
            current_ratio = (current_assets / current_liabilities).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            quick_ratio = (max(Decimal("0"), current_assets - inventory) / current_liabilities).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            cash_ratio = (cash_equivalents / current_liabilities).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            has_no_liabilities = False
        else:
            current_ratio = Decimal("0.00")
            quick_ratio = Decimal("0.00")
            cash_ratio = Decimal("0.00")
            has_no_liabilities = True

        if has_no_liabilities:
            current_ratio_status = "success"
            quick_ratio_status = "success"
            cash_ratio_status = "success"
        else:
            current_ratio_status = "success" if current_ratio >= Decimal("1.50") else ("warning" if current_ratio >= Decimal("1.00") else "danger")
            quick_ratio_status = "success" if quick_ratio >= Decimal("1.00") else ("warning" if quick_ratio >= Decimal("0.70") else "danger")
            cash_ratio_status = "success" if cash_ratio >= Decimal("0.20") else ("warning" if cash_ratio >= Decimal("0.10") else "danger")

        return {
            "current_assets": current_assets,
            "current_liabilities": current_liabilities,
            "cash_equivalents": cash_equivalents,
            "inventory": inventory,
            "net_working_capital": net_working_capital,
            "current_ratio": current_ratio,
            "quick_ratio": quick_ratio,
            "cash_ratio": cash_ratio,
            "has_no_liabilities": has_no_liabilities,
            "current_ratio_status": current_ratio_status,
            "quick_ratio_status": quick_ratio_status,
            "cash_ratio_status": cash_ratio_status,
        }

    @classmethod
    def _calculate_profitability_ratios(
        cls,
        net_sales: Decimal,
        net_cogs: Decimal,
        gross_profit: Decimal,
        operating_profit: Decimal,
        net_income: Decimal,
        days_in_period: int,
        avg_assets: Decimal,
        avg_equity: Decimal,
        ending_equity: Decimal,
    ) -> Dict[str, Any]:
        """
        2. حساب نسب الربحية ومعدلات العائد المحولة سنوياً
        """
        if net_sales > 0:
            gross_margin = ((gross_profit / net_sales) * 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            operating_margin = ((operating_profit / net_sales) * 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            net_margin = ((net_income / net_sales) * 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        else:
            gross_margin = Decimal("0.00")
            operating_margin = Decimal("0.00")
            net_margin = Decimal("0.00")

        if avg_assets > 0:
            periodic_roa = ((net_income / avg_assets) * 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            annualized_roa = (periodic_roa * (Decimal("365") / Decimal(days_in_period))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        else:
            periodic_roa = Decimal("0.00")
            annualized_roa = Decimal("0.00")

        is_negative_equity = bool(avg_equity <= 0 or ending_equity <= 0)
        if not is_negative_equity and avg_equity > 0:
            periodic_roe = ((net_income / avg_equity) * 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            annualized_roe = (periodic_roe * (Decimal("365") / Decimal(days_in_period))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            roe_status = "success" if annualized_roe >= Decimal("15.00") else ("warning" if annualized_roe >= Decimal("5.00") else "danger")
        else:
            periodic_roe = Decimal("0.00")
            annualized_roe = Decimal("0.00")
            roe_status = "danger"

        return {
            "gross_profit": gross_profit,
            "operating_profit": operating_profit,
            "net_income": net_income,
            "gross_margin": gross_margin,
            "operating_margin": operating_margin,
            "net_margin": net_margin,
            "periodic_roa": periodic_roa,
            "annualized_roa": annualized_roa,
            "periodic_roe": periodic_roe,
            "annualized_roe": annualized_roe,
            "is_negative_equity": is_negative_equity,
            "roe_status": roe_status,
        }

    @classmethod
    def _calculate_solvency_ratios(cls, ending_assets: Decimal, ending_liabilities: Decimal, ending_equity: Decimal) -> Dict[str, Any]:
        """
        3. حساب نسب الملاءة والرفع المالي والهيكل التمويلي
        """
        is_negative_equity = bool(ending_equity <= 0)
        has_no_debt = bool(ending_liabilities <= 0)

        if not is_negative_equity and ending_equity > 0:
            if not has_no_debt:
                debt_to_equity = (ending_liabilities / ending_equity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                de_status = "success" if debt_to_equity <= Decimal("1.50") else ("warning" if debt_to_equity <= Decimal("2.50") else "danger")
            else:
                debt_to_equity = Decimal("0.00")
                de_status = "success"
            equity_multiplier = (ending_assets / ending_equity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        else:
            debt_to_equity = Decimal("0.00")
            de_status = "danger"
            equity_multiplier = Decimal("1.00")

        if ending_assets > 0:
            debt_ratio = ((ending_liabilities / ending_assets) * 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            dr_status = "success" if debt_ratio <= Decimal("60.00") else ("warning" if debt_ratio <= Decimal("80.00") else "danger")
        else:
            debt_ratio = Decimal("0.00")
            dr_status = "warning"

        return {
            "total_assets": ending_assets,
            "total_liabilities": ending_liabilities,
            "total_equity": ending_equity,
            "debt_to_equity": debt_to_equity,
            "debt_ratio": debt_ratio,
            "equity_multiplier": equity_multiplier,
            "is_negative_equity": is_negative_equity,
            "has_no_debt": has_no_debt,
            "de_status": de_status,
            "dr_status": dr_status,
        }

    @classmethod
    def _calculate_activity_ratios(
        cls,
        date_from: date,
        date_to: date,
        days_in_period: int,
        net_sales: Decimal,
        net_cogs: Decimal,
        avg_assets: Decimal,
        status_list: List[str]
    ) -> Dict[str, Any]:
        """
        4. حساب نسب النشاط والكفاءة ودورة التحول النقدي (Cash Conversion Cycle)
        """
        prev_day = date_from - timedelta(days=1)

        # متوسط العملاء (112)
        ar_qs = ChartOfAccounts.objects.filter(code__startswith="112", is_active=True, is_leaf=True)
        ar_beg = JournalEntryLine.objects.filter(account__in=ar_qs, journal_entry__status__in=status_list, journal_entry__date__lte=prev_day).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        ar_end = JournalEntryLine.objects.filter(account__in=ar_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        b_ar = max(Decimal("0"), ar_beg["d"] - ar_beg["c"])
        e_ar = max(Decimal("0"), ar_end["d"] - ar_end["c"])
        avg_ar = (b_ar + e_ar) / 2 if (b_ar + e_ar) > 0 else e_ar

        # متوسط المخزون (113)
        inv_qs = ChartOfAccounts.objects.filter(code__startswith="113", is_active=True, is_leaf=True)
        inv_beg = JournalEntryLine.objects.filter(account__in=inv_qs, journal_entry__status__in=status_list, journal_entry__date__lte=prev_day).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        inv_end = JournalEntryLine.objects.filter(account__in=inv_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        b_inv = max(Decimal("0"), inv_beg["d"] - inv_beg["c"])
        e_inv = max(Decimal("0"), inv_end["d"] - inv_end["c"])
        avg_inv = (b_inv + e_inv) / 2 if (b_inv + e_inv) > 0 else e_inv

        # متوسط الموردين (211)
        ap_qs = ChartOfAccounts.objects.filter(code__startswith="211", is_active=True, is_leaf=True)
        ap_beg = JournalEntryLine.objects.filter(account__in=ap_qs, journal_entry__status__in=status_list, journal_entry__date__lte=prev_day).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        ap_end = JournalEntryLine.objects.filter(account__in=ap_qs, journal_entry__status__in=status_list, journal_entry__date__lte=date_to).exclude(journal_entry__entry_type="closing").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        b_ap = max(Decimal("0"), ap_beg["c"] - ap_beg["d"])
        e_ap = max(Decimal("0"), ap_end["c"] - ap_end["d"])
        avg_ap = (b_ap + e_ap) / 2 if (b_ap + e_ap) > 0 else e_ap

        # (أ) دورة العملاء (DSO)
        if net_sales > 0 and avg_ar > 0:
            ar_turnover = (net_sales / avg_ar).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            dso_days = ((avg_ar / net_sales) * Decimal(days_in_period)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
        else:
            ar_turnover = Decimal("0.00")
            dso_days = Decimal("0.0")

        # (ب) دورة المخزون (DSI)
        if net_cogs > 0 and avg_inv > 0:
            inv_turnover = (net_cogs / avg_inv).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            dsi_days = ((avg_inv / net_cogs) * Decimal(days_in_period)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
        else:
            inv_turnover = Decimal("0.00")
            dsi_days = Decimal("0.0")

        # (ج) دورة الموردين (DPO)
        cogs_base = net_cogs if net_cogs > 0 else net_sales
        if cogs_base > 0 and avg_ap > 0:
            ap_turnover = (cogs_base / avg_ap).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            dpo_days = ((avg_ap / cogs_base) * Decimal(days_in_period)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
        else:
            ap_turnover = Decimal("0.00")
            dpo_days = Decimal("0.0")

        # (د) دورة التحول النقدي الكاملة (Cash Conversion Cycle = DSO + DSI - DPO)
        ccc_days = (dso_days + dsi_days - dpo_days).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)

        # (هـ) دوران إجمالي الأصول (Asset Turnover)
        if avg_assets > 0 and net_sales > 0:
            asset_turnover = (net_sales / avg_assets).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        else:
            asset_turnover = Decimal("0.00")

        return {
            "avg_receivables": avg_ar,
            "avg_inventory": avg_inv,
            "avg_payables": avg_ap,
            "ar_turnover": ar_turnover,
            "dso_days": dso_days,
            "inv_turnover": inv_turnover,
            "dsi_days": dsi_days,
            "ap_turnover": ap_turnover,
            "dpo_days": dpo_days,
            "ccc_days": ccc_days,
            "asset_turnover": asset_turnover,
        }

    @classmethod
    def _calculate_dupont_model(cls, net_margin: Decimal, asset_turnover: Decimal, leverage: Decimal, is_negative_equity: bool) -> Dict[str, Any]:
        """
        5. نموذج دوبونت الاستراتيجي الثلاثي (DuPont 3-Step Analysis)
        ROE = (Net Margin) * (Asset Turnover) * (Equity Multiplier)
        """
        if not is_negative_equity:
            dupont_roe = (net_margin * asset_turnover * leverage).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        else:
            dupont_roe = Decimal("0.00")

        return {
            "net_margin": net_margin,
            "asset_turnover": asset_turnover,
            "leverage": leverage,
            "dupont_roe": dupont_roe,
            "is_negative_equity": is_negative_equity,
        }

    @classmethod
    def _calculate_altman_z_score(
        cls,
        working_capital: Decimal,
        retained_earnings: Decimal,
        ebit: Decimal,
        total_equity: Decimal,
        total_liabilities: Decimal,
        net_sales: Decimal,
        total_assets: Decimal,
    ) -> Dict[str, Any]:
        """
        حساب مؤشر ألتمان للسلامة والتعثر المالي (Altman Z-Score for Emerging Markets & Private Firms)
        Z' = 0.717 X1 + 0.847 X2 + 3.107 X3 + 0.420 X4 + 0.998 X5
        """
        if total_assets <= 0:
            return {
                "z_score": Decimal("0.00"),
                "zone": "distress",
                "zone_name": _("منطقة الخطر والتعثر المالي (Distress Zone)"),
                "zone_class": "danger",
                "x1": Decimal("0.00"),
                "x2": Decimal("0.00"),
                "x3": Decimal("0.00"),
                "x4": Decimal("0.00"),
                "x5": Decimal("0.00"),
            }

        x1 = working_capital / total_assets
        x2 = retained_earnings / total_assets
        x3 = ebit / total_assets

        # X4 = Equity / Liabilities (مع حماية انعدام الديون)
        if total_liabilities > 0:
            x4 = total_equity / total_liabilities
        else:
            x4 = Decimal("10.0") if total_equity > 0 else Decimal("0.0")

        x5 = net_sales / total_assets

        z_score = (
            (Decimal("0.717") * x1)
            + (Decimal("0.847") * x2)
            + (Decimal("3.107") * x3)
            + (Decimal("0.420") * x4)
            + (Decimal("0.998") * x5)
        ).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        # تصنيف منطقة الأمان المالي
        if z_score >= Decimal("2.90"):
            zone = "safe"
            zone_name = _("منطقة الأمان المالي والاستقرار (Safe Zone)")
            zone_class = "success"
            zone_desc = _("المنشأة تتمتع بصلابة مالية عالية وملاءة ممتازة، ومخاطر التعثر المالي شبه منعدمة.")
        elif z_score >= Decimal("1.23"):
            zone = "grey"
            zone_name = _("المنطقة الرمادية / حذر متوسط (Grey / Caution Zone)")
            zone_class = "warning"
            zone_desc = _("المركز المالي مستقر مع وجود مؤشرات تتطلب المتابعة والتحكم في الالتزامات والسيولة.")
        else:
            zone = "distress"
            zone_name = _("منطقة الخطر والتعثر المالي (Distress Zone)")
            zone_class = "danger"
            zone_desc = _("المنشأة تواجه ضغوط سيولة أو مديونية مرتفعة تتطلب إجراءات فورية لإعادة الهيكلة.")

        return {
            "z_score": z_score,
            "zone": zone,
            "zone_name": zone_name,
            "zone_class": zone_class,
            "zone_desc": zone_desc,
            "x1": x1.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
            "x2": x2.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
            "x3": x3.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
            "x4": x4.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
            "x5": x5.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP),
        }

    @classmethod
    def _calculate_health_score(cls, liquidity: Dict, profitability: Dict, solvency: Dict, activity: Dict) -> Dict[str, Any]:
        """
        حساب بطاقة ملخص الصحة المالية العامة (Executive Health Scorecard من 100 نقطة)
        """
        score = 0

        # 1. السيولة (25 نقطة)
        cr = liquidity["current_ratio"]
        if liquidity.get("has_no_liabilities") or cr >= Decimal("1.50"): score += 25
        elif cr >= Decimal("1.10"): score += 18
        elif cr >= Decimal("0.90"): score += 10

        # 2. الربحية (25 نقطة)
        nm = profitability["net_margin"]
        if nm >= Decimal("20.0"): score += 25
        elif nm >= Decimal("10.0"): score += 20
        elif nm > Decimal("0.0"): score += 12

        # 3. الملاءة والرفع المالي (20 نقطة)
        if not solvency["is_negative_equity"]:
            dr = solvency["debt_ratio"]
            if dr <= Decimal("50.0"): score += 20
            elif dr <= Decimal("70.0"): score += 14
            elif dr <= Decimal("85.0"): score += 8

        # 4. الكفاءة والنشاط (15 نقطة)
        dso = activity["dso_days"]
        if dso > 0:
            if dso <= Decimal("45.0"): score += 15
            elif dso <= Decimal("75.0"): score += 10
            elif dso <= Decimal("120.0"): score += 5
        else:
            score += 10

        # 5. دوران الأصول (15 نقطة)
        at = activity["asset_turnover"]
        if at >= Decimal("1.00"): score += 15
        elif at >= Decimal("0.50"): score += 10
        elif at > Decimal("0.00"): score += 5

        if score >= 80:
            grade = _("ممتاز وقوي جداً (Strong & Resilient)")
            grade_class = "success"
            icon = "fa-shield-halved"
        elif score >= 65:
            grade = _("جيد ومستقر (Good & Stable)")
            grade_class = "info"
            icon = "fa-check-circle"
        elif score >= 50:
            grade = _("متوسط ويحتاج متابعة (Fair / Needs Monitoring)")
            grade_class = "warning"
            icon = "fa-triangle-exclamation"
        else:
            grade = _("حرج ومخاطر مرتفعة (High Financial Risk)")
            grade_class = "danger"
            icon = "fa-circle-radiation"

        return {
            "score": score,
            "grade": grade,
            "grade_class": grade_class,
            "icon": icon,
        }

    @classmethod
    def _merge_comparative_deltas(cls, curr: Dict[str, Any], comp: Dict[str, Any]) -> Dict[str, Any]:
        """
        دمج ومقارنة مؤشرات الفترتين واستخراج الفروقات ونسب التغير
        """
        def calc_delta(cur_val, cmp_val):
            diff = cur_val - cmp_val
            pct = ((diff / abs(cmp_val)) * 100).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP) if cmp_val != 0 else Decimal("0.0")
            return {
                "curr": cur_val,
                "comp": cmp_val,
                "diff": diff,
                "pct": pct,
                "is_positive": diff >= 0,
            }

        # دمج basic_metrics
        curr["basic_metrics"]["comp_monthly_income"] = comp["basic_metrics"]["monthly_income"]
        curr["basic_metrics"]["delta_income"] = calc_delta(curr["basic_metrics"]["monthly_income"], comp["basic_metrics"]["monthly_income"])

        curr["basic_metrics"]["comp_net_profit"] = comp["basic_metrics"]["net_profit"]
        curr["basic_metrics"]["delta_profit"] = calc_delta(curr["basic_metrics"]["net_profit"], comp["basic_metrics"]["net_profit"])

        curr["basic_metrics"]["comp_profit_margin"] = comp["basic_metrics"]["profit_margin"]
        curr["basic_metrics"]["delta_margin"] = calc_delta(curr["basic_metrics"]["profit_margin"], comp["basic_metrics"]["profit_margin"])

        curr["basic_metrics"]["comp_current_ratio"] = comp["basic_metrics"]["current_ratio"]
        curr["basic_metrics"]["delta_current_ratio"] = calc_delta(curr["basic_metrics"]["current_ratio"], comp["basic_metrics"]["current_ratio"])

        # دمج السيولة
        curr["liquidity"]["comp_current_ratio"] = comp["liquidity"]["current_ratio"]
        curr["liquidity"]["delta_current_ratio"] = calc_delta(curr["liquidity"]["current_ratio"], comp["liquidity"]["current_ratio"])

        curr["liquidity"]["comp_quick_ratio"] = comp["liquidity"]["quick_ratio"]
        curr["liquidity"]["delta_quick_ratio"] = calc_delta(curr["liquidity"]["quick_ratio"], comp["liquidity"]["quick_ratio"])

        curr["liquidity"]["comp_cash_ratio"] = comp["liquidity"]["cash_ratio"]
        curr["liquidity"]["delta_cash_ratio"] = calc_delta(curr["liquidity"]["cash_ratio"], comp["liquidity"]["cash_ratio"])

        curr["liquidity"]["comp_net_working_capital"] = comp["liquidity"]["net_working_capital"]
        curr["liquidity"]["delta_net_working_capital"] = calc_delta(curr["liquidity"]["net_working_capital"], comp["liquidity"]["net_working_capital"])

        # دمج الربحية
        curr["profitability"]["comp_gross_margin"] = comp["profitability"]["gross_margin"]
        curr["profitability"]["delta_gross_margin"] = calc_delta(curr["profitability"]["gross_margin"], comp["profitability"]["gross_margin"])

        curr["profitability"]["comp_operating_margin"] = comp["profitability"]["operating_margin"]
        curr["profitability"]["delta_operating_margin"] = calc_delta(curr["profitability"]["operating_margin"], comp["profitability"]["operating_margin"])

        curr["profitability"]["comp_net_margin"] = comp["profitability"]["net_margin"]
        curr["profitability"]["delta_net_margin"] = calc_delta(curr["profitability"]["net_margin"], comp["profitability"]["net_margin"])

        curr["profitability"]["comp_annualized_roa"] = comp["profitability"]["annualized_roa"]
        curr["profitability"]["delta_annualized_roa"] = calc_delta(curr["profitability"]["annualized_roa"], comp["profitability"]["annualized_roa"])

        curr["profitability"]["comp_annualized_roe"] = comp["profitability"]["annualized_roe"]
        curr["profitability"]["delta_annualized_roe"] = calc_delta(curr["profitability"]["annualized_roe"], comp["profitability"]["annualized_roe"])

        # دمج الملاءة
        curr["solvency"]["comp_debt_to_equity"] = comp["solvency"]["debt_to_equity"]
        curr["solvency"]["delta_debt_to_equity"] = calc_delta(curr["solvency"]["debt_to_equity"], comp["solvency"]["debt_to_equity"])

        curr["solvency"]["comp_debt_ratio"] = comp["solvency"]["debt_ratio"]
        curr["solvency"]["delta_debt_ratio"] = calc_delta(curr["solvency"]["debt_ratio"], comp["solvency"]["debt_ratio"])

        # دمج النشاط
        curr["activity"]["comp_dso_days"] = comp["activity"]["dso_days"]
        curr["activity"]["delta_dso_days"] = calc_delta(curr["activity"]["dso_days"], comp["activity"]["dso_days"])

        curr["activity"]["comp_dsi_days"] = comp["activity"]["dsi_days"]
        curr["activity"]["delta_dsi_days"] = calc_delta(curr["activity"]["dsi_days"], comp["activity"]["dsi_days"])

        curr["activity"]["comp_dpo_days"] = comp["activity"]["dpo_days"]
        curr["activity"]["delta_dpo_days"] = calc_delta(curr["activity"]["dpo_days"], comp["activity"]["dpo_days"])

        curr["activity"]["comp_ccc_days"] = comp["activity"]["ccc_days"]
        curr["activity"]["delta_ccc_days"] = calc_delta(curr["activity"]["ccc_days"], comp["activity"]["ccc_days"])

        # دمج ألتمان
        curr["altman_z"]["comp_z_score"] = comp["altman_z"]["z_score"]
        curr["altman_z"]["delta_z_score"] = calc_delta(curr["altman_z"]["z_score"], comp["altman_z"]["z_score"])

        return curr

    @classmethod
    def _get_twelve_months_trends(cls, date_to: date, status_list: List[str]) -> Dict[str, Any]:
        """
        جلب اتجاهات الإيرادات والمصروفات والأرباح لـ 12 شهراً باستعلام SQL مجمع واحد عبر TruncMonth
        """
        twelve_months_ago = (date_to.replace(day=1) - timedelta(days=365)).replace(day=1)

        monthly_raw = JournalEntryLine.objects.filter(
            journal_entry__status__in=status_list,
            journal_entry__date__gte=twelve_months_ago,
            journal_entry__date__lte=date_to,
            account__account_type__category__in=["revenue", "expense"]
        ).exclude(
            journal_entry__entry_type="closing"
        ).annotate(
            month_date=TruncMonth("journal_entry__date")
        ).values(
            "month_date", "account__account_type__category"
        ).annotate(
            total_debit=Coalesce(Sum("debit"), Decimal("0")),
            total_credit=Coalesce(Sum("credit"), Decimal("0")),
        ).order_by("month_date")

        month_map = {}
        curr_dt = twelve_months_ago
        while curr_dt <= date_to:
            m_key = curr_dt.strftime("%Y-%m")
            month_map[m_key] = {
                "label": curr_dt.strftime("%b %Y"),
                "revenue": Decimal("0"),
                "expense": Decimal("0"),
                "profit": Decimal("0"),
            }
            next_m = curr_dt.month % 12 + 1
            next_y = curr_dt.year + (1 if curr_dt.month == 12 else 0)
            curr_dt = date(next_y, next_m, 1)

        for row in monthly_raw:
            m_dt = row["month_date"]
            if not m_dt:
                continue
            m_key = m_dt.strftime("%Y-%m")
            if m_key in month_map:
                cat = row["account__account_type__category"]
                d = row["total_debit"]
                c = row["total_credit"]
                if cat == "revenue":
                    month_map[m_key]["revenue"] += (c - d)
                elif cat == "expense":
                    month_map[m_key]["expense"] += (d - c)

        labels = []
        revenues = []
        expenses = []
        profits = []

        for k in sorted(month_map.keys()):
            v = month_map[k]
            p = v["revenue"] - v["expense"]
            labels.append(v["label"])
            revenues.append(float(v["revenue"]))
            expenses.append(float(v["expense"]))
            profits.append(float(p))

        return {
            "labels": labels,
            "revenues": revenues,
            "expenses": expenses,
            "profits": profits,
        }

    @classmethod
    def _get_expense_distribution(
        cls,
        date_from: date,
        date_to: date,
        cost_center_id: Optional[Union[int, str]],
        status_list: List[str]
    ) -> Dict[str, Any]:
        """
        توزيع هيكل المصروفات والتكاليف حسب المجموعات المحاسبية الرئيسية
        """
        lines_qs = JournalEntryLine.objects.filter(
            journal_entry__status__in=status_list,
            journal_entry__date__gte=date_from,
            journal_entry__date__lte=date_to,
            account__account_type__category="expense"
        ).exclude(journal_entry__entry_type="closing")

        if cost_center_id:
            lines_qs = lines_qs.filter(cost_center_id=cost_center_id)

        cogs_agg = lines_qs.filter(account__code__startswith="51").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        admin_agg = lines_qs.filter(account__code__startswith="52", account__code__lt="52800").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        depr_agg = lines_qs.filter(account__code__startswith="528").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))
        fin_agg = lines_qs.filter(account__code__startswith="54").aggregate(d=Coalesce(Sum("debit"), Decimal("0")), c=Coalesce(Sum("credit"), Decimal("0")))

        cogs_val = max(Decimal("0"), cogs_agg["d"] - cogs_agg["c"])
        admin_val = max(Decimal("0"), admin_agg["d"] - admin_agg["c"])
        depr_val = max(Decimal("0"), depr_agg["d"] - depr_agg["c"])
        fin_val = max(Decimal("0"), fin_agg["d"] - fin_agg["c"])

        labels = [_("تكلفة المبيعات والنشاط"), _("مصروفات إدارية وعمومية"), _("إهلاك واستهلاك أصول"), _("مصروفات تمويلية وأخرى")]
        data = [float(cogs_val), float(admin_val), float(depr_val), float(fin_val)]

        return {
            "labels": labels,
            "data": data,
            "total_expenses": float(cogs_val + admin_val + depr_val + fin_val),
        }

    @classmethod
    def export_to_excel(
        cls,
        date_from: Optional[Union[date, str]] = None,
        date_to: Optional[Union[date, str]] = None,
        comp_date_from: Optional[Union[date, str]] = None,
        comp_date_to: Optional[Union[date, str]] = None,
        cost_center_id: Optional[Union[int, str]] = None,
        include_unposted: bool = False,
    ) -> bytes:
        """
        تصدير لوحة التحليلات والمؤشرات المالية إلى ملف Excel معتمد متعدد الأعمدة.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            analytics = cls.get_complete_analytics(
                date_from=date_from,
                date_to=date_to,
                comp_date_from=comp_date_from,
                comp_date_to=comp_date_to,
                cost_center_id=cost_center_id,
                include_unposted=include_unposted,
            )

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "لوحة التحليلات والمؤشرات المالية"
            ws.views.sheetView[0].rightToLeft = True

            font_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
            font_subtitle = Font(name="Segoe UI", size=10, bold=False, color="FFFFFF")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_section = Font(name="Segoe UI", size=11, bold=True, color="1E3A8A")
            font_row = Font(name="Segoe UI", size=10, bold=False, color="1F2937")

            fill_navy = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            fill_group = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

            border_thin = Border(
                left=Side(style="thin", color="CBD5E1"),
                right=Side(style="thin", color="CBD5E1"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            )

            ws.merge_cells("A1:E1")
            ws["A1"] = "مركز التحليلات والمؤشرات المالية التنفيذي - MWHEBA ERP"
            ws["A1"].font = font_title
            ws["A1"].fill = fill_navy
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells("A2:E2")
            ws["A2"] = f"تقرير التحليل المالي للفترة من {analytics['date_from']} إلى {analytics['date_to']} (مقارنة بـ {analytics['comp_date_from']} إلى {analytics['comp_date_to']})"
            ws["A2"].font = font_subtitle
            ws["A2"].fill = fill_navy
            ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

            headers = ["المؤشر المالي", "الفترة الحالية", "فترة المقارنة", "نسبة التغير", "التقييم والمعيار"]
            for col_idx, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=col_idx, value=h)
                c.font = font_header
                c.fill = fill_navy
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border_thin

            row = 5
            def write_item(name, curr_v, comp_v, delta_pct, rating):
                nonlocal row
                c1 = ws.cell(row=row, column=1, value=name)
                c2 = ws.cell(row=row, column=2, value=curr_v)
                c3 = ws.cell(row=row, column=3, value=comp_v)
                c4 = ws.cell(row=row, column=4, value=f"{delta_pct}%")
                c5 = ws.cell(row=row, column=5, value=rating)
                for c in [c1, c2, c3, c4, c5]:
                    c.font = font_row
                    c.border = border_thin
                c1.alignment = Alignment(horizontal="right", vertical="center")
                c2.alignment = Alignment(horizontal="center", vertical="center")
                c3.alignment = Alignment(horizontal="center", vertical="center")
                c4.alignment = Alignment(horizontal="center", vertical="center")
                c5.alignment = Alignment(horizontal="center", vertical="center")
                row += 1

            def write_section(title):
                nonlocal row
                ws.merge_cells(f"A{row}:E{row}")
                c = ws.cell(row=row, column=1, value=title)
                c.font = font_section
                c.fill = fill_group
                c.alignment = Alignment(horizontal="right", vertical="center")
                row += 1

            # 1. السيولة
            write_section("1. مؤشرات السيولة ورأس المال العامل")
            write_item("نسبة التداول الحالية (Current Ratio)", f"{analytics['liquidity']['current_ratio']}", f"{analytics['liquidity']['comp_current_ratio']}", f"{analytics['liquidity']['delta_current_ratio']['pct']}", "آمن >= 1.5")
            write_item("نسبة السيولة السريعة (Quick Ratio)", f"{analytics['liquidity']['quick_ratio']}", f"{analytics['liquidity']['comp_quick_ratio']}", f"{analytics['liquidity']['delta_quick_ratio']['pct']}", "آمن >= 1.0")
            write_item("نسبة النقدية الفورية (Cash Ratio)", f"{analytics['liquidity']['cash_ratio']}", f"{analytics['liquidity']['comp_cash_ratio']}", f"{analytics['liquidity']['delta_cash_ratio']['pct']}", "آمن >= 0.2")
            write_item("صافي رأس المال العامل", f"{float(analytics['liquidity']['net_working_capital']):,.2f}", f"{float(analytics['liquidity']['comp_net_working_capital']):,.2f}", f"{analytics['liquidity']['delta_net_working_capital']['pct']}", "")

            # 2. الربحية
            write_section("2. مؤشرات الربحية وهوامش العائد")
            write_item("هامش مجمل الربح (Gross Margin)", f"{analytics['profitability']['gross_margin']}%", f"{analytics['profitability']['comp_gross_margin']}%", f"{analytics['profitability']['delta_gross_margin']['diff']}", "")
            write_item("هامش التشغيل (Operating Margin)", f"{analytics['profitability']['operating_margin']}%", f"{analytics['profitability']['comp_operating_margin']}%", f"{analytics['profitability']['delta_operating_margin']['diff']}", "")
            write_item("هامش صافي الربح (Net Profit Margin)", f"{analytics['profitability']['net_margin']}%", f"{analytics['profitability']['comp_net_margin']}%", f"{analytics['profitability']['delta_net_margin']['diff']}", "")
            write_item("العائد على الأصول السنوي (ROA)", f"{analytics['profitability']['annualized_roa']}%", f"{analytics['profitability']['comp_annualized_roa']}%", f"{analytics['profitability']['delta_annualized_roa']['diff']}", "")
            write_item("العائد على الملكية السنوي (ROE)", f"{analytics['profitability']['annualized_roe']}%", f"{analytics['profitability']['comp_annualized_roe']}%", f"{analytics['profitability']['delta_annualized_roe']['diff']}", "")

            # 3. الملاءة والرفع المالي
            write_section("3. مؤشرات الملاءة والرفع المالي")
            write_item("نسبة المديونية للملكية (Debt-to-Equity)", f"{analytics['solvency']['debt_to_equity']}", f"{analytics['solvency']['comp_debt_to_equity']}", f"{analytics['solvency']['delta_debt_to_equity']['pct']}", "آمن <= 1.5")
            write_item("نسبة تغطية الأصول (Debt Ratio)", f"{analytics['solvency']['debt_ratio']}%", f"{analytics['solvency']['comp_debt_ratio']}%", f"{analytics['solvency']['delta_debt_ratio']['diff']}", "آمن <= 60%")
            write_item("مضاعف الرفع المالي (Equity Multiplier)", f"{analytics['solvency']['equity_multiplier']}", "-", "-", "")

            # 4. النشاط والكفاءة
            write_section("4. مؤشرات الكفاءة ودورة التحول النقدي")
            write_item("متوسط فترة تحصيل العملاء (DSO)", f"{analytics['activity']['dso_days']} يوم", f"{analytics['activity']['comp_dso_days']} يوم", f"{analytics['activity']['delta_dso_days']['pct']}", "جيد <= 45 يوم")
            write_item("متوسط فترة دوران المخزون (DSI)", f"{analytics['activity']['dsi_days']} يوم", f"{analytics['activity']['comp_dsi_days']} يوم", f"{analytics['activity']['delta_dsi_days']['pct']}", "")
            write_item("متوسط فترة سداد الموردين (DPO)", f"{analytics['activity']['dpo_days']} يوم", f"{analytics['activity']['comp_dpo_days']} يوم", f"{analytics['activity']['delta_dpo_days']['pct']}", "")
            write_item("دورة التحول النقدي (Cash Conversion Cycle)", f"{analytics['activity']['ccc_days']} يوم", f"{analytics['activity']['comp_ccc_days']} يوم", f"{analytics['activity']['delta_ccc_days']['diff']}", "DSO + DSI - DPO")

            # 5. مؤشر ألتمان للسلامة والتعثر
            write_section("5. مؤشر ألتمان للسلامة والتعثر المالي (Altman Z-Score)")
            write_item("نتيجة مؤشر ألتمان (Z-Score)", f"{analytics['altman_z']['z_score']}", f"{analytics['altman_z']['comp_z_score']}", f"{analytics['altman_z']['delta_z_score']['pct']}", f"{analytics['altman_z']['zone_name']}")

            ws.column_dimensions["A"].width = 45
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 35

            output = BytesIO()
            wb.save(output)
            return output.getvalue()

        except Exception as e:
            logger.error(f"Error exporting Financial Analytics to Excel: {str(e)}", exc_info=True)
            raise
