# financial/services/trial_balance_service.py
"""
خدمة ميزان المراجعة المعياري - Enterprise Trial Balance Service (v10.0)
تطبيق شامل لمعايير المحاسبة الدولية (IAS/IFRS) مع دعم ميزان الـ 6 أعمدة والـ 2 عمود،
التجميع الشجري O(N)، عزل السنوات المالية، التصفية الصاعدة للحسابات الصفرية، وتصدير Excel الرسمي.
"""

import logging
from decimal import Decimal
from datetime import date, datetime
from typing import Dict, List, Optional, Any, Union
from django.db.models import Sum, Q
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.translation import gettext as _

from financial.models.chart_of_accounts import ChartOfAccounts
from financial.models.journal_entry import JournalEntryLine
from financial.models.fiscal_year import FiscalYear
from financial.services.exchange_rate_service import ExchangeRateService

logger = logging.getLogger(__name__)


class TrialBalanceService:
    """
    خدمة ميزان المراجعة المؤسسي الشامل
    """

    @classmethod
    def generate_trial_balance(
        cls,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        display_mode: str = '6_columns',
        account_level: Optional[Union[int, str]] = None,
        hide_zero_balances: bool = False,
        group_by_type: bool = True,
        fiscal_year_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        إنشاء ميزان المراجعة الكامل والدقيق

        Args:
            date_from: من تاريخ (الافتراضي: بداية السنة المالية الحالية)
            date_to: إلى تاريخ (الافتراضي: تاريخ اليوم)
            display_mode: نمط العرض ('6_columns' بالمجاميع والأرصدة أو '2_columns' بالأرصدة الختامية فقط)
            account_level: تصفية مستوى الشجرة (None / 1 / 2 / 3 / 4 / 'leaf')
            hide_zero_balances: استبعاد الحسابات الصفرية غير النشطة؟
            group_by_type: تجميع الحسابات حسب التصنيف الرئيسي (أصول، خصوم، حقوق ملكية، إيرادات، مصروفات)؟
            fiscal_year_id: معرف سنة مالية محددة (اختياري)

        Returns:
            Dict يحتوي على بيانات الميزان الكاملة مع كروت الإحصائيات وحالة التوازن
        """
        try:
            # 1. تحديد العملة الوظيفية للمؤسسة ديناميكياً
            functional_currency = ExchangeRateService.get_functional_currency()
            currency_code = functional_currency.code if functional_currency else "EGP"
            currency_symbol = functional_currency.symbol or currency_code if functional_currency else "ج.م"

            # 2. تحديد النطاق الزمني الذكي والسنة المالية
            today = timezone.now().date()
            if date_to is None:
                date_to = today

            current_fiscal_year = None
            if fiscal_year_id:
                current_fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
            elif date_from:
                current_fiscal_year = FiscalYear.objects.filter(
                    start_date__lte=date_from,
                    end_date__gte=date_from
                ).first()
            else:
                current_fiscal_year = FiscalYear.objects.filter(
                    start_date__lte=today,
                    end_date__gte=today
                ).first()

            if date_from is None:
                if current_fiscal_year and current_fiscal_year.start_date:
                    date_from = current_fiscal_year.start_date
                else:
                    date_from = date(date_to.year, 1, 1)

            # تاريخ بداية السنة المالية لعزل حسابات الإيرادات والمصروفات
            fy_start_date = current_fiscal_year.start_date if current_fiscal_year else date(date_from.year, 1, 1)

            # 3. الاستعلام الأول: جلب كافة حسابات الدليل النشطة أو التي لها حركات/أرصدة
            accounts_qs = ChartOfAccounts.objects.filter(
                Q(is_active=True) | Q(journal_lines__journal_entry__status='posted') | Q(opening_balance__gt=0)
            ).distinct().select_related('account_type', 'parent', 'currency').order_by('code')

            accounts_list = list(accounts_qs)
            if not accounts_list:
                return cls._empty_trial_balance_response(date_from, date_to, display_mode, currency_code, currency_symbol)

            acc_map = {acc.id: acc for acc in accounts_list}
            leaf_ids = {acc.id for acc in accounts_list if acc.is_leaf}

            # 4. الاستعلام الثاني: تجميع حركات ما قبل تاريخ البداية (لحساب رصيد أول المدة)
            # لحسابات الميزانية: كل الحركات قبل date_from
            # لحسابات قائمة الدخل (إيرادات ومصروفات): الحركات بين fy_start_date و date_from
            prior_balances = {}
            if date_from:
                # أ. حسابات الميزانية (أصول، خصوم، حقوق ملكية)
                bs_query = Q(
                    journal_entry__status='posted',
                    journal_entry__date__lt=date_from,
                    account__account_type__category__in=['asset', 'liability', 'equity']
                )
                bs_totals = (
                    JournalEntryLine.objects.filter(bs_query)
                    .values('account_id')
                    .annotate(
                        sum_debit=Coalesce(Sum('debit'), Decimal('0.00')),
                        sum_credit=Coalesce(Sum('credit'), Decimal('0.00'))
                    )
                )
                for row in bs_totals:
                    prior_balances[row['account_id']] = {
                        'debit': row['sum_debit'],
                        'credit': row['sum_credit']
                    }

                # ب. حسابات قائمة الدخل (إيرادات ومصروفات) محصورة بالسنة المالية الحالية
                if date_from > fy_start_date:
                    pl_query = Q(
                        journal_entry__status='posted',
                        journal_entry__date__gte=fy_start_date,
                        journal_entry__date__lt=date_from,
                        account__account_type__category__in=['revenue', 'expense']
                    )
                    pl_totals = (
                        JournalEntryLine.objects.filter(pl_query)
                        .values('account_id')
                        .annotate(
                            sum_debit=Coalesce(Sum('debit'), Decimal('0.00')),
                            sum_credit=Coalesce(Sum('credit'), Decimal('0.00'))
                        )
                    )
                    for row in pl_totals:
                        prior_balances[row['account_id']] = {
                            'debit': row['sum_debit'],
                            'credit': row['sum_credit']
                        }

                # ج. تسوية أرباح وخسائر السنوات السابقة غير المقفلة تلقائياً (Implicit Retained Earnings)
                if fy_start_date:
                    unclosed_pl_lines = JournalEntryLine.objects.filter(
                        journal_entry__status='posted',
                        journal_entry__date__lt=fy_start_date,
                        account__account_type__category__in=['revenue', 'expense']
                    ).aggregate(
                        sum_dr=Coalesce(Sum('debit'), Decimal('0.00')),
                        sum_cr=Coalesce(Sum('credit'), Decimal('0.00'))
                    )
                    net_unclosed_pl = (unclosed_pl_lines['sum_cr'] or Decimal('0.00')) - (unclosed_pl_lines['sum_dr'] or Decimal('0.00'))
                    if net_unclosed_pl != Decimal('0.00'):
                        from financial.services.role_registry import AccountRoleRegistry
                        retained_acc = AccountRoleRegistry.get_account_by_role("RETAINED_EARNINGS")
                        if not retained_acc:
                            retained_acc = accounts_qs.filter(code__in=['30200', '3020', '30000', '31000'], is_active=True).first()
                        if not retained_acc:
                            retained_acc = accounts_qs.filter(account_type__category='equity').first()

                        if retained_acc:
                            cur_prior = prior_balances.get(retained_acc.id, {'debit': Decimal('0.00'), 'credit': Decimal('0.00')})
                            if net_unclosed_pl > Decimal('0.00'):
                                cur_prior['credit'] += net_unclosed_pl
                            else:
                                cur_prior['debit'] += abs(net_unclosed_pl)
                            prior_balances[retained_acc.id] = cur_prior

            # 5. الاستعلام الثالث: تجميع حركات الفترة بين date_from و date_to
            period_query = Q(
                journal_entry__status='posted',
                journal_entry__date__gte=date_from,
                journal_entry__date__lte=date_to
            )
            period_totals = (
                JournalEntryLine.objects.filter(period_query)
                .values('account_id')
                .annotate(
                    sum_debit=Coalesce(Sum('debit'), Decimal('0.00')),
                    sum_credit=Coalesce(Sum('credit'), Decimal('0.00'))
                )
            )
            period_map = {row['account_id']: row for row in period_totals}

            # فحص الحسابات التي لها قيود افتتاحية مسجلة في الدفاتر لمنع التكرار المزدوج
            opening_entries_account_ids = set(
                JournalEntryLine.objects.filter(
                    journal_entry__status='posted',
                    journal_entry__entry_type='opening'
                ).values_list('account_id', flat=True).distinct()
            )

            # 6. بناء بيانات الحسابات المباشرة (Direct Account Balances)
            account_nodes: Dict[int, Dict[str, Any]] = {}
            for acc in accounts_list:
                # حركة ما قبل الفترة
                prior_data = prior_balances.get(acc.id, {'debit': Decimal('0.00'), 'credit': Decimal('0.00')})
                prior_dr = prior_data['debit']
                prior_cr = prior_data['credit']

                # إضافة الرصيد الافتتاحي المعرف على الحساب كـ Fallback فقط إذا لم يكن له قيد افتتاحي مرحل
                if acc.id not in opening_entries_account_ids and acc.opening_balance:
                    op_val = acc.opening_balance
                    if acc.nature == 'debit':
                        prior_dr += op_val
                    else:
                        prior_cr += op_val

                # احتساب صافي رصيد أول المدة
                net_opening = prior_dr - prior_cr
                if net_opening > Decimal('0.00'):
                    opening_dr = net_opening
                    opening_cr = Decimal('0.00')
                elif net_opening < Decimal('0.00'):
                    opening_dr = Decimal('0.00')
                    opening_cr = abs(net_opening)
                else:
                    opening_dr = Decimal('0.00')
                    opening_cr = Decimal('0.00')

                # حركات الفترة
                p_row = period_map.get(acc.id, {})
                period_dr = p_row.get('sum_debit') or Decimal('0.00')
                period_cr = p_row.get('sum_credit') or Decimal('0.00')

                # احتساب صافي رصيد الإقفال
                # Net = (Opening Dr - Opening Cr) + (Period Dr - Period Cr)
                net_closing = (opening_dr - opening_cr) + (period_dr - period_cr)
                if net_closing > Decimal('0.00'):
                    closing_dr = net_closing
                    closing_cr = Decimal('0.00')
                elif net_closing < Decimal('0.00'):
                    closing_dr = Decimal('0.00')
                    closing_cr = abs(net_closing)
                else:
                    closing_dr = Decimal('0.00')
                    closing_cr = Decimal('0.00')

                account_nodes[acc.id] = {
                    'account': acc,
                    'id': acc.id,
                    'code': acc.code,
                    'name': acc.name,
                    'category': acc.account_type.category if acc.account_type else 'asset',
                    'category_name': acc.account_type.name if acc.account_type else '',
                    'nature': acc.nature,
                    'level': acc.level,
                    'is_leaf': acc.is_leaf,
                    'parent_id': acc.parent_id,
                    # أرقام مباشرة
                    'opening_debit': opening_dr.quantize(Decimal('0.01')),
                    'opening_credit': opening_cr.quantize(Decimal('0.01')),
                    'period_debit': period_dr.quantize(Decimal('0.01')),
                    'period_credit': period_cr.quantize(Decimal('0.01')),
                    'closing_debit': closing_dr.quantize(Decimal('0.01')),
                    'closing_credit': closing_cr.quantize(Decimal('0.01')),
                    'net_closing': net_closing.quantize(Decimal('0.01')),
                    # تجميعات صاعدة (Roll-up) للأمهات
                    'rolled_opening_dr': opening_dr,
                    'rolled_opening_cr': opening_cr,
                    'rolled_period_dr': period_dr,
                    'rolled_period_cr': period_cr,
                    'rolled_closing_dr': closing_dr,
                    'rolled_closing_cr': closing_cr,
                    'has_activity': bool(
                        opening_dr > 0 or opening_cr > 0 or
                        period_dr > 0 or period_cr > 0 or
                        closing_dr > 0 or closing_cr > 0
                    ),
                    'active_children_count': 0
                }

            # 7. التجميع الشجري الصاعد في الذاكرة O(N) (Post-Order Roll-up)
            # ترتيب الحسابات من الأسفل للأعلى (من أعمق مستوى لأعلى مستوى)
            sorted_by_depth_desc = sorted(
                accounts_list,
                key=lambda a: (getattr(a, 'level', 1), len(getattr(a, 'code', ''))),
                reverse=True
            )

            for acc in sorted_by_depth_desc:
                if acc.parent_id and acc.parent_id in account_nodes:
                    child_node = account_nodes[acc.id]
                    parent_node = account_nodes[acc.parent_id]

                    parent_node['rolled_opening_dr'] += child_node['rolled_opening_dr']
                    parent_node['rolled_opening_cr'] += child_node['rolled_opening_cr']
                    parent_node['rolled_period_dr'] += child_node['rolled_period_dr']
                    parent_node['rolled_period_cr'] += child_node['rolled_period_cr']
                    parent_node['rolled_closing_dr'] += child_node['rolled_closing_dr']
                    parent_node['rolled_closing_cr'] += child_node['rolled_closing_cr']

                    if child_node['has_activity']:
                        parent_node['has_activity'] = True
                        parent_node['active_children_count'] += 1

            # 8. حساب الإجماليات العامة للميزان (Grand Totals)
            # مطابقة تامة مع دفتر الأستاذ العام عبر تجميع كافة الحسابات المباشرة
            total_opening_debit = Decimal('0.00')
            total_opening_credit = Decimal('0.00')
            total_period_debit = Decimal('0.00')
            total_period_credit = Decimal('0.00')
            total_closing_debit = Decimal('0.00')
            total_closing_credit = Decimal('0.00')
            active_accounts_count = 0

            for node in account_nodes.values():
                total_opening_debit += node['opening_debit']
                total_opening_credit += node['opening_credit']
                total_period_debit += node['period_debit']
                total_period_credit += node['period_credit']
                total_closing_debit += node['closing_debit']
                total_closing_credit += node['closing_credit']

                if node['has_activity']:
                    active_accounts_count += 1

            # التحقق من التوازن المحاسبي الثلاثي
            diff_opening = (total_opening_debit - total_opening_credit).quantize(Decimal('0.01'))
            diff_period = (total_period_debit - total_period_credit).quantize(Decimal('0.01'))
            diff_closing = (total_closing_debit - total_closing_credit).quantize(Decimal('0.01'))

            is_balanced = (
                abs(diff_closing) <= Decimal('0.05') and
                abs(diff_period) <= Decimal('0.05') and
                abs(diff_opening) <= Decimal('0.05')
            )

            # 9. تصفية الحسابات للعرض بناءً على الفلاتر (Level & Zero Balances)
            target_level = None
            if account_level not in [None, '', 'all', 'ALL']:
                if str(account_level).lower() == 'leaf':
                    target_level = 'leaf'
                else:
                    try:
                        target_level = int(account_level)
                    except ValueError:
                        pass

            display_accounts = []
            for acc in accounts_list:
                node = account_nodes[acc.id]

                # فلتر إخفاء الحسابات الصفرية (Bottom-up cascade check)
                if hide_zero_balances and not node['has_activity']:
                    continue

                # فلتر المستويات
                if target_level == 'leaf':
                    if not node['is_leaf']:
                        continue
                elif isinstance(target_level, int):
                    if node['level'] > target_level:
                        continue

                # للأمهات نستخدم القيم المجمعة صعوداً، وللفرعيات نستخدم القيم المباشرة
                if not node['is_leaf']:
                    row_data = {
                        'account': acc,
                        'id': acc.id,
                        'code': acc.code,
                        'name': acc.name,
                        'category': node['category'],
                        'category_name': node['category_name'],
                        'level': node['level'],
                        'is_leaf': False,
                        'parent_id': node['parent_id'],
                        'opening_debit': node['rolled_opening_dr'].quantize(Decimal('0.01')),
                        'opening_credit': node['rolled_opening_cr'].quantize(Decimal('0.01')),
                        'period_debit': node['rolled_period_dr'].quantize(Decimal('0.01')),
                        'period_credit': node['rolled_period_cr'].quantize(Decimal('0.01')),
                        'closing_debit': node['rolled_closing_dr'].quantize(Decimal('0.01')),
                        'closing_credit': node['rolled_closing_cr'].quantize(Decimal('0.01')),
                        'net_closing': (node['rolled_closing_dr'] - node['rolled_closing_cr']).quantize(Decimal('0.01')),
                        'direct_opening_debit': node['opening_debit'],
                        'direct_opening_credit': node['opening_credit'],
                        'direct_period_debit': node['period_debit'],
                        'direct_period_credit': node['period_credit'],
                        'direct_closing_debit': node['closing_debit'],
                        'direct_closing_credit': node['closing_credit'],
                        'is_parent': True,
                    }
                else:
                    row_data = {
                        'account': acc,
                        'id': acc.id,
                        'code': acc.code,
                        'name': acc.name,
                        'category': node['category'],
                        'category_name': node['category_name'],
                        'level': node['level'],
                        'is_leaf': True,
                        'parent_id': node['parent_id'],
                        'opening_debit': node['opening_debit'],
                        'opening_credit': node['opening_credit'],
                        'period_debit': node['period_debit'],
                        'period_credit': node['period_credit'],
                        'closing_debit': node['closing_debit'],
                        'closing_credit': node['closing_credit'],
                        'net_closing': node['net_closing'],
                        'direct_opening_debit': node['opening_debit'],
                        'direct_opening_credit': node['opening_credit'],
                        'direct_period_debit': node['period_debit'],
                        'direct_period_credit': node['period_credit'],
                        'direct_closing_debit': node['closing_debit'],
                        'direct_closing_credit': node['closing_credit'],
                        'is_parent': False,
                    }

                display_accounts.append(row_data)

            # 10. التجميع حسب التصنيف الرئيسي
            grouped_data = {}
            if group_by_type:
                grouped_data = cls._group_accounts_by_category(display_accounts)

            return {
                'accounts': display_accounts,
                'grouped': grouped_data,
                'total_opening_debit': total_opening_debit.quantize(Decimal('0.01')),
                'total_opening_credit': total_opening_credit.quantize(Decimal('0.01')),
                'total_period_debit': total_period_debit.quantize(Decimal('0.01')),
                'total_period_credit': total_period_credit.quantize(Decimal('0.01')),
                'total_closing_debit': total_closing_debit.quantize(Decimal('0.01')),
                'total_closing_credit': total_closing_credit.quantize(Decimal('0.01')),
                # للتوافق العكسي مع الشاشات السابقة
                'total_debit': total_closing_debit.quantize(Decimal('0.01')),
                'total_credit': total_closing_credit.quantize(Decimal('0.01')),
                'is_balanced': is_balanced,
                'difference': diff_closing,
                'diff_opening': diff_opening,
                'diff_period': diff_period,
                'diff_closing': diff_closing,
                'date_from': date_from,
                'date_to': date_to,
                'display_mode': display_mode,
                'account_level': account_level,
                'hide_zero_balances': hide_zero_balances,
                'group_by_type': group_by_type,
                'accounts_count': len(display_accounts),
                'active_accounts_count': active_accounts_count,
                'total_accounts_count': len(accounts_list),
                'currency_code': currency_code,
                'currency_symbol': currency_symbol,
                'generated_at': timezone.now(),
                'fiscal_year': current_fiscal_year,
            }

        except Exception as e:
            logger.exception(f"خطأ غير متوقع في إنشاء ميزان المراجعة: {e}")
            return cls._empty_trial_balance_response(
                date_from, date_to, display_mode, "EGP", "ج.م", error=str(e)
            )

    @classmethod
    def _group_accounts_by_category(cls, accounts_data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        تجميع الحسابات حسب التصنيف المحاسبي القياسي
        """
        category_order = {
            'asset': 1,
            'liability': 2,
            'equity': 3,
            'revenue': 4,
            'expense': 5,
        }
        category_names = {
            'asset': _('الأصول'),
            'liability': _('الخصوم والالتزامات'),
            'equity': _('حقوق الملكية'),
            'revenue': _('الإيرادات'),
            'expense': _('المصروفات'),
        }

        grouped = {}
        for item in accounts_data:
            cat = item.get('category', 'asset')
            if cat not in grouped:
                grouped[cat] = {
                    'name': category_names.get(cat, cat),
                    'order': category_order.get(cat, 99),
                    'accounts': [],
                    'total_opening_debit': Decimal('0.00'),
                    'total_opening_credit': Decimal('0.00'),
                    'total_period_debit': Decimal('0.00'),
                    'total_period_credit': Decimal('0.00'),
                    'total_closing_debit': Decimal('0.00'),
                    'total_closing_credit': Decimal('0.00'),
                }

            grouped[cat]['accounts'].append(item)

            # نجمع في إجمالي المجموعة الحركات المباشرة لكافة حسابات المجموعة
            grouped[cat]['total_opening_debit'] += item.get('direct_opening_debit', item.get('opening_debit', Decimal('0.00')))
            grouped[cat]['total_opening_credit'] += item.get('direct_opening_credit', item.get('opening_credit', Decimal('0.00')))
            grouped[cat]['total_period_debit'] += item.get('direct_period_debit', item.get('period_debit', Decimal('0.00')))
            grouped[cat]['total_period_credit'] += item.get('direct_period_credit', item.get('period_credit', Decimal('0.00')))
            grouped[cat]['total_closing_debit'] += item.get('direct_closing_debit', item.get('closing_debit', Decimal('0.00')))
            grouped[cat]['total_closing_credit'] += item.get('direct_closing_credit', item.get('closing_credit', Decimal('0.00')))

        # ترتيب المجموعات وفق التسلسل المحاسبي
        return dict(sorted(grouped.items(), key=lambda x: x[1]['order']))

    @classmethod
    def _empty_trial_balance_response(
        cls,
        date_from: Optional[date],
        date_to: Optional[date],
        display_mode: str,
        currency_code: str,
        currency_symbol: str,
        error: Optional[str] = None
    ) -> Dict[str, Any]:
        """توليد هيكل استجابة فارغ آمن"""
        return {
            'accounts': [],
            'grouped': {},
            'total_opening_debit': Decimal('0.00'),
            'total_opening_credit': Decimal('0.00'),
            'total_period_debit': Decimal('0.00'),
            'total_period_credit': Decimal('0.00'),
            'total_closing_debit': Decimal('0.00'),
            'total_closing_credit': Decimal('0.00'),
            'total_debit': Decimal('0.00'),
            'total_credit': Decimal('0.00'),
            'is_balanced': True if not error else False,
            'difference': Decimal('0.00'),
            'diff_opening': Decimal('0.00'),
            'diff_period': Decimal('0.00'),
            'diff_closing': Decimal('0.00'),
            'date_from': date_from,
            'date_to': date_to,
            'display_mode': display_mode,
            'account_level': None,
            'hide_zero_balances': False,
            'group_by_type': True,
            'accounts_count': 0,
            'active_accounts_count': 0,
            'total_accounts_count': 0,
            'currency_code': currency_code,
            'currency_symbol': currency_symbol,
            'generated_at': timezone.now(),
            'error': error
        }

    @classmethod
    def export_to_excel(
        cls,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        display_mode: str = '6_columns',
        account_level: Optional[Union[int, str]] = None,
        hide_zero_balances: bool = False,
        group_by_type: bool = True
    ) -> bytes:
        """
        تصدير ميزان المراجعة إلى مصنف Excel رسمي متكامل (RTL & Formatted)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from io import BytesIO

            # إنشاء ميزان المراجعة
            tb_data = cls.generate_trial_balance(
                date_from=date_from,
                date_to=date_to,
                display_mode=display_mode,
                account_level=account_level,
                hide_zero_balances=hide_zero_balances,
                group_by_type=group_by_type
            )

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ميزان المراجعة"

            # 1. ضبط اتجاه الورقة RTL
            ws.views.sheetView[0].rightToLeft = True

            # 2. تعريف الأنماط والألوان الهادئة
            font_title = Font(name="Arial", size=15, bold=True, color="1F2937")
            font_subtitle = Font(name="Arial", size=10, italic=True, color="4B5563")
            font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            font_group = Font(name="Arial", size=11, bold=True, color="111827")
            font_data = Font(name="Arial", size=10, color="1F2937")
            font_total = Font(name="Arial", size=11, bold=True, color="FFFFFF")

            fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            fill_sub_header = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            fill_group = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            fill_total = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

            border_thin = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            align_right = Alignment(horizontal='right', vertical='center')
            align_number = Alignment(horizontal='right', vertical='center')

            number_format = '#,##0.00'

            # 3. ترويسة التقرير
            ws.merge_cells('A1:I1' if display_mode == '6_columns' else 'A1:E1')
            ws['A1'] = "تقرير ميزان المراجعة"
            ws['A1'].font = font_title
            ws['A1'].alignment = align_center

            ws.merge_cells('A2:I2' if display_mode == '6_columns' else 'A2:E2')
            period_str = f"عن الفترة من {tb_data['date_from'].strftime('%Y-%m-%d')} إلى {tb_data['date_to'].strftime('%Y-%m-%d')} | العملة: {tb_data['currency_code']}"
            ws['A2'] = period_str
            ws['A2'].font = font_subtitle
            ws['A2'].alignment = align_center

            ws.merge_cells('A3:I3' if display_mode == '6_columns' else 'A3:E3')
            status_str = "الحالة: متوازن محاسبياً ✓" if tb_data['is_balanced'] else f"الحالة: غير متوازن (الفرق: {float(tb_data['difference'])}) ⚠"
            ws['A3'] = f"تاريخ الاستخراج: {tb_data['generated_at'].strftime('%Y-%m-%d %H:%M')} | {status_str}"
            ws['A3'].font = font_subtitle
            ws['A3'].alignment = align_center

            row = 5
            # 4. بناء رؤوس الأعمدة
            if display_mode == '6_columns':
                # السطر الأول من الرؤوس
                ws.merge_cells(f'A{row}:A{row+1}')
                ws[f'A{row}'] = "كود الحساب"
                ws.merge_cells(f'B{row}:B{row+1}')
                ws[f'B{row}'] = "اسم الحساب"
                ws.merge_cells(f'C{row}:C{row+1}')
                ws[f'C{row}'] = "النوع / المستوى"

                ws.merge_cells(f'D{row}:E{row}')
                ws[f'D{row}'] = "رصيد أول المدة"
                ws.merge_cells(f'F{row}:G{row}')
                ws[f'F{row}'] = "حركات الفترة"
                ws.merge_cells(f'H{row}:I{row}')
                ws[f'H{row}'] = "رصيد نهاية المدة"

                for col_idx in range(1, 10):
                    cell = ws.cell(row=row, column=col_idx)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = align_center

                # السطر الثاني من الرؤوس
                row += 1
                sub_headers = ["", "", "", "مدين", "دائن", "مدين", "دائن", "مدين", "دائن"]
                for col_idx in range(4, 10):
                    cell = ws.cell(row=row, column=col_idx, value=sub_headers[col_idx-1])
                    cell.fill = fill_sub_header
                    cell.font = font_header
                    cell.alignment = align_center
                    cell.border = border_thin

                ws.freeze_panes = 'A7'
            else:
                # نمط 2 عمود
                headers_2 = ["كود الحساب", "اسم الحساب", "النوع", "مدين (ختامي)", "دائن (ختامي)"]
                for col_idx, h in enumerate(headers_2, 1):
                    cell = ws.cell(row=row, column=col_idx, value=h)
                    cell.fill = fill_header
                    cell.font = font_header
                    cell.alignment = align_center
                    cell.border = border_thin
                ws.freeze_panes = 'A6'

            row += 1

            # 5. ملء أسطر البيانات
            for item in tb_data['accounts']:
                is_parent = item.get('is_parent', False)
                indent = "  " * max(0, item['level'] - 1)

                if display_mode == '6_columns':
                    ws.cell(row=row, column=1, value=item['code']).alignment = align_center
                    ws.cell(row=row, column=2, value=f"{indent}{item['name']}").alignment = align_right
                    ws.cell(row=row, column=3, value=item.get('category_name', '')).alignment = align_center

                    c4 = ws.cell(row=row, column=4, value=float(item['opening_debit']))
                    c5 = ws.cell(row=row, column=5, value=float(item['opening_credit']))
                    c6 = ws.cell(row=row, column=6, value=float(item['period_debit']))
                    c7 = ws.cell(row=row, column=7, value=float(item['period_credit']))
                    c8 = ws.cell(row=row, column=8, value=float(item['closing_debit']))
                    c9 = ws.cell(row=row, column=9, value=float(item['closing_credit']))

                    for c in [c4, c5, c6, c7, c8, c9]:
                        c.number_format = number_format
                        c.alignment = align_number

                    for col_idx in range(1, 10):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.border = border_thin
                        if is_parent:
                            cell.font = font_group
                            cell.fill = fill_group
                        else:
                            cell.font = font_data
                else:
                    ws.cell(row=row, column=1, value=item['code']).alignment = align_center
                    ws.cell(row=row, column=2, value=f"{indent}{item['name']}").alignment = align_right
                    ws.cell(row=row, column=3, value=item.get('category_name', '')).alignment = align_center

                    c4 = ws.cell(row=row, column=4, value=float(item['closing_debit']))
                    c5 = ws.cell(row=row, column=5, value=float(item['closing_credit']))
                    c4.number_format = number_format
                    c5.number_format = number_format
                    c4.alignment = align_number
                    c5.alignment = align_number

                    for col_idx in range(1, 6):
                        cell = ws.cell(row=row, column=col_idx)
                        cell.border = border_thin
                        if is_parent:
                            cell.font = font_group
                            cell.fill = fill_group
                        else:
                            cell.font = font_data

                row += 1

            # 6. الإجمالي النهائي
            if display_mode == '6_columns':
                ws.merge_cells(f'A{row}:C{row}')
                ws[f'A{row}'] = "الإجمالي النهائي"
                ws[f'A{row}'].font = font_total
                ws[f'A{row}'].alignment = align_center

                t_vals = [
                    float(tb_data['total_opening_debit']),
                    float(tb_data['total_opening_credit']),
                    float(tb_data['total_period_debit']),
                    float(tb_data['total_period_credit']),
                    float(tb_data['total_closing_debit']),
                    float(tb_data['total_closing_credit'])
                ]
                for idx, val in enumerate(t_vals, 4):
                    cell = ws.cell(row=row, column=idx, value=val)
                    cell.number_format = number_format
                    cell.font = font_total
                    cell.alignment = align_number

                for col_idx in range(1, 10):
                    ws.cell(row=row, column=col_idx).fill = fill_total
                    ws.cell(row=row, column=col_idx).border = border_thin
            else:
                ws.merge_cells(f'A{row}:C{row}')
                ws[f'A{row}'] = "الإجمالي النهائي"
                ws[f'A{row}'].font = font_total
                ws[f'A{row}'].alignment = align_center

                c4 = ws.cell(row=row, column=4, value=float(tb_data['total_closing_debit']))
                c5 = ws.cell(row=row, column=5, value=float(tb_data['total_closing_credit']))
                c4.number_format = number_format
                c5.number_format = number_format
                c4.font = font_total
                c5.font = font_total

                for col_idx in range(1, 6):
                    ws.cell(row=row, column=col_idx).fill = fill_total
                    ws.cell(row=row, column=col_idx).border = border_thin

            # 7. ضبط أبعاد الأعمدة
            max_col = 9 if display_mode == '6_columns' else 5
            for idx in range(1, max_col + 1):
                col_letter = get_column_letter(idx)
                if idx == 1:
                    ws.column_dimensions[col_letter].width = 16
                elif idx == 2:
                    ws.column_dimensions[col_letter].width = 38
                elif idx == 3:
                    ws.column_dimensions[col_letter].width = 18
                else:
                    ws.column_dimensions[col_letter].width = 16

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except Exception as e:
            logger.exception(f"خطأ في تصدير Excel لميزان المراجعة: {e}")
            raise
