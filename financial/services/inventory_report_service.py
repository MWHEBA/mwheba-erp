"""
خدمة تقرير المخزون المالي
توفر تحليل شامل لحالة المخزون وقيمته
"""

from decimal import Decimal
from datetime import datetime, timedelta
from django.db.models import Sum, Count, Q
from django.utils import timezone
from financial.models.chart_of_accounts import ChartOfAccounts
from financial.models.journal_entry import JournalEntry, JournalEntryLine


class InventoryReportService:
    """
    خدمة تقرير المخزون المالي
    """

    # الكلمات المفتاحية للبحث عن حسابات المخزون
    INVENTORY_KEYWORDS = ['مخزون', 'بضاعة', 'مواد', 'منتجات', 'سلع', 'خامات', 'inventory', 'stock']

    def __init__(self, report_date=None):
        """
        تهيئة الخدمة
        
        Args:
            report_date: تاريخ التقرير (افتراضي: اليوم)
        """
        self.report_date = report_date or timezone.now().date()

    def get_inventory_accounts(self):
        """
        الحصول على حسابات المخزون
        
        Returns:
            QuerySet: حسابات المخزون
        """
        # جلب حسابات الأصول
        asset_accounts = ChartOfAccounts.objects.filter(
            account_type__category="asset",
            is_active=True
        )

        # فلترة الحسابات التي تحتوي على كلمات مفتاحية للمخزون
        inventory_accounts = []
        for account in asset_accounts:
            # التحقق من وجود كلمات مفتاحية في اسم الحساب
            is_inventory = any(
                keyword.lower() in account.name.lower() 
                for keyword in self.INVENTORY_KEYWORDS
            )
            
            if is_inventory:
                inventory_accounts.append(account)

        return inventory_accounts

    def get_account_balance(self, account):
        """
        حساب رصيد حساب معين حتى تاريخ التقرير
        
        Args:
            account: الحساب
            
        Returns:
            Decimal: الرصيد
        """
        journal_lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__date__lte=self.report_date,
            journal_entry__status='posted'
        )

        totals = journal_lines.aggregate(
            total_debit=Sum("debit"),
            total_credit=Sum("credit")
        )

        total_debit = Decimal(str(totals["total_debit"])) if totals["total_debit"] is not None else Decimal("0")
        total_credit = Decimal(str(totals["total_credit"])) if totals["total_credit"] is not None else Decimal("0")

        # الأصول لها رصيد مدين
        balance = total_debit - total_credit

        return balance

    def get_account_movements(self, account, days=30):
        """
        حساب عدد الحركات الأخيرة لحساب معين
        
        Args:
            account: الحساب
            days: عدد الأيام (افتراضي: 30)
            
        Returns:
            int: عدد الحركات
        """
        recent_date = self.report_date - timedelta(days=days)
        
        movements = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__date__gte=recent_date,
            journal_entry__date__lte=self.report_date,
            journal_entry__status='posted'
        ).count()

        return movements

    def get_account_transactions_count(self, account):
        """
        حساب إجمالي عدد المعاملات لحساب معين
        
        Args:
            account: الحساب
            
        Returns:
            int: عدد المعاملات
        """
        count = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__date__lte=self.report_date,
            journal_entry__status='posted'
        ).count()

        return count

    def get_inventory_by_account(self):
        """
        الحصول على المخزون مجمعاً حسب الحساب
        
        Returns:
            dict: قائمة المخزون حسب الحساب
        """
        inventory_accounts = self.get_inventory_accounts()
        inventory_data = []
        total_inventory_value = Decimal("0")

        for account in inventory_accounts:
            # حساب الرصيد
            balance = self.get_account_balance(account)

            # عرض الحسابات ذات الرصيد فقط
            if balance != 0:
                # حساب عدد الحركات الأخيرة
                recent_movements = self.get_account_movements(account, days=30)
                
                # حساب إجمالي عدد المعاملات
                transactions_count = self.get_account_transactions_count(account)

                inventory_data.append({
                    "account": account,
                    "account_code": account.code,
                    "account_name": account.name,
                    "balance": balance,
                    "recent_movements": recent_movements,
                    "transactions_count": transactions_count,
                    "percentage": Decimal("0"),  # سيتم حسابها لاحقاً
                    "status": "active" if recent_movements > 0 else "stagnant"
                })
                total_inventory_value += balance

        # حساب النسب المئوية
        if total_inventory_value > 0:
            for item in inventory_data:
                try:
                    item["percentage"] = (item["balance"] / total_inventory_value * 100).quantize(Decimal("0.01"))
                except:
                    item["percentage"] = Decimal("0")

        # ترتيب حسب القيمة (الأعلى أولاً)
        inventory_data.sort(key=lambda x: x["balance"], reverse=True)

        return {
            "inventory_data": inventory_data,
            "total_inventory_value": total_inventory_value
        }

    def get_inventory_statistics(self):
        """
        حساب إحصائيات المخزون
        
        Returns:
            dict: الإحصائيات
        """
        result = self.get_inventory_by_account()
        total_inventory_value = result["total_inventory_value"]
        inventory_data = result["inventory_data"]

        # عدد الحسابات
        total_accounts = len(inventory_data)

        # متوسط قيمة الحساب
        avg_account_value = total_inventory_value / total_accounts if total_accounts > 0 else Decimal("0")

        # الحسابات النشطة
        active_accounts = len([item for item in inventory_data if item["status"] == "active"])

        # الحسابات الراكدة
        stagnant_accounts = total_accounts - active_accounts

        # أعلى حساب قيمة
        top_account = inventory_data[0] if inventory_data else None

        # أقل حساب قيمة
        lowest_account = inventory_data[-1] if inventory_data else None

        return {
            "total_inventory_value": total_inventory_value,
            "total_accounts": total_accounts,
            "avg_account_value": avg_account_value,
            "active_accounts": active_accounts,
            "stagnant_accounts": stagnant_accounts,
            "top_account": top_account,
            "lowest_account": lowest_account,
        }

    def get_inventory_by_category(self):
        """
        تجميع المخزون حسب نوع الحساب
        
        Returns:
            dict: المخزون حسب الفئة
        """
        inventory_accounts = self.get_inventory_accounts()
        categories = {}

        for account in inventory_accounts:
            category = account.account_type.name
            balance = self.get_account_balance(account)

            if balance > 0:
                if category not in categories:
                    categories[category] = Decimal("0")
                categories[category] += balance

        # تحويل إلى قوائم
        labels = list(categories.keys())
        data = [float(categories[key]) for key in labels]

        return {
            "labels": labels,
            "data": data
        }

    def get_inventory_turnover_analysis(self):
        """
        تحليل معدل دوران المخزون
        
        Returns:
            dict: تحليل معدل الدوران
        """
        result = self.get_inventory_by_account()
        inventory_data = result["inventory_data"]

        # تصنيف الحسابات حسب النشاط
        high_turnover = []  # حركة عالية (> 10 حركات في 30 يوم)
        medium_turnover = []  # حركة متوسطة (5-10 حركات)
        low_turnover = []  # حركة منخفضة (1-4 حركات)
        no_turnover = []  # راكد (0 حركات)

        for item in inventory_data:
            movements = item["recent_movements"]
            if movements > 10:
                high_turnover.append(item)
            elif movements >= 5:
                medium_turnover.append(item)
            elif movements >= 1:
                low_turnover.append(item)
            else:
                no_turnover.append(item)

        return {
            "high_turnover": high_turnover,
            "medium_turnover": medium_turnover,
            "low_turnover": low_turnover,
            "no_turnover": no_turnover,
            "high_turnover_count": len(high_turnover),
            "medium_turnover_count": len(medium_turnover),
            "low_turnover_count": len(low_turnover),
            "no_turnover_count": len(no_turnover),
        }

    def export_to_excel(self):
        """
        تصدير التقرير إلى Excel
        
        Returns:
            bytes: محتوى ملف Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO

            # إنشاء workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "تقرير المخزون"

            # تنسيق العنوان
            title_font = Font(size=16, bold=True)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")

            # العنوان
            ws.merge_cells('A1:G1')
            ws['A1'] = f"تقرير المخزون المالي - {self.report_date}"
            ws['A1'].font = title_font
            ws['A1'].alignment = center_alignment

            # رأس الجدول
            headers = ["كود الحساب", "اسم الحساب", "القيمة الحالية", "عدد المعاملات", "الحركات الأخيرة (30 يوم)", "النسبة %", "الحالة"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # البيانات
            result = self.get_inventory_by_account()
            inventory_data = result["inventory_data"]
            total_inventory_value = result["total_inventory_value"]

            row = 4
            for item in inventory_data:
                ws.cell(row=row, column=1, value=item["account_code"])
                ws.cell(row=row, column=2, value=item["account_name"])
                ws.cell(row=row, column=3, value=float(item["balance"]))
                ws.cell(row=row, column=4, value=item["transactions_count"])
                ws.cell(row=row, column=5, value=item["recent_movements"])
                ws.cell(row=row, column=6, value=float(item["percentage"]))
                ws.cell(row=row, column=7, value="نشط" if item["status"] == "active" else "راكد")
                row += 1

            # الإجمالي
            ws.cell(row=row, column=1, value="الإجمالي").font = Font(bold=True)
            ws.cell(row=row, column=3, value=float(total_inventory_value)).font = Font(bold=True)
            ws.cell(row=row, column=6, value=100.0).font = Font(bold=True)

            # تعديل عرض الأعمدة
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12

            # حفظ إلى BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except ImportError:
            return None

    def get_complete_report(self):
        """
        الحصول على التقرير الكامل
        
        Returns:
            dict: التقرير الكامل
        """
        inventory_by_account = self.get_inventory_by_account()
        statistics = self.get_inventory_statistics()
        inventory_by_category = self.get_inventory_by_category()
        turnover_analysis = self.get_inventory_turnover_analysis()

        return {
            "inventory_data": inventory_by_account["inventory_data"],
            "total_inventory_value": inventory_by_account["total_inventory_value"],
            "summary": statistics,  # alias for tests
            "statistics": statistics,
            "inventory_by_category": inventory_by_category,
            "turnover_analysis": turnover_analysis,
            "report_date": self.report_date,
        }
    
    def get_inventory_report(self):
        """
        Alias for get_complete_report() for backward compatibility
        """
        return self.get_complete_report()
