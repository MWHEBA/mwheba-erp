"""
خدمة تحليل ABC للمخزون
تصنيف المخزون حسب الأهمية (A, B, C)
"""

from decimal import Decimal
from datetime import datetime, timedelta
from django.db.models import Sum, Count, Q, F
from django.utils import timezone
from financial.models.chart_of_accounts import ChartOfAccounts
from financial.models.journal_entry import JournalEntry, JournalEntryLine


class ABCAnalysisService:
    """
    خدمة تحليل ABC للمخزون
    
    تصنيف المخزون إلى ثلاث فئات:
    - A: 20% من الأصناف تمثل 80% من القيمة (عالية الأهمية)
    - B: 30% من الأصناف تمثل 15% من القيمة (متوسطة الأهمية)
    - C: 50% من الأصناف تمثل 5% من القيمة (منخفضة الأهمية)
    """

    # الكلمات المفتاحية للبحث عن حسابات المخزون
    INVENTORY_KEYWORDS = ['مخزون', 'بضاعة', 'مواد', 'منتجات', 'سلع', 'خامات', 'inventory', 'stock']

    # معايير التصنيف (يمكن تخصيصها)
    CATEGORY_A_PERCENTAGE = 80  # نسبة القيمة لفئة A
    CATEGORY_B_PERCENTAGE = 15  # نسبة القيمة لفئة B
    CATEGORY_C_PERCENTAGE = 5   # نسبة القيمة لفئة C

    def __init__(self, analysis_date=None, days_period=365):
        """
        تهيئة الخدمة
        
        Args:
            analysis_date: تاريخ التحليل (افتراضي: اليوم)
            days_period: فترة التحليل بالأيام (افتراضي: 365 يوم)
        """
        self.analysis_date = analysis_date or timezone.now().date()
        self.days_period = days_period
        self.date_from = self.analysis_date - timedelta(days=days_period)

    def get_inventory_accounts(self):
        """
        الحصول على حسابات المخزون
        
        Returns:
            QuerySet: حسابات المخزون
        """
        asset_accounts = ChartOfAccounts.objects.filter(
            account_type__category="asset",
            is_active=True
        )

        inventory_accounts = []
        for account in asset_accounts:
            is_inventory = any(
                keyword.lower() in account.name.lower() 
                for keyword in self.INVENTORY_KEYWORDS
            )
            
            if is_inventory:
                inventory_accounts.append(account)

        return inventory_accounts

    def get_account_value(self, account):
        """
        حساب قيمة حساب المخزون
        
        Args:
            account: الحساب
            
        Returns:
            Decimal: القيمة
        """
        journal_lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__date__lte=self.analysis_date,
            journal_entry__status='posted'
        )

        totals = journal_lines.aggregate(
            total_debit=Sum("debit"),
            total_credit=Sum("credit")
        )

        total_debit = Decimal(str(totals["total_debit"])) if totals["total_debit"] is not None else Decimal("0")
        total_credit = Decimal(str(totals["total_credit"])) if totals["total_credit"] is not None else Decimal("0")

        # الأصول لها رصيد مدين
        balance = total_debit - total_credit

        return balance if balance > 0 else Decimal("0")

    def get_account_transactions_count(self, account):
        """
        حساب عدد المعاملات لحساب معين في الفترة
        
        Args:
            account: الحساب
            
        Returns:
            int: عدد المعاملات
        """
        count = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__date__gte=self.date_from,
            journal_entry__date__lte=self.analysis_date,
            journal_entry__status='posted'
        ).count()

        return count

    def get_account_turnover(self, account):
        """
        حساب معدل دوران حساب المخزون
        
        Args:
            account: الحساب
            
        Returns:
            Decimal: معدل الدوران
        """
        # حساب إجمالي الحركة (مدين + دائن) في الفترة
        journal_lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__date__gte=self.date_from,
            journal_entry__date__lte=self.analysis_date,
            journal_entry__status='posted'
        )

        totals = journal_lines.aggregate(
            total_debit=Sum("debit"),
            total_credit=Sum("credit")
        )

        total_debit = Decimal(str(totals["total_debit"])) if totals["total_debit"] is not None else Decimal("0")
        total_credit = Decimal(str(totals["total_credit"])) if totals["total_credit"] is not None else Decimal("0")

        # معدل الدوران = إجمالي الحركة
        turnover = total_debit + total_credit

        return turnover

    def classify_inventory(self):
        """
        تصنيف المخزون حسب تحليل ABC
        
        Returns:
            dict: بيانات التصنيف
        """
        inventory_accounts = self.get_inventory_accounts()
        inventory_data = []
        total_value = Decimal("0")

        # جمع البيانات
        for account in inventory_accounts:
            value = self.get_account_value(account)
            
            if value > 0:
                transactions_count = self.get_account_transactions_count(account)
                turnover = self.get_account_turnover(account)
                
                inventory_data.append({
                    "account": account,
                    "account_code": account.code,
                    "account_name": account.name,
                    "value": value,
                    "transactions_count": transactions_count,
                    "turnover": turnover,
                    "percentage": Decimal("0"),
                    "cumulative_percentage": Decimal("0"),
                    "category": "",
                })
                total_value += value

        # ترتيب حسب القيمة (الأعلى أولاً)
        inventory_data.sort(key=lambda x: x["value"], reverse=True)

        # حساب النسب المئوية والنسب التراكمية
        cumulative_value = Decimal("0")
        for item in inventory_data:
            if total_value > 0:
                item["percentage"] = (item["value"] / total_value * 100).quantize(Decimal("0.01"))
                cumulative_value += item["value"]
                item["cumulative_percentage"] = (cumulative_value / total_value * 100).quantize(Decimal("0.01"))

        # تصنيف ABC
        for item in inventory_data:
            cumulative = float(item["cumulative_percentage"])
            if cumulative <= self.CATEGORY_A_PERCENTAGE:
                item["category"] = "A"
            elif cumulative <= (self.CATEGORY_A_PERCENTAGE + self.CATEGORY_B_PERCENTAGE):
                item["category"] = "B"
            else:
                item["category"] = "C"

        return {
            "inventory_data": inventory_data,
            "total_value": total_value,
        }

    def get_category_statistics(self):
        """
        حساب إحصائيات كل فئة
        
        Returns:
            dict: إحصائيات الفئات
        """
        result = self.classify_inventory()
        inventory_data = result["inventory_data"]
        total_value = result["total_value"]

        # تجميع حسب الفئة
        category_a = [item for item in inventory_data if item["category"] == "A"]
        category_b = [item for item in inventory_data if item["category"] == "B"]
        category_c = [item for item in inventory_data if item["category"] == "C"]

        # حساب الإحصائيات
        def calculate_stats(category_items):
            if not category_items:
                return {
                    "count": 0,
                    "value": Decimal("0"),
                    "percentage_count": Decimal("0"),
                    "percentage_value": Decimal("0"),
                    "avg_value": Decimal("0"),
                }
            
            count = len(category_items)
            value = sum(item["value"] for item in category_items)
            
            return {
                "count": count,
                "value": value,
                "percentage_count": (Decimal(count) / len(inventory_data) * 100).quantize(Decimal("0.01")) if inventory_data else Decimal("0"),
                "percentage_value": (value / total_value * 100).quantize(Decimal("0.01")) if total_value > 0 else Decimal("0"),
                "avg_value": (value / count).quantize(Decimal("0.01")) if count > 0 else Decimal("0"),
            }

        return {
            "category_a": calculate_stats(category_a),
            "category_b": calculate_stats(category_b),
            "category_c": calculate_stats(category_c),
            "total_items": len(inventory_data),
            "total_value": total_value,
        }

    def get_recommendations(self):
        """
        الحصول على توصيات بناءً على التحليل
        
        Returns:
            dict: التوصيات
        """
        result = self.classify_inventory()
        inventory_data = result["inventory_data"]

        recommendations = {
            "category_a": [],
            "category_b": [],
            "category_c": [],
        }

        for item in inventory_data:
            account_name = item["account_name"]
            category = item["category"]
            
            if category == "A":
                recommendations["category_a"].append({
                    "account": account_name,
                    "recommendation": "مراقبة دقيقة - مخزون عالي القيمة",
                    "action": "جرد دوري متكرر، تحكم صارم في المخزون",
                })
            elif category == "B":
                recommendations["category_b"].append({
                    "account": account_name,
                    "recommendation": "مراقبة متوسطة - مخزون متوسط القيمة",
                    "action": "جرد دوري منتظم، تحكم معتدل في المخزون",
                })
            else:
                recommendations["category_c"].append({
                    "account": account_name,
                    "recommendation": "مراقبة بسيطة - مخزون منخفض القيمة",
                    "action": "جرد دوري سنوي، تحكم بسيط في المخزون",
                })

        return recommendations

    def export_to_excel(self):
        """
        تصدير التحليل إلى Excel
        
        Returns:
            bytes: محتوى ملف Excel
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO

            wb = Workbook()
            ws = wb.active
            ws.title = "تحليل ABC"

            # تنسيق
            title_font = Font(size=16, bold=True)
            header_font = Font(bold=True, color="FFFFFF")
            center_alignment = Alignment(horizontal="center", vertical="center")

            # ألوان الفئات
            category_colors = {
                "A": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # أحمر
                "B": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),  # برتقالي
                "C": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # أصفر
            }

            # العنوان
            ws.merge_cells('A1:H1')
            ws['A1'] = f"تحليل ABC للمخزون - {self.analysis_date}"
            ws['A1'].font = title_font
            ws['A1'].alignment = center_alignment

            # رأس الجدول
            headers = ["كود الحساب", "اسم الحساب", "القيمة", "النسبة %", "النسبة التراكمية %", "عدد المعاملات", "معدل الدوران", "الفئة"]
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # البيانات
            result = self.classify_inventory()
            inventory_data = result["inventory_data"]

            row = 4
            for item in inventory_data:
                ws.cell(row=row, column=1, value=item["account_code"])
                ws.cell(row=row, column=2, value=item["account_name"])
                ws.cell(row=row, column=3, value=float(item["value"]))
                ws.cell(row=row, column=4, value=float(item["percentage"]))
                ws.cell(row=row, column=5, value=float(item["cumulative_percentage"]))
                ws.cell(row=row, column=6, value=item["transactions_count"])
                ws.cell(row=row, column=7, value=float(item["turnover"]))
                
                category_cell = ws.cell(row=row, column=8, value=item["category"])
                category_cell.fill = category_colors.get(item["category"], PatternFill())
                category_cell.font = Font(bold=True)
                category_cell.alignment = center_alignment
                
                row += 1

            # تعديل عرض الأعمدة
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 18
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 10

            # حفظ
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except ImportError:
            return None

    def get_complete_analysis(self):
        """
        الحصول على التحليل الكامل
        
        Returns:
            dict: التحليل الكامل
        """
        classification = self.classify_inventory()
        statistics = self.get_category_statistics()
        recommendations = self.get_recommendations()

        return {
            "inventory_data": classification["inventory_data"],
            "total_value": classification["total_value"],
            "statistics": statistics,
            "recommendations": recommendations,
            "analysis_date": self.analysis_date,
            "days_period": self.days_period,
            "date_from": self.date_from,
        }
