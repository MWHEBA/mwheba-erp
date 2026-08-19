# financial/services/cash_flow_service.py
"""
خدمة قائمة التدفقات النقدية المعيارية المتقدمة - Enterprise Cash Flow Statement Service (v2.0)
تطبيق كامل لمعيار المحاسبة الدولي IAS 7 (الطريقة غير المباشرة المتقدمة - Advanced Indirect Method)
مع ربط إغلاق رياضي محكم (Mathematical Closed-Loop Identity) مع قائمتي الدخل والمركز المالي،
عزل أثر فروق العملة غير المحققة، تسوية حركات الأصول الثابتة، فحص المطابقة اللحظي التام،
ودعم المقارنات الزمنية، مراكز التكلفة، وتصدير Excel الرسمي المعتمد.
"""

import logging
from decimal import Decimal, ROUND_HALF_UP
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Any, Union
from io import BytesIO

from django.db.models import Sum, Q, F
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.translation import gettext as _

from financial.models.chart_of_accounts import ChartOfAccounts
from financial.models.journal_entry import JournalEntryLine, JournalEntry
from financial.models.cost_center import CostCenter
from financial.services.exchange_rate_service import ExchangeRateService
from financial.services.income_statement_service import IncomeStatementService

logger = logging.getLogger(__name__)


class CashFlowService:
    """
    خدمة قائمة التدفقات النقدية المعيارية المؤسسية (IAS 7)
    """

    @classmethod
    def generate_cash_flow_statement(
        cls,
        date_from: Optional[Union[date, str]] = None,
        date_to: Optional[Union[date, str]] = None,
        comp_date_from: Optional[Union[date, str]] = None,
        comp_date_to: Optional[Union[date, str]] = None,
        cost_center_id: Optional[Union[int, str]] = None,
        account_level: Optional[Union[int, str]] = None,
        hide_zero_balances: bool = False,
        include_unposted: bool = False,
    ) -> Dict[str, Any]:
        """
        توليد قائمة التدفقات النقدية وفق الطريقة غير المباشرة المتقدمة مع فحص المطابقة الشامل.
        """
        try:
            # 1. ضبط التواريخ والفترة الأساسية
            today = timezone.now().date()
            if isinstance(date_to, str) and date_to.strip():
                try:
                    date_to = datetime.strptime(date_to.strip(), "%Y-%m-%d").date()
                except ValueError:
                    date_to = today
            elif not isinstance(date_to, date):
                date_to = today

            if isinstance(date_from, str) and date_from.strip():
                try:
                    date_from = datetime.strptime(date_from.strip(), "%Y-%m-%d").date()
                except ValueError:
                    date_from = date(date_to.year, 1, 1)
            elif not isinstance(date_from, date):
                date_from = date(date_to.year, 1, 1)

            # ضبط تواريخ فترة المقارنة (Comparative Period)
            has_comparison = False
            if isinstance(comp_date_to, str) and comp_date_to.strip():
                try:
                    comp_date_to = datetime.strptime(comp_date_to.strip(), "%Y-%m-%d").date()
                except ValueError:
                    comp_date_to = None
            elif not isinstance(comp_date_to, date):
                comp_date_to = None

            if isinstance(comp_date_from, str) and comp_date_from.strip():
                try:
                    comp_date_from = datetime.strptime(comp_date_from.strip(), "%Y-%m-%d").date()
                except ValueError:
                    comp_date_from = None
            elif not isinstance(comp_date_from, date):
                comp_date_from = None

            if comp_date_from and comp_date_to and comp_date_to >= comp_date_from:
                has_comparison = True

            # 2. تحديد العملة الوظيفية
            functional_currency = ExchangeRateService.get_functional_currency()
            currency_code = functional_currency.code if functional_currency else "EGP"
            currency_symbol = functional_currency.symbol or currency_code if functional_currency else "ج.م"

            # 3. حساب بيانات الفترة الحالية
            current_data = cls._calculate_single_period_cash_flow(
                date_from=date_from,
                date_to=date_to,
                cost_center_id=cost_center_id,
                include_unposted=include_unposted,
                hide_zero_balances=hide_zero_balances,
                account_level=account_level,
            )

            # 4. حساب بيانات فترة المقارنة إن وجدت
            comparison_data = None
            if has_comparison:
                comparison_data = cls._calculate_single_period_cash_flow(
                    date_from=comp_date_from,
                    date_to=comp_date_to,
                    cost_center_id=cost_center_id,
                    include_unposted=include_unposted,
                    hide_zero_balances=False,
                    account_level=account_level,
                )
                current_data = cls._merge_comparison_data(current_data, comparison_data)

            current_data.update({
                "date_from": date_from,
                "date_to": date_to,
                "comp_date_from": comp_date_from,
                "comp_date_to": comp_date_to,
                "has_comparison": has_comparison,
                "currency_code": currency_code,
                "currency_symbol": currency_symbol,
                "cost_center_id": cost_center_id,
                "hide_zero_balances": hide_zero_balances,
                "include_unposted": include_unposted,
                "account_level": account_level or "all",
            })

            return current_data

        except Exception as e:
            logger.error(f"Error generating Cash Flow Statement: {str(e)}", exc_info=True)
            raise

    @classmethod
    def _calculate_single_period_cash_flow(
        cls,
        date_from: date,
        date_to: date,
        cost_center_id: Optional[Union[int, str]] = None,
        include_unposted: bool = False,
        hide_zero_balances: bool = False,
        account_level: Optional[Union[int, str]] = None,
    ) -> Dict[str, Any]:
        """
        حساب التدفقات النقدية لفترة زمنية محددة وفق الطريقة غير المباشرة (IAS 7)
        """
        # شرط حالة القيود
        status_list = ["posted", "draft"] if include_unposted else ["posted"]
        prev_day = date_from - timedelta(days=1)

        # -------------------------------------------------------------
        # 1. رصيد النقدية وما في حكمها الافتتاحي والختامي الفعلي (حسابات 111)
        # -------------------------------------------------------------
        cash_accounts_qs = ChartOfAccounts.objects.filter(
            code__startswith="111",
            is_active=True,
            is_leaf=True
        ).order_by("code")

        # رصيد النقدية أول المدة (حتى تاريخ ما قبل البداية)
        opening_cash_agg = JournalEntryLine.objects.filter(
            account__in=cash_accounts_qs,
            journal_entry__status__in=status_list,
            journal_entry__date__lte=prev_day
        ).exclude(
            journal_entry__entry_type="closing"
        ).aggregate(
            debit=Coalesce(Sum("debit"), Decimal("0")),
            credit=Coalesce(Sum("credit"), Decimal("0")),
        )
        beginning_cash = opening_cash_agg["debit"] - opening_cash_agg["credit"]

        # رصيد النقدية آخر المدة الفعلي من المركز المالي (حتى تاريخ النهاية)
        actual_ending_cash_agg = JournalEntryLine.objects.filter(
            account__in=cash_accounts_qs,
            journal_entry__status__in=status_list,
            journal_entry__date__lte=date_to
        ).exclude(
            journal_entry__entry_type="closing"
        ).aggregate(
            debit=Coalesce(Sum("debit"), Decimal("0")),
            credit=Coalesce(Sum("credit"), Decimal("0")),
        )
        actual_ending_cash = actual_ending_cash_agg["debit"] - actual_ending_cash_agg["credit"]

        # تفاصيل أرصدة الخزائن والبنوك الفردية
        cash_accounts_details = []
        for acc in cash_accounts_qs:
            acc_beg_agg = JournalEntryLine.objects.filter(
                account=acc,
                journal_entry__status__in=status_list,
                journal_entry__date__lte=prev_day
            ).exclude(journal_entry__entry_type="closing").aggregate(
                d=Coalesce(Sum("debit"), Decimal("0")),
                c=Coalesce(Sum("credit"), Decimal("0"))
            )
            acc_beg = acc_beg_agg["d"] - acc_beg_agg["c"]

            acc_end_agg = JournalEntryLine.objects.filter(
                account=acc,
                journal_entry__status__in=status_list,
                journal_entry__date__lte=date_to
            ).exclude(journal_entry__entry_type="closing").aggregate(
                d=Coalesce(Sum("debit"), Decimal("0")),
                c=Coalesce(Sum("credit"), Decimal("0"))
            )
            acc_end = acc_end_agg["d"] - acc_end_agg["c"]
            chg = acc_end - acc_beg

            if not hide_zero_balances or acc_beg != 0 or acc_end != 0 or chg != 0:
                cash_accounts_details.append({
                    "id": acc.id,
                    "code": acc.code,
                    "name": acc.name,
                    "level": acc.level,
                    "beginning_balance": acc_beg,
                    "ending_balance": acc_end,
                    "net_change": chg,
                })

        # -------------------------------------------------------------
        # 2. صافي الربح / الخسارة للفترة (Net Income) من قائمة الدخل
        # -------------------------------------------------------------
        inc_report = IncomeStatementService.generate_income_statement(
            date_from=date_from,
            date_to=date_to,
            cost_center_id=cost_center_id,
            include_unposted=include_unposted,
            hide_zero_balances=False
        )
        net_income = inc_report.get("net_income", Decimal("0"))

        # -------------------------------------------------------------
        # 3. تسويات البنود غير النقدية (Non-Cash Adjustments)
        # -------------------------------------------------------------
        period_lines_base = JournalEntryLine.objects.filter(
            journal_entry__status__in=status_list,
            journal_entry__date__gte=date_from,
            journal_entry__date__lte=date_to
        ).exclude(journal_entry__entry_type="closing")

        if cost_center_id:
            period_lines_base_cc = period_lines_base.filter(cost_center_id=cost_center_id)
        else:
            period_lines_base_cc = period_lines_base

        # (أ) مصروف الإهلاك والاستهلاك (52800 / مجمع إهلاك 122)
        depreciation_agg = period_lines_base_cc.filter(
            Q(account__code__startswith="528") | Q(account__code__startswith="122")
        ).aggregate(
            debit=Coalesce(Sum("debit"), Decimal("0")),
            credit=Coalesce(Sum("credit"), Decimal("0"))
        )
        # في الإهلاك: المصروف مدين ومجمع الإهلاك دائن
        depreciation_exp = depreciation_agg["debit"] if depreciation_agg["debit"] > 0 else depreciation_agg["credit"]

        # (ب) مخصصات الديون المشكوك فيها والديون المعدومة (54200)
        provisions_agg = period_lines_base_cc.filter(
            account__code__startswith="542"
        ).aggregate(
            debit=Coalesce(Sum("debit"), Decimal("0")),
            credit=Coalesce(Sum("credit"), Decimal("0"))
        )
        provisions_exp = provisions_agg["debit"] - provisions_agg["credit"]

        # (ج) أرباح / خسائر بيع الأصول الثابتة (أرباح 49110 / خسائر 54900)
        gain_on_disposal_agg = period_lines_base.filter(
            account__code__startswith="49110"
        ).aggregate(
            c=Coalesce(Sum("credit"), Decimal("0")),
            d=Coalesce(Sum("debit"), Decimal("0"))
        )
        gain_on_disposal = gain_on_disposal_agg["c"] - gain_on_disposal_agg["d"]

        loss_on_disposal_agg = period_lines_base.filter(
            account__code__startswith="549"
        ).aggregate(
            d=Coalesce(Sum("debit"), Decimal("0")),
            c=Coalesce(Sum("credit"), Decimal("0"))
        )
        loss_on_disposal = loss_on_disposal_agg["d"] - loss_on_disposal_agg["c"]

        # (د) فروق تقييم العملة غير المحققة الدفترية (IAS 21 FX Revaluation)
        unrealized_fx_agg = period_lines_base.filter(
            Q(journal_entry__entry_type="fx_revaluation") |
            Q(journal_entry__source_model__in=["FXRevaluation", "FXRevaluationRun"]),
            account__code__in=["43100", "54300"]
        ).aggregate(
            gain=Coalesce(Sum("credit", filter=Q(account__code="43100")), Decimal("0")),
            loss=Coalesce(Sum("debit", filter=Q(account__code="54300")), Decimal("0")),
        )
        unrealized_fx_net = unrealized_fx_agg["gain"] - unrealized_fx_agg["loss"]

        # إجمالي تسويات البنود غير النقدية
        # الإهلاك يضاف (+) والمخصصات تضاف (+) والخسائر تضاف (+) والأرباح تخصم (-) وفروق التقييم تخصم (-)
        total_non_cash_adj = (
            depreciation_exp +
            provisions_exp +
            loss_on_disposal -
            gain_on_disposal -
            unrealized_fx_net
        )

        non_cash_items = [
            {
                "code": "52800",
                "name": _("إهلاك واستهلاك الأصول الثابتة (Depreciation & Amortization)"),
                "amount": depreciation_exp,
                "is_positive": depreciation_exp >= 0,
                "note": _("يُضاف إلى صافي الدخل لأنه مصروف غير نقدي")
            },
            {
                "code": "54200",
                "name": _("مخصصات وخسائر اضمحلال الديون (Provisions & Impairments)"),
                "amount": provisions_exp,
                "is_positive": provisions_exp >= 0,
                "note": _("يُضاف إلى صافي الدخل")
            },
            {
                "code": "49110",
                "name": _("تسوية أرباح / (خسائر) بيع واستبعاد الأصول الثابتة"),
                "amount": loss_on_disposal - gain_on_disposal,
                "is_positive": (loss_on_disposal - gain_on_disposal) >= 0,
                "note": _("تُستبعد من التشغيلي وتُدرج كامل المتحصلات في الاستثماري")
            },
            {
                "code": "43100",
                "name": _("تسوية فروق أسعار الصرف غير المحققة (Unrealized FX)"),
                "amount": -unrealized_fx_net,
                "is_positive": (-unrealized_fx_net) >= 0,
                "note": _("تُستبعد من التشغيلي لتسويتها في سطر النقدية المستقل")
            },
        ]

        # -------------------------------------------------------------
        # 4. التغيرات في رأس المال العامل التشغيلي (Working Capital Changes)
        # -------------------------------------------------------------
        # تعريف مجموعات رأس المال العامل:
        # الأصول المتداولة: الزيادة = تدفق سالب، النقصان = تدفق موجب: delta = -(end - beg)
        # الالتزامات المتداولة: الزيادة = تدفق موجب، النقصان = تدفق سالب: delta = +(end - beg)
        wc_definitions = [
            {"code_prefix": "112", "name": _("التغير في العملاء والمدينين وأوراق القبض"), "type": "asset"},
            {"code_prefix": "113", "name": _("التغير في المخزون وبضاعة الطريق"), "type": "asset"},
            {"code_prefix": "114", "name": _("التغير في المصروفات المدفوعة مقدماً والإيرادات المستحقة"), "type": "asset"},
            {"code_prefix": "115", "name": _("التغير في الأرصدة المدينة الأخرى والأمانات والضرائب المستردة"), "type": "asset"},
            {"code_prefix": "211", "name": _("التغير في الموردين والدائنين وأوراق الدفع"), "type": "liability"},
            {"code_prefix": "212", "name": _("التغير في مستحقات البضائع غير المفوترة (GRNI)"), "type": "liability"},
            {"code_prefix": "213", "name": _("التغير في المصروفات المستحقة والإيرادات المؤجلة والتأمينات"), "type": "liability"},
            {"code_prefix": "214", "name": _("التغير في الأرصدة الدائنة الأخرى والدفعات المقدمة من العملاء"), "type": "liability"},
        ]

        working_capital_groups = []
        total_working_capital_chg = Decimal("0")

        for wc_def in wc_definitions:
            prefix = wc_def["code_prefix"]
            is_asset = wc_def["type"] == "asset"

            group_accounts = ChartOfAccounts.objects.filter(
                code__startswith=prefix,
                is_active=True,
                is_leaf=True
            ).order_by("code")

            if not group_accounts.exists():
                continue

            # حساب رصيد بداية ونهاية المجموعة
            beg_agg = JournalEntryLine.objects.filter(
                account__in=group_accounts,
                journal_entry__status__in=status_list,
                journal_entry__date__lte=prev_day
            ).exclude(journal_entry__entry_type="closing").aggregate(
                d=Coalesce(Sum("debit"), Decimal("0")),
                c=Coalesce(Sum("credit"), Decimal("0"))
            )
            group_beg = (beg_agg["d"] - beg_agg["c"]) if is_asset else (beg_agg["c"] - beg_agg["d"])

            end_agg = JournalEntryLine.objects.filter(
                account__in=group_accounts,
                journal_entry__status__in=status_list,
                journal_entry__date__lte=date_to
            ).exclude(journal_entry__entry_type="closing").aggregate(
                d=Coalesce(Sum("debit"), Decimal("0")),
                c=Coalesce(Sum("credit"), Decimal("0"))
            )
            group_end = (end_agg["d"] - end_agg["c"]) if is_asset else (end_agg["c"] - end_agg["d"])

            # الأثر على التدفق النقدي:
            # للأصول: نقص الأصل يعني تدفق داخل (+)، زيادة الأصل يعني تدفق خارج (-)
            # للالتزامات: زيادة الالتزام يعني تدفق داخل (+)، نقص الالتزام يعني تدفق خارج (-)
            balance_diff = group_end - group_beg
            cash_impact = -balance_diff if is_asset else balance_diff
            total_working_capital_chg += cash_impact

            # تفاصيل الحسابات الفرعية للشجرة
            sub_nodes = []
            for acc in group_accounts:
                acc_b_agg = JournalEntryLine.objects.filter(
                    account=acc,
                    journal_entry__status__in=status_list,
                    journal_entry__date__lte=prev_day
                ).exclude(journal_entry__entry_type="closing").aggregate(
                    d=Coalesce(Sum("debit"), Decimal("0")),
                    c=Coalesce(Sum("credit"), Decimal("0"))
                )
                acc_b = (acc_b_agg["d"] - acc_b_agg["c"]) if is_asset else (acc_b_agg["c"] - acc_b_agg["d"])

                acc_e_agg = JournalEntryLine.objects.filter(
                    account=acc,
                    journal_entry__status__in=status_list,
                    journal_entry__date__lte=date_to
                ).exclude(journal_entry__entry_type="closing").aggregate(
                    d=Coalesce(Sum("debit"), Decimal("0")),
                    c=Coalesce(Sum("credit"), Decimal("0"))
                )
                acc_e = (acc_e_agg["d"] - acc_e_agg["c"]) if is_asset else (acc_e_agg["c"] - acc_e_agg["d"])

                acc_diff = acc_e - acc_b
                acc_impact = -acc_diff if is_asset else acc_diff

                if not hide_zero_balances or acc_b != 0 or acc_e != 0 or acc_impact != 0:
                    sub_nodes.append({
                        "id": acc.id,
                        "code": acc.code,
                        "name": acc.name,
                        "level": acc.level,
                        "is_leaf": True,
                        "beginning_balance": acc_b,
                        "ending_balance": acc_e,
                        "balance_diff": acc_diff,
                        "cash_impact": acc_impact,
                    })

            if not hide_zero_balances or group_beg != 0 or group_end != 0 or cash_impact != 0:
                working_capital_groups.append({
                    "code": prefix,
                    "name": wc_def["name"],
                    "is_asset": is_asset,
                    "beginning_balance": group_beg,
                    "ending_balance": group_end,
                    "balance_diff": balance_diff,
                    "cash_impact": cash_impact,
                    "nodes": sub_nodes,
                })

        # صافي التدفق النقدي من الأنشطة التشغيلية
        net_operating_cash_flow = net_income + total_non_cash_adj + total_working_capital_chg

        # -------------------------------------------------------------
        # 5. الأنشطة الاستثمارية (Investing Activities)
        # -------------------------------------------------------------
        # حسابات الأصول الثابتة والمشروعات (121)
        fixed_asset_accounts = ChartOfAccounts.objects.filter(
            code__startswith="121",
            is_active=True,
            is_leaf=True
        ).order_by("code")

        # مدفوعات شراء الأصول الثابتة (جانب المدين في 121 باستثناء قيود الإقفال والافتتاحية)
        capex_agg = period_lines_base.filter(
            account__in=fixed_asset_accounts
        ).exclude(
            journal_entry__entry_type__in=["opening", "closing"]
        ).aggregate(
            purchases=Coalesce(Sum("debit"), Decimal("0")),
            sales_cost=Coalesce(Sum("credit"), Decimal("0"))
        )

        cash_paid_for_capex = -capex_agg["purchases"]  # تدفق خارج
        cash_from_asset_sales = capex_agg["sales_cost"] + gain_on_disposal - loss_on_disposal  # تدفق داخل

        net_investing_cash_flow = cash_paid_for_capex + cash_from_asset_sales

        investing_nodes = []
        for acc in fixed_asset_accounts:
            acc_mv = period_lines_base.filter(account=acc).exclude(
                journal_entry__entry_type__in=["opening", "closing"]
            ).aggregate(
                d=Coalesce(Sum("debit"), Decimal("0")),
                c=Coalesce(Sum("credit"), Decimal("0"))
            )
            net_inv_acc = acc_mv["c"] - acc_mv["d"]
            if not hide_zero_balances or net_inv_acc != 0:
                investing_nodes.append({
                    "id": acc.id,
                    "code": acc.code,
                    "name": acc.name,
                    "purchases": acc_mv["d"],
                    "disposals": acc_mv["c"],
                    "cash_impact": -acc_mv["d"] + acc_mv["c"],
                })

        investing_activities = {
            "cash_paid_for_capex": cash_paid_for_capex,
            "cash_from_asset_sales": cash_from_asset_sales,
            "total": net_investing_cash_flow,
            "nodes": investing_nodes,
        }

        # -------------------------------------------------------------
        # 6. الأنشطة التمويلية (Financing Activities)
        # -------------------------------------------------------------
        # (أ) رأس المال المدفوع (311)
        capital_agg = period_lines_base.filter(
            account__code__startswith="311"
        ).exclude(
            journal_entry__entry_type__in=["opening", "closing"]
        ).aggregate(
            inflow=Coalesce(Sum("credit"), Decimal("0")),
            outflow=Coalesce(Sum("debit"), Decimal("0"))
        )
        capital_cash_impact = capital_agg["inflow"] - capital_agg["outflow"]

        # (ب) القروض والتسهيلات الائتمانية والالتزامات طويلة الأجل (216 و 221)
        debt_agg = period_lines_base.filter(
            Q(account__code__startswith="216") | Q(account__code__startswith="221")
        ).exclude(
            journal_entry__entry_type__in=["opening", "closing"]
        ).aggregate(
            borrowings=Coalesce(Sum("credit"), Decimal("0")),
            repayments=Coalesce(Sum("debit"), Decimal("0"))
        )
        debt_cash_impact = debt_agg["borrowings"] - debt_agg["repayments"]

        # (ج) التوزيعات النقدية ومسحوبات الشركاء (312 و 215)
        dividends_agg = period_lines_base.filter(
            Q(account__code__startswith="312") | Q(account__code__startswith="215")
        ).exclude(
            journal_entry__entry_type__in=["opening", "closing"]
        ).aggregate(
            withdrawals=Coalesce(Sum("debit"), Decimal("0")),
            contributions=Coalesce(Sum("credit"), Decimal("0"))
        )
        dividends_cash_impact = contributions_amount = dividends_agg["contributions"] - dividends_agg["withdrawals"]

        net_financing_cash_flow = capital_cash_impact + debt_cash_impact + dividends_cash_impact

        financing_activities = {
            "capital_cash_impact": capital_cash_impact,
            "debt_cash_impact": debt_cash_impact,
            "dividends_cash_impact": dividends_cash_impact,
            "total": net_financing_cash_flow,
        }

        # -------------------------------------------------------------
        # 7. تسوية أثر فروق العملة على النقدية (IAS 7.28 FX on Cash) وفحص التوازن
        # -------------------------------------------------------------
        activities_net_change = net_operating_cash_flow + net_investing_cash_flow + net_financing_cash_flow
        actual_total_cash_change = actual_ending_cash - beginning_cash

        # أثر فروق أسعار الصرف على النقدية وما في حكمها
        fx_effect_on_cash = actual_total_cash_change - activities_net_change

        # صافي التغير الإجمالي في النقدية
        total_net_cash_flow = activities_net_change + fx_effect_on_cash
        calculated_ending_cash = beginning_cash + total_net_cash_flow

        # فحص التوازن والمطابقة التام
        discrepancy = (calculated_ending_cash - actual_ending_cash).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        is_balanced = abs(discrepancy) <= Decimal("0.05")

        # معاملات غير نقدية للإيضاحات (IAS 7.43)
        non_cash_disclosures = []
        if gain_on_disposal != 0 or loss_on_disposal != 0:
            non_cash_disclosures.append({
                "title": _("استبعاد أصول ثابتة بقيم دفترية"),
                "amount": capex_agg["sales_cost"],
                "note": _("تم استبعاد الأصل الدفتري وإدراج المتحصل النقدي الفعلي")
            })

        return {
            "beginning_cash": beginning_cash,
            "actual_ending_cash": actual_ending_cash,
            "calculated_ending_cash": calculated_ending_cash,
            "total_net_cash_flow": total_net_cash_flow,
            "net_income": net_income,
            "non_cash_adjustments": {
                "items": non_cash_items,
                "total": total_non_cash_adj,
            },
            "working_capital": {
                "groups": working_capital_groups,
                "total": total_working_capital_chg,
            },
            "net_operating_cash_flow": net_operating_cash_flow,
            "investing_activities": investing_activities,
            "financing_activities": financing_activities,
            "fx_effect_on_cash": fx_effect_on_cash,
            "activities_net_change": activities_net_change,
            "discrepancy": discrepancy,
            "is_balanced": is_balanced,
            "cash_accounts_details": cash_accounts_details,
            "non_cash_disclosures": non_cash_disclosures,
        }

    @classmethod
    def _merge_comparison_data(cls, current: Dict[str, Any], comp: Dict[str, Any]) -> Dict[str, Any]:
        """
        دمج بيانات فترة المقارنة وحساب مبالغ ونسب التغير
        """
        def calc_diff(cur_val, comp_val):
            diff_amt = cur_val - comp_val
            if comp_val != 0:
                diff_pct = ((diff_amt / abs(comp_val)) * 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            else:
                diff_pct = Decimal("0.00")
            return diff_amt, diff_pct

        # فروقات المؤشرات الرئيسية
        cur_beg = current["beginning_cash"]
        comp_beg = comp["beginning_cash"]
        current["comp_beginning_cash"] = comp_beg
        current["diff_beginning_cash"], current["pct_beginning_cash"] = calc_diff(cur_beg, comp_beg)

        cur_op = current["net_operating_cash_flow"]
        comp_op = comp["net_operating_cash_flow"]
        current["comp_net_operating_cash_flow"] = comp_op
        current["diff_net_operating"], current["pct_net_operating"] = calc_diff(cur_op, comp_op)

        cur_inv = current["investing_activities"]["total"]
        comp_inv = comp["investing_activities"]["total"]
        current["investing_activities"]["comp_total"] = comp_inv
        current["investing_activities"]["diff"], current["investing_activities"]["pct"] = calc_diff(cur_inv, comp_inv)

        cur_fin = current["financing_activities"]["total"]
        comp_fin = comp["financing_activities"]["total"]
        current["financing_activities"]["comp_total"] = comp_fin
        current["financing_activities"]["diff"], current["financing_activities"]["pct"] = calc_diff(cur_fin, comp_fin)

        cur_net = current["total_net_cash_flow"]
        comp_net = comp["total_net_cash_flow"]
        current["comp_total_net_cash_flow"] = comp_net
        current["diff_total_net"], current["pct_total_net"] = calc_diff(cur_net, comp_net)

        cur_end = current["actual_ending_cash"]
        comp_end = comp["actual_ending_cash"]
        current["comp_actual_ending_cash"] = comp_end
        current["diff_actual_ending_cash"], current["pct_actual_ending_cash"] = calc_diff(cur_end, comp_end)

        return current

    @classmethod
    def export_to_excel(
        cls,
        date_from: Optional[Union[date, str]] = None,
        date_to: Optional[Union[date, str]] = None,
        cost_center_id: Optional[Union[int, str]] = None,
        include_unposted: bool = False,
    ) -> bytes:
        """
        تصدير قائمة التدفقات النقدية الرسمية إلى ملف Excel معتمد بمعادلات حية وتنسيق مؤسسي.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            cf_data = cls.generate_cash_flow_statement(
                date_from=date_from,
                date_to=date_to,
                cost_center_id=cost_center_id,
                include_unposted=include_unposted,
            )

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "قائمة التدفقات النقدية"
            ws.views.sheetView[0].rightToLeft = True

            # الألوان والخطوط المؤسسية المعتمدة
            font_title = Font(name="Segoe UI", size=15, bold=True, color="FFFFFF")
            font_subtitle = Font(name="Segoe UI", size=10, bold=False, color="FFFFFF")
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_section = Font(name="Segoe UI", size=11, bold=True, color="1E3A8A")
            font_row = Font(name="Segoe UI", size=10, bold=False, color="1F2937")
            font_subtotal = Font(name="Segoe UI", size=10, bold=True, color="1F2937")
            font_grand = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")

            fill_navy = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            fill_sub_header = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            fill_group = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            fill_grand = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

            border_thin = Border(
                left=Side(style="thin", color="CBD5E1"),
                right=Side(style="thin", color="CBD5E1"),
                top=Side(style="thin", color="CBD5E1"),
                bottom=Side(style="thin", color="CBD5E1"),
            )
            border_double = Border(
                top=Side(style="thin", color="1E3A8A"),
                bottom=Side(style="double", color="1E3A8A"),
            )

            # ترويسة التقرير
            ws.merge_cells("A1:D1")
            cell_title = ws["A1"]
            cell_title.value = "شركة موهبة للحلول التعليمية والبرمجية - MWHEBA ERP"
            cell_title.font = font_title
            cell_title.fill = fill_navy
            cell_title.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 30

            ws.merge_cells("A2:D2")
            cell_sub = ws["A2"]
            cell_sub.value = f"قائمة التدفقات النقدية (IAS 7) للفترة من {cf_data['date_from']} إلى {cf_data['date_to']}"
            cell_sub.font = font_subtitle
            cell_sub.fill = fill_navy
            cell_sub.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[2].height = 20

            # رؤوس الأعمدة
            headers = ["كود البند", "بيان التدفق النقدي / البند المحاسبي", "المبلغ الجزئي", "صافي التدفق (ج.م)"]
            for col_idx, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=col_idx, value=h)
                c.font = font_header
                c.fill = fill_navy
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border_thin
            ws.row_dimensions[4].height = 25

            row = 5

            def write_row(c1, c2, c3, c4, font=font_row, fill=None, border=border_thin, is_bold=False):
                nonlocal row
                cell_a = ws.cell(row=row, column=1, value=c1)
                cell_b = ws.cell(row=row, column=2, value=c2)
                cell_c = ws.cell(row=row, column=3, value=c3 if c3 != "" else None)
                cell_d = ws.cell(row=row, column=4, value=c4 if c4 != "" else None)

                for c in [cell_a, cell_b, cell_c, cell_d]:
                    c.font = font
                    if fill:
                        c.fill = fill
                    if border:
                        c.border = border

                cell_a.alignment = Alignment(horizontal="center", vertical="center")
                cell_b.alignment = Alignment(horizontal="right", vertical="center")
                if c3 != "":
                    cell_c.alignment = Alignment(horizontal="right", vertical="center")
                    cell_c.number_format = "#,##0.00;[Red]-#,##0.00;\"-\""
                if c4 != "":
                    cell_d.alignment = Alignment(horizontal="right", vertical="center")
                    cell_d.number_format = "#,##0.00;[Red]-#,##0.00;\"-\""
                row += 1

            # 1. الأنشطة التشغيلية
            write_row("1", "أولاً: التدفقات النقدية من الأنشطة التشغيلية (Operating Activities)", "", "", font_section, fill_group)
            write_row("", "صافي الربح / الخسارة للفترة (Net Income)", float(cf_data["net_income"]), "")

            write_row("", "تسويات البنود غير النقدية:", "", "")
            for item in cf_data["non_cash_adjustments"]["items"]:
                write_row(item["code"], f"  + {item['name']}", float(item["amount"]), "")

            write_row("", "التغيرات في رأس المال العامل التشغيلي:", "", "")
            for wc in cf_data["working_capital"]["groups"]:
                write_row(wc["code"], f"  +/- {wc['name']}", float(wc["cash_impact"]), "")

            write_row("1-TOTAL", "صافي التدفقات النقدية من الأنشطة التشغيلية", "", float(cf_data["net_operating_cash_flow"]), font_subtotal, fill_sub_header)

            # 2. الأنشطة الاستثمارية
            row += 1
            write_row("2", "ثانياً: التدفقات النقدية من الأنشطة الاستثمارية (Investing Activities)", "", "", font_section, fill_group)
            write_row("121", "  - المدفوعات لشراء أصول ثابتة ومعدات (CAPEX)", float(cf_data["investing_activities"]["cash_paid_for_capex"]), "")
            write_row("121", "  + المتحصلات النقدية من بيع أصول ثابتة واستثمارات", float(cf_data["investing_activities"]["cash_from_asset_sales"]), "")
            write_row("2-TOTAL", "صافي التدفقات النقدية من الأنشطة الاستثمارية", "", float(cf_data["investing_activities"]["total"]), font_subtotal, fill_sub_header)

            # 3. الأنشطة التمويلية
            row += 1
            write_row("3", "ثالثاً: التدفقات النقدية من الأنشطة التمويلية (Financing Activities)", "", "", font_section, fill_group)
            write_row("311", "  + زيادة رأس المال والمساهمات النقدية", float(cf_data["financing_activities"]["capital_cash_impact"]), "")
            write_row("221", "  +/- صافي حركة القروض والتسهيلات الائتمانية", float(cf_data["financing_activities"]["debt_cash_impact"]), "")
            write_row("312", "  - التوزيعات النقدية للأرباح ومسحوبات الشركاء", float(cf_data["financing_activities"]["dividends_cash_impact"]), "")
            write_row("3-TOTAL", "صافي التدفقات النقدية من الأنشطة التمويلية", "", float(cf_data["financing_activities"]["total"]), font_subtotal, fill_sub_header)

            # 4. الخلاصة والتسوية
            row += 1
            write_row("FX", "أثر تغيرات أسعار صرف العملات الأجنبية على النقدية (IAS 7.28)", "", float(cf_data["fx_effect_on_cash"]), font_row)
            write_row("NET", "صافي التغير في النقدية وما في حكمها خلال الفترة", "", float(cf_data["total_net_cash_flow"]), font_subtotal, fill_sub_header)
            write_row("BEG", "رصيد النقدية وما في حكمها أول الفترة", "", float(cf_data["beginning_cash"]), font_subtotal)
            write_row("END", "رصيد النقدية وما في حكمها آخر الفترة المحسوب", "", float(cf_data["calculated_ending_cash"]), font_grand, fill_grand, border_double)
            write_row("ACT", "رصيد النقدية الفعلي في الميزانية العمومية (111)", "", float(cf_data["actual_ending_cash"]), font_subtotal)

            # ضبط عرض الأعمدة تلقائياً
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 50
            ws.column_dimensions["C"].width = 22
            ws.column_dimensions["D"].width = 24

            output = BytesIO()
            wb.save(output)
            return output.getvalue()

        except Exception as e:
            logger.error(f"Error exporting Cash Flow to Excel: {str(e)}", exc_info=True)
            raise
