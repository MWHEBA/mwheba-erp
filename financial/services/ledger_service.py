# financial/services/ledger_service.py
"""
خدمة دفتر الأستاذ - Ledger Service
توفر جميع العمليات المتعلقة بدفتر الأستاذ بشكل احترافي وديناميكي
"""

from django.db.models import Sum, Q, F, Value, CharField
from django.db.models.functions import Coalesce
from django.utils import timezone
from decimal import Decimal
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Tuple
import logging

from ..models import (
    ChartOfAccounts,
    JournalEntry,
    JournalEntryLine,
    AccountType,
)

logger = logging.getLogger(__name__)


class LedgerService:
    """
    خدمة دفتر الأستاذ - توفر جميع العمليات المتعلقة بالحسابات
    """

    @staticmethod
    def get_opening_balance(
        account: ChartOfAccounts,
        as_of_date: date
    ) -> Decimal:
        """
        حساب الرصيد الافتتاحي للحساب قبل تاريخ معين عبر LedgerQueryService
        """
        try:
            from .ledger_query_service import LedgerQueryService
            res = LedgerQueryService.get_account_balance(
                account_or_id=account,
                as_of_date=as_of_date - timedelta(days=1)
            )
            return res.get('balance', Decimal('0.00'))
        except Exception as e:
            logger.error(f"خطأ في حساب الرصيد الافتتاحي للحساب {account.code}: {e}")
            return Decimal('0.00')

    @staticmethod
    def get_account_transactions(
        account: ChartOfAccounts,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        include_unposted: bool = False
    ) -> List[Dict]:
        """
        جلب جميع معاملات الحساب مع حساب الرصيد التراكمي عبر LedgerQueryService
        """
        try:
            from .ledger_query_service import LedgerQueryService
            statement = LedgerQueryService.get_account_statement(
                account_or_id=account,
                start_date=date_from,
                end_date=date_to,
                include_unposted=include_unposted
            )
            return statement.get('transactions', [])
        except Exception as e:
            logger.error(f"خطأ في جلب معاملات الحساب {account.code}: {e}")
            return []

    @staticmethod
    def get_ledger_report(
        account_id: int,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None
    ) -> Dict:
        """
        توليد تقرير دفتر الأستاذ لحساب معين عبر LedgerQueryService
        """
        try:
            account = ChartOfAccounts.objects.get(id=account_id)
            from .ledger_query_service import LedgerQueryService
            statement = LedgerQueryService.get_account_statement(
                account_or_id=account,
                start_date=start_date,
                end_date=end_date
            )
            
            return {
                'account': account,
                'transactions': statement.get('transactions', []),
                'summary': statement,
                'start_date': start_date,
                'end_date': end_date,
            }
            
        except ChartOfAccounts.DoesNotExist:
            logger.error(f"الحساب {account_id} غير موجود")
            return {
                'error': 'الحساب غير موجود',
                'transactions': [],
                'summary': {}
            }
        except Exception as e:
            logger.error(f"خطأ في توليد تقرير دفتر الأستاذ: {e}")
            return {
                'error': str(e),
                'transactions': [],
                'summary': {}
            }

    @staticmethod
    def get_account_summary(
        account: ChartOfAccounts,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None
    ) -> Dict:
        """
        حساب ملخص الحساب (مدين، دائن، رصيد)
        
        Args:
            account: الحساب
            date_from: من تاريخ (اختياري)
            date_to: إلى تاريخ (اختياري)
            
        Returns:
            ملخص الحساب
        """
        try:
            # بناء الاستعلام
            query = Q(account=account, journal_entry__status='posted')
            
            if date_from:
                query &= Q(journal_entry__date__gte=date_from)
            if date_to:
                query &= Q(journal_entry__date__lte=date_to)
            
            # حساب المجاميع
            totals = JournalEntryLine.objects.filter(query).aggregate(
                total_debit=Coalesce(Sum('debit'), Decimal('0')),
                total_credit=Coalesce(Sum('credit'), Decimal('0')),
                count=Coalesce(Sum(Value(1)), 0)
            )
            
            total_debit = totals['total_debit']
            total_credit = totals['total_credit']
            transaction_count = totals['count']
            
            # حساب الرصيد الافتتاحي
            opening_balance = Decimal('0')
            if date_from:
                opening_balance = LedgerService.get_opening_balance(account, date_from)
            
            # حساب الرصيد الختامي
            if account.account_type.nature == 'debit':
                period_movement = total_debit - total_credit
                closing_balance = opening_balance + period_movement
            else:
                period_movement = total_credit - total_debit
                closing_balance = opening_balance + period_movement
            
            return {
                'account': account,
                'opening_balance': opening_balance,
                'total_debit': total_debit,
                'total_credit': total_credit,
                'period_movement': period_movement,
                'closing_balance': closing_balance,
                'transaction_count': transaction_count,
                'account_nature': account.account_type.nature,
            }
            
        except Exception as e:
            logger.error(f"خطأ في حساب ملخص الحساب {account.code}: {e}")
            return {
                'account': account,
                'opening_balance': Decimal('0'),
                'total_debit': Decimal('0'),
                'total_credit': Decimal('0'),
                'period_movement': Decimal('0'),
                'closing_balance': Decimal('0'),
                'transaction_count': 0,
                'account_nature': 'debit',
                'error': str(e)
            }

    @staticmethod
    def get_all_accounts_summary(
        date_from: Optional[date] = None,
        date_to: Optional[date] = None,
        account_type: Optional[str] = None,
        only_active: bool = True
    ) -> List[Dict]:
        """
        حساب ملخص جميع الحسابات
        
        Args:
            date_from: من تاريخ (اختياري)
            date_to: إلى تاريخ (اختياري)
            account_type: نوع الحساب (اختياري)
            only_active: فقط الحسابات النشطة؟
            
        Returns:
            قائمة ملخصات الحسابات
        """
        try:
            # جلب الحسابات
            accounts_query = ChartOfAccounts.objects.filter(
                is_leaf=True  # فقط الحسابات الفرعية (ليس المجموعات)
            ).select_related('account_type')
            
            if only_active:
                accounts_query = accounts_query.filter(is_active=True)
            
            if account_type:
                accounts_query = accounts_query.filter(
                    account_type__category=account_type
                )
            
            accounts_query = accounts_query.order_by(
                'account_type__category',
                'code'
            )
            
            # حساب ملخص كل حساب
            summaries = []
            for account in accounts_query:
                summary = LedgerService.get_account_summary(
                    account,
                    date_from,
                    date_to
                )
                
                # فقط الحسابات التي لها حركة
                if only_active and summary['transaction_count'] == 0:
                    continue
                
                summaries.append(summary)
            
            return summaries
            
        except Exception as e:
            logger.error(f"خطأ في حساب ملخص جميع الحسابات: {e}")
            return []

    @staticmethod
    def export_to_excel(
        account: Optional[ChartOfAccounts] = None,
        date_from: Optional[date] = None,
        date_to: Optional[date] = None
    ) -> bytes:
        """
        تصدير دفتر الأستاذ إلى Excel
        
        Args:
            account: الحساب (اختياري - إذا لم يحدد يصدر كل الحسابات)
            date_from: من تاريخ
            date_to: إلى تاريخ
            
        Returns:
            ملف Excel كـ bytes
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from io import BytesIO
            
            # إنشاء workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "كشف حساب"
            
            # تنسيق العنوان
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            if account:
                # تصدير حساب واحد
                ws['A1'] = f"كشف حساب - {account.name}"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:G1')
                
                # معلومات الحساب
                ws['A2'] = f"الكود: {account.code}"
                ws['C2'] = f"النوع: {account.account_type.name}"
                ws['E2'] = f"الفترة: {date_from or 'البداية'} - {date_to or 'النهاية'}"
                
                # العناوين
                headers = ['التاريخ', 'النوع', 'البيان', 'المرجع', 'مدين', 'دائن', 'الرصيد']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=4, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                
                # البيانات
                transactions = LedgerService.get_account_transactions(account, date_from, date_to)
                summary = LedgerService.get_account_summary(account, date_from, date_to)
                
                # الرصيد الافتتاحي
                row = 5
                ws.cell(row=row, column=1, value="الرصيد الافتتاحي")
                ws.cell(row=row, column=7, value=float(summary['opening_balance']))
                ws.cell(row=row, column=7).font = Font(bold=True)
                
                # المعاملات
                for trans in transactions:
                    row += 1
                    ws.cell(row=row, column=1, value=trans['date'].strftime('%Y-%m-%d'))
                    ws.cell(row=row, column=2, value=trans.get('entry_type_display') or trans.get('journal_number', ''))
                    ws.cell(row=row, column=3, value=trans['description'])
                    ws.cell(row=row, column=4, value=trans['reference'])
                    ws.cell(row=row, column=5, value=float(trans['debit']))
                    ws.cell(row=row, column=6, value=float(trans['credit']))
                    ws.cell(row=row, column=7, value=float(trans['balance']))
                
                # الإجمالي
                row += 1
                ws.cell(row=row, column=1, value="الإجمالي").font = Font(bold=True)
                ws.cell(row=row, column=5, value=float(summary['total_debit'])).font = Font(bold=True)
                ws.cell(row=row, column=6, value=float(summary['total_credit'])).font = Font(bold=True)
                ws.cell(row=row, column=7, value=float(summary['closing_balance'])).font = Font(bold=True)
                
            else:
                # تصدير جميع الحسابات
                ws['A1'] = "كشف حركة وأرصدة الحسابات العامة"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:F1')
                
                # العناوين
                headers = ['الكود', 'الحساب', 'النوع', 'مدين', 'دائن', 'الرصيد']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                
                # البيانات
                summaries = LedgerService.get_all_accounts_summary(date_from, date_to)
                
                row = 4
                for summary in summaries:
                    ws.cell(row=row, column=1, value=summary['account'].code)
                    ws.cell(row=row, column=2, value=summary['account'].name)
                    ws.cell(row=row, column=3, value=summary['account'].account_type.name)
                    ws.cell(row=row, column=4, value=float(summary['total_debit']))
                    ws.cell(row=row, column=5, value=float(summary['total_credit']))
                    ws.cell(row=row, column=6, value=float(summary['closing_balance']))
                    row += 1
            
            # ضبط عرض الأعمدة
            from openpyxl.utils import get_column_letter
            for idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(idx)
                for cell in ws[column_letter]:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # حفظ في BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"خطأ في تصدير Excel: {e}")
            raise
