# financial/services/balance_sheet_service.py
"""
خدمة الميزانية العمومية المعيارية - Enterprise Balance Sheet Service (v2.0)
تطبيق كامل لمعايير المحاسبة الدولية (IAS 1 / IAS 21) مع التصنيف الخماسي المعياري،
عزل السنوات المالية، منع ازدواجية الرصيد الافتتاحي، احتساب النسب المالية المحمية،
التجميع الشجري O(N)، ودعم المقارنة الزمنية وتصدير Excel الرسمي المعتمد.
"""

import logging
from decimal import Decimal
from datetime import date, datetime
from typing import Dict, List, Optional, Any, Union
from io import BytesIO

from django.db.models import Sum, Q
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.translation import gettext as _

from financial.models.chart_of_accounts import ChartOfAccounts
from financial.models.journal_entry import JournalEntryLine, AccountingPeriod
from financial.models.fiscal_year import FiscalYear
from financial.services.exchange_rate_service import ExchangeRateService
from financial.services.role_registry import AccountRoleRegistry, AccountRoleNames

logger = logging.getLogger(__name__)


class BalanceSheetService:
    """
    خدمة الميزانية العمومية والمركز المالي المؤسسي
    """

    @classmethod
    def generate_balance_sheet(
        cls,
        as_of_date: Optional[Union[date, str]] = None,
        comparison_date: Optional[Union[date, str]] = None,
        account_level: Optional[Union[int, str]] = None,
        hide_zero_balances: bool = False,
        group_by_subtype: bool = True,
        fiscal_year_id: Optional[int] = None,
    ) -> Dict[str, Any]:
        """
        إنشاء الميزانية العمومية المعيارية الكاملة طبقاً لمعيار IAS 1
        """
        try:
            # 1. تحويل وضبط التواريخ
            if isinstance(as_of_date, str):
                try:
                    as_of_date = datetime.strptime(as_of_date, "%Y-%m-%d").date()
                except ValueError:
                    as_of_date = timezone.now().date()
            elif as_of_date is None:
                as_of_date = timezone.now().date()

            if isinstance(comparison_date, str) and comparison_date.strip():
                try:
                    comparison_date = datetime.strptime(comparison_date.strip(), "%Y-%m-%d").date()
                except ValueError:
                    comparison_date = None
            elif not isinstance(comparison_date, date):
                comparison_date = None

            # 2. تحديد العملة الوظيفية
            functional_currency = ExchangeRateService.get_functional_currency()
            currency_code = functional_currency.code if functional_currency else "EGP"
            currency_symbol = functional_currency.symbol or currency_code if functional_currency else "ج.م"

            # 3. تحديد السنة المالية وتاريخ بدايتها
            current_fiscal_year = None
            if fiscal_year_id:
                current_fiscal_year = FiscalYear.objects.filter(id=fiscal_year_id).first()
            if not current_fiscal_year:
                current_fiscal_year = FiscalYear.objects.filter(
                    start_date__lte=as_of_date,
                    end_date__gte=as_of_date
                ).first()

            fy_start_date = current_fiscal_year.start_date if current_fiscal_year else date(as_of_date.year, 1, 1)

            # فحص حالة الفترة المحاسبية (مقفلة/مفتوحة)
            period = AccountingPeriod.objects.filter(
                start_date__lte=as_of_date,
                end_date__gte=as_of_date
            ).first()
            is_period_closed = bool(period and getattr(period, 'is_closed', False))

            # 4. الاستعلام الأول: جلب شجرة الحسابات بالكامل مع الأنواع والآباء
            accounts_qs = ChartOfAccounts.objects.filter(
                Q(is_active=True) | Q(journal_lines__journal_entry__status='posted') | Q(opening_balance__gt=0)
            ).distinct().select_related('account_type', 'parent', 'currency').order_by('code')

            accounts_list = list(accounts_qs)
            if not accounts_list:
                return cls._empty_balance_sheet_response(as_of_date, comparison_date, currency_code, currency_symbol)

            acc_map = {acc.id: acc for acc in accounts_list}

            # 5. الاستعلام الثاني: تجميع حركات القيود المرحلة حتى as_of_date
            as_of_query = Q(
                journal_entry__status='posted',
                journal_entry__date__lte=as_of_date
            )
            as_of_totals = (
                JournalEntryLine.objects.filter(as_of_query)
                .values('account_id')
                .annotate(
                    sum_debit=Coalesce(Sum('debit'), Decimal('0.00')),
                    sum_credit=Coalesce(Sum('credit'), Decimal('0.00'))
                )
            )
            as_of_map = {row['account_id']: row for row in as_of_totals}

            # استعلام المقارنة إن وجد
            comp_map = {}
            if comparison_date:
                comp_query = Q(
                    journal_entry__status='posted',
                    journal_entry__date__lte=comparison_date
                )
                comp_totals = (
                    JournalEntryLine.objects.filter(comp_query)
                    .values('account_id')
                    .annotate(
                        sum_debit=Coalesce(Sum('debit'), Decimal('0.00')),
                        sum_credit=Coalesce(Sum('credit'), Decimal('0.00'))
                    )
                )
                comp_map = {row['account_id']: row for row in comp_totals}

            # فحص القيود الافتتاحية المسجلة لمنع ازدواجية حقل opening_balance
            opening_entries_account_ids = set(
                JournalEntryLine.objects.filter(
                    journal_entry__status='posted',
                    journal_entry__entry_type='opening'
                ).values_list('account_id', flat=True).distinct()
            )

            # 6. تسوية أرباح/خسائر السنوات السابقة غير المقفلة تلقائياً (Implicit Retained Earnings)
            net_unclosed_prior_pl = Decimal('0.00')
            if fy_start_date:
                unclosed_pl = JournalEntryLine.objects.filter(
                    journal_entry__status='posted',
                    journal_entry__date__lt=fy_start_date,
                    account__account_type__category__in=['revenue', 'expense']
                ).aggregate(
                    sum_dr=Coalesce(Sum('debit'), Decimal('0.00')),
                    sum_cr=Coalesce(Sum('credit'), Decimal('0.00'))
                )
                net_unclosed_prior_pl = (unclosed_pl['sum_cr'] or Decimal('0.00')) - (unclosed_pl['sum_dr'] or Decimal('0.00'))

            # تسوية المقارنة للسنوات السابقة غير المقفلة
            comp_unclosed_prior_pl = Decimal('0.00')
            if comparison_date:
                comp_fy = FiscalYear.objects.filter(start_date__lte=comparison_date, end_date__gte=comparison_date).first()
                comp_fy_start = comp_fy.start_date if comp_fy else date(comparison_date.year, 1, 1)
                if comp_fy_start:
                    comp_unclosed = JournalEntryLine.objects.filter(
                        journal_entry__status='posted',
                        journal_entry__date__lt=comp_fy_start,
                        account__account_type__category__in=['revenue', 'expense']
                    ).aggregate(
                        sum_dr=Coalesce(Sum('debit'), Decimal('0.00')),
                        sum_cr=Coalesce(Sum('credit'), Decimal('0.00'))
                    )
                    comp_unclosed_prior_pl = (comp_unclosed['sum_cr'] or Decimal('0.00')) - (comp_unclosed['sum_dr'] or Decimal('0.00'))

            # 7. حساب صافي رصيد كل حساب نهائي مباشر (Direct Balances)
            direct_balances = {}
            comp_direct_balances = {}

            for acc in accounts_list:
                row = as_of_map.get(acc.id, {})
                dr = row.get('sum_debit') or Decimal('0.00')
                cr = row.get('sum_credit') or Decimal('0.00')

                # إضافة الرصيد الافتتاحي من الحساب كـ Fallback فقط إذا لم يكن له قيد افتتاحي
                if acc.id not in opening_entries_account_ids and acc.opening_balance:
                    if acc.account_type.nature == 'debit':
                        dr += acc.opening_balance
                    else:
                        cr += acc.opening_balance

                # احتساب الرصيد حسب طبيعة الحساب
                if acc.account_type.nature == 'debit':
                    bal = dr - cr
                else:
                    bal = cr - dr

                direct_balances[acc.id] = bal

                # حساب رصيد المقارنة
                if comparison_date:
                    c_row = comp_map.get(acc.id, {})
                    c_dr = c_row.get('sum_debit') or Decimal('0.00')
                    c_cr = c_row.get('sum_credit') or Decimal('0.00')
                    if acc.id not in opening_entries_account_ids and acc.opening_balance:
                        if acc.account_type.nature == 'debit':
                            c_dr += acc.opening_balance
                        else:
                            c_cr += acc.opening_balance
                    if acc.account_type.nature == 'debit':
                        c_bal = c_dr - c_cr
                    else:
                        c_bal = c_cr - c_dr
                    comp_direct_balances[acc.id] = c_bal
                else:
                    comp_direct_balances[acc.id] = Decimal('0.00')

            # 8. حساب صافي ربح/خسارة الفترة الحالية (Current Period Net Income)
            # الإيرادات والمصروفات بين fy_start_date و as_of_date
            current_period_pl_query = Q(
                journal_entry__status='posted',
                journal_entry__date__gte=fy_start_date,
                journal_entry__date__lte=as_of_date,
                account__account_type__category__in=['revenue', 'expense']
            )
            cur_pl_totals = JournalEntryLine.objects.filter(current_period_pl_query).aggregate(
                sum_dr=Coalesce(Sum('debit'), Decimal('0.00')),
                sum_cr=Coalesce(Sum('credit'), Decimal('0.00'))
            )
            # الربح = الدائن (إيرادات) - المدين (مصروفات)
            current_net_income = (cur_pl_totals['sum_cr'] or Decimal('0.00')) - (cur_pl_totals['sum_dr'] or Decimal('0.00'))

            # صافي ربح فترة المقارنة
            comp_net_income = Decimal('0.00')
            if comparison_date:
                comp_fy = FiscalYear.objects.filter(start_date__lte=comparison_date, end_date__gte=comparison_date).first()
                comp_fy_start = comp_fy.start_date if comp_fy else date(comparison_date.year, 1, 1)
                comp_pl_query = Q(
                    journal_entry__status='posted',
                    journal_entry__date__gte=comp_fy_start,
                    journal_entry__date__lte=comparison_date,
                    account__account_type__category__in=['revenue', 'expense']
                )
                comp_pl_totals = JournalEntryLine.objects.filter(comp_pl_query).aggregate(
                    sum_dr=Coalesce(Sum('debit'), Decimal('0.00')),
                    sum_cr=Coalesce(Sum('credit'), Decimal('0.00'))
                )
                comp_net_income = (comp_pl_totals['sum_cr'] or Decimal('0.00')) - (comp_pl_totals['sum_dr'] or Decimal('0.00'))

            # 9. تدوير أرباح السنوات السابقة غير المقفلة إلى حساب الأرباح المرحلة
            retained_earnings_account = AccountRoleRegistry.get_account_by_role("RETAINED_EARNINGS")
            if not retained_earnings_account:
                retained_earnings_account = ChartOfAccounts.objects.filter(code__in=['31410', '3201', '3141', '31400', '32000'], is_active=True).first()
            if not retained_earnings_account:
                retained_earnings_account = ChartOfAccounts.objects.filter(account_type__category='equity', code__startswith='314').first()

            if retained_earnings_account and net_unclosed_prior_pl != Decimal('0.00'):
                direct_balances[retained_earnings_account.id] = direct_balances.get(retained_earnings_account.id, Decimal('0.00')) + net_unclosed_prior_pl
            if retained_earnings_account and comp_unclosed_prior_pl != Decimal('0.00'):
                comp_direct_balances[retained_earnings_account.id] = comp_direct_balances.get(retained_earnings_account.id, Decimal('0.00')) + comp_unclosed_prior_pl

            # 10. التجميع الشجري التصاعدي للأبناء في الآباء (Bottom-Up Tree Rollup)
            rolled_balances = dict(direct_balances)
            comp_rolled_balances = dict(comp_direct_balances)

            max_level = max((acc.level for acc in accounts_list), default=1)
            for lvl in range(max_level, 1, -1):
                for acc in accounts_list:
                    if acc.level == lvl and acc.parent_id:
                        rolled_balances[acc.parent_id] = rolled_balances.get(acc.parent_id, Decimal('0.00')) + rolled_balances.get(acc.id, Decimal('0.00'))
                        comp_rolled_balances[acc.parent_id] = comp_rolled_balances.get(acc.parent_id, Decimal('0.00')) + comp_rolled_balances.get(acc.id, Decimal('0.00'))

            # 11. تبويب بنود الميزانية وفق معيار IAS 1 الخماسي
            # أ. الأصول المتداولة (11)
            current_assets_nodes = []
            total_current_assets = Decimal('0.00')
            comp_total_current_assets = Decimal('0.00')
            cash_and_bank_total = Decimal('0.00')
            inventory_total = Decimal('0.00')

            # ب. الأصول الثابتة / غير المتداولة (12)
            non_current_assets_nodes = []
            gross_fixed_assets = Decimal('0.00')
            accumulated_depreciation = Decimal('0.00')
            total_non_current_assets = Decimal('0.00')
            comp_total_non_current_assets = Decimal('0.00')

            # ج. الخصوم المتداولة (21)
            current_liab_nodes = []
            total_current_liabilities = Decimal('0.00')
            comp_total_current_liabilities = Decimal('0.00')

            # د. الخصوم غير المتداولة (22)
            non_current_liab_nodes = []
            total_non_current_liabilities = Decimal('0.00')
            comp_total_non_current_liabilities = Decimal('0.00')

            # هـ. رأس المال وحقوق الملكية (31)
            equity_nodes = []
            total_equity_before_net_income = Decimal('0.00')
            comp_total_equity_before_net_income = Decimal('0.00')

            for acc in accounts_list:
                cat = acc.account_type.category
                code = str(acc.code)
                is_leaf = acc.is_leaf
                bal = rolled_balances.get(acc.id, Decimal('0.00'))
                comp_bal = comp_rolled_balances.get(acc.id, Decimal('0.00'))

                # حساب التغير
                diff_amount = bal - comp_bal
                if comp_bal != Decimal('0.00'):
                    diff_pct = (diff_amount / abs(comp_bal)) * Decimal('100.00')
                    is_new = False
                elif bal != Decimal('0.00'):
                    diff_pct = Decimal('100.00')
                    is_new = True
                else:
                    diff_pct = Decimal('0.00')
                    is_new = False

                node = {
                    'account': acc,
                    'id': acc.id,
                    'code': acc.code,
                    'name': acc.name,
                    'level': acc.level,
                    'is_leaf': is_leaf,
                    'parent_id': acc.parent_id,
                    'balance': bal,
                    'comparison_balance': comp_bal,
                    'diff_amount': diff_amount,
                    'diff_percent': diff_pct.quantize(Decimal('0.01')),
                    'is_new': is_new,
                    'is_abnormal': (acc.account_type.nature == 'debit' and bal < 0) or (acc.account_type.nature == 'credit' and bal < 0),
                    'has_children': not is_leaf,
                }

                # تصنيف الأصول
                if cat == 'asset':
                    if code.startswith('11'):
                        current_assets_nodes.append(node)
                        if is_leaf:
                            total_current_assets += direct_balances.get(acc.id, Decimal('0.00'))
                            comp_total_current_assets += comp_direct_balances.get(acc.id, Decimal('0.00'))
                            if code.startswith(('111', '1111', '1116')):
                                cash_and_bank_total += direct_balances.get(acc.id, Decimal('0.00'))
                            if code.startswith(('113', '1131')):
                                inventory_total += direct_balances.get(acc.id, Decimal('0.00'))
                    elif code.startswith('12') or code.startswith('1'):
                        non_current_assets_nodes.append(node)
                        if is_leaf:
                            acc_bal = direct_balances.get(acc.id, Decimal('0.00'))
                            c_acc_bal = comp_direct_balances.get(acc.id, Decimal('0.00'))
                            # معالجة مجمع الإهلاك (122) كحساب مقابل
                            if code.startswith(('122', '129')):
                                accumulated_depreciation += abs(acc_bal)
                                total_non_current_assets -= abs(acc_bal)
                                comp_total_non_current_assets -= abs(c_acc_bal)
                            else:
                                gross_fixed_assets += acc_bal
                                total_non_current_assets += acc_bal
                                comp_total_non_current_assets += c_acc_bal

                # تصنيف الخصوم
                elif cat == 'liability':
                    if code.startswith('21'):
                        current_liab_nodes.append(node)
                        if is_leaf:
                            total_current_liabilities += direct_balances.get(acc.id, Decimal('0.00'))
                            comp_total_current_liabilities += comp_direct_balances.get(acc.id, Decimal('0.00'))
                    elif code.startswith('22') or code.startswith('2'):
                        non_current_liab_nodes.append(node)
                        if is_leaf:
                            total_non_current_liabilities += direct_balances.get(acc.id, Decimal('0.00'))
                            comp_total_non_current_liabilities += comp_direct_balances.get(acc.id, Decimal('0.00'))

                # تصنيف حقوق الملكية
                elif cat == 'equity':
                    equity_nodes.append(node)
                    if is_leaf:
                        total_equity_before_net_income += direct_balances.get(acc.id, Decimal('0.00'))
                        comp_total_equity_before_net_income += comp_direct_balances.get(acc.id, Decimal('0.00'))

            # 12. الإجماليات والمجاميع الكبرى (Grand Totals)
            total_assets = total_current_assets + total_non_current_assets
            comp_total_assets = comp_total_current_assets + comp_total_non_current_assets

            total_liabilities = total_current_liabilities + total_non_current_liabilities
            comp_total_liabilities = comp_total_current_liabilities + comp_total_non_current_liabilities

            total_equity = total_equity_before_net_income + current_net_income
            comp_total_equity = comp_total_equity_before_net_income + comp_net_income

            total_liabilities_and_equity = total_liabilities + total_equity
            comp_total_liabilities_and_equity = comp_total_liabilities + comp_total_equity

            # التحقق المحاسبي الصارم من توازن الميزانية (Assets == Liabilities + Equity)
            difference = (total_assets - total_liabilities_and_equity).quantize(Decimal('0.01'))
            is_balanced = abs(difference) <= Decimal('0.05')

            comp_difference = (comp_total_assets - comp_total_liabilities_and_equity).quantize(Decimal('0.01')) if comparison_date else Decimal('0.00')

            # 13. محرك النسب والمؤشرات المالية المحمى (Guarded Financial KPIs)
            financial_ratios = cls._calculate_guarded_ratios(
                total_current_assets=total_current_assets,
                total_current_liabilities=total_current_liabilities,
                total_assets=total_assets,
                total_liabilities=total_liabilities,
                total_equity=total_equity,
                cash_and_bank=cash_and_bank_total,
                inventory=inventory_total,
                net_income=current_net_income
            )

            # تصفية الحسابات الصفرية وتحديد المستويات إن طلب
            if hide_zero_balances:
                current_assets_nodes = [n for n in current_assets_nodes if n['balance'] != 0 or n['comparison_balance'] != 0]
                non_current_assets_nodes = [n for n in non_current_assets_nodes if n['balance'] != 0 or n['comparison_balance'] != 0]
                current_liab_nodes = [n for n in current_liab_nodes if n['balance'] != 0 or n['comparison_balance'] != 0]
                non_current_liab_nodes = [n for n in non_current_liab_nodes if n['balance'] != 0 or n['comparison_balance'] != 0]
                equity_nodes = [n for n in equity_nodes if n['balance'] != 0 or n['comparison_balance'] != 0]

            if account_level and str(account_level).isdigit():
                lvl_val = int(account_level)
                current_assets_nodes = [n for n in current_assets_nodes if n['level'] <= lvl_val]
                non_current_assets_nodes = [n for n in non_current_assets_nodes if n['level'] <= lvl_val]
                current_liab_nodes = [n for n in current_liab_nodes if n['level'] <= lvl_val]
                non_current_liab_nodes = [n for n in non_current_liab_nodes if n['level'] <= lvl_val]
                equity_nodes = [n for n in equity_nodes if n['level'] <= lvl_val]

            return {
                'as_of_date': as_of_date,
                'comparison_date': comparison_date,
                'is_period_closed': is_period_closed,
                'period_name': period.name if period else "",
                'currency_code': currency_code,
                'currency_symbol': currency_symbol,
                'generated_at': timezone.now(),

                # أقسام الأصول
                'current_assets': {
                    'nodes': current_assets_nodes,
                    'total': total_current_assets,
                    'comp_total': comp_total_current_assets,
                    'diff_amount': total_current_assets - comp_total_current_assets,
                },
                'non_current_assets': {
                    'nodes': non_current_assets_nodes,
                    'gross_fixed_assets': gross_fixed_assets,
                    'accumulated_depreciation': accumulated_depreciation,
                    'total': total_non_current_assets,
                    'comp_total': comp_total_non_current_assets,
                    'diff_amount': total_non_current_assets - comp_total_non_current_assets,
                },
                'total_assets': total_assets,
                'comp_total_assets': comp_total_assets,
                'assets_diff': total_assets - comp_total_assets,

                # أقسام الخصوم
                'current_liabilities': {
                    'nodes': current_liab_nodes,
                    'total': total_current_liabilities,
                    'comp_total': comp_total_current_liabilities,
                    'diff_amount': total_current_liabilities - comp_total_current_liabilities,
                },
                'non_current_liabilities': {
                    'nodes': non_current_liab_nodes,
                    'total': total_non_current_liabilities,
                    'comp_total': comp_total_non_current_liabilities,
                    'diff_amount': total_non_current_liabilities - comp_total_non_current_liabilities,
                },
                'total_liabilities': total_liabilities,
                'comp_total_liabilities': comp_total_liabilities,
                'liabilities_diff': total_liabilities - comp_total_liabilities,

                # حقوق الملكية
                'equity': {
                    'nodes': equity_nodes,
                    'total_before_net_income': total_equity_before_net_income,
                    'current_net_income': current_net_income,
                    'comp_net_income': comp_net_income,
                    'net_income_diff': current_net_income - comp_net_income,
                    'total': total_equity,
                    'comp_total': comp_total_equity,
                    'diff_amount': total_equity - comp_total_equity,
                },
                'total_equity': total_equity,
                'comp_total_equity': comp_total_equity,

                # الإجماليات المشتركة
                'total_liabilities_equity': total_liabilities_and_equity,
                'comp_total_liabilities_equity': comp_total_liabilities_and_equity,
                'liabilities_equity_diff': total_liabilities_and_equity - comp_total_liabilities_and_equity,

                # حالة التوازن
                'difference': difference,
                'is_balanced': is_balanced,
                'comp_difference': comp_difference,

                # المؤشرات المالية
                'financial_ratios': financial_ratios,

                # backwards compatibility fields for legacy templates/tests
                'assets': {
                    'accounts': current_assets_nodes + non_current_assets_nodes,
                    'total': total_assets,
                },
                'liabilities': {
                    'accounts': current_liab_nodes + non_current_liab_nodes,
                    'total': total_liabilities,
                },
            }

        except Exception as e:
            logger.error(f"خطأ في توليد الميزانية العمومية: {e}", exc_info=True)
            return cls._empty_balance_sheet_response(
                as_of_date or timezone.now().date(),
                comparison_date,
                "EGP",
                "ج.م",
                error=str(e)
            )

    @classmethod
    def _calculate_guarded_ratios(
        cls,
        total_current_assets: Decimal,
        total_current_liabilities: Decimal,
        total_assets: Decimal,
        total_liabilities: Decimal,
        total_equity: Decimal,
        cash_and_bank: Decimal,
        inventory: Decimal,
        net_income: Decimal
    ) -> Dict[str, Any]:
        """
        احتساب النسب المالية المحمية من أخطاء القسمة على صفر وعجز حقوق الملكية
        """
        ratios: Dict[str, Any] = {}

        # 1. رأس المال العامل (Working Capital)
        working_capital = total_current_assets - total_current_liabilities
        ratios['working_capital'] = working_capital

        # 2. نسبة التداول (Current Ratio)
        if total_current_liabilities > Decimal('0.00'):
            ratios['current_ratio'] = (total_current_assets / total_current_liabilities).quantize(Decimal('0.01'))
        elif total_current_assets > Decimal('0.00'):
            ratios['current_ratio'] = Decimal('999.99')  # ملاءة كاملة بدون التزامات متداولة
        else:
            ratios['current_ratio'] = None

        # 3. نسبة السيولة السريعة (Quick Ratio)
        if total_current_liabilities > Decimal('0.00'):
            ratios['quick_ratio'] = ((total_current_assets - inventory) / total_current_liabilities).quantize(Decimal('0.01'))
        else:
            ratios['quick_ratio'] = None

        # 4. نسبة السيولة النقدية (Cash Ratio)
        if total_current_liabilities > Decimal('0.00'):
            ratios['cash_ratio'] = (cash_and_bank / total_current_liabilities).quantize(Decimal('0.01'))
        else:
            ratios['cash_ratio'] = None

        # 5. نسبة المديونية الكلية (Debt Ratio)
        if total_assets > Decimal('0.00'):
            ratios['debt_ratio'] = ((total_liabilities / total_assets) * Decimal('100.00')).quantize(Decimal('0.01'))
            ratios['equity_ratio'] = ((total_equity / total_assets) * Decimal('100.00')).quantize(Decimal('0.01'))
        else:
            ratios['debt_ratio'] = None
            ratios['equity_ratio'] = None

        # 6. نسبة المديونية لحقوق الملكية (Debt to Equity)
        if total_equity > Decimal('0.00'):
            ratios['debt_to_equity'] = ((total_liabilities / total_equity) * Decimal('100.00')).quantize(Decimal('0.01'))
            ratios['is_negative_equity'] = False
        else:
            ratios['debt_to_equity'] = None
            ratios['is_negative_equity'] = True if total_equity < Decimal('0.00') else False

        # 7. معدلات العائد (ROA & ROE)
        if total_assets > Decimal('0.00'):
            ratios['roa'] = ((net_income / total_assets) * Decimal('100.00')).quantize(Decimal('0.01'))
        else:
            ratios['roa'] = None

        if total_equity > Decimal('0.00'):
            ratios['roe'] = ((net_income / total_equity) * Decimal('100.00')).quantize(Decimal('0.01'))
        else:
            ratios['roe'] = None

        return ratios

    @classmethod
    def _empty_balance_sheet_response(
        cls,
        as_of_date: date,
        comparison_date: Optional[date],
        currency_code: str,
        currency_symbol: str,
        error: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        استجابة فارغة آمنة عند حدوث خطأ أو عدم وجود حسابات
        """
        return {
            'as_of_date': as_of_date,
            'comparison_date': comparison_date,
            'is_period_closed': False,
            'period_name': "",
            'currency_code': currency_code,
            'currency_symbol': currency_symbol,
            'generated_at': timezone.now(),
            'current_assets': {'nodes': [], 'total': Decimal('0.00'), 'comp_total': Decimal('0.00'), 'diff_amount': Decimal('0.00')},
            'non_current_assets': {'nodes': [], 'gross_fixed_assets': Decimal('0.00'), 'accumulated_depreciation': Decimal('0.00'), 'total': Decimal('0.00'), 'comp_total': Decimal('0.00'), 'diff_amount': Decimal('0.00')},
            'total_assets': Decimal('0.00'),
            'comp_total_assets': Decimal('0.00'),
            'assets_diff': Decimal('0.00'),
            'current_liabilities': {'nodes': [], 'total': Decimal('0.00'), 'comp_total': Decimal('0.00'), 'diff_amount': Decimal('0.00')},
            'non_current_liabilities': {'nodes': [], 'total': Decimal('0.00'), 'comp_total': Decimal('0.00'), 'diff_amount': Decimal('0.00')},
            'total_liabilities': Decimal('0.00'),
            'comp_total_liabilities': Decimal('0.00'),
            'liabilities_diff': Decimal('0.00'),
            'equity': {'nodes': [], 'total_before_net_income': Decimal('0.00'), 'current_net_income': Decimal('0.00'), 'comp_net_income': Decimal('0.00'), 'net_income_diff': Decimal('0.00'), 'total': Decimal('0.00'), 'comp_total': Decimal('0.00'), 'diff_amount': Decimal('0.00')},
            'total_equity': Decimal('0.00'),
            'comp_total_equity': Decimal('0.00'),
            'total_liabilities_equity': Decimal('0.00'),
            'comp_total_liabilities_equity': Decimal('0.00'),
            'liabilities_equity_diff': Decimal('0.00'),
            'difference': Decimal('0.00'),
            'is_balanced': True,
            'comp_difference': Decimal('0.00'),
            'financial_ratios': {},
            'assets': {'accounts': [], 'total': Decimal('0.00')},
            'liabilities': {'accounts': [], 'total': Decimal('0.00')},
            'error': error
        }

    @classmethod
    def calculate_financial_ratios(cls, balance_sheet_data: Dict) -> Dict:
        """توافقية رجعية لدوال احتساب النسب"""
        if 'financial_ratios' in balance_sheet_data:
            return balance_sheet_data['financial_ratios']
        return cls._calculate_guarded_ratios(
            total_current_assets=balance_sheet_data.get('current_assets', {}).get('total', Decimal('0.00')),
            total_current_liabilities=balance_sheet_data.get('current_liabilities', {}).get('total', Decimal('0.00')),
            total_assets=balance_sheet_data.get('total_assets', Decimal('0.00')),
            total_liabilities=balance_sheet_data.get('total_liabilities', Decimal('0.00')),
            total_equity=balance_sheet_data.get('total_equity', Decimal('0.00')),
            cash_and_bank=Decimal('0.00'),
            inventory=Decimal('0.00'),
            net_income=balance_sheet_data.get('equity', {}).get('current_net_income', Decimal('0.00'))
        )

    @classmethod
    def export_to_excel(
        cls,
        as_of_date: Optional[Union[date, str]] = None,
        comparison_date: Optional[Union[date, str]] = None,
        account_level: Optional[Union[int, str]] = None,
        hide_zero_balances: bool = False,
    ) -> bytes:
        """
        تصدير الميزانية العمومية الرسمية إلى Excel بمعادلات حية وتنسيق محاسبي معتمد
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            bs = cls.generate_balance_sheet(
                as_of_date=as_of_date,
                comparison_date=comparison_date,
                account_level=account_level,
                hide_zero_balances=hide_zero_balances
            )

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "الميزانية العمومية"
            ws.views.sheetView[0].rightToLeft = True

            # الأنماط والتنسيقات
            font_title = Font(name="Calibri", size=16, bold=True, color="1F2937")
            font_subtitle = Font(name="Calibri", size=11, color="4B5563")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_group_1 = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
            font_group_2 = Font(name="Calibri", size=11, bold=True, color="1F2937")
            font_regular = Font(name="Calibri", size=10, color="374151")
            font_total = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

            fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            fill_group_1 = PatternFill(start_color="EEF4FF", end_color="EEF4FF", fill_type="solid")
            fill_group_2 = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            fill_total = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            fill_subtotal = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

            border_thin = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            # الترويسة الرئيسية
            ws['A1'] = "تقرير الميزانية العمومية والمركز المالي (IAS 1 Statement of Financial Position)"
            ws['A1'].font = font_title
            ws.merge_cells('A1:F1')

            date_str = bs['as_of_date'].strftime('%Y-%m-%d')
            comp_str = f" | فترة المقارنة: {bs['comparison_date'].strftime('%Y-%m-%d')}" if bs['comparison_date'] else ""
            status_str = " (فترة مقفلة ومعتمدة)" if bs['is_period_closed'] else " (فترة تشغيلية مفتوحة)"
            ws['A2'] = f"كما في: {date_str}{comp_str}{status_str} - العملة: {bs['currency_code']}"
            ws['A2'].font = font_subtitle
            ws.merge_cells('A2:F2')

            # رأس أعمدة الجدول
            headers = ["كود الحساب", "اسم الحساب / البند المحاسبي", f"الرصيد في {date_str}"]
            if bs['comparison_date']:
                headers.extend([f"الرصيد في {bs['comparison_date'].strftime('%Y-%m-%d')}", "مبلغ التغير", "نسبة التغير %"])

            row_idx = 4
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_thin

            row_idx += 1

            def write_section(title, nodes, subtotal, comp_subtotal=Decimal('0.00')):
                nonlocal row_idx
                # عنوان القسم
                cell_title = ws.cell(row=row_idx, column=1, value=title)
                cell_title.font = font_group_1
                cell_title.fill = fill_group_1
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
                row_idx += 1

                for n in nodes:
                    indent = "    " * (n['level'] - 1)
                    ws.cell(row=row_idx, column=1, value=str(n['code'])).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=2, value=f"{indent}{n['name']}")

                    cell_bal = ws.cell(row=row_idx, column=3, value=float(n['balance']))
                    cell_bal.number_format = '#,##0.00'

                    if bs['comparison_date']:
                        cell_c = ws.cell(row=row_idx, column=4, value=float(n['comparison_balance']))
                        cell_c.number_format = '#,##0.00'

                        cell_diff = ws.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}")
                        cell_diff.number_format = '#,##0.00'

                        cell_pct = ws.cell(row=row_idx, column=6, value=f"=IF(D{row_idx}=0, 1, (C{row_idx}-D{row_idx})/ABS(D{row_idx}))")
                        cell_pct.number_format = '0.00%'

                    for c in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=c)
                        cell.border = border_thin
                        if n['level'] == 1:
                            cell.font = font_group_1
                            cell.fill = fill_group_1
                        elif n['level'] == 2:
                            cell.font = font_group_2
                            cell.fill = fill_group_2
                        else:
                            cell.font = font_regular
                    row_idx += 1

                # صف إجمالي القسم
                ws.cell(row=row_idx, column=1, value="")
                ws.cell(row=row_idx, column=2, value=f"إجمالي {title}")
                cell_sub = ws.cell(row=row_idx, column=3, value=float(subtotal))
                cell_sub.number_format = '#,##0.00'

                if bs['comparison_date']:
                    cell_c_sub = ws.cell(row=row_idx, column=4, value=float(comp_subtotal))
                    cell_c_sub.number_format = '#,##0.00'
                    cell_d = ws.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}")
                    cell_d.number_format = '#,##0.00'
                    cell_p = ws.cell(row=row_idx, column=6, value=f"=IF(D{row_idx}=0, 1, (C{row_idx}-D{row_idx})/ABS(D{row_idx}))")
                    cell_p.number_format = '0.00%'

                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=c)
                    cell.font = font_group_2
                    cell.fill = fill_subtotal
                    cell.border = border_thin
                row_idx += 2

            # 1. الأصول المتداولة
            write_section(
                "الأصول المتداولة (Current Assets)",
                bs['current_assets']['nodes'],
                bs['current_assets']['total'],
                bs['current_assets']['comp_total']
            )

            # 2. الأصول الثابتة
            write_section(
                "الأصول الثابتة وغير المتداولة (Non-Current Assets)",
                bs['non_current_assets']['nodes'],
                bs['non_current_assets']['total'],
                bs['non_current_assets']['comp_total']
            )

            # صف إجمالي الأصول الكلي
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value="إجمالي الأصول (TOTAL ASSETS)")
            ws.cell(row=row_idx, column=3, value=float(bs['total_assets'])).number_format = '#,##0.00'
            if bs['comparison_date']:
                ws.cell(row=row_idx, column=4, value=float(bs['comp_total_assets'])).number_format = '#,##0.00'
                ws.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}").number_format = '#,##0.00'
                ws.cell(row=row_idx, column=6, value=f"=IF(D{row_idx}=0, 1, (C{row_idx}-D{row_idx})/ABS(D{row_idx}))").number_format = '0.00%'

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_thin
            row_idx += 2

            # 3. الخصوم المتداولة
            write_section(
                "الخصوم المتداولة (Current Liabilities)",
                bs['current_liabilities']['nodes'],
                bs['current_liabilities']['total'],
                bs['current_liabilities']['comp_total']
            )

            # 4. الخصوم غير المتداولة
            write_section(
                "الخصوم غير المتداولة (Non-Current Liabilities)",
                bs['non_current_liabilities']['nodes'],
                bs['non_current_liabilities']['total'],
                bs['non_current_liabilities']['comp_total']
            )

            # 5. حقوق الملكية
            write_section(
                "رأس المال وحقوق الملكية (Equity)",
                bs['equity']['nodes'],
                bs['total_equity'],
                bs['comp_total_equity']
            )

            # صف إجمالي الخصوم وحقوق الملكية
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value="إجمالي الخصوم وحقوق الملكية (TOTAL LIABILITIES & EQUITY)")
            ws.cell(row=row_idx, column=3, value=float(bs['total_liabilities_equity'])).number_format = '#,##0.00'
            if bs['comparison_date']:
                ws.cell(row=row_idx, column=4, value=float(bs['comp_total_liabilities_equity'])).number_format = '#,##0.00'
                ws.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}").number_format = '#,##0.00'
                ws.cell(row=row_idx, column=6, value=f"=IF(D{row_idx}=0, 1, (C{row_idx}-D{row_idx})/ABS(D{row_idx}))").number_format = '0.00%'

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_thin
            row_idx += 2

            # ضبط عرض الأعمدة تلقائياً
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 22
            if bs['comparison_date']:
                ws.column_dimensions['D'].width = 22
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 16

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except Exception as e:
            logger.error(f"خطأ في تصدير Excel للميزانية: {e}", exc_info=True)
            raise
