# financial/services/income_statement_service.py
"""
خدمة قائمة الدخل والأرباح والخسائر المعيارية - Enterprise Income Statement Service (v2.0)
تطبيق كامل لمعايير المحاسبة الدولية (IAS 1 / IAS 21) بهيكل متعدد المراحل (Multi-Step)،
عزل قيود الإقفال السنوية، المعالجة الدقيقة للحسابات المقابلة (مردودات وخصومات)،
احتساب الهوامش والنسب المالية المحمية، فلترة مراكز التكلفة، ودعم المقارنات الزمنية وتصدير Excel الرسمي.
"""

import logging
from decimal import Decimal
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional, Any, Union
from io import BytesIO

from django.db.models import Sum, Q
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.utils.translation import gettext as _

from financial.models.chart_of_accounts import ChartOfAccounts
from financial.models.journal_entry import JournalEntryLine, AccountingPeriod
from financial.models.fiscal_year import FiscalYear
from financial.models.cost_center import CostCenter
from financial.services.exchange_rate_service import ExchangeRateService

logger = logging.getLogger(__name__)


class IncomeStatementService:
    """
    خدمة قائمة الدخل والأرباح والخسائر المؤسسية
    """

    @classmethod
    def generate_income_statement(
        cls,
        date_from: Optional[Union[date, str]] = None,
        date_to: Optional[Union[date, str]] = None,
        comp_date_from: Optional[Union[date, str]] = None,
        comp_date_to: Optional[Union[date, str]] = None,
        cost_center_id: Optional[Union[int, str]] = None,
        account_level: Optional[Union[int, str]] = None,
        hide_zero_balances: bool = False,
        include_unposted: bool = False,
    ) -> Dict[str, Any]:
        """
        إنشاء قائمة الدخل المعيارية متعددة المراحل طبقاً لمعيار IAS 1
        """
        try:
            # 1. ضبط التواريخ والفترة
            today = timezone.now().date()
            if isinstance(date_to, str) and date_to.strip():
                try:
                    date_to = datetime.strptime(date_to.strip(), "%Y-%m-%d").date()
                except ValueError:
                    date_to = today
            elif not isinstance(date_to, date):
                date_to = today

            if isinstance(date_from, str) and date_from.strip():
                try:
                    date_from = datetime.strptime(date_from.strip(), "%Y-%m-%d").date()
                except ValueError:
                    date_from = date(date_to.year, 1, 1)
            elif not isinstance(date_from, date):
                date_from = date(date_to.year, 1, 1)

            # ضبط تواريخ فترة المقارنة
            has_comparison = False
            if isinstance(comp_date_to, str) and comp_date_to.strip():
                try:
                    comp_date_to = datetime.strptime(comp_date_to.strip(), "%Y-%m-%d").date()
                except ValueError:
                    comp_date_to = None
            elif not isinstance(comp_date_to, date):
                comp_date_to = None

            if isinstance(comp_date_from, str) and comp_date_from.strip():
                try:
                    comp_date_from = datetime.strptime(comp_date_from.strip(), "%Y-%m-%d").date()
                except ValueError:
                    comp_date_from = None
            elif not isinstance(comp_date_from, date):
                comp_date_from = None

            if comp_date_from and comp_date_to and comp_date_to >= comp_date_from:
                has_comparison = True

            # 2. تحديد العملة الوظيفية
            functional_currency = ExchangeRateService.get_functional_currency()
            currency_code = functional_currency.code if functional_currency else "EGP"
            currency_symbol = functional_currency.symbol or currency_code if functional_currency else "ج.م"

            # 3. فحص تطابق الفترة مع بداية السنة المالية (YTD Indicator)
            fiscal_year = FiscalYear.objects.filter(start_date__lte=date_to, end_date__gte=date_to).first()
            is_ytd = bool(fiscal_year and date_from == fiscal_year.start_date)

            # 4. جلب شجرة حسابات الإيرادات والمصروفات
            accounts_qs = ChartOfAccounts.objects.filter(
                account_type__category__in=['revenue', 'expense']
            ).select_related('account_type', 'parent', 'currency').order_by('code')

            accounts_list = list(accounts_qs)
            if not accounts_list:
                return cls._empty_income_statement_response(date_from, date_to, currency_code, currency_symbol)

            # 5. الاستعلام التجميعي الأساسي للفترة الحالية (مع استبعاد قيود الإقفال)
            cur_query = Q(
                journal_entry__date__gte=date_from,
                journal_entry__date__lte=date_to,
                account__account_type__category__in=['revenue', 'expense']
            )
            # استبعاد قيود الإقفال السنوية
            cur_query &= ~Q(journal_entry__entry_type='closing')

            if not include_unposted:
                cur_query &= Q(journal_entry__status='posted')

            if cost_center_id and str(cost_center_id).isdigit():
                cur_query &= Q(cost_center_id=int(cost_center_id))

            cur_totals = (
                JournalEntryLine.objects.filter(cur_query)
                .values('account_id')
                .annotate(
                    sum_debit=Coalesce(Sum('debit'), Decimal('0.00')),
                    sum_credit=Coalesce(Sum('credit'), Decimal('0.00'))
                )
            )
            cur_map = {row['account_id']: row for row in cur_totals}

            # 6. استعلام فترة المقارنة إن وجدت
            comp_map = {}
            if has_comparison:
                comp_query = Q(
                    journal_entry__date__gte=comp_date_from,
                    journal_entry__date__lte=comp_date_to,
                    account__account_type__category__in=['revenue', 'expense']
                )
                comp_query &= ~Q(journal_entry__entry_type='closing')
                if not include_unposted:
                    comp_query &= Q(journal_entry__status='posted')
                if cost_center_id and str(cost_center_id).isdigit():
                    comp_query &= Q(cost_center_id=int(cost_center_id))

                comp_totals = (
                    JournalEntryLine.objects.filter(comp_query)
                    .values('account_id')
                    .annotate(
                        sum_debit=Coalesce(Sum('debit'), Decimal('0.00')),
                        sum_credit=Coalesce(Sum('credit'), Decimal('0.00'))
                    )
                )
                comp_map = {row['account_id']: row for row in comp_totals}

            # 7. حساب صافي رصيد كل حساب نهائي مباشر (Direct Balances with Contra Math)
            direct_balances = {}
            comp_direct_balances = {}
            activity_flags = {}

            for acc in accounts_list:
                row = cur_map.get(acc.id, {})
                dr = row.get('sum_debit') or Decimal('0.00')
                cr = row.get('sum_credit') or Decimal('0.00')
                has_activity = (dr != Decimal('0.00') or cr != Decimal('0.00'))

                # حساب الرصيد حسب طبيعة الحساب
                if acc.account_type.category == 'revenue':
                    # الإيرادات: الدائن موجب، والمدين (مردودات/خصومات) يطرح
                    bal = cr - dr
                else:
                    # المصروفات: المدين موجب، والدائن (مردودات مشتريات/خصم مكتسب) يطرح
                    bal = dr - cr

                direct_balances[acc.id] = bal
                activity_flags[acc.id] = has_activity

                # فترة المقارنة
                if has_comparison:
                    c_row = comp_map.get(acc.id, {})
                    c_dr = c_row.get('sum_debit') or Decimal('0.00')
                    c_cr = c_row.get('sum_credit') or Decimal('0.00')
                    if acc.account_type.category == 'revenue':
                        c_bal = c_cr - c_dr
                    else:
                        c_bal = c_dr - c_cr
                    comp_direct_balances[acc.id] = c_bal
                else:
                    comp_direct_balances[acc.id] = Decimal('0.00')

            # 8. التجميع الشجري التصاعدي للأبناء في الآباء (Bottom-Up Tree Rollup)
            rolled_balances = dict(direct_balances)
            comp_rolled_balances = dict(comp_direct_balances)

            max_level = max((acc.level for acc in accounts_list), default=1)
            for lvl in range(max_level, 1, -1):
                for acc in accounts_list:
                    if acc.level == lvl and acc.parent_id:
                        rolled_balances[acc.parent_id] = rolled_balances.get(acc.parent_id, Decimal('0.00')) + rolled_balances.get(acc.id, Decimal('0.00'))
                        comp_rolled_balances[acc.parent_id] = comp_rolled_balances.get(acc.parent_id, Decimal('0.00')) + comp_rolled_balances.get(acc.id, Decimal('0.00'))
                        if activity_flags.get(acc.id, False):
                            activity_flags[acc.parent_id] = True

            # 9. التبويب الهيكلي متعدد المراحل (Multi-Step Sections Categorization)
            # أ. إيرادات النشاط التشغيلي (41)
            operating_revenue_nodes = []
            gross_sales = Decimal('0.00')
            sales_deductions = Decimal('0.00')
            net_operating_revenue = Decimal('0.00')
            comp_net_operating_revenue = Decimal('0.00')

            # ب. تكلفة المبيعات (51)
            cogs_nodes = []
            gross_cogs = Decimal('0.00')
            purchase_deductions = Decimal('0.00')
            net_cogs = Decimal('0.00')
            comp_net_cogs = Decimal('0.00')

            # ج. المصروفات التشغيلية والعمومية والإدارية والتسويقية (52)
            operating_expense_nodes = []
            total_operating_expenses = Decimal('0.00')
            comp_total_operating_expenses = Decimal('0.00')

            # د. الإيرادات الأخرى وأرباح فروق العملة (42, 43, 49)
            other_revenue_nodes = []
            total_other_revenue = Decimal('0.00')
            comp_total_other_revenue = Decimal('0.00')

            # هـ. المصروفات التمويلية والأخرى وخسائر العملة (54, 55, 59)
            other_expense_nodes = []
            total_other_expenses = Decimal('0.00')
            comp_total_other_expenses = Decimal('0.00')

            fx_gains = Decimal('0.00')
            fx_losses = Decimal('0.00')

            for acc in accounts_list:
                cat = acc.account_type.category
                code = str(acc.code)
                is_leaf = acc.is_leaf
                bal = rolled_balances.get(acc.id, Decimal('0.00'))
                comp_bal = comp_rolled_balances.get(acc.id, Decimal('0.00'))

                # حساب التغير
                diff_amount = bal - comp_bal
                if comp_bal != Decimal('0.00'):
                    diff_pct = (diff_amount / abs(comp_bal)) * Decimal('100.00')
                    is_new = False
                elif bal != Decimal('0.00'):
                    diff_pct = Decimal('100.00')
                    is_new = True
                else:
                    diff_pct = Decimal('0.00')
                    is_new = False

                node = {
                    'account': acc,
                    'id': acc.id,
                    'code': acc.code,
                    'name': acc.name,
                    'level': acc.level,
                    'is_leaf': is_leaf,
                    'parent_id': acc.parent_id,
                    'balance': bal,
                    'comparison_balance': comp_bal,
                    'diff_amount': diff_amount,
                    'diff_percent': diff_pct.quantize(Decimal('0.01')),
                    'is_new': is_new,
                    'is_abnormal': bal < 0,
                    'has_children': not is_leaf,
                    'has_activity': activity_flags.get(acc.id, False),
                }

                # تصنيف الحسابات في المراحل المحددة
                if cat == 'revenue':
                    if code.startswith('41'):
                        operating_revenue_nodes.append(node)
                        if is_leaf:
                            acc_bal = direct_balances.get(acc.id, Decimal('0.00'))
                            c_acc_bal = comp_direct_balances.get(acc.id, Decimal('0.00'))
                            if code.startswith(('4191', '4193', '419')):
                                sales_deductions += abs(acc_bal)
                            else:
                                gross_sales += acc_bal
                            net_operating_revenue += acc_bal
                            comp_net_operating_revenue += c_acc_bal
                    else:
                        other_revenue_nodes.append(node)
                        if is_leaf:
                            acc_bal = direct_balances.get(acc.id, Decimal('0.00'))
                            total_other_revenue += acc_bal
                            comp_total_other_revenue += comp_direct_balances.get(acc.id, Decimal('0.00'))
                            if code.startswith(('4310', '431', '43')):
                                fx_gains += acc_bal

                elif cat == 'expense':
                    if code.startswith('51'):
                        cogs_nodes.append(node)
                        if is_leaf:
                            acc_bal = direct_balances.get(acc.id, Decimal('0.00'))
                            c_acc_bal = comp_direct_balances.get(acc.id, Decimal('0.00'))
                            if code.startswith(('5191', '5193', '519')):
                                purchase_deductions += abs(acc_bal)
                            else:
                                gross_cogs += acc_bal
                            net_cogs += acc_bal
                            comp_net_cogs += c_acc_bal
                    elif code.startswith('52'):
                        operating_expense_nodes.append(node)
                        if is_leaf:
                            acc_bal = direct_balances.get(acc.id, Decimal('0.00'))
                            total_operating_expenses += acc_bal
                            comp_total_operating_expenses += comp_direct_balances.get(acc.id, Decimal('0.00'))
                    else:
                        other_expense_nodes.append(node)
                        if is_leaf:
                            acc_bal = direct_balances.get(acc.id, Decimal('0.00'))
                            total_other_expenses += acc_bal
                            comp_total_other_expenses += comp_direct_balances.get(acc.id, Decimal('0.00'))
                            if code.startswith(('5430', '543')):
                                fx_losses += acc_bal

            # 10. حساب النتائج المرحلية الكبرى والهوامش المحمية (Multi-Step Results & Margins)
            # مجمل الربح (Gross Profit)
            gross_profit = net_operating_revenue - net_cogs
            comp_gross_profit = comp_net_operating_revenue - comp_net_cogs
            gross_profit_diff = gross_profit - comp_gross_profit

            # الربح التشغيلي (Operating Profit / EBIT)
            operating_profit = gross_profit - total_operating_expenses
            comp_operating_profit = comp_gross_profit - comp_total_operating_expenses
            operating_profit_diff = operating_profit - comp_operating_profit

            # صافي أثر تقلبات العملة (Net FX Impact)
            net_fx_impact = fx_gains - fx_losses

            # صافي الإيرادات / المصروفات الأخرى
            net_other_items = total_other_revenue - total_other_expenses
            comp_net_other_items = comp_total_other_revenue - comp_total_other_expenses

            # صافي الربح / الخسارة النهائي (Net Profit / Loss)
            net_income = operating_profit + net_other_items
            comp_net_income = comp_operating_profit + comp_net_other_items
            net_income_diff = net_income - comp_net_income

            # حساب الهوامش المحمية
            margins = cls._calculate_guarded_margins(
                net_revenue=net_operating_revenue,
                gross_profit=gross_profit,
                operating_profit=operating_profit,
                net_income=net_income
            )

            comp_margins = cls._calculate_guarded_margins(
                net_revenue=comp_net_operating_revenue,
                gross_profit=comp_gross_profit,
                operating_profit=comp_operating_profit,
                net_income=comp_net_income
            ) if has_comparison else {}

            # تصفية الحسابات بدون حركة أو ذات الرصيد الصفري
            if hide_zero_balances:
                operating_revenue_nodes = [n for n in operating_revenue_nodes if n['balance'] != 0 or n['comparison_balance'] != 0 or n['has_activity']]
                cogs_nodes = [n for n in cogs_nodes if n['balance'] != 0 or n['comparison_balance'] != 0 or n['has_activity']]
                operating_expense_nodes = [n for n in operating_expense_nodes if n['balance'] != 0 or n['comparison_balance'] != 0 or n['has_activity']]
                other_revenue_nodes = [n for n in other_revenue_nodes if n['balance'] != 0 or n['comparison_balance'] != 0 or n['has_activity']]
                other_expense_nodes = [n for n in other_expense_nodes if n['balance'] != 0 or n['comparison_balance'] != 0 or n['has_activity']]

            if account_level and str(account_level).isdigit():
                lvl_val = int(account_level)
                operating_revenue_nodes = [n for n in operating_revenue_nodes if n['level'] <= lvl_val]
                cogs_nodes = [n for n in cogs_nodes if n['level'] <= lvl_val]
                operating_expense_nodes = [n for n in operating_expense_nodes if n['level'] <= lvl_val]
                other_revenue_nodes = [n for n in other_revenue_nodes if n['level'] <= lvl_val]
                other_expense_nodes = [n for n in other_expense_nodes if n['level'] <= lvl_val]

            # معلومات مركز التكلفة إن وجد
            cost_center_obj = None
            if cost_center_id and str(cost_center_id).isdigit():
                cost_center_obj = CostCenter.objects.filter(id=int(cost_center_id)).first()

            return {
                'date_from': date_from,
                'date_to': date_to,
                'comp_date_from': comp_date_from,
                'comp_date_to': comp_date_to,
                'has_comparison': has_comparison,
                'is_ytd': is_ytd,
                'cost_center': cost_center_obj,
                'currency_code': currency_code,
                'currency_symbol': currency_symbol,
                'generated_at': timezone.now(),

                # 1. إيرادات النشاط التشغيلي
                'operating_revenues': {
                    'nodes': operating_revenue_nodes,
                    'gross_sales': gross_sales,
                    'sales_deductions': sales_deductions,
                    'total': net_operating_revenue,
                    'comp_total': comp_net_operating_revenue,
                    'diff_amount': net_operating_revenue - comp_net_operating_revenue,
                },

                # 2. تكلفة المبيعات
                'cogs': {
                    'nodes': cogs_nodes,
                    'gross_cogs': gross_cogs,
                    'purchase_deductions': purchase_deductions,
                    'total': net_cogs,
                    'comp_total': comp_net_cogs,
                    'diff_amount': net_cogs - comp_net_cogs,
                },

                # 3. مجمل الربح
                'gross_profit': gross_profit,
                'comp_gross_profit': comp_gross_profit,
                'gross_profit_diff': gross_profit_diff,
                'gross_margin': margins.get('gross_margin', Decimal('0.00')),

                # 4. المصروفات التشغيلية
                'operating_expenses': {
                    'nodes': operating_expense_nodes,
                    'total': total_operating_expenses,
                    'comp_total': comp_total_operating_expenses,
                    'diff_amount': total_operating_expenses - comp_total_operating_expenses,
                },

                # 5. الربح التشغيلي
                'operating_profit': operating_profit,
                'comp_operating_profit': comp_operating_profit,
                'operating_profit_diff': operating_profit_diff,
                'operating_margin': margins.get('operating_margin', Decimal('0.00')),

                # 6. الإيرادات والمصروفات الأخرى وفروق العملة
                'other_revenues': {
                    'nodes': other_revenue_nodes,
                    'total': total_other_revenue,
                    'comp_total': comp_total_other_revenue,
                    'diff_amount': total_other_revenue - comp_total_other_revenue,
                },
                'other_expenses': {
                    'nodes': other_expense_nodes,
                    'total': total_other_expenses,
                    'comp_total': comp_total_other_expenses,
                    'diff_amount': total_other_expenses - comp_total_other_expenses,
                },
                'net_fx_impact': net_fx_impact,
                'net_other_items': net_other_items,

                # 7. صافي الربح النهائي
                'net_income': net_income,
                'comp_net_income': comp_net_income,
                'net_income_diff': net_income_diff,
                'net_margin': margins.get('net_margin', Decimal('0.00')),
                'is_profit': net_income >= Decimal('0.00'),

                # الهوامش والمؤشرات
                'margins': margins,
                'comp_margins': comp_margins,

                # Backward compatibility
                'total_revenue': net_operating_revenue + total_other_revenue,
                'total_expense': net_cogs + total_operating_expenses + total_other_expenses,
                'revenues': operating_revenue_nodes + other_revenue_nodes,
                'expenses': cogs_nodes + operating_expense_nodes + other_expense_nodes,
            }

        except Exception as e:
            logger.error(f"خطأ في إنشاء قائمة الدخل: {e}", exc_info=True)
            return cls._empty_income_statement_response(
                date_from or timezone.now().date(),
                date_to or timezone.now().date(),
                "EGP",
                "ج.م",
                error=str(e)
            )

    @classmethod
    def _calculate_guarded_margins(
        cls,
        net_revenue: Decimal,
        gross_profit: Decimal,
        operating_profit: Decimal,
        net_income: Decimal
    ) -> Dict[str, Decimal]:
        """
        احتساب هوامش الربحية المحمية من أخطاء القسمة على صفر
        """
        margins: Dict[str, Decimal] = {}

        if net_revenue > Decimal('0.00'):
            margins['gross_margin'] = ((gross_profit / net_revenue) * Decimal('100.00')).quantize(Decimal('0.01'))
            margins['operating_margin'] = ((operating_profit / net_revenue) * Decimal('100.00')).quantize(Decimal('0.01'))
            margins['net_margin'] = ((net_income / net_revenue) * Decimal('100.00')).quantize(Decimal('0.01'))
        else:
            margins['gross_margin'] = Decimal('0.00')
            margins['operating_margin'] = Decimal('0.00')
            margins['net_margin'] = Decimal('0.00')

        return margins

    @classmethod
    def _empty_income_statement_response(
        cls,
        date_from: date,
        date_to: date,
        currency_code: str,
        currency_symbol: str,
        error: Optional[str] = None
    ) -> Dict[str, Any]:
        """استجابة فارغة آمنة عند حدوث خطأ"""
        return {
            'date_from': date_from,
            'date_to': date_to,
            'comp_date_from': None,
            'comp_date_to': None,
            'has_comparison': False,
            'is_ytd': False,
            'cost_center': None,
            'currency_code': currency_code,
            'currency_symbol': currency_symbol,
            'generated_at': timezone.now(),
            'operating_revenues': {'nodes': [], 'gross_sales': Decimal('0.00'), 'sales_deductions': Decimal('0.00'), 'total': Decimal('0.00'), 'comp_total': Decimal('0.00'), 'diff_amount': Decimal('0.00')},
            'cogs': {'nodes': [], 'gross_cogs': Decimal('0.00'), 'purchase_deductions': Decimal('0.00'), 'total': Decimal('0.00'), 'comp_total': Decimal('0.00'), 'diff_amount': Decimal('0.00')},
            'gross_profit': Decimal('0.00'),
            'comp_gross_profit': Decimal('0.00'),
            'gross_profit_diff': Decimal('0.00'),
            'gross_margin': Decimal('0.00'),
            'operating_expenses': {'nodes': [], 'total': Decimal('0.00'), 'comp_total': Decimal('0.00'), 'diff_amount': Decimal('0.00')},
            'operating_profit': Decimal('0.00'),
            'comp_operating_profit': Decimal('0.00'),
            'operating_profit_diff': Decimal('0.00'),
            'operating_margin': Decimal('0.00'),
            'other_revenues': {'nodes': [], 'total': Decimal('0.00'), 'comp_total': Decimal('0.00'), 'diff_amount': Decimal('0.00')},
            'other_expenses': {'nodes': [], 'total': Decimal('0.00'), 'comp_total': Decimal('0.00'), 'diff_amount': Decimal('0.00')},
            'net_fx_impact': Decimal('0.00'),
            'net_other_items': Decimal('0.00'),
            'net_income': Decimal('0.00'),
            'comp_net_income': Decimal('0.00'),
            'net_income_diff': Decimal('0.00'),
            'net_margin': Decimal('0.00'),
            'is_profit': True,
            'margins': {'gross_margin': Decimal('0.00'), 'operating_margin': Decimal('0.00'), 'net_margin': Decimal('0.00')},
            'comp_margins': {},
            'total_revenue': Decimal('0.00'),
            'total_expense': Decimal('0.00'),
            'revenues': [],
            'expenses': [],
            'error': error,
        }

    @classmethod
    def export_to_excel(
        cls,
        date_from: Optional[Union[date, str]] = None,
        date_to: Optional[Union[date, str]] = None,
        comp_date_from: Optional[Union[date, str]] = None,
        comp_date_to: Optional[Union[date, str]] = None,
        cost_center_id: Optional[Union[int, str]] = None,
        account_level: Optional[Union[int, str]] = None,
        hide_zero_balances: bool = False,
    ) -> bytes:
        """
        تصدير قائمة الدخل الرسمية إلى Excel بمعادلات حية متعددة المراحل
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            inc = cls.generate_income_statement(
                date_from=date_from,
                date_to=date_to,
                comp_date_from=comp_date_from,
                comp_date_to=comp_date_to,
                cost_center_id=cost_center_id,
                account_level=account_level,
                hide_zero_balances=hide_zero_balances
            )

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "قائمة الدخل"
            ws.views.sheetView[0].rightToLeft = True

            # الأنماط والتنسيقات
            font_title = Font(name="Calibri", size=16, bold=True, color="1F2937")
            font_subtitle = Font(name="Calibri", size=11, color="4B5563")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_group_1 = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
            font_group_2 = Font(name="Calibri", size=11, bold=True, color="1F2937")
            font_regular = Font(name="Calibri", size=10, color="374151")
            font_total = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

            fill_header = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            fill_group_1 = PatternFill(start_color="EEF4FF", end_color="EEF4FF", fill_type="solid")
            fill_group_2 = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            fill_total = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            fill_subtotal = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")

            border_thin = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )

            # الترويسة الرئيسية
            ws['A1'] = "تقرير قائمة الدخل والأرباح والخسائر (IAS 1 Statement of Profit or Loss)"
            ws['A1'].font = font_title
            ws.merge_cells('A1:F1')

            d_from_str = inc['date_from'].strftime('%Y-%m-%d')
            d_to_str = inc['date_to'].strftime('%Y-%m-%d')
            cc_str = f" | مركز التكلفة: {inc['cost_center'].name}" if inc['cost_center'] else ""
            comp_str = f" | المقارنة من {inc['comp_date_from'].strftime('%Y-%m-%d')} إلى {inc['comp_date_to'].strftime('%Y-%m-%d')}" if inc['has_comparison'] else ""
            ws['A2'] = f"عن الفترة من: {d_from_str} إلى: {d_to_str}{comp_str}{cc_str} - العملة: {inc['currency_code']}"
            ws['A2'].font = font_subtitle
            ws.merge_cells('A2:F2')

            # رأس أعمدة الجدول
            headers = ["كود الحساب", "اسم الحساب / البند المحاسبي", f"الفترة الحالية ({d_from_str} - {d_to_str})"]
            if inc['has_comparison']:
                headers.extend([f"فترة المقارنة ({inc['comp_date_from'].strftime('%Y-%m-%d')} - {inc['comp_date_to'].strftime('%Y-%m-%d')})", "مبلغ التغير", "نسبة التغير %"])

            row_idx = 4
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_thin

            row_idx += 1

            def write_section(title, nodes, subtotal, comp_subtotal=Decimal('0.00')):
                nonlocal row_idx
                # عنوان القسم
                cell_title = ws.cell(row=row_idx, column=1, value=title)
                cell_title.font = font_group_1
                cell_title.fill = fill_group_1
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
                row_idx += 1

                for n in nodes:
                    indent = "    " * (n['level'] - 1)
                    ws.cell(row=row_idx, column=1, value=str(n['code'])).alignment = Alignment(horizontal="center")
                    ws.cell(row=row_idx, column=2, value=f"{indent}{n['name']}")

                    cell_bal = ws.cell(row=row_idx, column=3, value=float(n['balance']))
                    cell_bal.number_format = '#,##0.00'

                    if inc['has_comparison']:
                        cell_c = ws.cell(row=row_idx, column=4, value=float(n['comparison_balance']))
                        cell_c.number_format = '#,##0.00'

                        cell_diff = ws.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}")
                        cell_diff.number_format = '#,##0.00'

                        cell_pct = ws.cell(row=row_idx, column=6, value=f"=IF(D{row_idx}=0, 1, (C{row_idx}-D{row_idx})/ABS(D{row_idx}))")
                        cell_pct.number_format = '0.00%'

                    for c in range(1, len(headers) + 1):
                        cell = ws.cell(row=row_idx, column=c)
                        cell.border = border_thin
                        if n['level'] == 1:
                            cell.font = font_group_1
                            cell.fill = fill_group_1
                        elif n['level'] == 2:
                            cell.font = font_group_2
                            cell.fill = fill_group_2
                        else:
                            cell.font = font_regular
                    row_idx += 1

                # صف إجمالي القسم
                ws.cell(row=row_idx, column=1, value="")
                ws.cell(row=row_idx, column=2, value=f"إجمالي {title}")
                cell_sub = ws.cell(row=row_idx, column=3, value=float(subtotal))
                cell_sub.number_format = '#,##0.00'

                if inc['has_comparison']:
                    cell_c_sub = ws.cell(row=row_idx, column=4, value=float(comp_subtotal))
                    cell_c_sub.number_format = '#,##0.00'
                    cell_d = ws.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}")
                    cell_d.number_format = '#,##0.00'
                    cell_p = ws.cell(row=row_idx, column=6, value=f"=IF(D{row_idx}=0, 1, (C{row_idx}-D{row_idx})/ABS(D{row_idx}))")
                    cell_p.number_format = '0.00%'

                for c in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=c)
                    cell.font = font_group_2
                    cell.fill = fill_subtotal
                    cell.border = border_thin
                subtotal_row_number = row_idx
                row_idx += 2
                return subtotal_row_number

            # 1. إيرادات النشاط
            rev_row = write_section(
                "إيرادات النشاط التشغيلي (Operating Revenue)",
                inc['operating_revenues']['nodes'],
                inc['operating_revenues']['total'],
                inc['operating_revenues']['comp_total']
            )

            # 2. تكلفة المبيعات
            cogs_row = write_section(
                "تكلفة المبيعات (Cost of Goods Sold)",
                inc['cogs']['nodes'],
                inc['cogs']['total'],
                inc['cogs']['comp_total']
            )

            # صف مجمل الربح (Gross Profit)
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value="مجمل الربح / الخسارة (GROSS PROFIT)")
            ws.cell(row=row_idx, column=3, value=f"=C{rev_row}-C{cogs_row}").number_format = '#,##0.00'
            if inc['has_comparison']:
                ws.cell(row=row_idx, column=4, value=f"=D{rev_row}-D{cogs_row}").number_format = '#,##0.00'
                ws.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}").number_format = '#,##0.00'
                ws.cell(row=row_idx, column=6, value=f"=IF(D{row_idx}=0, 1, (C{row_idx}-D{row_idx})/ABS(D{row_idx}))").number_format = '0.00%'

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_thin
            gross_profit_row = row_idx
            row_idx += 2

            # 3. المصروفات التشغيلية والعمومية
            op_exp_row = write_section(
                "المصروفات التشغيلية والعمومية والإدارية (Operating Expenses)",
                inc['operating_expenses']['nodes'],
                inc['operating_expenses']['total'],
                inc['operating_expenses']['comp_total']
            )

            # صف الربح التشغيلي (Operating Profit)
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value="الربح / الخسارة التشغيلية (OPERATING PROFIT / EBIT)")
            ws.cell(row=row_idx, column=3, value=f"=C{gross_profit_row}-C{op_exp_row}").number_format = '#,##0.00'
            if inc['has_comparison']:
                ws.cell(row=row_idx, column=4, value=f"=D{gross_profit_row}-D{op_exp_row}").number_format = '#,##0.00'
                ws.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}").number_format = '#,##0.00'
                ws.cell(row=row_idx, column=6, value=f"=IF(D{row_idx}=0, 1, (C{row_idx}-D{row_idx})/ABS(D{row_idx}))").number_format = '0.00%'

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_thin
            operating_profit_row = row_idx
            row_idx += 2

            # 4. الإيرادات الأخرى
            other_rev_row = write_section(
                "الإيرادات الأخرى وأرباح فروق العملة (Other Revenues & FX Gains)",
                inc['other_revenues']['nodes'],
                inc['other_revenues']['total'],
                inc['other_revenues']['comp_total']
            )

            # 5. المصروفات الأخرى
            other_exp_row = write_section(
                "المصروفات التمويلية والأخرى وخسائر العملة (Other Expenses & FX Losses)",
                inc['other_expenses']['nodes'],
                inc['other_expenses']['total'],
                inc['other_expenses']['comp_total']
            )

            # صف صافي الربح النهائي (NET PROFIT / LOSS)
            ws.cell(row=row_idx, column=1, value="")
            ws.cell(row=row_idx, column=2, value="صافي الربح / الخسارة النهائي (NET PROFIT / LOSS)")
            ws.cell(row=row_idx, column=3, value=f"=C{operating_profit_row}+C{other_rev_row}-C{other_exp_row}").number_format = '#,##0.00'
            if inc['has_comparison']:
                ws.cell(row=row_idx, column=4, value=f"=D{operating_profit_row}+D{other_rev_row}-D{other_exp_row}").number_format = '#,##0.00'
                ws.cell(row=row_idx, column=5, value=f"=C{row_idx}-D{row_idx}").number_format = '#,##0.00'
                ws.cell(row=row_idx, column=6, value=f"=IF(D{row_idx}=0, 1, (C{row_idx}-D{row_idx})/ABS(D{row_idx}))").number_format = '0.00%'

            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=c)
                cell.font = font_total
                cell.fill = fill_total
                cell.border = border_thin
            row_idx += 2

            # ضبط عرض الأعمدة
            ws.column_dimensions['A'].width = 16
            ws.column_dimensions['B'].width = 45
            ws.column_dimensions['C'].width = 24
            if inc['has_comparison']:
                ws.column_dimensions['D'].width = 24
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 16

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()

        except Exception as e:
            logger.error(f"خطأ في تصدير Excel لقائمة الدخل: {e}", exc_info=True)
            raise
