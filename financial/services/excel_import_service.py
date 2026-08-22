import logging
from decimal import Decimal
from typing import Dict, Any, List, Tuple
from django.core.exceptions import ValidationError
from django.utils.translation import gettext as _

from financial.models.opening_balance import OpeningBalanceBatch, OpeningBalanceLine, OpeningBalanceImportBatch
from financial.models.chart_of_accounts import ChartOfAccounts
from financial.models.currency import Currency

logger = logging.getLogger("financial.excel_import_service")


class ExcelImportService:
    """
    خدمة استيراد شيتات الأرصدة الافتتاحية مجمعة (4-Step Pipeline: Parse -> Validate -> Preview -> Commit)
    """

    @classmethod
    def parse(cls, file_obj, template_version: str = 'v1.0') -> List[Dict[str, Any]]:
        """
        1. قراءة واستخراج البيانات من ملف Excel أو CSV
        """
        rows = []
        filename = getattr(file_obj, 'name', 'import.csv').lower()

        if filename.endswith('.csv'):
            import csv
            import io
            content = file_obj.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for r in reader:
                rows.append(r)
        else:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_obj, data_only=True)
                ws = wb.active
                headers = [str(cell.value or '').strip() for cell in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                    rows.append(row_dict)
            except Exception as e:
                logger.error(f"Error parsing Excel file: {e}")
                raise ValidationError(_("فشل قراءة ملف Excel: {}").format(str(e)))

        return rows

    @classmethod
    def validate_rows(cls, raw_rows: List[Dict[str, Any]], batch: OpeningBalanceBatch = None) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        """
        2. فحص والتحقق من صحة كود الحساب والعملة والمبالغ
        """
        valid_rows = []
        invalid_rows = []

        all_accounts = {a.code: a for a in ChartOfAccounts.objects.filter(is_active=True, is_leaf=True)}
        all_currencies = {c.code.upper(): c for c in Currency.objects.filter(is_active=True)}
        
        try:
            from client.models import Customer
            all_customers = {c.code: c for c in Customer.objects.filter(is_active=True) if hasattr(c, 'code') and c.code}
            all_customers.update({str(c.id): c for c in Customer.objects.filter(is_active=True)})
        except Exception:
            all_customers = {}

        try:
            from supplier.models import Supplier
            all_suppliers = {s.code: s for s in Supplier.objects.filter(is_active=True) if hasattr(s, 'code') and s.code}
            all_suppliers.update({str(s.id): s for s in Supplier.objects.filter(is_active=True)})
        except Exception:
            all_suppliers = {}

        opening_date = batch.opening_date if batch else None

        for idx, row in enumerate(raw_rows, start=2):
            acc_code = str(row.get('account_code') or row.get('كود الحساب') or '').strip()
            debit_val = Decimal(str(row.get('debit') or row.get('مدين') or '0.00'))
            credit_val = Decimal(str(row.get('credit') or row.get('دائن') or '0.00'))
            line_type = str(row.get('line_type') or row.get('نوع السطر') or 'GL').strip().upper()
            curr_code = str(row.get('currency_code') or row.get('currency') or row.get('العملة') or 'EGP').strip().upper()
            rate_raw = row.get('exchange_rate') or row.get('سعر الصرف') or row.get('rate')

            customer_ref = str(row.get('customer_code') or row.get('كود العميل') or row.get('العميل') or '').strip()
            supplier_ref = str(row.get('supplier_code') or row.get('كود المورد') or row.get('المورد') or '').strip()

            customer_obj = all_customers.get(customer_ref) if customer_ref else None
            supplier_obj = all_suppliers.get(supplier_ref) if supplier_ref else None

            # Auto-resolve account if subledger provided
            account = all_accounts.get(acc_code)
            if not account:
                if line_type == 'AR' and customer_obj:
                    if not getattr(customer_obj, 'financial_account', None):
                        from client.services.customer_service import CustomerService
                        account = CustomerService.create_financial_account_for_customer(customer_obj)
                    else:
                        account = customer_obj.financial_account
                elif line_type == 'AP' and supplier_obj:
                    if not getattr(supplier_obj, 'financial_account', None):
                        from supplier.services.supplier_service import SupplierService
                        account = SupplierService.create_financial_account_for_supplier(supplier_obj)
                    else:
                        account = supplier_obj.financial_account
                elif line_type == 'INVENTORY':
                    from financial.services.role_registry import AccountRoleRegistry
                    account = AccountRoleRegistry.get_account("INVENTORY_GENERAL")

            errors = []

            if not account:
                errors.append(_("كود الحساب ({}) غير موجود أو غير نشط").format(acc_code))
            if debit_val < 0 or credit_val < 0:
                errors.append(_("المبالغ لا يمكن أن تكون بالسالب"))

            curr_obj = None
            rate = Decimal('1.000000')
            debit_foreign = Decimal('0.00')
            credit_foreign = Decimal('0.00')
            debit_func = debit_val
            credit_func = credit_val

            if curr_code and curr_code != 'EGP':
                curr_obj = all_currencies.get(curr_code)
                if not curr_obj:
                    errors.append(_("رمز العملة ({}) غير معرف بالنظام").format(curr_code))
                else:
                    if rate_raw:
                        try:
                            rate = Decimal(str(rate_raw))
                            if rate <= Decimal('0.00'):
                                errors.append(_("سعر الصرف يجب أن يكون أكبر من الصفر"))
                        except Exception:
                            errors.append(_("قيمة سعر الصرف غير صالحة"))
                    else:
                        from financial.services.exchange_rate_service import ExchangeRateService
                        try:
                            rate = ExchangeRateService.get_rate(curr_code, 'EGP', date=opening_date)
                        except Exception:
                            rate = Decimal('1.000000')

                    debit_foreign = debit_val
                    credit_foreign = credit_val
                    debit_func = (debit_foreign * rate).quantize(Decimal('0.01'))
                    credit_func = (credit_foreign * rate).quantize(Decimal('0.01'))

            if errors:
                invalid_rows.append({'row_number': idx, 'data': row, 'errors': errors})
            else:
                valid_rows.append({
                    'row_number': idx,
                    'account': account,
                    'line_type': line_type,
                    'currency': curr_obj,
                    'exchange_rate': rate,
                    'debit_foreign': debit_foreign,
                    'credit_foreign': credit_foreign,
                    'debit': debit_func,
                    'credit': credit_func,
                    'customer': customer_obj,
                    'supplier': supplier_obj
                })

        return valid_rows, invalid_rows

    @classmethod
    def commit(cls, batch: OpeningBalanceBatch, valid_rows: List[Dict[str, Any]], user, filename: str = 'import.xlsx', template_version: str = 'v1.0') -> OpeningBalanceImportBatch:
        """
        4. حفظ وحقن الأسطر الصحيحة بالدفعة المسودة وتوثيق سجل الاستيراد
        """
        line_objects = []
        for v in valid_rows:
            line_objects.append(OpeningBalanceLine(
                batch=batch,
                account=v['account'],
                line_type=v.get('line_type', 'GL'),
                currency=v.get('currency'),
                exchange_rate=v.get('exchange_rate', Decimal('1.000000')),
                debit_foreign=v.get('debit_foreign', Decimal('0.00')),
                credit_foreign=v.get('credit_foreign', Decimal('0.00')),
                debit=v['debit'],
                credit=v['credit'],
                customer=v.get('customer'),
                supplier=v.get('supplier')
            ))

        OpeningBalanceLine.objects.bulk_create(line_objects)

        import_record = OpeningBalanceImportBatch.objects.create(
            file_name=filename,
            template_version=template_version,
            uploaded_by=user,
            total_rows=len(valid_rows),
            valid_rows=len(valid_rows),
            invalid_rows=0
        )
        return import_record
