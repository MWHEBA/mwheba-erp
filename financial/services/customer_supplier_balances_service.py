"""
FIN-CORE-REPORTING: CustomerSupplierBalancesService
خدمة تقارير أعمار الديون ومطابقة أرصدة العملاء والموردين (AR / AP Aging & Balances)
تحليل الفواتير والقيود المحاسبية حسب فترات الاستحقاق وتصدير الإكسيل.
"""

from django.db.models import Sum, Q, F, Case, When, Value, DecimalField
from django.db.models.functions import Coalesce
from django.db import models
from django.utils import timezone
from decimal import Decimal
from datetime import date, timedelta
from typing import Dict, List, Optional, Any
import logging

logger = logging.getLogger(__name__)


class CustomerSupplierBalancesService:
    """خدمة تقارير أعمار ديون وأرصدة العملاء والموردين"""

    def __init__(self, as_of_date: Optional[date] = None):
        """
        تهيئة الخدمة
        Args:
            as_of_date: تاريخ التقرير (افتراضي: اليوم)
        """
        self.as_of_date = as_of_date or timezone.now().date()

    def generate_customer_balances_report(self) -> Dict[str, Any]:
        """
        إنشاء تقرير أرصدة العملاء (AR Aging Report)
        """
        try:
            from customer.models import Customer
        except ImportError:
            return {
                "error": "نماذج العملاء غير متوفرة",
                "accounts": [],
                "due_periods": {},
                "summary": {},
            }

        # جلب جميع العملاء
        customers = Customer.objects.all().select_related("financial_account")

        accounts_data = []
        total_current = Decimal("0")
        total_30 = Decimal("0")
        total_60 = Decimal("0")
        total_90 = Decimal("0")
        total_over_90 = Decimal("0")

        for customer in customers:
            account_data = self._calculate_customer_balance(customer)

            if account_data["total_balance"] > 0:
                accounts_data.append(account_data)
                total_current += account_data["current"]
                total_30 += account_data["days_1_30"]
                total_60 += account_data["days_31_60"]
                total_90 += account_data["days_61_90"]
                total_over_90 += account_data["over_90"]

        accounts_data.sort(key=lambda x: x["total_balance"], reverse=True)
        total_balance = total_current + total_30 + total_60 + total_90 + total_over_90

        due_periods = {
            "current": {
                "amount": total_current,
                "percentage": (total_current / total_balance * 100) if total_balance > 0 else 0,
                "label": "حالي (0-30 يوم)",
                "days": "0-30",
            },
            "days_1_30": {
                "amount": total_30,
                "percentage": (total_30 / total_balance * 100) if total_balance > 0 else 0,
                "label": "31-60 يوم",
                "days": "31-60",
            },
            "days_31_60": {
                "amount": total_60,
                "percentage": (total_60 / total_balance * 100) if total_balance > 0 else 0,
                "label": "61-90 يوم",
                "days": "61-90",
            },
            "days_61_90": {
                "amount": total_90,
                "percentage": (total_90 / total_balance * 100) if total_balance > 0 else 0,
                "label": "91-120 يوم",
                "days": "91-120",
            },
            "over_90": {
                "amount": total_over_90,
                "percentage": (total_over_90 / total_balance * 100) if total_balance > 0 else 0,
                "label": "أكثر من 120 يوم",
                "days": "120+",
            },
        }

        summary = {
            "total_balance": total_balance,
            "total_accounts": len(accounts_data),
            "as_of_date": self.as_of_date,
        }

        return {
            "accounts": accounts_data,
            "due_periods": due_periods,
            "summary": summary,
            "as_of_date": self.as_of_date,
        }

    def generate_supplier_balances_report(self) -> Dict[str, Any]:
        """
        إنشاء تقرير أرصدة الموردين (AP Aging Report)
        """
        try:
            from supplier.models import Supplier
        except ImportError:
            return {
                "error": "نماذج الموردين غير متوفرة",
                "accounts": [],
                "due_periods": {},
                "summary": {},
            }

        suppliers = Supplier.objects.filter(is_active=True).select_related("financial_account")

        accounts_data = []
        total_current = Decimal("0")
        total_30 = Decimal("0")
        total_60 = Decimal("0")
        total_90 = Decimal("0")
        total_over_90 = Decimal("0")

        for supplier in suppliers:
            account_data = self._calculate_supplier_balance(supplier)

            if account_data["total_balance"] > 0:
                accounts_data.append(account_data)
                total_current += account_data["current"]
                total_30 += account_data["days_1_30"]
                total_60 += account_data["days_31_60"]
                total_90 += account_data["days_61_90"]
                total_over_90 += account_data["over_90"]

        accounts_data.sort(key=lambda x: x["total_balance"], reverse=True)
        total_balance = total_current + total_30 + total_60 + total_90 + total_over_90

        due_periods = {
            "current": {
                "amount": total_current,
                "percentage": (total_current / total_balance * 100) if total_balance > 0 else 0,
                "label": "حالي (0-30 يوم)",
                "days": "0-30",
            },
            "days_1_30": {
                "amount": total_30,
                "percentage": (total_30 / total_balance * 100) if total_balance > 0 else 0,
                "label": "31-60 يوم",
                "days": "31-60",
            },
            "days_31_60": {
                "amount": total_60,
                "percentage": (total_60 / total_balance * 100) if total_balance > 0 else 0,
                "label": "61-90 يوم",
                "days": "61-90",
            },
            "days_61_90": {
                "amount": total_90,
                "percentage": (total_90 / total_balance * 100) if total_balance > 0 else 0,
                "label": "91-120 يوم",
                "days": "91-120",
            },
            "over_90": {
                "amount": total_over_90,
                "percentage": (total_over_90 / total_balance * 100) if total_balance > 0 else 0,
                "label": "أكثر من 120 يوم",
                "days": "120+",
            },
        }

        summary = {
            "total_balance": total_balance,
            "total_accounts": len(accounts_data),
            "as_of_date": self.as_of_date,
        }

        return {
            "accounts": accounts_data,
            "due_periods": due_periods,
            "summary": summary,
            "as_of_date": self.as_of_date,
        }

    def _calculate_customer_balance(self, customer) -> Dict[str, Any]:
        """حساب رصيد العميل حسب فترات الاستحقاق"""
        try:
            from sale.models import Sale
        except ImportError:
            return self._empty_balance(customer.code if hasattr(customer, 'code') else '', customer.name)

        sales = Sale.objects.filter(
            customer=customer,
            issue_date__lte=self.as_of_date
        ).exclude(payment_status='paid')

        current = Decimal("0")
        days_1_30 = Decimal("0")
        days_31_60 = Decimal("0")
        days_61_90 = Decimal("0")
        over_90 = Decimal("0")

        for sale in sales:
            due_date = sale.due_date or sale.issue_date
            remaining = getattr(sale, 'remaining_amount', getattr(sale, 'grand_total', Decimal("0"))) or Decimal("0")

            if remaining <= 0:
                continue

            if due_date > self.as_of_date:
                current += remaining
            else:
                days_overdue = (self.as_of_date - due_date).days
                if days_overdue <= 30:
                    days_1_30 += remaining
                elif days_overdue <= 60:
                    days_31_60 += remaining
                elif days_overdue <= 90:
                    days_61_90 += remaining
                else:
                    over_90 += remaining

        total_balance = current + days_1_30 + days_31_60 + days_61_90 + over_90
        account_code = customer.financial_account.code if getattr(customer, "financial_account", None) else (customer.code or "")

        return {
            "account_code": account_code,
            "account_name": customer.name,
            "current": current,
            "days_1_30": days_1_30,
            "days_31_60": days_31_60,
            "days_61_90": days_61_90,
            "over_90": over_90,
            "total_balance": total_balance,
        }

    def _calculate_supplier_balance(self, supplier) -> Dict[str, Any]:
        """حساب رصيد المورد حسب فترات الاستحقاق"""
        try:
            from purchase.models import Purchase
        except ImportError:
            return self._empty_balance(supplier.code if hasattr(supplier, 'code') else '', supplier.name)

        purchases = Purchase.objects.filter(
            supplier=supplier,
            date__lte=self.as_of_date
        ).exclude(payment_status='paid')

        current = Decimal("0")
        days_1_30 = Decimal("0")
        days_31_60 = Decimal("0")
        days_61_90 = Decimal("0")
        over_90 = Decimal("0")

        for purchase in purchases:
            due_date = purchase.due_date or purchase.date
            remaining = getattr(purchase, 'remaining_amount', getattr(purchase, 'total', Decimal("0"))) or Decimal("0")

            if remaining <= 0:
                continue

            if due_date > self.as_of_date:
                current += remaining
            else:
                days_overdue = (self.as_of_date - due_date).days
                if days_overdue <= 30:
                    days_1_30 += remaining
                elif days_overdue <= 60:
                    days_31_60 += remaining
                elif days_overdue <= 90:
                    days_61_90 += remaining
                else:
                    over_90 += remaining

        total_balance = current + days_1_30 + days_31_60 + days_61_90 + over_90
        account_code = supplier.financial_account.code if getattr(supplier, "financial_account", None) else (supplier.code or "")

        return {
            "account_code": account_code,
            "account_name": supplier.name,
            "current": current,
            "days_1_30": days_1_30,
            "days_31_60": days_31_60,
            "days_61_90": days_61_90,
            "over_90": over_90,
            "total_balance": total_balance,
        }

    def _empty_balance(self, code: str, name: str) -> Dict[str, Any]:
        return {
            "account_code": code,
            "account_name": name,
            "current": Decimal("0"),
            "days_1_30": Decimal("0"),
            "days_31_60": Decimal("0"),
            "days_61_90": Decimal("0"),
            "over_90": Decimal("0"),
            "total_balance": Decimal("0"),
        }

    def export_to_excel(self, report_data: Dict[str, Any], report_type: str = "ar") -> bytes:
        """
        تصدير التقرير إلى Excel
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO

            wb = openpyxl.Workbook()
            ws = wb.active

            title = "تقرير أرصدة العملاء" if report_type == "ar" else "تقرير أرصدة الموردين"
            ws.title = title

            title_font = Font(name="Arial", size=16, bold=True)
            header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

            # العنوان والتاريخ
            ws["A1"] = title
            ws["A1"].font = title_font
            ws.merge_cells("A1:H1")

            ws["A2"] = f"كما في: {self.as_of_date}"
            ws.merge_cells("A2:H2")

            # العناوين
            row = 4
            headers = ["الكود", "الاسم", "حالي (0-30)", "31-60 يوم", "61-90 يوم", "91-120 يوم", "+120 يوم", "الإجمالي"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # البيانات
            for account in report_data.get("accounts", []):
                row += 1
                ws.cell(row=row, column=1, value=account["account_code"])
                ws.cell(row=row, column=2, value=account["account_name"])
                ws.cell(row=row, column=3, value=float(account["current"]))
                ws.cell(row=row, column=4, value=float(account["days_1_30"]))
                ws.cell(row=row, column=5, value=float(account["days_31_60"]))
                ws.cell(row=row, column=6, value=float(account["days_61_90"]))
                ws.cell(row=row, column=7, value=float(account["over_90"]))
                ws.cell(row=row, column=8, value=float(account["total_balance"]))

            # الإجماليات
            row += 2
            ws.cell(row=row, column=2, value="الإجمالي").font = Font(bold=True)
            due_periods = report_data.get("due_periods", {})
            if due_periods:
                ws.cell(row=row, column=3, value=float(due_periods.get("current", {}).get("amount", 0))).font = Font(bold=True)
                ws.cell(row=row, column=4, value=float(due_periods.get("days_1_30", {}).get("amount", 0))).font = Font(bold=True)
                ws.cell(row=row, column=5, value=float(due_periods.get("days_31_60", {}).get("amount", 0))).font = Font(bold=True)
                ws.cell(row=row, column=6, value=float(due_periods.get("days_61_90", {}).get("amount", 0))).font = Font(bold=True)
                ws.cell(row=row, column=7, value=float(due_periods.get("over_90", {}).get("amount", 0))).font = Font(bold=True)
            summary = report_data.get("summary", {})
            ws.cell(row=row, column=8, value=float(summary.get("total_balance", 0))).font = Font(bold=True)

            # تنسيق عرض الأعمدة
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 32
            for col in ["C", "D", "E", "F", "G", "H"]:
                ws.column_dimensions[col].width = 16

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        except ImportError:
            logger.error("❌ مكتبة openpyxl غير مثبتة لتصدير الإكسيل.")
            return b""
