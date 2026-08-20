from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, Avg
from django.db import transaction
from django.http import JsonResponse
from django.utils import timezone
from django.utils.translation import gettext as _
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.contenttypes.models import ContentType
from datetime import date, datetime, timedelta
from decimal import Decimal
from django.db import models
import json
import logging

logger = logging.getLogger(__name__)


def generate_balance_sheet_optimized(balance_date, group_by_subtype=True):
    """تفويض مباشر إلى BalanceSheetService"""
    from financial.services.balance_sheet_service import BalanceSheetService
    return BalanceSheetService.generate_balance_sheet(as_of_date=balance_date, group_by_subtype=group_by_subtype)


def calculate_financial_ratios_optimized(balance_sheet_data):
    """تفويض مباشر إلى BalanceSheetService"""
    from financial.services.balance_sheet_service import BalanceSheetService
    return BalanceSheetService.calculate_financial_ratios(balance_sheet_data)


def generate_balance_sheet_excel_optimized(balance_date, group_by_subtype=True):
    """تفويض مباشر إلى BalanceSheetService"""
    from financial.services.balance_sheet_service import BalanceSheetService
    return BalanceSheetService.export_to_excel(as_of_date=balance_date)


def handle_progressive_ledger_load(request, account_id, date_from, date_to, page_size, cost_center=None, include_unposted=False):
    """
    معالجة التحميل التدريجي لبيانات كشف الحساب
    """
    try:
        page = int(request.GET.get('page', 1))
        account = get_object_or_404(ChartOfAccounts, id=account_id)
        
        # جلب المعاملات بشكل دقيق عبر LedgerQueryService
        transactions, _ = get_account_transactions_optimized(
            account, date_from, date_to, cost_center=cost_center, include_unposted=include_unposted
        )
        
        # تطبيق Pagination
        paginator = Paginator(transactions, page_size)
        page_obj = paginator.get_page(page)
        
        # تحويل البيانات إلى JSON
        transactions_data = []
        for transaction in page_obj:
            transactions_data.append({
                'date': transaction.get('date', '').strftime('%Y-%m-%d') if transaction.get('date') else '',
                'reference': transaction.get('reference', ''),
                'description': transaction.get('description', ''),
                'debit': float(transaction.get('debit', 0)),
                'credit': float(transaction.get('credit', 0)),
                'balance': float(transaction.get('balance', 0)),
                'currency': transaction.get('currency', 'EGP'),
                'foreign_debit': float(transaction.get('foreign_debit', 0)),
                'foreign_credit': float(transaction.get('foreign_credit', 0)),
                'cost_center_name': transaction.get('cost_center_name', ''),
                'is_reversal': transaction.get('is_reversal', False),
            })
        
        return JsonResponse({
            'success': True,
            'transactions': transactions_data,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'page_number': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_count': paginator.count,
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def get_account_transactions_optimized(account, date_from=None, date_to=None, cost_center=None, currency=None, include_unposted=False):
    """
    جلب معاملات كشف الحساب بدقة محاسبية 100% عبر LedgerQueryService
    مع الدعم الكامل للعملات الأجنبية ومعيار IAS 21
    """
    from ..services.ledger_query_service import LedgerQueryService
    statement_data = LedgerQueryService.get_account_statement(
        account_or_id=account,
        start_date=date_from,
        end_date=date_to,
        cost_center=cost_center,
        currency=currency,
        include_unposted=include_unposted
    )
    transactions = statement_data['transactions']
    summary = {
        'opening_balance': statement_data['opening_balance'],
        'opening_balance_foreign': statement_data.get('opening_balance_foreign', Decimal('0.00')),
        'total_debit': statement_data['total_debit'],
        'total_credit': statement_data['total_credit'],
        'total_foreign_debit': statement_data.get('total_foreign_debit', Decimal('0.00')),
        'total_foreign_credit': statement_data.get('total_foreign_credit', Decimal('0.00')),
        'closing_balance': statement_data['closing_balance'],
        'closing_balance_foreign': statement_data.get('closing_balance_foreign', Decimal('0.00')),
        'period_movement': statement_data['period_movement'],
        'period_foreign_movement': statement_data.get('period_foreign_movement', Decimal('0.00')),
        'transaction_count': statement_data['transaction_count'],
        'account_currency': statement_data.get('account_currency', 'EGP'),
        'account_currency_symbol': statement_data.get('account_currency_symbol', 'ج.م'),
        'is_foreign_account': statement_data.get('is_foreign_account', False),
        'is_consolidated': statement_data.get('is_consolidated', False),
        'functional_currency': statement_data.get('functional_currency', 'EGP'),
        'functional_currency_symbol': statement_data.get('functional_currency_symbol', 'ج.م'),
        'by_currency_breakdown': statement_data.get('by_currency_breakdown', {}),
    }
    return transactions, summary


def get_all_accounts_summary_optimized(date_from=None, date_to=None):
    """
    جلب ملخص جميع الحسابات بطريقة محسنة
    """
    from django.db.models import Sum, Count, Q
    from decimal import Decimal
    
    # جلب الحسابات النهائية النشطة
    accounts = ChartOfAccounts.objects.filter(
        is_leaf=True,
        is_active=True
    ).select_related('account_type').order_by('code')
    
    summaries = []
    
    for account in accounts:
        # استخدام الطريقة المحسنة لحساب الرصيد
        current_balance = account.get_balance_optimized(
            date_from=date_from,
            date_to=date_to,
            use_cache=True
        )
        
        # جلب إحصائيات المعاملات
        from ..models import JournalEntryLine
        
        query = Q(account=account, journal_entry__status="posted")
        if date_from:
            query &= Q(journal_entry__date__gte=date_from)
        if date_to:
            query &= Q(journal_entry__date__lte=date_to)
        
        stats = JournalEntryLine.objects.filter(query).aggregate(
            total_debit=Sum('debit'),
            total_credit=Sum('credit'),
            transaction_count=Count('id')
        )
        
        # إضافة الملخص فقط للحسابات التي لها حركة أو رصيد
        if (current_balance != 0 or 
            stats['transaction_count'] > 0 or 
            (account.opening_balance and account.opening_balance != 0)):
            
            summaries.append({
                'account': account,
                'current_balance': current_balance,
                'total_debit': stats['total_debit'] or Decimal('0'),
                'total_credit': stats['total_credit'] or Decimal('0'),
                'transaction_count': stats['transaction_count'] or 0,
                'net_movement': (stats['total_debit'] or Decimal('0')) - (stats['total_credit'] or Decimal('0')),
                'account_type': account.account_type.name,
                'category': account.category
            })
    
    return summaries


def generate_ledger_excel_optimized(account, date_from=None, date_to=None, cost_center=None, currency=None, include_unposted=False):
    """
    تصدير كشف الحساب إلى Excel بطريقة محسنة مع الدعم الكامل للعملات المتعددة
    """
    try:
        import re
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        title_font = Font(bold=True, size=14, color="1E3A8A")
        subheader_font = Font(bold=True, size=11)
        normal_font = Font(size=10)
        border = Border(
            left=Side(style='thin', color="D1D5DB"),
            right=Side(style='thin', color="D1D5DB"),
            top=Side(style='thin', color="D1D5DB"),
            bottom=Side(style='thin', color="D1D5DB")
        )
        
        if account:
            raw_title = f"كشف حساب - {account.code}"
            ws.title = re.sub(r'[\/\\\*\?\:\[\]]', '_', raw_title)[:31]
            
            # جلب المعاملات
            transactions, summary = get_account_transactions_optimized(
                account, date_from, date_to, cost_center=cost_center, currency=currency, include_unposted=include_unposted
            )
            
            # العنوان الرئيسي
            ws.merge_cells('A1:G1')
            ws['A1'] = f"كشف حساب - {account.name} ({account.code})"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # معلومات الفترة
            period_text = f"الفترة: من {date_from or 'البداية'} إلى {date_to or 'النهاية'}"
            ws.merge_cells('A2:G2')
            ws['A2'] = period_text
            ws['A2'].font = subheader_font
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            # رؤوس الأعمدة
            headers = ['التاريخ', 'النوع', 'البيان', 'المرجع', 'مدين', 'دائن', 'الرصيد']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row = 5
            # سطر الرصيد الافتتاحي
            ws.cell(row=row, column=1, value=date_from.strftime('%Y-%m-%d') if date_from else '')
            ws.cell(row=row, column=2, value="-")
            ws.cell(row=row, column=3, value="الرصيد الافتتاحي / المنقول")
            ws.cell(row=row, column=4, value="-")
            ws.cell(row=row, column=5, value="-")
            ws.cell(row=row, column=6, value="-")
            try:
                op_val = float(summary.get('opening_balance') or 0)
            except (ValueError, TypeError):
                op_val = 0.0
            cell_op = ws.cell(row=row, column=7, value=op_val)
            cell_op.number_format = '#,##0.00'
            ws.cell(row=row, column=3).font = subheader_font
            cell_op.font = subheader_font
            
            # البيانات
            for transaction in transactions:
                row += 1
                d_val = transaction.get('date')
                if hasattr(d_val, 'strftime'):
                    d_str = d_val.strftime('%Y-%m-%d')
                else:
                    d_str = str(d_val or '')
                ws.cell(row=row, column=1, value=d_str)
                
                type_display = str(transaction.get('entry_type_display') or transaction.get('journal_number') or '')
                ws.cell(row=row, column=2, value=type_display)
                
                desc = str(transaction.get('description') or '')
                ws.cell(row=row, column=3, value=desc)
                
                ref = str(transaction.get('reference') or '')
                ws.cell(row=row, column=4, value=ref)
                
                try:
                    d_amount = float(transaction.get('debit') or 0)
                except (ValueError, TypeError):
                    d_amount = 0.0
                    
                try:
                    c_amount = float(transaction.get('credit') or 0)
                except (ValueError, TypeError):
                    c_amount = 0.0
                    
                bal_raw = transaction.get('running_balance') if transaction.get('running_balance') is not None else transaction.get('balance', 0)
                try:
                    b_amount = float(bal_raw or 0)
                except (ValueError, TypeError):
                    b_amount = 0.0
                
                cell_d = ws.cell(row=row, column=5, value=d_amount)
                cell_c = ws.cell(row=row, column=6, value=c_amount)
                cell_b = ws.cell(row=row, column=7, value=b_amount)
                
                cell_d.number_format = '#,##0.00'
                cell_c.number_format = '#,##0.00'
                cell_b.number_format = '#,##0.00'
                
                for c in range(1, 8):
                    ws.cell(row=row, column=c).border = border
                    ws.cell(row=row, column=c).font = normal_font
            
            # الإجمالي
            row += 1
            ws.cell(row=row, column=3, value="الإجمالي:").font = subheader_font
            try:
                tot_d = float(summary.get('total_debit') or 0)
            except (ValueError, TypeError):
                tot_d = 0.0

            try:
                tot_c = float(summary.get('total_credit') or 0)
            except (ValueError, TypeError):
                tot_c = 0.0

            try:
                tot_b = float(summary.get('closing_balance') or 0)
            except (ValueError, TypeError):
                tot_b = 0.0

            cell_td = ws.cell(row=row, column=5, value=tot_d)
            cell_tc = ws.cell(row=row, column=6, value=tot_c)
            cell_cb = ws.cell(row=row, column=7, value=tot_b)
            
            cell_td.font = subheader_font
            cell_tc.font = subheader_font
            cell_cb.font = subheader_font
            cell_td.number_format = '#,##0.00'
            cell_tc.number_format = '#,##0.00'
            cell_cb.number_format = '#,##0.00'
            
        else:
            # تصدير ملخص جميع الحسابات
            ws.title = "كشف حركة الحسابات"
            
            summaries = get_all_accounts_summary_optimized(date_from, date_to)
            
            # العنوان
            ws.merge_cells('A1:G1')
            ws['A1'] = "كشف حركة وأرصدة الحسابات العامة"
            ws['A1'].font = title_font
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # رؤوس الأعمدة
            headers = ['كود الحساب', 'اسم الحساب', 'نوع الحساب', 'إجمالي مدين', 'إجمالي دائن', 'الرصيد الحالي', 'عدد المعاملات']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            row = 4
            for s in summaries:
                acc_obj = s.get('account')
                ws.cell(row=row, column=1, value=str(getattr(acc_obj, 'code', '')))
                ws.cell(row=row, column=2, value=str(getattr(acc_obj, 'name', '')))
                ws.cell(row=row, column=3, value=str(s.get('account_type') or ''))
                
                try:
                    cd = float(s.get('total_debit') or 0)
                except (ValueError, TypeError):
                    cd = 0.0
                try:
                    cc = float(s.get('total_credit') or 0)
                except (ValueError, TypeError):
                    cc = 0.0
                try:
                    cb = float(s.get('current_balance') or 0)
                except (ValueError, TypeError):
                    cb = 0.0
                try:
                    tc = int(s.get('transaction_count') or 0)
                except (ValueError, TypeError):
                    tc = 0
                
                cell_d = ws.cell(row=row, column=4, value=cd)
                cell_c = ws.cell(row=row, column=5, value=cc)
                cell_b = ws.cell(row=row, column=6, value=cb)
                ws.cell(row=row, column=7, value=tc)
                
                cell_d.number_format = '#,##0.00'
                cell_c.number_format = '#,##0.00'
                cell_b.number_format = '#,##0.00'
                
                for c in range(1, 8):
                    ws.cell(row=row, column=c).border = border
                    ws.cell(row=row, column=c).font = normal_font
                row += 1
        
        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        output = BytesIO()
        wb.save(output)
        return output.getvalue()
        
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Error generating Excel statement: {e}", exc_info=True)
        return None


# إضافة دالة محسنة لتصدير التقارير
def optimize_report_export(report_data, report_type="balance_sheet"):
    """
    تحسين تصدير التقارير مع ضغط البيانات الكبيرة
    """
    try:
        import gzip
        import json
        
        # تحويل البيانات إلى JSON
        json_data = json.dumps(report_data, default=str, ensure_ascii=False)
        
        # ضغط البيانات إذا كانت كبيرة (أكثر من 1MB)
        if len(json_data.encode('utf-8')) > 1024 * 1024:
            compressed_data = gzip.compress(json_data.encode('utf-8'))
            return compressed_data, True  # مضغوط
        else:
            return json_data.encode('utf-8'), False  # غير مضغوط
            
    except Exception:
        return None, False

# استيراد النماذج والخدمات الجديدة
from ..forms.expense_forms import ExpenseForm, ExpenseEditForm, ExpenseFilterForm
from ..forms.income_forms import IncomeForm, IncomeEditForm, IncomeFilterForm
from ..services.expense_income_service import ExpenseIncomeService
from ..services.account_helper import AccountHelperService

# استيراد النماذج الأساسية (موجودة بالتأكيد)
from ..models import (
    AccountType,
    ChartOfAccounts,
    AccountingPeriod,
    JournalEntry,
    JournalEntryLine,
)

# استيراد النماذج الاختيارية
try:
    from ..models import (
        AccountGroup,
        JournalEntryTemplate,
        JournalEntryTemplateLine,
        BalanceSnapshot,
        AccountBalanceCache,
        BalanceAuditLog,
        PaymentSyncOperation,
        PaymentSyncLog,
    )
except ImportError:
    # في حالة عدم توفر بعض النماذج الاختيارية
    AccountGroup = None
    JournalEntryTemplate = None
    JournalEntryTemplateLine = None
    BalanceSnapshot = None
    AccountBalanceCache = None
    BalanceAuditLog = None
    PaymentSyncOperation = None
    PaymentSyncLog = None

# استيراد النماذج القديمة للتوافق (اختيارية)
try:
    from ..models import Transaction, Account, TransactionLine, TransactionForm
except ImportError:
    # في حالة عدم توفر النماذج القديمة، إنشاء نماذج وهمية
    class Transaction:
        objects = type(
            "MockManager",
            (),
            {
                "filter": lambda *args, **kwargs: type(
                    "MockQuerySet",
                    (),
                    {
                        "order_by": lambda *args: [],
                        "aggregate": lambda *args: {"amount__sum": 0, "total": 0},
                        "count": lambda: 0,
                        "exists": lambda: False,
                    },
                )(),
                "create": lambda *args, **kwargs: None,
                "all": lambda: type(
                    "MockQuerySet", (), {"order_by": lambda *args: []}
                )(),
            },
        )()

    Account = ChartOfAccounts  # استخدام النموذج الجديد
    TransactionLine = JournalEntryLine
    TransactionForm = None

@login_required
def api_expense_accounts(request):
    """API لجلب حسابات المصروفات النهائية فقط (باستثناء تكلفة البضاعة المباعة)"""
    try:
        # جلب الحسابات النهائية (الفرعية) فقط من فئة المصروفات
        # استثناء: تكلفة البضاعة المباعة (خاصة بالمبيعات فقط)
        accounts = ChartOfAccounts.objects.filter(
            is_active=True, 
            is_leaf=True,  # الحسابات النهائية فقط
            account_type__category="expense"  # فئة المصروفات
        ).exclude(
            code__startswith='51'  # استثناء تكلفة البضاعة المباعة (51xxx)
        ).values('id', 'name', 'code').order_by('code')
        
        return JsonResponse(list(accounts), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_financial_categories(request):
    """API لجلب التصنيفات المالية النشطة للمصروفات والإيرادات مع الفرعية"""
    try:
        from ..models import FinancialCategory, FinancialSubcategory
        
        category_type = request.GET.get('type', 'expense')
        
        # جلب التصنيفات المالية النشطة
        categories_qs = FinancialCategory.objects.filter(is_active=True)
        
        if category_type == 'income':
            # للإيرادات: التصنيفات التي لها حساب إيراد افتراضي
            categories_qs = categories_qs.filter(default_revenue_account__isnull=False)
        else:
            # للمصروفات: التصنيفات التي لها حساب مصروف افتراضي
            categories_qs = categories_qs.filter(default_expense_account__isnull=False)
        
        categories_qs = categories_qs.order_by('display_order', 'name')
        
        # بناء القائمة مع التصنيفات الفرعية
        categories_list = []
        for category in categories_qs:
            cat_data = {
                'id': category.id,
                'name': category.name,
                'code': category.code,
                'subcategories': []
            }
            
            # جلب التصنيفات الفرعية النشطة
            subcategories = category.subcategories.filter(is_active=True).order_by('display_order', 'name')
            for subcat in subcategories:
                cat_data['subcategories'].append({
                    'id': subcat.id,
                    'name': subcat.name,
                    'code': subcat.code
                })
            
            categories_list.append(cat_data)
        
        return JsonResponse({'categories': categories_list}, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required 
def api_payment_accounts(request):
    """API لجلب حسابات الخزينة (نقدية وبنكية) النهائية فقط"""
    try:
        # جلب الحسابات النهائية للخزينة والبنوك فقط
        accounts = ChartOfAccounts.objects.filter(
            is_active=True, 
            is_leaf=True,  # الحسابات النهائية فقط
            account_type__category="asset"  # من فئة الأصول
        ).filter(
            models.Q(is_cash_account=True) |  # حسابات نقدية
            models.Q(is_bank_account=True) |  # حسابات بنكية
            models.Q(account_type__code="CASH") |  # نوع الخزينة
            models.Q(account_type__code="BANK")   # نوع البنوك
        ).values('id', 'name', 'code').order_by('code')
        
        return JsonResponse(list(accounts), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_income_accounts(request):
    """API لجلب حسابات الإيرادات النهائية فقط (باستثناء إيرادات المبيعات)"""
    try:
        # جلب الحسابات النهائية (الفرعية) فقط من فئة الإيرادات
        # استثناء: إيرادات المبيعات (خاصة بالمبيعات فقط)
        accounts = ChartOfAccounts.objects.filter(
            is_active=True, 
            is_leaf=True,  # الحسابات النهائية فقط
            account_type__category="revenue"  # فئة الإيرادات
        ).exclude(
            code__startswith='41'  # استثناء إيرادات المبيعات (41xxx)
        ).values('id', 'name', 'code').order_by('code')
        
        return JsonResponse(list(accounts), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def export_transactions(request):
    """
    تصدير المعاملات المالية
    """
    try:
        from .models.transactions import FinancialTransaction

        transactions = FinancialTransaction.objects.all().order_by("-date", "-id")
    except ImportError:
        # استخدام القيود المحاسبية كبديل
        transactions = JournalEntry.objects.all().order_by("-date", "-id")

    # تطبيق الفلترة إذا كانت موجودة
    account_id = request.GET.get("account")
    trans_type = request.GET.get("type")
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")

    if account_id:
        account = get_object_or_404(Account, id=account_id)
        transactions = transactions.filter(Q(account=account) | Q(to_account=account))

    if trans_type:
        transactions = transactions.filter(transaction_type=trans_type)

    if date_from:
        date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        transactions = transactions.filter(date__gte=date_from)

    if date_to:
        date_to = datetime.strptime(date_to, "%Y-%m-%d").date()
        transactions = transactions.filter(date__lte=date_to)

    # إنشاء ملف CSV
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="transactions.csv"'

    writer = csv.writer(response)
    writer.writerow(
        ["ID", "التاريخ", "النوع", "الحساب", "الوصف", "المبلغ", "الرقم المرجعي"]
    )

    for transaction in transactions:
        writer.writerow(
            [
                transaction.id,
                transaction.date,
                transaction.get_transaction_type_display(),
                transaction.account.name,
                transaction.description,
                transaction.amount,
                transaction.reference_number or "",
            ]
        )

    return response


@login_required
def ledger_report(request):
    """
    تقرير كشف الحسابات المالي - مطابق للمعايير المحاسبية وقواعد المشروع
    يستخدم محرك LedgerQueryService مع دعم البحث التفاعلي والعملات الأجنبية ومراكز التكلفة
    """
    from django.http import HttpResponse, JsonResponse
    from django.core.paginator import Paginator
    from django.template.loader import render_to_string
    from ..models import CostCenter
    
    # معالجة الفلاتر
    account_id = request.GET.get("account")
    date_from_str = request.GET.get("date_from")
    date_to_str = request.GET.get("date_to")
    cost_center_id = request.GET.get("cost_center")
    currency = request.GET.get("currency")
    include_unposted = request.GET.get("include_unposted") == "1"
    export_format = request.GET.get("export")  # excel / pdf
    page_size = int(request.GET.get("page_size", "50"))
    progressive_load = request.GET.get("progressive", "0") == "1"
    
    # تحويل التواريخ
    date_from = None
    date_to = None
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
        except ValueError:
            messages.warning(request, "تنسيق تاريخ البداية غير صحيح")
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except ValueError:
            messages.warning(request, "تنسيق تاريخ النهاية غير صحيح")
    
    # جلب مركز التكلفة إذا تم اختياره
    selected_cost_center = None
    if cost_center_id:
        try:
            selected_cost_center = CostCenter.objects.get(id=cost_center_id)
        except (CostCenter.DoesNotExist, ValueError):
            pass

    cost_centers = CostCenter.objects.filter(is_active=True).order_by('code')
    
    from ..models.currency import Currency
    currencies = Currency.objects.filter(is_active=True).order_by("-is_functional", "code")

    # جلب جميع الحسابات مع تجميعها للفئات
    accounts = ChartOfAccounts.objects.filter(
        is_active=True
    ).select_related('account_type', 'currency').only(
        'id', 'code', 'name', 'is_leaf', 'currency__code', 'currency__symbol', 'account_type__name', 'account_type__category'
    ).order_by('code')
    
    categories_def = [
        ('asset', 'الأصول'),
        ('liability', 'الالتزامات والخصوم'),
        ('equity', 'حقوق الملكية'),
        ('revenue', 'الإيرادات'),
        ('expense', 'المصروفات'),
    ]
    grouped_accounts = []
    for cat_key, cat_name in categories_def:
        cat_accs = [acc for acc in accounts if getattr(getattr(acc, 'account_type', None), 'category', '') == cat_key]
        if cat_accs:
            grouped_accounts.append({
                'category_key': cat_key,
                'category_name': cat_name,
                'accounts': cat_accs
            })
    
    # معالجة التصدير
    if export_format == 'excel':
        try:
            account = None
            if account_id:
                account = get_object_or_404(ChartOfAccounts, id=account_id)
            
            excel_data = generate_ledger_excel_optimized(
                account, date_from, date_to, cost_center=selected_cost_center, currency=currency, include_unposted=include_unposted
            )
            
            if excel_data:
                response = HttpResponse(
                    excel_data,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"account_statement_{account.code if account else 'all'}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            else:
                messages.error(request, "تعذر إنشاء ملف Excel")
        except Exception as e:
            messages.error(request, f"خطأ في تصدير Excel: {e}")

    elif export_format == 'pdf':
        try:
            from django.template.loader import render_to_string
            from utils.pdf_utils import generate_pdf_from_html, get_base64_encoded_file
            from core.models import SystemSetting
            import os

            # صياغة نص الفترة الزمنية
            if date_from and date_to:
                period_label = f"الفترة: من {date_from.strftime('%d/%m/%Y')} إلى {date_to.strftime('%d/%m/%Y')}"
            elif date_from:
                period_label = f"الفترة: من {date_from.strftime('%d/%m/%Y')} حتى تاريخه"
            elif date_to:
                period_label = f"الفترة: حتى تاريخ {date_to.strftime('%d/%m/%Y')}"
            else:
                period_label = "الفترة: كافة الحركات المالية حتى تاريخه"

            # جلب هوية وبيانات المنشأة
            company_name = SystemSetting.get_setting('company_name', 'مؤسسة موهبة للتجارة')
            company_tax_number = SystemSetting.get_setting('tax_number', '') or SystemSetting.get_setting('company_tax_number', '')
            company_cr = SystemSetting.get_setting('commercial_register', '') or SystemSetting.get_setting('company_cr', '')
            company_phone = SystemSetting.get_setting('company_phone', '')
            company_address = SystemSetting.get_setting('company_address', '')
            
            logo_rel = SystemSetting.get_setting('company_logo', '')
            company_logo = None
            if logo_rel:
                logo_path = os.path.join(settings.MEDIA_ROOT, str(logo_rel).lstrip('/'))
                if not os.path.exists(logo_path):
                    logo_path = os.path.join(settings.BASE_DIR, 'static', str(logo_rel).lstrip('/'))
                if os.path.exists(logo_path):
                    company_logo = get_base64_encoded_file(logo_path)

            pdf_context = {
                "company_name": company_name,
                "company_tax_number": company_tax_number,
                "company_cr": company_cr,
                "company_phone": company_phone,
                "company_address": company_address,
                "company_logo": company_logo,
                "period_label": period_label,
                "currency_symbol_active": SystemSetting.get_currency_symbol() if hasattr(SystemSetting, 'get_currency_symbol') else 'ج.م',
                "generated_at": timezone.now().strftime('%d/%m/%Y %H:%M'),
                "generated_by": request.user.get_full_name() or request.user.username,
                "date_from": date_from,
                "date_to": date_to,
            }

            if account_id:
                # تصدير كشف حساب تفصيلي للحساب المحدد
                account = get_object_or_404(ChartOfAccounts, id=account_id)
                transactions, summary = get_account_transactions_optimized(
                    account, date_from, date_to,
                    cost_center=selected_cost_center, currency=currency, include_unposted=include_unposted
                )
                pdf_context.update({
                    "account": account,
                    "account_name": account.name,
                    "account_code": account.code,
                    "summary": summary,
                    "transactions": transactions,
                    "cost_center_name": selected_cost_center.name if selected_cost_center else "",
                    "document_title": f"كشف حساب - {account.name}",
                })
                html_content = render_to_string("financial/reports/pdf/account_statement_pdf.html", pdf_context, request=request)
                safe_filename = f"statement_{account.code}_{timezone.now().strftime('%Y%m%d')}.pdf"
                return generate_pdf_from_html(html_content, request=request, filename=safe_filename, doc_type="account_statement", context=pdf_context)
            else:
                # تصدير ملخص حركة وأرصدة كافة الحسابات
                account_summaries = get_all_accounts_summary_optimized(date_from, date_to)
                pdf_context.update({
                    "accounts_summary": account_summaries,
                    "document_title": "ملخص حركة وأرصدة الحسابات المالية",
                })
                html_content = render_to_string("financial/reports/pdf/accounts_summary_pdf.html", pdf_context, request=request)
                safe_filename = f"accounts_summary_{timezone.now().strftime('%Y%m%d')}.pdf"
                return generate_pdf_from_html(html_content, request=request, filename=safe_filename, doc_type="accounts_summary", context=pdf_context)

        except Exception as e:
            logger.error(f"خطأ في تصدير PDF لكشف الحسابات: {e}", exc_info=True)
            messages.error(request, f"خطأ في تصدير ملف PDF: {e}")
    
    # معالجة طلبات AJAX للتحميل التدريجي
    if progressive_load and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return handle_progressive_ledger_load(
            request, account_id, date_from, date_to, page_size,
            cost_center=selected_cost_center, include_unposted=include_unposted
        )
    
    # عرض تفاصيل حساب معين
    if account_id:
        try:
            account = get_object_or_404(ChartOfAccounts, id=account_id)
            
            # جلب المعاملات والملخص مباشرة من محرك الحسابات المالي
            transactions, summary = get_account_transactions_optimized(
                account, date_from, date_to,
                cost_center=selected_cost_center, currency=currency, include_unposted=include_unposted
            )
            
            # Pagination للمعاملات
            paginator = Paginator(transactions, page_size)
            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)
            
            # حساب الرصيد المنقول من الصفحة السابقة (المحلي والأجنبي)
            page_brought_forward = summary['opening_balance']
            page_brought_forward_foreign = summary.get('opening_balance_foreign', Decimal('0.00'))
            if page_obj.number > 1 and len(transactions) > 0:
                prev_idx = (page_obj.number - 1) * page_size - 1
                if 0 <= prev_idx < len(transactions):
                    page_brought_forward = transactions[prev_idx]['balance']
                    page_brought_forward_foreign = transactions[prev_idx].get('foreign_balance', Decimal('0.00'))
            
            # بناء أزرار الهيدر التفاعلية (Dynamic Export Dispatchers)
            header_buttons = [
                {
                    "onclick": "window.exportReport('pdf')",
                    "icon": "fa-file-pdf",
                    "text": "تصدير PDF",
                    "class": "btn-outline-danger",
                },
                {
                    "onclick": "window.exportReport('excel')",
                    "icon": "fa-file-excel",
                    "text": "تصدير Excel",
                    "class": "btn-success",
                },
            ]
            
            context = {
                "page_title": f"كشف حساب - {account.name}",
                "page_subtitle": f"كشف تفصيلي لحركات الحساب المالي: {account.code} - {account.name}",
                "page_icon": "fas fa-file-invoice-dollar",
                "header_buttons": header_buttons,
                "breadcrumb_items": [
                    {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
                    {"title": "الإدارة المالية", "icon": "fas fa-money-bill-wave"},
                    {"title": "التقارير", "icon": "fas fa-chart-bar"},
                    {"title": "كشف الحسابات", "active": True},
                ],
                "account": account,
                "summary": summary,
                "page_obj": page_obj,
                "transactions": transactions,
                "accounts": accounts,
                "grouped_accounts": grouped_accounts,
                "cost_centers": cost_centers,
                "currencies": currencies,
                "selected_currency": currency,
                "selected_cost_center_id": int(cost_center_id) if cost_center_id and cost_center_id.isdigit() else None,
                "include_unposted": include_unposted,
                "date_from": date_from,
                "date_to": date_to,
                "selected_account_id": int(account_id),
                "page_size": page_size,
                "page_brought_forward": page_brought_forward,
                "page_brought_forward_foreign": page_brought_forward_foreign,
            }
            
        except ChartOfAccounts.DoesNotExist:
            messages.error(request, "الحساب المطلوب غير موجود")
            return redirect('financial:ledger_report')
        except Exception as e:
            logger.error(f"خطأ في تحميل تفاصيل الحساب: {e}")
            messages.error(request, f"خطأ في تحميل تفاصيل الحساب: {e}")
            return redirect('financial:ledger_report')
    
    else:
        # عرض ملخص جميع الحسابات
        try:
            account_summaries = get_all_accounts_summary_optimized(date_from, date_to)
            
            paginator = Paginator(account_summaries, page_size)
            page_number = request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)
            
            header_buttons = [
                {
                    "onclick": "window.exportReport('pdf')",
                    "icon": "fa-file-pdf",
                    "text": "تصدير PDF",
                    "class": "btn-outline-danger",
                },
                {
                    "onclick": "window.exportReport('excel')",
                    "icon": "fa-file-excel",
                    "text": "تصدير Excel",
                    "class": "btn-success",
                },
            ]
            
            context = {
                "page_title": "كشف حركة وأرصدة الحسابات",
                "page_subtitle": "ملخص شامل لأرصدة وحركات الحسابات المالية في دليل الحسابات",
                "page_icon": "fas fa-table",
                "header_buttons": header_buttons,
                "breadcrumb_items": [
                    {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
                    {"title": "الإدارة المالية", "icon": "fas fa-money-bill-wave"},
                    {"title": "التقارير", "icon": "fas fa-chart-bar"},
                    {"title": "كشف الحسابات", "active": True},
                ],
                "page_obj": page_obj,
                "account_summaries": account_summaries,
                "accounts": accounts,
                "grouped_accounts": grouped_accounts,
                "cost_centers": cost_centers,
                "date_from": date_from,
                "date_to": date_to,
                "page_size": page_size,
            }
            
        except Exception as e:
            messages.error(request, f"خطأ في تحميل ملخص الحسابات: {e}")
            context = {
                "page_title": "كشف حركة وأرصدة الحسابات",
                "page_subtitle": "ملخص شامل لأرصدة وحركات الحسابات المالية في دليل الحسابات",
                "page_icon": "fas fa-table",
                "breadcrumb_items": [
                    {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
                    {"title": "الإدارة المالية", "icon": "fas fa-money-bill-wave"},
                    {"title": "التقارير", "icon": "fas fa-chart-bar"},
                    {"title": "كشف الحسابات", "active": True},
                ],
                "accounts": accounts,
                "grouped_accounts": grouped_accounts,
                "cost_centers": cost_centers,
                "date_from": date_from,
                "date_to": date_to,
                "error": str(e)
            }
    
    # دعم طلبات AJAX الديناميكية لاستبدال المحتوى
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest' and not progressive_load:
        # Partial rendering
        return render(request, "financial/reports/ledger_report.html", context)
        
    return render(request, "financial/reports/ledger_report.html", context)


@login_required
def balance_sheet(request):
    """
    تقرير الميزانية العمومية والمركز المالي (IAS 1 Statement of Financial Position)
    مرتبط بالكامل بـ BalanceSheetService المعياري مع دعم التخزين المؤقت، المقارنة الزمنية، وتصدير Excel الرسمي.
    """
    from django.http import HttpResponse
    from django.core.cache import cache
    import hashlib
    from financial.services.balance_sheet_service import BalanceSheetService

    # 1. معالجة الفلاتر والمعايير
    date_str = request.GET.get("date") or request.GET.get("date_to") or request.GET.get("as_of_date")
    comparison_date_str = request.GET.get("comparison_date")
    account_level = request.GET.get("account_level")
    hide_zero_balances = request.GET.get("hide_zero_balances", "0") in ["1", "true", "True"]
    export_format = request.GET.get("export")
    use_cache = request.GET.get("use_cache", "1") == "1"

    # معالجة التواريخ
    as_of_date = None
    if date_str:
        try:
            as_of_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            messages.warning(request, "تنسيق التاريخ غير صحيح، تم استخدام تاريخ اليوم.")
            as_of_date = timezone.now().date()
    else:
        as_of_date = timezone.now().date()

    comparison_date = None
    if comparison_date_str:
        try:
            comparison_date = datetime.strptime(comparison_date_str, "%Y-%m-%d").date()
        except ValueError:
            comparison_date = None

    # 2. معالجة تصدير Excel الرسمي المعتمد
    if export_format == 'excel':
        try:
            excel_data = BalanceSheetService.export_to_excel(
                as_of_date=as_of_date,
                comparison_date=comparison_date,
                account_level=account_level,
                hide_zero_balances=hide_zero_balances
            )
            response = HttpResponse(
                excel_data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"balance_sheet_{as_of_date.strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            logger.error(f"خطأ في تصدير Excel للميزانية: {e}", exc_info=True)
            messages.error(request, f"خطأ في تصدير ملف Excel: {e}")

    # 3. إدارة التخزين المؤقت (Cache Key)
    cache_key_data = f"bs_v2_{as_of_date}_{comparison_date}_{account_level}_{hide_zero_balances}"
    cache_key = f"report_bs_{hashlib.md5(cache_key_data.encode()).hexdigest()}"

    cached_data = None
    if use_cache and not export_format:
        cached_data = cache.get(cache_key)

    if cached_data:
        try:
            balance_sheet_data = cached_data
            financial_ratios = balance_sheet_data.get('financial_ratios', {})
        except Exception:
            cached_data = None

    if not cached_data:
        balance_sheet_data = BalanceSheetService.generate_balance_sheet(
            as_of_date=as_of_date,
            comparison_date=comparison_date,
            account_level=account_level,
            hide_zero_balances=hide_zero_balances
        )
        financial_ratios = balance_sheet_data.get('financial_ratios', {})

        if use_cache:
            try:
                cache.set(cache_key, balance_sheet_data, 300)  # 5 دقائق TTL
            except Exception:
                pass

    # 4. بناء أزرار الهيدر المركزي
    header_buttons = [
        {
            "onclick": "window.print()",
            "icon": "fa-print",
            "text": "طباعة",
            "class": "btn-outline-secondary",
        }
    ]

    # زر تصدير Excel
    export_params = [f"date={as_of_date.strftime('%Y-%m-%d')}"]
    if comparison_date:
        export_params.append(f"comparison_date={comparison_date.strftime('%Y-%m-%d')}")
    if account_level:
        export_params.append(f"account_level={account_level}")
    if hide_zero_balances:
        export_params.append("hide_zero_balances=1")
    export_params.append("export=excel")
    export_url = "?" + "&".join(export_params)

    header_buttons.append({
        "url": export_url,
        "icon": "fa-file-excel",
        "text": "تصدير Excel",
        "class": "btn-success",
    })

    # زر التحديث المباشر (تجاوز الكاش)
    refresh_params = [f"date={as_of_date.strftime('%Y-%m-%d')}"]
    if comparison_date:
        refresh_params.append(f"comparison_date={comparison_date.strftime('%Y-%m-%d')}")
    if account_level:
        refresh_params.append(f"account_level={account_level}")
    if hide_zero_balances:
        refresh_params.append("hide_zero_balances=1")
    refresh_params.append("use_cache=0")
    refresh_url = "?" + "&".join(refresh_params)

    header_buttons.append({
        "url": refresh_url,
        "icon": "fa-sync",
        "text": "تحديث البيانات",
        "class": "btn-outline-primary",
    })

    context = {
        "page_title": "الميزانية العمومية والمركز المالي",
        "page_subtitle": f"تقرير المركز المالي كما في {as_of_date.strftime('%d/%m/%Y')}",
        "page_icon": "fas fa-balance-scale",
        "breadcrumb_items": [
            {"title": "الرئيسية", "url": reverse("core:dashboard"), "icon": "fas fa-home"},
            {"title": "الإدارة المالية", "url": reverse("financial:chart_of_accounts_list"), "icon": "fas fa-calculator"},
            {"title": "التقارير المالية", "url": "#", "icon": "fas fa-chart-line"},
            {"title": "الميزانية العمومية", "active": True},
        ],
        "header_buttons": header_buttons,
        "bs": balance_sheet_data,
        "balance_sheet_data": balance_sheet_data,
        "financial_ratios": financial_ratios,
        "as_of_date": as_of_date,
        "comparison_date": comparison_date,
        "account_level": account_level or "all",
        "hide_zero_balances": hide_zero_balances,
        "is_cached": cached_data is not None,
    }

    return render(request, "financial/reports/balance_sheet.html", context)


@login_required
def income_statement(request):
    """
    تقرير قائمة الدخل والأرباح والخسائر (IAS 1 Statement of Profit or Loss)
    مرتبط بالكامل بـ IncomeStatementService المعياري مع دعم المقارنة الزمنية، مراكز التكلفة، وتصدير Excel الرسمي.
    """
    from django.http import HttpResponse
    from django.core.cache import cache
    import hashlib
    from financial.services.income_statement_service import IncomeStatementService
    from financial.models.cost_center import CostCenter

    # 1. معالجة الفلاتر والمعايير
    preset = request.GET.get("preset")
    date_from_str = request.GET.get("date_from")
    date_to_str = request.GET.get("date_to")
    comp_date_from_str = request.GET.get("comp_date_from")
    comp_date_to_str = request.GET.get("comp_date_to")
    cost_center_id = request.GET.get("cost_center") or request.GET.get("cost_center_id")
    account_level = request.GET.get("account_level")
    hide_zero_balances = request.GET.get("hide_zero_balances", "0") in ["1", "true", "True"]
    include_unposted = request.GET.get("include_unposted", "0") in ["1", "true", "True"]
    export_format = request.GET.get("export")
    use_cache = request.GET.get("use_cache", "1") == "1"

    # تحويل التواريخ وفق الفترات السريعة أو المدخلة يدوياً
    today = timezone.now().date()
    date_from = None
    date_to = None

    if preset == 'this_month':
        date_from = date(today.year, today.month, 1)
        next_month = today.month % 12 + 1
        next_month_year = today.year + (1 if today.month == 12 else 0)
        date_to = date(next_month_year, next_month, 1) - timedelta(days=1)
    elif preset == 'this_quarter':
        q_start_month = ((today.month - 1) // 3) * 3 + 1
        date_from = date(today.year, q_start_month, 1)
        q_end_month = q_start_month + 2
        next_q_month = q_end_month % 12 + 1
        next_q_year = today.year + (1 if q_end_month == 12 else 0)
        date_to = date(next_q_year, next_q_month, 1) - timedelta(days=1)
    elif preset == 'ytd':
        date_from = date(today.year, 1, 1)
        date_to = today
    elif preset == 'last_year':
        date_from = date(today.year - 1, 1, 1)
        date_to = date(today.year - 1, 12, 31)
    else:
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
            except ValueError:
                date_to = today
        else:
            date_to = today

        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            except ValueError:
                date_from = date(date_to.year, 1, 1)
        else:
            date_from = date(date_to.year, 1, 1)

    # تحديد الفترة السريعة النشطة لتفعيل الزر المناسب في الواجهة
    active_preset = preset
    if not active_preset:
        month_start = date(today.year, today.month, 1)
        next_m = today.month % 12 + 1
        next_m_y = today.year + (1 if today.month == 12 else 0)
        month_end = date(next_m_y, next_m, 1) - timedelta(days=1)

        q_s_m = ((today.month - 1) // 3) * 3 + 1
        quarter_start = date(today.year, q_s_m, 1)
        q_e_m = q_s_m + 2
        next_q_m = q_e_m % 12 + 1
        next_q_y = today.year + (1 if q_e_m == 12 else 0)
        quarter_end = date(next_q_y, next_q_m, 1) - timedelta(days=1)

        if date_from == date(today.year, 1, 1) and date_to == today:
            active_preset = 'ytd'
        elif date_from == month_start and date_to == month_end:
            active_preset = 'this_month'
        elif date_from == quarter_start and date_to == quarter_end:
            active_preset = 'this_quarter'
        elif date_from == date(today.year - 1, 1, 1) and date_to == date(today.year - 1, 12, 31):
            active_preset = 'last_year'

    comp_date_from = None
    if comp_date_from_str:
        try:
            comp_date_from = datetime.strptime(comp_date_from_str, "%Y-%m-%d").date()
        except ValueError:
            comp_date_from = None

    comp_date_to = None
    if comp_date_to_str:
        try:
            comp_date_to = datetime.strptime(comp_date_to_str, "%Y-%m-%d").date()
        except ValueError:
            comp_date_to = None

    # 2. معالجة تصدير Excel الرسمي المعتمد
    if export_format == 'excel':
        try:
            excel_data = IncomeStatementService.export_to_excel(
                date_from=date_from,
                date_to=date_to,
                comp_date_from=comp_date_from,
                comp_date_to=comp_date_to,
                cost_center_id=cost_center_id,
                account_level=account_level,
                hide_zero_balances=hide_zero_balances
            )
            response = HttpResponse(
                excel_data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"income_statement_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            logger.error(f"خطأ في تصدير Excel لقائمة الدخل: {e}", exc_info=True)
            messages.error(request, f"خطأ في تصدير ملف Excel: {e}")

    # 3. إدارة التخزين المؤقت
    cache_key_data = f"inc_v2_{date_from}_{date_to}_{comp_date_from}_{comp_date_to}_{cost_center_id}_{account_level}_{hide_zero_balances}_{include_unposted}"
    cache_key = f"report_inc_{hashlib.md5(cache_key_data.encode()).hexdigest()}"

    cached_data = None
    if use_cache and not export_format and not include_unposted:
        cached_data = cache.get(cache_key)

    if cached_data:
        try:
            income_statement_data = cached_data
        except Exception:
            cached_data = None

    if not cached_data:
        income_statement_data = IncomeStatementService.generate_income_statement(
            date_from=date_from,
            date_to=date_to,
            comp_date_from=comp_date_from,
            comp_date_to=comp_date_to,
            cost_center_id=cost_center_id,
            account_level=account_level,
            hide_zero_balances=hide_zero_balances,
            include_unposted=include_unposted
        )
        if use_cache and not include_unposted:
            try:
                cache.set(cache_key, income_statement_data, 300)
            except Exception:
                pass

    # 4. بناء أزرار الهيدر المركزي
    header_buttons = [
        {
            "onclick": "window.print()",
            "icon": "fa-print",
            "text": "طباعة",
            "class": "btn-outline-secondary",
        }
    ]

    # زر تصدير Excel
    export_params = [f"date_from={date_from.strftime('%Y-%m-%d')}", f"date_to={date_to.strftime('%Y-%m-%d')}"]
    if comp_date_from and comp_date_to:
        export_params.extend([f"comp_date_from={comp_date_from.strftime('%Y-%m-%d')}", f"comp_date_to={comp_date_to.strftime('%Y-%m-%d')}"])
    if cost_center_id:
        export_params.append(f"cost_center={cost_center_id}")
    if account_level:
        export_params.append(f"account_level={account_level}")
    if hide_zero_balances:
        export_params.append("hide_zero_balances=1")
    export_params.append("export=excel")
    export_url = "?" + "&".join(export_params)

    header_buttons.append({
        "url": export_url,
        "icon": "fa-file-excel",
        "text": "تصدير Excel",
        "class": "btn-success",
    })

    # زر التحديث المباشر
    refresh_params = [f"date_from={date_from.strftime('%Y-%m-%d')}", f"date_to={date_to.strftime('%Y-%m-%d')}"]
    if comp_date_from and comp_date_to:
        refresh_params.extend([f"comp_date_from={comp_date_from.strftime('%Y-%m-%d')}", f"comp_date_to={comp_date_to.strftime('%Y-%m-%d')}"])
    if cost_center_id:
        refresh_params.append(f"cost_center={cost_center_id}")
    if account_level:
        refresh_params.append(f"account_level={account_level}")
    if hide_zero_balances:
        refresh_params.append("hide_zero_balances=1")
    if include_unposted:
        refresh_params.append("include_unposted=1")
    refresh_params.append("use_cache=0")
    refresh_url = "?" + "&".join(refresh_params)

    header_buttons.append({
        "url": refresh_url,
        "icon": "fa-sync",
        "text": "تحديث البيانات",
        "class": "btn-outline-primary",
    })

    # قائمة مراكز التكلفة
    cost_centers_list = CostCenter.objects.filter(is_active=True).order_by('code')

    context = {
        "page_title": "قائمة الدخل والأرباح والخسائر",
        "page_subtitle": f"تقرير نتائج الأعمال للفترة من {date_from.strftime('%d/%m/%Y')} إلى {date_to.strftime('%d/%m/%Y')}",
        "page_icon": "fas fa-chart-line",
        "breadcrumb_items": [
            {"title": "الرئيسية", "url": reverse("core:dashboard"), "icon": "fas fa-home"},
            {"title": "الإدارة المالية", "url": reverse("financial:chart_of_accounts_list"), "icon": "fas fa-calculator"},
            {"title": "التقارير المالية", "url": "#", "icon": "fas fa-chart-line"},
            {"title": "قائمة الدخل", "active": True},
        ],
        "header_buttons": header_buttons,
        "inc": income_statement_data,
        "income_statement_data": income_statement_data,
        "date_from": date_from,
        "date_to": date_to,
        "comp_date_from": comp_date_from,
        "comp_date_to": comp_date_to,
        "cost_centers_list": cost_centers_list,
        "selected_cost_center_id": str(cost_center_id) if cost_center_id else "",
        "account_level": account_level or "all",
        "hide_zero_balances": hide_zero_balances,
        "include_unposted": include_unposted,
        "active_preset": active_preset,
        "is_cached": cached_data is not None,
    }

    return render(request, "financial/reports/income_statement.html", context)


@login_required
def cash_flow_statement(request):
    """
    تقرير قائمة التدفقات النقدية (IAS 7 Statement of Cash Flows)
    مرتبط بالكامل بـ CashFlowService المعياري مع دعم الطريقة غير المباشرة المتقدمة،
    المقارنة الزمنية، مراكز التكلفة، وتصدير Excel الرسمي.
    """
    from django.http import HttpResponse
    from django.core.cache import cache
    import hashlib
    from financial.services.cash_flow_service import CashFlowService
    from financial.models.cost_center import CostCenter

    # 1. معالجة الفلاتر والمعايير
    preset = request.GET.get("preset")
    date_from_str = request.GET.get("date_from")
    date_to_str = request.GET.get("date_to")
    comp_date_from_str = request.GET.get("comp_date_from")
    comp_date_to_str = request.GET.get("comp_date_to")
    cost_center_id = request.GET.get("cost_center") or request.GET.get("cost_center_id")
    account_level = request.GET.get("account_level")
    hide_zero_balances = request.GET.get("hide_zero_balances", "0") in ["1", "true", "True"]
    include_unposted = request.GET.get("include_unposted", "0") in ["1", "true", "True"]
    export_format = request.GET.get("export")
    use_cache = request.GET.get("use_cache", "1") == "1"

    # تحويل التواريخ وفق الفترات السريعة أو المدخلة يدوياً
    today = timezone.now().date()
    date_from = None
    date_to = None

    if preset == 'this_month':
        date_from = date(today.year, today.month, 1)
        next_month = today.month % 12 + 1
        next_month_year = today.year + (1 if today.month == 12 else 0)
        date_to = date(next_month_year, next_month, 1) - timedelta(days=1)
    elif preset == 'this_quarter':
        q_start_month = ((today.month - 1) // 3) * 3 + 1
        date_from = date(today.year, q_start_month, 1)
        q_end_month = q_start_month + 2
        next_q_month = q_end_month % 12 + 1
        next_q_year = today.year + (1 if q_end_month == 12 else 0)
        date_to = date(next_q_year, next_q_month, 1) - timedelta(days=1)
    elif preset == 'ytd':
        date_from = date(today.year, 1, 1)
        date_to = today
    elif preset == 'last_year':
        date_from = date(today.year - 1, 1, 1)
        date_to = date(today.year - 1, 12, 31)
    else:
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
            except ValueError:
                date_to = today
        else:
            date_to = today

        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            except ValueError:
                date_from = date(date_to.year, 1, 1)
        else:
            date_from = date(date_to.year, 1, 1)

    # تحديد الفترة السريعة النشطة لتفعيل الزر المناسب في الواجهة
    active_preset = preset
    if not active_preset:
        month_start = date(today.year, today.month, 1)
        next_m = today.month % 12 + 1
        next_m_y = today.year + (1 if today.month == 12 else 0)
        month_end = date(next_m_y, next_m, 1) - timedelta(days=1)

        q_s_m = ((today.month - 1) // 3) * 3 + 1
        quarter_start = date(today.year, q_s_m, 1)
        q_e_m = q_s_m + 2
        next_q_m = q_e_m % 12 + 1
        next_q_y = today.year + (1 if q_e_m == 12 else 0)
        quarter_end = date(next_q_y, next_q_m, 1) - timedelta(days=1)

        if date_from == date(today.year, 1, 1) and date_to == today:
            active_preset = 'ytd'
        elif date_from == month_start and date_to == month_end:
            active_preset = 'this_month'
        elif date_from == quarter_start and date_to == quarter_end:
            active_preset = 'this_quarter'
        elif date_from == date(today.year - 1, 1, 1) and date_to == date(today.year - 1, 12, 31):
            active_preset = 'last_year'

    comp_date_from = None
    if comp_date_from_str:
        try:
            comp_date_from = datetime.strptime(comp_date_from_str, "%Y-%m-%d").date()
        except ValueError:
            comp_date_from = None

    comp_date_to = None
    if comp_date_to_str:
        try:
            comp_date_to = datetime.strptime(comp_date_to_str, "%Y-%m-%d").date()
        except ValueError:
            comp_date_to = None

    # 2. معالجة تصدير Excel الرسمي المعتمد
    if export_format == 'excel':
        try:
            excel_data = CashFlowService.export_to_excel(
                date_from=date_from,
                date_to=date_to,
                cost_center_id=cost_center_id,
                include_unposted=include_unposted,
            )
            response = HttpResponse(
                excel_data,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="Cash_Flow_IAS7_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.xlsx"'
            return response
        except Exception as e:
            logger.error(f"Error exporting Cash Flow to Excel: {e}", exc_info=True)
            messages.error(request, f"حدث خطأ أثناء تصدير ملف Excel: {e}")

    # 3. جلب بيانات التدفقات النقدية مع التخزين المؤقت
    cache_key = None
    cached_data = None
    if use_cache:
        cache_key_raw = f"cf_stmt_{date_from}_{date_to}_{comp_date_from}_{comp_date_to}_{cost_center_id}_{account_level}_{hide_zero_balances}_{include_unposted}"
        cache_key = f"cf_report_{hashlib.md5(cache_key_raw.encode()).hexdigest()}"
        cached_data = cache.get(cache_key)

    if cached_data:
        cash_flow_data = cached_data
    else:
        try:
            cash_flow_data = CashFlowService.generate_cash_flow_statement(
                date_from=date_from,
                date_to=date_to,
                comp_date_from=comp_date_from,
                comp_date_to=comp_date_to,
                cost_center_id=cost_center_id,
                account_level=account_level,
                hide_zero_balances=hide_zero_balances,
                include_unposted=include_unposted,
            )
            if use_cache and cache_key:
                cache.set(cache_key, cash_flow_data, 180)  # 3 minutes
        except Exception as e:
            logger.error(f"Error generating Cash Flow data: {e}", exc_info=True)
            messages.error(request, f"حدث خطأ أثناء احتساب قائمة التدفقات النقدية: {e}")
            cash_flow_data = {}

    # 4. أزرار الهيدر المعيارية
    header_buttons = [
        {
            "onclick": "window.print()",
            "icon": "fa-print",
            "text": "طباعة",
            "class": "btn-outline-secondary",
        }
    ]

    # رابط تصدير Excel
    export_params = [f"date_from={date_from.strftime('%Y-%m-%d')}", f"date_to={date_to.strftime('%Y-%m-%d')}"]
    if comp_date_from and comp_date_to:
        export_params.extend([f"comp_date_from={comp_date_from.strftime('%Y-%m-%d')}", f"comp_date_to={comp_date_to.strftime('%Y-%m-%d')}"])
    if cost_center_id:
        export_params.append(f"cost_center={cost_center_id}")
    if account_level:
        export_params.append(f"account_level={account_level}")
    if hide_zero_balances:
        export_params.append("hide_zero_balances=1")
    if include_unposted:
        export_params.append("include_unposted=1")
    export_params.append("export=excel")
    export_url = "?" + "&".join(export_params)

    header_buttons.append({
        "url": export_url,
        "icon": "fa-file-excel",
        "text": "تصدير Excel",
        "class": "btn-success",
    })

    # زر التحديث المباشر
    refresh_params = [f"date_from={date_from.strftime('%Y-%m-%d')}", f"date_to={date_to.strftime('%Y-%m-%d')}"]
    if comp_date_from and comp_date_to:
        refresh_params.extend([f"comp_date_from={comp_date_from.strftime('%Y-%m-%d')}", f"comp_date_to={comp_date_to.strftime('%Y-%m-%d')}"])
    if cost_center_id:
        refresh_params.append(f"cost_center={cost_center_id}")
    if account_level:
        refresh_params.append(f"account_level={account_level}")
    if hide_zero_balances:
        refresh_params.append("hide_zero_balances=1")
    if include_unposted:
        refresh_params.append("include_unposted=1")
    refresh_params.append("use_cache=0")
    refresh_url = "?" + "&".join(refresh_params)

    header_buttons.append({
        "url": refresh_url,
        "icon": "fa-sync",
        "text": "تحديث البيانات",
        "class": "btn-outline-primary",
    })

    # قائمة مراكز التكلفة
    cost_centers_list = CostCenter.objects.filter(is_active=True).order_by('code')

    context = {
        "page_title": "قائمة التدفقات النقدية (IAS 7)",
        "page_subtitle": f"تقرير التدفقات النقدية للفترة من {date_from.strftime('%d/%m/%Y')} إلى {date_to.strftime('%d/%m/%Y')}",
        "page_icon": "fas fa-money-bill-wave",
        "breadcrumb_items": [
            {"title": "الرئيسية", "url": reverse("core:dashboard"), "icon": "fas fa-home"},
            {"title": "الإدارة المالية", "url": reverse("financial:chart_of_accounts_list"), "icon": "fas fa-calculator"},
            {"title": "التقارير المالية", "url": "#", "icon": "fas fa-chart-line"},
            {"title": "قائمة التدفقات النقدية", "active": True},
        ],
        "header_buttons": header_buttons,
        "cf": cash_flow_data,
        "cash_flow_data": cash_flow_data,
        "date_from": date_from,
        "date_to": date_to,
        "comp_date_from": comp_date_from,
        "comp_date_to": comp_date_to,
        "cost_centers_list": cost_centers_list,
        "selected_cost_center_id": str(cost_center_id) if cost_center_id else "",
        "account_level": account_level or "all",
        "hide_zero_balances": hide_zero_balances,
        "include_unposted": include_unposted,
        "active_preset": active_preset,
        "is_cached": cached_data is not None,
    }

    return render(request, "financial/reports/cash_flow_statement.html", context)


@login_required
def customer_supplier_balances_report(request, account_type):
    """
    تقرير أرصدة أولياء الأمور والموردين
    account_type: 'customers' أو 'suppliers'
    """
    from ..services.parent_supplier_balances_service import ParentSupplierBalancesService
    from django.http import HttpResponse
    
    # تحديد تاريخ التقرير
    as_of_date = request.GET.get("as_of_date")
    
    if as_of_date:
        as_of_date = datetime.strptime(as_of_date, "%Y-%m-%d").date()
    else:
        # افتراضيًا، تاريخ اليوم
        as_of_date = timezone.now().date()
    
    try:
        # إنشاء خدمة تقارير الأرصدة
        balances_service = CustomerSupplierBalancesService(as_of_date=as_of_date)
        
        # التحقق من طلب التصدير
        if request.GET.get('export') == 'excel':
            # إنشاء التقرير
            if account_type == "customers":
                report_data = balances_service.generate_customer_balances_report()
                report_type = 'ar'
                filename = f'parent_balances_{as_of_date}.xlsx'
            else:
                report_data = balances_service.generate_supplier_balances_report()
                report_type = 'ap'
                filename = f'supplier_balances_{as_of_date}.xlsx'
            
            # تصدير إلى Excel
            excel_content = balances_service.export_to_excel(report_data, report_type)
            
            if excel_content:
                response = HttpResponse(
                    excel_content,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            else:
                messages.warning(request, "تصدير Excel غير متاح. يرجى تثبيت openpyxl")
        
        # إنشاء التقرير
        if account_type == "customers":
            report_data = balances_service.generate_customer_balances_report()
        else:
            report_data = balances_service.generate_supplier_balances_report()
        
        # التحقق من وجود خطأ
        if "error" in report_data:
            messages.error(request, report_data["error"])
            return redirect("core:dashboard")
        
        # تحديد العنوان والأيقونة حسب النوع
        if account_type == "customers":
            page_title = "تقرير أرصدة أولياء الأمور"
            page_icon = "fas fa-users"
        else:
            page_title = "تقرير أرصدة الموردين"
            page_icon = "fas fa-truck"
        
        context = {
            "page_title": page_title,
            "page_subtitle": f"عرض الأرصدة المستحقة والمدفوعة حتى {as_of_date}",
            "page_icon": page_icon,
            "header_buttons": [
                {
                    "onclick": "window.print()",
                    "icon": "fa-print",
                    "text": "طباعة",
                    "class": "btn-outline-secondary",
                },
                {
                    "url": f"?as_of_date={as_of_date}&export=excel",
                    "icon": "fa-file-excel",
                    "text": "تصدير Excel",
                    "class": "btn-success",
                },
            ],
            "breadcrumb_items": [
                {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
                {"title": "الإدارة المالية", "icon": "fas fa-money-bill-wave"},
                {"title": "التقارير", "icon": "fas fa-chart-bar"},
                {"title": page_title, "active": True},
            ],
            "report_data": report_data,
            "as_of_date": as_of_date,
            "account_type": account_type,
        }
        
        return render(request, "financial/reports/customer_supplier_balances.html", context)
    
    except Exception as e:
        messages.error(request, f"خطأ في تحميل تقرير الأرصدة: {e}")
        return redirect("core:dashboard")


@login_required
def financial_analytics(request):
    """
    عرض لوحة التحليلات والمؤشرات المالية التنفيذية (CFO Executive Analytics Command Center v3.0)
    متكاملة مع المحاور الخمسة، مؤشر ألتمان للسلامة والتعثر، بطاقة الصحة المالية، المقارنة بالفترة السابقة، وتصدير Excel المعتمد.
    """
    import json
    import hashlib
    from django.http import HttpResponse
    from django.core.cache import cache
    from financial.services.financial_analytics_service import FinancialAnalyticsService
    from financial.models.cost_center import CostCenter

    # 1. معالجة الفلاتر والمعايير
    preset = request.GET.get("preset")
    date_from_str = request.GET.get("date_from")
    date_to_str = request.GET.get("date_to")
    comp_date_from_str = request.GET.get("comp_date_from")
    comp_date_to_str = request.GET.get("comp_date_to")
    cost_center_id = request.GET.get("cost_center") or request.GET.get("cost_center_id")
    include_unposted = request.GET.get("include_unposted", "0") in ["1", "true", "True"]
    export_format = request.GET.get("export")
    use_cache = request.GET.get("use_cache", "1") == "1"

    today = timezone.now().date()
    date_from = None
    date_to = None

    if preset == 'this_month':
        date_from = date(today.year, today.month, 1)
        next_month = today.month % 12 + 1
        next_month_year = today.year + (1 if today.month == 12 else 0)
        date_to = date(next_month_year, next_month, 1) - timedelta(days=1)
    elif preset == 'this_quarter':
        q_start_month = ((today.month - 1) // 3) * 3 + 1
        date_from = date(today.year, q_start_month, 1)
        q_end_month = q_start_month + 2
        next_q_month = q_end_month % 12 + 1
        next_q_year = today.year + (1 if q_end_month == 12 else 0)
        date_to = date(next_q_year, next_q_month, 1) - timedelta(days=1)
    elif preset == 'ytd':
        date_from = date(today.year, 1, 1)
        date_to = today
    elif preset == 'last_year':
        date_from = date(today.year - 1, 1, 1)
        date_to = date(today.year - 1, 12, 31)
    else:
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
            except ValueError:
                date_to = today
        else:
            date_to = today

        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
            except ValueError:
                date_from = date(date_to.year, 1, 1)
        else:
            date_from = date(date_to.year, 1, 1)

    # تحديد الفترة السريعة النشطة
    active_preset = preset
    if not active_preset:
        month_start = date(today.year, today.month, 1)
        next_m = today.month % 12 + 1
        next_m_y = today.year + (1 if today.month == 12 else 0)
        month_end = date(next_m_y, next_m, 1) - timedelta(days=1)

        q_s_m = ((today.month - 1) // 3) * 3 + 1
        quarter_start = date(today.year, q_s_m, 1)
        q_e_m = q_s_m + 2
        next_q_m = q_e_m % 12 + 1
        next_q_y = today.year + (1 if q_e_m == 12 else 0)
        quarter_end = date(next_q_y, next_q_m, 1) - timedelta(days=1)

        if date_from == date(today.year, 1, 1) and date_to == today:
            active_preset = 'ytd'
        elif date_from == month_start and date_to == month_end:
            active_preset = 'this_month'
        elif date_from == quarter_start and date_to == quarter_end:
            active_preset = 'this_quarter'
        elif date_from == date(today.year - 1, 1, 1) and date_to == date(today.year - 1, 12, 31):
            active_preset = 'last_year'

    comp_date_from = None
    if comp_date_from_str:
        try:
            comp_date_from = datetime.strptime(comp_date_from_str, "%Y-%m-%d").date()
        except ValueError:
            comp_date_from = None

    comp_date_to = None
    if comp_date_to_str:
        try:
            comp_date_to = datetime.strptime(comp_date_to_str, "%Y-%m-%d").date()
        except ValueError:
            comp_date_to = None

    # 2. معالجة تصدير Excel الرسمي
    if export_format == 'excel':
        try:
            excel_data = FinancialAnalyticsService.export_to_excel(
                date_from=date_from,
                date_to=date_to,
                comp_date_from=comp_date_from,
                comp_date_to=comp_date_to,
                cost_center_id=cost_center_id,
                include_unposted=include_unposted,
            )
            response = HttpResponse(
                excel_data,
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="Financial_Analytics_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.xlsx"'
            return response
        except Exception as e:
            logger.error(f"Error exporting Financial Analytics to Excel: {e}", exc_info=True)
            messages.error(request, f"حدث خطأ أثناء تصدير ملف Excel: {e}")

    # 3. جلب بيانات التحليلات مع الكاش
    cache_key = None
    cached_data = None
    if use_cache:
        cache_key_raw = f"fa_stmt_{date_from}_{date_to}_{comp_date_from}_{comp_date_to}_{cost_center_id}_{include_unposted}"
        cache_key = f"fa_report_{hashlib.md5(cache_key_raw.encode()).hexdigest()}"
        cached_data = cache.get(cache_key)

    if cached_data:
        analytics = cached_data
    else:
        try:
            analytics = FinancialAnalyticsService.get_complete_analytics(
                date_from=date_from,
                date_to=date_to,
                comp_date_from=comp_date_from,
                comp_date_to=comp_date_to,
                cost_center_id=cost_center_id,
                include_unposted=include_unposted,
            )
            if use_cache and cache_key:
                cache.set(cache_key, analytics, 180)  # 3 minutes
        except Exception as e:
            logger.error(f"Error generating Financial Analytics data: {e}", exc_info=True)
            messages.error(request, f"حدث خطأ أثناء احتساب المؤشرات المالية: {e}")
            analytics = {}

    monthly_trends_json = json.dumps(analytics.get("monthly_trends", {}))
    expense_distribution_json = json.dumps(analytics.get("expense_distribution", {}))

    # 4. أزرار الهيدر
    header_buttons = [
        {
            "onclick": "window.print()",
            "icon": "fa-print",
            "text": "طباعة",
            "class": "btn-outline-secondary",
        }
    ]

    export_params = [f"date_from={date_from.strftime('%Y-%m-%d')}", f"date_to={date_to.strftime('%Y-%m-%d')}"]
    if analytics.get("comp_date_from") and analytics.get("comp_date_to"):
        export_params.extend([
            f"comp_date_from={analytics['comp_date_from'].strftime('%Y-%m-%d')}",
            f"comp_date_to={analytics['comp_date_to'].strftime('%Y-%m-%d')}"
        ])
    if cost_center_id:
        export_params.append(f"cost_center={cost_center_id}")
    if include_unposted:
        export_params.append("include_unposted=1")
    export_params.append("export=excel")
    export_url = "?" + "&".join(export_params)

    header_buttons.append({
        "url": export_url,
        "icon": "fa-file-excel",
        "text": "تصدير Excel",
        "class": "btn-success",
    })

    refresh_params = [f"date_from={date_from.strftime('%Y-%m-%d')}", f"date_to={date_to.strftime('%Y-%m-%d')}"]
    if analytics.get("comp_date_from") and analytics.get("comp_date_to"):
        refresh_params.extend([
            f"comp_date_from={analytics['comp_date_from'].strftime('%Y-%m-%d')}",
            f"comp_date_to={analytics['comp_date_to'].strftime('%Y-%m-%d')}"
        ])
    if cost_center_id:
        refresh_params.append(f"cost_center={cost_center_id}")
    if include_unposted:
        refresh_params.append("include_unposted=1")
    refresh_params.append("use_cache=0")
    refresh_url = "?" + "&".join(refresh_params)

    header_buttons.append({
        "url": refresh_url,
        "icon": "fa-sync",
        "text": "تحديث البيانات",
        "class": "btn-outline-primary",
    })

    cost_centers_list = CostCenter.objects.filter(is_active=True).order_by('code')

    context = {
        "page_title": "لوحة التحليلات والمؤشرات المالية التنفيذية",
        "page_subtitle": f"مؤشرات الأداء المالي للفترة من {date_from.strftime('%d/%m/%Y')} إلى {date_to.strftime('%d/%m/%Y')}",
        "page_icon": "fas fa-chart-pie",
        "breadcrumb_items": [
            {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
            {"title": "الإدارة المالية", "url": reverse('financial:chart_of_accounts_list'), "icon": "fas fa-calculator"},
            {"title": "التقارير المالية", "url": "#", "icon": "fas fa-chart-line"},
            {"title": "التحليلات والمؤشرات المالية", "active": True},
        ],
        "header_buttons": header_buttons,
        "analytics": analytics,
        "liquidity": analytics.get("liquidity", {}),
        "profitability": analytics.get("profitability", {}),
        "solvency": analytics.get("solvency", {}),
        "activity": analytics.get("activity", {}),
        "dupont": analytics.get("dupont", {}),
        "altman_z": analytics.get("altman_z", {}),
        "health_scorecard": analytics.get("health_scorecard", {}),
        "monthly_trends_json": monthly_trends_json,
        "expense_distribution_json": expense_distribution_json,
        "date_from": date_from,
        "date_to": date_to,
        "comp_date_from": analytics.get("comp_date_from"),
        "comp_date_to": analytics.get("comp_date_to"),
        "has_comparison": analytics.get("has_comparison", False),
        "cost_centers_list": cost_centers_list,
        "selected_cost_center_id": str(cost_center_id) if cost_center_id else "",
        "cost_center_name": analytics.get("cost_center_name", ""),
        "include_unposted": include_unposted,
        "active_preset": active_preset,
        "is_cached": cached_data is not None,
    }
    return render(request, "financial/reports/analytics.html", context)


@login_required
@require_http_methods(["POST"])
def payment_sync_retry_failed_api(request):
    """
    API لإعادة محاولة العمليات الفاشلة
    """
    try:
        from .models.payment_sync import PaymentSyncOperation
        from django.db import models

        # العمليات الفاشلة القابلة لإعادة المحاولة
        failed_operations = PaymentSyncOperation.objects.filter(
            status="failed", retry_count__lt=models.F("max_retries")
        )

        count = 0
        for operation in failed_operations:
            operation.status = "pending"
            operation.retry_count += 1
            operation.save()
            count += 1

        return JsonResponse(
            {
                "success": True,
                "count": count,
                "message": f"تم إعادة تعيين {count} عملية للمحاولة مرة أخرى",
            }
        )

    except ImportError:
        return JsonResponse({"success": False, "message": "نماذج التزامن غير متاحة"})
    except Exception as e:
        logger.error(f"Error in views.py: {str(e)}", exc_info=True)
        return JsonResponse({"success": False, "message": "حدث خطأ غير متوقع"})


@login_required
@require_http_methods(["POST"])
def payment_sync_resolve_errors_api(request):
    """
    API لحل الأخطاء القديمة
    """
    try:
        from .models.payment_sync import PaymentSyncError
        from django.utils import timezone
        from datetime import timedelta

        # حل الأخطاء المتعلقة بالاستيراد (تم إصلاحها)
        import_errors = PaymentSyncError.objects.filter(
            error_message__icontains="import", is_resolved=False
        )

        import_count = import_errors.update(
            is_resolved=True,
            resolved_at=timezone.now(),
            resolution_notes="تم إنشاء النماذج المفقودة",
        )

        # حل الأخطاء القديمة (أكثر من 7 أيام)
        old_errors = PaymentSyncError.objects.filter(
            occurred_at__lt=timezone.now() - timedelta(days=7), is_resolved=False
        )

        old_count = old_errors.update(
            is_resolved=True,
            resolved_at=timezone.now(),
            resolution_notes="حل تلقائي للأخطاء القديمة",
        )

        total_count = import_count + old_count

        return JsonResponse(
            {
                "success": True,
                "count": total_count,
                "message": f"تم حل {total_count} خطأ ({import_count} أخطاء استيراد + {old_count} أخطاء قديمة)",
            }
        )

    except ImportError:
        return JsonResponse({"success": False, "message": "نماذج الأخطاء غير متاحة"})
    except Exception as e:
        logger.error(f"Error in views.py: {str(e)}", exc_info=True)
        return JsonResponse({"success": False, "message": "حدث خطأ غير متوقع"})


@login_required
def trial_balance_report(request):
    """
    تقرير ميزان المراجعة المؤسسي الشامل - Trial Balance Report (v10.0)
    يدعم ميزان الـ 6 أعمدة والـ 2 عمود، التدرج الشجري، فلترة المستويات، وتصدير Excel الرسمي
    """
    from ..services.trial_balance_service import TrialBalanceService
    from django.http import HttpResponse
    from django.urls import reverse

    # معالجة الفلاتر
    date_from_str = request.GET.get("date_from")
    date_to_str = request.GET.get("date_to")
    display_mode = request.GET.get("display_mode", "6_columns")
    account_level = request.GET.get("account_level")
    hide_zero = request.GET.get("hide_zero", "0") == "1"
    group_by_type = request.GET.get("group_by_type", "1") == "1"
    export_format = request.GET.get("export")

    # تحويل التواريخ
    date_from = None
    date_to = None

    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, "%Y-%m-%d").date()
        except ValueError:
            messages.warning(request, "تنسيق تاريخ البداية غير صحيح")

    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, "%Y-%m-%d").date()
        except ValueError:
            messages.warning(request, "تنسيق تاريخ النهاية غير صحيح")

    # معالجة التصدير لـ Excel
    if export_format == 'excel':
        try:
            excel_data = TrialBalanceService.export_to_excel(
                date_from=date_from,
                date_to=date_to,
                display_mode=display_mode,
                account_level=account_level,
                hide_zero_balances=hide_zero,
                group_by_type=group_by_type
            )

            response = HttpResponse(
                excel_data,
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"trial_balance_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except Exception as e:
            messages.error(request, f"خطأ في تصدير Excel: {e}")

    # إنشاء ميزان المراجعة
    try:
        trial_balance_data = TrialBalanceService.generate_trial_balance(
            date_from=date_from,
            date_to=date_to,
            display_mode=display_mode,
            account_level=account_level,
            hide_zero_balances=hide_zero,
            group_by_type=group_by_type
        )

        # بناء رابط التصدير للـ Excel مع الحفاظ على كافة الفلاتر
        export_params = []
        if trial_balance_data.get('date_from'):
            export_params.append(f"date_from={trial_balance_data['date_from'].strftime('%Y-%m-%d')}")
        if trial_balance_data.get('date_to'):
            export_params.append(f"date_to={trial_balance_data['date_to'].strftime('%Y-%m-%d')}")
        if display_mode:
            export_params.append(f"display_mode={display_mode}")
        if account_level:
            export_params.append(f"account_level={account_level}")
        if hide_zero:
            export_params.append("hide_zero=1")
        if group_by_type:
            export_params.append("group_by_type=1")
        else:
            export_params.append("group_by_type=0")
        export_params.append("export=excel")
        export_url = "?" + "&".join(export_params)

        context = {
            "page_title": "ميزان المراجعة",
            "page_subtitle": f"عن الفترة من {trial_balance_data['date_from'].strftime('%Y-%m-%d')} إلى {trial_balance_data['date_to'].strftime('%Y-%m-%d')}",
            "page_icon": "fas fa-balance-scale",
            "header_buttons": [
                {
                    "onclick": "window.print()",
                    "icon": "fa-print",
                    "text": "طباعة",
                    "class": "btn-outline-secondary",
                },
                {
                    "url": export_url,
                    "icon": "fa-file-excel",
                    "text": "تصدير Excel",
                    "class": "btn-primary",
                },
            ],
            "breadcrumb_items": [
                {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fa-home"},
                {"title": "الإدارة المالية", "url": reverse('financial:chart_of_accounts_list'), "icon": "fa-calculator"},
                {"title": "ميزان المراجعة", "active": True},
            ],
            "trial_balance_data": trial_balance_data,
            "date_from": trial_balance_data['date_from'],
            "date_to": trial_balance_data['date_to'],
            "display_mode": display_mode,
            "account_level": account_level or '',
            "hide_zero": hide_zero,
            "group_by_type": group_by_type,
        }

    except Exception as e:
        logger.exception(f"خطأ في صفحة ميزان المراجعة: {e}")
        messages.error(request, f"خطأ في إنشاء ميزان المراجعة: {e}")
        context = {
            "page_title": "ميزان المراجعة",
            "page_icon": "fas fa-balance-scale",
            "breadcrumb_items": [
                {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fa-home"},
                {"title": "الإدارة المالية", "url": reverse('financial:chart_of_accounts_list'), "icon": "fa-calculator"},
                {"title": "ميزان المراجعة", "active": True},
            ],
            "trial_balance_data": {
                'accounts': [],
                'grouped': {},
                'total_opening_debit': Decimal('0.00'),
                'total_opening_credit': Decimal('0.00'),
                'total_period_debit': Decimal('0.00'),
                'total_period_credit': Decimal('0.00'),
                'total_closing_debit': Decimal('0.00'),
                'total_closing_credit': Decimal('0.00'),
                'is_balanced': False,
                'error': str(e)
            },
            "date_from": date_from,
            "date_to": date_to,
            "display_mode": display_mode,
            "account_level": account_level or '',
            "hide_zero": hide_zero,
            "group_by_type": group_by_type,
        }

    return render(request, "financial/reports/trial_balance_report.html", context)


@login_required
def sales_report(request):
    """
    تقرير المبيعات - محدّث ✅
    بناءً على حسابات الإيرادات مع تحليلات متقدمة
    """
    from financial.services.sales_report_service import SalesReportService
    import json

    # تحديد فترة التقرير
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        except ValueError:
            date_from = None
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            date_to = None
    
    # إنشاء خدمة التقرير
    sales_service = SalesReportService(
        date_from=date_from,
        date_to=date_to
    )
    
    # الحصول على التقرير الكامل
    report = sales_service.get_complete_report()
    
    # تحويل البيانات إلى JSON للاستخدام في JavaScript
    from decimal import Decimal
    
    class DecimalEncoder(json.JSONEncoder):
        def default(self, obj):
            if isinstance(obj, Decimal):
                return float(obj)
            return super(DecimalEncoder, self).default(obj)
    
    daily_trend_json = json.dumps(report["daily_trend"], cls=DecimalEncoder)
    monthly_comparison_json = json.dumps(report["monthly_comparison"], cls=DecimalEncoder)
    sales_by_category_json = json.dumps(report["sales_by_category"], cls=DecimalEncoder)
    
    context = {
        "page_title": "تقرير المبيعات",
        "page_subtitle": "تحليل شامل للإيرادات والمبيعات",
        "page_icon": "fas fa-chart-line",
        "header_buttons": [
            {
                "onclick": "window.print()",
                "icon": "fa-print",
                "text": "طباعة",
                "class": "btn-outline-secondary",
            },
        ],
        "breadcrumb_items": [
            {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
            {"title": "الإدارة المالية", "icon": "fas fa-money-bill-wave"},
            {"title": "التقارير", "icon": "fas fa-chart-bar"},
            {"title": "تقرير المبيعات", "active": True},
        ],
        "sales_data": report["sales_data"],
        "total_sales": report["total_sales"],
        "statistics": report["statistics"],
        "daily_trend": daily_trend_json,
        "monthly_comparison": monthly_comparison_json,
        "sales_by_category": sales_by_category_json,
        "date_from": report["date_from"],
        "date_to": report["date_to"],
        # للتوافق مع القالب القديم
        "avg_daily_sales": report["statistics"]["avg_daily_sales"],
        "days_count": report["statistics"]["days_count"],
    }
    return render(request, "financial/reports/sales_report.html", context)


@login_required
def purchases_report(request):
    """
    تقرير المشتريات - محدّث ✅
    بناءً على حسابات المصروفات مع تحليلات متقدمة
    """
    from financial.services.purchases_report_service import PurchasesReportService
    import json

    # تحديد فترة التقرير
    date_from = request.GET.get("date_from")
    date_to = request.GET.get("date_to")
    
    if date_from:
        try:
            date_from = datetime.strptime(date_from, "%Y-%m-%d").date()
        except ValueError:
            date_from = None
    
    if date_to:
        try:
            date_to = datetime.strptime(date_to, "%Y-%m-%d").date()
        except ValueError:
            date_to = None
    
    # إنشاء خدمة التقرير
    purchases_service = PurchasesReportService(
        date_from=date_from,
        date_to=date_to
    )
    
    # الحصول على التقرير الكامل
    report = purchases_service.get_complete_report()
    
    # تحويل البيانات إلى JSON للاستخدام في JavaScript
    from decimal import Decimal
    
    class DecimalEncoder(json.JSONEncoder):
        def default(self, obj):
            if isinstance(obj, Decimal):
                return float(obj)
            return super(DecimalEncoder, self).default(obj)
    
    monthly_comparison_json = json.dumps(report["monthly_comparison"], cls=DecimalEncoder)
    purchases_by_category_json = json.dumps(report["purchases_by_category"], cls=DecimalEncoder)
    
    context = {
        "page_title": "تقرير المشتريات",
        "page_subtitle": "تحليل شامل للمصروفات والمشتريات",
        "page_icon": "fas fa-shopping-cart",
        "header_buttons": [
            {
                "onclick": "window.print()",
                "icon": "fa-print",
                "text": "طباعة",
                "class": "btn-outline-secondary",
            },
        ],
        "breadcrumb_items": [
            {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
            {"title": "الإدارة المالية", "icon": "fas fa-money-bill-wave"},
            {"title": "التقارير", "icon": "fas fa-chart-bar"},
            {"title": "تقرير المشتريات", "active": True},
        ],
        "purchases_data": report["purchases_data"],
        "total_purchases": report["total_purchases"],
        "statistics": report["statistics"],
        "monthly_comparison": monthly_comparison_json,
        "purchases_by_category": purchases_by_category_json,
        "date_from": report["date_from"],
        "date_to": report["date_to"],
        # للتوافق مع القالب القديم
        "avg_daily_purchases": report["statistics"]["avg_daily_purchases"],
        "days_count": report["statistics"]["days_count"],
    }
    return render(request, "financial/reports/purchases_report.html", context)


@login_required
def inventory_report(request):
    """
    تقرير المخزون - محدّث ✅
    بناءً على حسابات الأصول المتعلقة بالمخزون مع تحليلات متقدمة
    """
    from financial.services.inventory_report_service import InventoryReportService
    from django.http import HttpResponse

    # تحديد تاريخ التقرير
    report_date = request.GET.get("date")
    if report_date:
        try:
            report_date = datetime.strptime(report_date, "%Y-%m-%d").date()
        except ValueError:
            report_date = None

    # إنشاء خدمة التقرير
    inventory_service = InventoryReportService(report_date=report_date)

    # التحقق من طلب التصدير
    if request.GET.get('export') == 'excel':
        try:
            excel_data = inventory_service.export_to_excel()
            
            if excel_data:
                response = HttpResponse(
                    excel_data,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"inventory_report_{inventory_service.report_date.strftime('%Y%m%d')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            else:
                messages.warning(request, "تصدير Excel غير متاح. يرجى تثبيت openpyxl")
        except Exception as e:
            messages.error(request, f"خطأ في تصدير Excel: {e}")

    # الحصول على التقرير الكامل
    try:
        report = inventory_service.get_complete_report()

        context = {
            "page_title": "تقرير المخزون",
            "page_subtitle": "تحليل شامل لقيمة وحركة المخزون",
            "page_icon": "fas fa-boxes",
            "header_buttons": [
                {
                    "onclick": "window.print()",
                    "icon": "fa-print",
                    "text": "طباعة",
                    "class": "btn-outline-secondary",
                },
                {
                    "url": f"?date={report['report_date']}&export=excel",
                    "icon": "fa-file-excel",
                    "text": "تصدير Excel",
                    "class": "btn-success",
                },
            ],
            "breadcrumb_items": [
                {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
                {"title": "الإدارة المالية", "icon": "fas fa-money-bill-wave"},
                {"title": "التقارير", "icon": "fas fa-chart-bar"},
                {"title": "تقرير المخزون", "active": True},
            ],
            "inventory_data": report["inventory_data"],
            "total_inventory_value": report["total_inventory_value"],
            "statistics": report["statistics"],
            "inventory_by_category": report["inventory_by_category"],
            "turnover_analysis": report["turnover_analysis"],
            "report_date": report["report_date"],
            # للتوافق مع القالب القديم
            "total_accounts": report["statistics"]["total_accounts"],
            "avg_account_value": report["statistics"]["avg_account_value"],
            "active_accounts": report["statistics"]["active_accounts"],
        }
    except Exception as e:
        messages.error(request, f"خطأ في تحميل تقرير المخزون: {e}")
        context = {
            "page_title": "تقرير المخزون",
            "page_icon": "fas fa-boxes",
            "breadcrumb_items": [
                {"title": "الرئيسية", "url": "/", "icon": "fas fa-home"},
                {"title": "الإدارة المالية", "url": "#", "icon": "fas fa-money-bill-wave"},
                {"title": "التقارير", "url": "#", "icon": "fas fa-chart-bar"},
                {"title": "تقرير المخزون", "active": True, "icon": "fas fa-boxes"},
            ],
            "inventory_data": [],
            "total_inventory_value": 0,
            "total_accounts": 0,
            "avg_account_value": 0,
            "active_accounts": 0,
            "report_date": inventory_service.report_date,
            "error": str(e)
        }

    return render(request, "financial/reports/inventory_report.html", context)


@login_required
def abc_analysis_report(request):
    """
    تقرير تحليل ABC - محدّث ✅
    تصنيف المخزون حسب الأهمية (A, B, C)
    """
    from financial.services.abc_analysis_service import ABCAnalysisService
    from django.http import HttpResponse

    # تحديد تاريخ التحليل وفترة التحليل
    analysis_date = request.GET.get("date")
    days_period = request.GET.get("days_period", "365")
    
    if analysis_date:
        try:
            analysis_date = datetime.strptime(analysis_date, "%Y-%m-%d").date()
        except ValueError:
            analysis_date = None
    
    try:
        days_period = int(days_period)
    except ValueError:
        days_period = 365

    # إنشاء خدمة التحليل
    abc_service = ABCAnalysisService(
        analysis_date=analysis_date,
        days_period=days_period
    )

    # التحقق من طلب التصدير
    if request.GET.get('export') == 'excel':
        try:
            excel_data = abc_service.export_to_excel()
            
            if excel_data:
                response = HttpResponse(
                    excel_data,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                filename = f"abc_analysis_{abc_service.analysis_date.strftime('%Y%m%d')}.xlsx"
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
            else:
                messages.warning(request, "تصدير Excel غير متاح. يرجى تثبيت openpyxl")
        except Exception as e:
            messages.error(request, f"خطأ في تصدير Excel: {e}")

    # الحصول على التحليل الكامل
    try:
        analysis = abc_service.get_complete_analysis()

        context = {
            "page_title": "تحليل ABC للمخزون",
            "page_subtitle": "تصنيف المخزون حسب الأهمية (A, B, C)",
            "page_icon": "fas fa-chart-pie",
            "header_buttons": [
                {
                    "onclick": "window.print()",
                    "icon": "fa-print",
                    "text": "طباعة",
                    "class": "btn-outline-secondary",
                },
                {
                    "url": f"?date={analysis['analysis_date']}&days_period={analysis['days_period']}&export=excel",
                    "icon": "fa-file-excel",
                    "text": "تصدير Excel",
                    "class": "btn-success",
                },
            ],
            "breadcrumb_items": [
                {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
                {"title": "الإدارة المالية", "icon": "fas fa-money-bill-wave"},
                {"title": "التقارير", "icon": "fas fa-chart-bar"},
                {"title": "تحليل ABC", "active": True},
            ],
            "inventory_data": analysis["inventory_data"],
            "total_value": analysis["total_value"],
            "statistics": analysis["statistics"],
            "recommendations": analysis["recommendations"],
            "analysis_date": analysis["analysis_date"],
            "days_period": analysis["days_period"],
            "date_from": analysis["date_from"],
        }
    except Exception as e:
        messages.error(request, f"خطأ في تحميل تحليل ABC: {e}")
        context = {
            "page_title": "تحليل ABC للمخزون",
            "page_icon": "fas fa-chart-pie",
            "breadcrumb_items": [
                {"title": "الرئيسية", "url": "/", "icon": "fas fa-home"},
                {"title": "الإدارة المالية", "url": "#", "icon": "fas fa-money-bill-wave"},
                {"title": "التقارير", "url": "#", "icon": "fas fa-chart-bar"},
                {"title": "تحليل ABC", "active": True, "icon": "fas fa-chart-pie"},
            ],
            "inventory_data": [],
            "total_value": 0,
            "statistics": {
                "category_a": {"count": 0, "value": 0, "percentage_count": 0, "percentage_value": 0, "avg_value": 0},
                "category_b": {"count": 0, "value": 0, "percentage_count": 0, "percentage_value": 0, "avg_value": 0},
                "category_c": {"count": 0, "value": 0, "percentage_count": 0, "percentage_value": 0, "avg_value": 0},
                "total_items": 0,
                "total_value": 0,
            },
            "recommendations": {"category_a": [], "category_b": [], "category_c": []},
            "analysis_date": abc_service.analysis_date,
            "days_period": days_period,
            "error": str(e)
        }

    return render(request, "financial/reports/abc_analysis.html", context)


@login_required

@login_required

@login_required
def data_integrity_check(request):
    """
    التحقق من سلامة البيانات - فحص شامل
    """
    from financial.models import JournalEntry, ChartOfAccounts
    from purchase.models import Purchase
    from product.models import Stock, StockMovement
    from client.models import Customer
    from supplier.models import Supplier
    from django.db.models import Sum, Q, Count
    from decimal import Decimal
    
    results = {
        'checks': [],
        'errors': [],
        'warnings': [],
        'summary': {
            'total_checks': 0,
            'passed': 0,
            'failed': 0,
            'warnings': 0
        }
    }
    
    if request.method == "POST":
        try:
            # ==================== 1. فحص القيود المحاسبية ====================
            check_name = "توازن القيود المحاسبية"
            results['summary']['total_checks'] += 1
            
            unbalanced_entries = []
            for entry in JournalEntry.objects.all():
                debits = entry.lines.aggregate(total=Sum('debit'))['total'] or Decimal('0')
                credits = entry.lines.aggregate(total=Sum('credit'))['total'] or Decimal('0')
                if debits != credits:
                    unbalanced_entries.append({
                        'entry': entry,
                        'difference': abs(debits - credits)
                    })
            
            if unbalanced_entries:
                results['errors'].append({
                    'check': check_name,
                    'count': len(unbalanced_entries),
                    'details': unbalanced_entries[:10],  # أول 10 فقط
                    'severity': 'high'
                })
                results['summary']['failed'] += 1
            else:
                results['checks'].append({'name': check_name, 'status': 'passed'})
                results['summary']['passed'] += 1
            
            # ==================== 2. فحص أرصدة العملاء ====================
            check_name = "أرصدة العملاء"
            results['summary']['total_checks'] += 1

            customer_issues = []
            try:
                from sale.models import Sale

                for customer in Customer.objects.all():
                    if not customer.financial_account:
                        continue

                    total_sales = Sale.objects.filter(
                        customer=customer, status='confirmed'
                    ).aggregate(total=Sum('total'))['total'] or Decimal('0')

                    total_paid = Sale.objects.filter(
                        customer=customer, status='confirmed'
                    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')

                    calculated_balance = total_sales - total_paid
                    account_balance = customer.financial_account.get_balance()

                    if abs(calculated_balance - account_balance) > Decimal('0.01'):
                        customer_issues.append({
                            'customer': customer,
                            'system_balance': account_balance,
                            'calculated_balance': calculated_balance,
                            'difference': abs(account_balance - calculated_balance),
                        })
            except Exception as e:
                logger.error(f"خطأ في فحص أرصدة العملاء: {str(e)}")

            if customer_issues:
                results['warnings'].append({
                    'check': check_name,
                    'count': len(customer_issues),
                    'details': customer_issues[:10],
                    'severity': 'medium'
                })
                results['summary']['warnings'] += 1
            else:
                results['checks'].append({'name': check_name, 'status': 'passed'})
                results['summary']['passed'] += 1
            
            # ==================== 3. فحص أرصدة الموردين ====================
            check_name = "أرصدة الموردين"
            results['summary']['total_checks'] += 1
            
            supplier_issues = []
            try:
                from purchase.models import PurchasePayment
                
                for supplier in Supplier.objects.all():
                    calculated_balance = Decimal('0')
                    
                    # حساب من المشتريات
                    purchases_total = Purchase.objects.filter(supplier=supplier).aggregate(
                        total=Sum('total')
                    )['total'] or Decimal('0')
                    
                    # حساب من الدفعات
                    if PurchasePayment:
                        payments_total = PurchasePayment.objects.filter(purchase__supplier=supplier).aggregate(
                            total=Sum('amount')
                        )['total'] or Decimal('0')
                    else:
                        payments_total = Decimal('0')
                    
                    calculated_balance = purchases_total - payments_total
                    
                    if abs(calculated_balance - supplier.balance) > Decimal('0.01'):
                        supplier_issues.append({
                            'supplier': supplier,
                            'system_balance': supplier.balance,
                            'calculated_balance': calculated_balance,
                            'difference': abs(supplier.balance - calculated_balance)
                        })
            except Exception as e:
                # تجاهل الخطأ إذا كان النموذج غير موجود
                pass
            
            if supplier_issues:
                results['warnings'].append({
                    'check': check_name,
                    'count': len(supplier_issues),
                    'details': supplier_issues[:10],
                    'severity': 'medium'
                })
                results['summary']['warnings'] += 1
            else:
                results['checks'].append({'name': check_name, 'status': 'passed'})
                results['summary']['passed'] += 1
            
            # ==================== 4. فحص المخزون ====================
            check_name = "أرصدة المخزون"
            results['summary']['total_checks'] += 1
            
            stock_issues = []
            for stock in Stock.objects.all():
                # حساب من حركات المخزون
                movements_in = StockMovement.objects.filter(
                    product=stock.product,
                    movement_type__in=['purchase', 'return_in', 'adjustment_in']
                ).aggregate(total=Sum('quantity'))['total'] or 0
                
                movements_out = StockMovement.objects.filter(
                    product=stock.product,
                    movement_type__in=['sale', 'return_out', 'adjustment_out']
                ).aggregate(total=Sum('quantity'))['total'] or 0
                
                calculated_quantity = movements_in - movements_out
                
                if abs(calculated_quantity - stock.quantity) > 0.01:
                    stock_issues.append({
                        'product': stock.product,
                        'system_quantity': stock.quantity,
                        'calculated_quantity': calculated_quantity,
                        'difference': abs(stock.quantity - calculated_quantity)
                    })
            
            if stock_issues:
                results['errors'].append({
                    'check': check_name,
                    'count': len(stock_issues),
                    'details': stock_issues[:10],
                    'severity': 'high'
                })
                results['summary']['failed'] += 1
            else:
                results['checks'].append({'name': check_name, 'status': 'passed'})
                results['summary']['passed'] += 1
            
            # ==================== 5. فحص القيود اليتيمة ====================
            check_name = "القيود بدون مستند مرجعي"
            results['summary']['total_checks'] += 1
            
            orphan_entries = JournalEntry.objects.filter(
                Q(reference_type__isnull=True) | Q(reference_id__isnull=True)
            ).exclude(entry_type='manual').count()
            
            if orphan_entries > 0:
                results['warnings'].append({
                    'check': check_name,
                    'count': orphan_entries,
                    'message': f"يوجد {orphan_entries} قيد بدون مستند مرجعي",
                    'severity': 'low'
                })
                results['summary']['warnings'] += 1
            else:
                results['checks'].append({'name': check_name, 'status': 'passed'})
                results['summary']['passed'] += 1
            
            # ==================== 6. فحص الحسابات المكررة ====================
            check_name = "الحسابات المكررة"
            results['summary']['total_checks'] += 1
            
            duplicate_accounts = ChartOfAccounts.objects.values('code').annotate(
                count=Count('id')
            ).filter(count__gt=1)
            
            if duplicate_accounts.exists():
                results['errors'].append({
                    'check': check_name,
                    'count': duplicate_accounts.count(),
                    'message': f"يوجد {duplicate_accounts.count()} رمز حساب مكرر",
                    'severity': 'high'
                })
                results['summary']['failed'] += 1
            else:
                results['checks'].append({'name': check_name, 'status': 'passed'})
                results['summary']['passed'] += 1
            
            # ==================== 7. فحص المخزون السالب ====================
            check_name = "المخزون السالب"
            results['summary']['total_checks'] += 1
            
            negative_stock = Stock.objects.filter(quantity__lt=0).count()
            
            if negative_stock > 0:
                results['warnings'].append({
                    'check': check_name,
                    'count': negative_stock,
                    'message': f"يوجد {negative_stock} منتج برصيد سالب",
                    'severity': 'medium'
                })
                results['summary']['warnings'] += 1
            else:
                results['checks'].append({'name': check_name, 'status': 'passed'})
                results['summary']['passed'] += 1
            
            # ==================== 8. فحص الفواتير المعلقة ====================
            check_name = "الفواتير غير المكتملة"
            results['summary']['total_checks'] += 1

            try:
                from sale.models import Sale
                from django.utils import timezone
                thirty_days_ago = timezone.now() - timedelta(days=30)

                overdue_sales = Sale.objects.filter(
                    status='pending',
                    created_at__lt=thirty_days_ago
                ).count()

                if overdue_sales > 0:
                    results['warnings'].append({
                        'check': check_name,
                        'count': overdue_sales,
                        'message': f"يوجد {overdue_sales} فاتورة معلقة لأكثر من 30 يوم",
                        'severity': 'medium'
                    })
                    results['summary']['warnings'] += 1
                else:
                    results['checks'].append({'name': check_name, 'status': 'passed'})
                    results['summary']['passed'] += 1
            except Exception as e:
                logger.error(f"خطأ في فحص الفواتير: {str(e)}")
                results['summary']['passed'] += 1
            
            # رسالة النجاح
            if results['summary']['failed'] == 0 and results['summary']['warnings'] == 0:
                messages.success(request, "✅ تم فحص سلامة البيانات بنجاح. جميع الفحوصات اجتازت بنجاح!")
            elif results['summary']['failed'] > 0:
                messages.error(request, f"⚠️ تم العثور على {results['summary']['failed']} مشكلة حرجة تحتاج إلى إصلاح فوري!")
            else:
                messages.warning(request, f"⚠️ تم العثور على {results['summary']['warnings']} تحذير يحتاج إلى مراجعة.")
                
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء فحص البيانات: {str(e)}")
            import traceback
            traceback.print_exc()

    context = {
        "title": "التحقق من سلامة البيانات",
        "subtitle": "فحص شامل للتأكد من توافق وسلامة البيانات",
        "icon": "fas fa-shield-alt",
        "header_buttons": [
            {
                "onclick": "submitIntegrityCheck()",
                "icon": "fa-sync",
                "text": "بدء الفحص",
                "class": "btn-primary",
            },
        ],
        "breadcrumb_items": [
            {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
            {"title": "الإدارة المالية", "url": reverse('financial:cash_and_bank_accounts_list'), "icon": "fas fa-money-bill-wave"},
            {"title": "الصيانة", "icon": "fas fa-tools"},
            {"title": "فحص سلامة البيانات", "active": True},
        ],
        "results": results,
        "active_menu": "financial",
    }
    return render(request, "financial/reports/data_integrity_check.html", context)


@require_http_methods(["GET"])
@login_required
def payment_sync_check_pending_api(request):
    """
    API لفحص العمليات المعلقة
    """
    try:
        from financial.models.payment_sync import PaymentSyncOperation
        from django.utils import timezone

        # العمليات المعلقة
        pending_ops = PaymentSyncOperation.objects.filter(status="pending")
        processing_ops = PaymentSyncOperation.objects.filter(status="processing")

        # العمليات العالقة (أكثر من 10 دقائق)
        ten_minutes_ago = timezone.now() - timedelta(minutes=10)
        stuck_pending = pending_ops.filter(created_at__lt=ten_minutes_ago).count()
        stuck_processing = processing_ops.filter(started_at__lt=ten_minutes_ago).count()

        return JsonResponse(
            {
                "success": True,
                "pending_count": pending_ops.count(),
                "processing_count": processing_ops.count(),
                "stuck_operations": stuck_pending + stuck_processing,
                "details": {
                    "stuck_pending": stuck_pending,
                    "stuck_processing": stuck_processing,
                },
            }
        )

    except ImportError:
        return JsonResponse({"success": False, "message": "نماذج التزامن غير متاحة"})
    except Exception as e:
        logger.error(f"Error in views.py: {str(e)}", exc_info=True)
        return JsonResponse({"success": False, "message": "حدث خطأ غير متوقع"})


@require_http_methods(["POST"])
@login_required
def payment_sync_process_pending_api(request):
    """
    API لتشغيل العمليات المعلقة
    """
    try:
        from financial.models.payment_sync import PaymentSyncOperation
        from financial.services.payment_sync_service import PaymentSyncService

        # جلب العمليات المعلقة
        pending_ops = PaymentSyncOperation.objects.filter(status="pending").order_by(
            "created_at"
        )

        if not pending_ops.exists():
            return JsonResponse(
                {
                    "success": True,
                    "message": "لا توجد عمليات معلقة",
                    "processed_count": 0,
                }
            )

        # تشغيل العمليات
        sync_service = PaymentSyncService()
        processed_count = 0

        for operation in pending_ops[:10]:  # معالجة 10 عمليات كحد أقصى
            try:
                # تحديث حالة العملية إلى قيد المعالجة
                operation.status = "processing"
                operation.started_at = timezone.now()
                operation.save()

                # محاولة تنفيذ العملية
                if operation.operation_type == "retry_failed":
                    # إعادة محاولة العملية الفاشلة
                    sync_service.retry_failed_operation(operation)
                elif operation.operation_type == "delete_payment":
                    # حذف دفعة
                    sync_service.process_payment_deletion(operation)
                else:
                    # عملية عامة
                    sync_service.process_operation(operation)

                processed_count += 1

            except Exception as e:
                # تسجيل فشل العملية
                operation.status = "failed"
                operation.error_message = str(e)
                operation.save()

        return JsonResponse(
            {
                "success": True,
                "message": f"تم تشغيل {processed_count} عملية",
                "processed_count": processed_count,
            }
        )

    except ImportError:
        return JsonResponse({"success": False, "message": "خدمة التزامن غير متاحة"})
    except Exception as e:
        logger.error(f"Error in views.py: {str(e)}", exc_info=True)
        return JsonResponse({"success": False, "message": "حدث خطأ غير متوقع"})


@login_required
def audit_trail_list(request):
    """
    قائمة سجل التدقيق
    """
    try:
        from financial.models import AuditTrail
        
        # الفلترة
        action_filter = request.GET.get('action', '')
        entity_type_filter = request.GET.get('entity_type', '')
        user_filter = request.GET.get('user', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # الاستعلام الأساسي
        audit_entries = AuditTrail.objects.select_related('user').order_by('-timestamp')
        
        # تطبيق الفلاتر
        if action_filter:
            audit_entries = audit_entries.filter(action=action_filter)
        
        if entity_type_filter:
            audit_entries = audit_entries.filter(entity_type=entity_type_filter)
        
        if user_filter:
            audit_entries = audit_entries.filter(user_id=user_filter)
        
        if date_from:
            audit_entries = audit_entries.filter(timestamp__date__gte=date_from)
        
        if date_to:
            audit_entries = audit_entries.filter(timestamp__date__lte=date_to)
        
        # الترقيم الصفحي
        paginator = Paginator(audit_entries, 50)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        # الإحصائيات
        from django.contrib.auth import get_user_model
        from datetime import datetime, timedelta
        
        User = get_user_model()
        today = datetime.now().date()
        
        total_entries = audit_entries.count()
        today_entries = AuditTrail.objects.filter(timestamp__date=today).count()
        active_users = AuditTrail.objects.filter(timestamp__date=today).values('user').distinct().count()
        delete_entries = AuditTrail.objects.filter(action='delete').count()
        
        # قائمة المستخدمين للفلترة
        users = User.objects.filter(
            id__in=AuditTrail.objects.values_list('user_id', flat=True).distinct()
        ).order_by('first_name', 'last_name', 'username')
        
        context = {
            "page_obj": page_obj,
            "total_entries": total_entries,
            "users": users,
            "summary": {
                "total_count": total_entries,
                "today_count": today_entries,
                "active_users": active_users,
                "delete_count": delete_entries,
            },
            "filters": {
                "action": action_filter,
                "entity_type": entity_type_filter,
                "user": user_filter,
                "date_from": date_from,
                "date_to": date_to,
            },
            "page_title": "سجل التدقيق",
            "page_subtitle": "تتبع جميع العمليات والتغييرات في النظام",
            "page_icon": "fas fa-clipboard-list",
            "header_buttons": [
                {
                    "onclick": "window.print()",
                    "icon": "fa-print",
                    "text": "طباعة",
                    "class": "btn-outline-secondary",
                },
                {
                    "onclick": "confirmCleanup()",
                    "icon": "fa-trash",
                    "text": "تنظيف السجل",
                    "class": "btn-outline-danger",
                },
            ],
            "breadcrumb_items": [
                {"title": "الرئيسية", "url": reverse('core:dashboard'), "icon": "fas fa-home"},
                {"title": "الإدارة المالية", "icon": "fas fa-money-bill-wave"},
                {"title": "التقارير", "icon": "fas fa-chart-bar"},
                {"title": "سجل التدقيق", "active": True},
            ],
            "action_filter": action_filter,
            "entity_type_filter": entity_type_filter,
        }

        return render(request, "financial/reports/audit_trail_list.html", context)

    except Exception as e:
        messages.error(request, f"خطأ في تحميل سجل التدقيق: {str(e)}")
        return render(request, "financial/reports/audit_trail_list.html", {"page_obj": None})


@login_required
@transaction.atomic
def audit_trail_cleanup(request):
    """
    تنظيف سجل التدقيق - حذف السجلات القديمة
    """
    if request.method == 'POST':
        try:
            from financial.models import AuditTrail
            from datetime import datetime
            
            cleanup_date = request.POST.get('cleanup_date')
            if not cleanup_date:
                messages.error(request, "يجب تحديد تاريخ التنظيف")
                return redirect("financial:audit_trail_list")
            
            # تحويل التاريخ
            cleanup_date = datetime.strptime(cleanup_date, '%Y-%m-%d').date()
            
            # فحص السجلات الموجودة للتشخيص
            total_records = AuditTrail.objects.count()
            records_before_date = AuditTrail.objects.filter(
                timestamp__date__lte=cleanup_date
            ).count()
            
            print(f"DEBUG: إجمالي السجلات: {total_records}")
            print(f"DEBUG: السجلات قبل {cleanup_date}: {records_before_date}")
            
            # حذف السجلات الأقدم من أو تساوي التاريخ المحدد
            records_to_delete = AuditTrail.objects.filter(
                timestamp__date__lte=cleanup_date
            )
            
            deleted_count = records_to_delete.count()
            print(f"DEBUG: سيتم حذف {deleted_count} سجل")
            
            # عرض بعض السجلات التي سيتم حذفها للتشخيص
            if deleted_count > 0:
                sample_records = records_to_delete[:5]
                print("DEBUG: عينة من السجلات التي سيتم حذفها:")
                for record in sample_records:
                    print(f"  - ID: {record.id}, التاريخ: {record.timestamp}, الوصف: {record.description[:50]}")
            
            if deleted_count > 0:
                # تنفيذ الحذف الفعلي
                deleted_result = records_to_delete.delete()
                actual_deleted = deleted_result[0]  # العدد الفعلي المحذوف
                
                print(f"DEBUG: تم حذف {actual_deleted} سجل فعلياً")
                
                # فحص السجلات المتبقية للتأكد
                remaining_records = AuditTrail.objects.count()
                print(f"DEBUG: السجلات المتبقية بعد الحذف: {remaining_records}")
                
                success_message = f"تم حذف {actual_deleted} سجل تدقيق أقدم من أو يساوي {cleanup_date}. السجلات المتبقية: {remaining_records}"
                
                # للطلبات AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True, 
                        'message': success_message,
                        'deleted_count': actual_deleted,
                        'remaining_count': remaining_records
                    })
                
                messages.success(request, success_message)
                
                # تسجيل عملية التنظيف في سجل التدقيق
                AuditTrail.log_action(
                    action='delete',
                    entity_type='audit_trail',
                    entity_id=0,
                    user=request.user,
                    description=f"تنظيف سجل التدقيق - حذف {actual_deleted} سجل أقدم من {cleanup_date}",
                    reason="تنظيف دوري للسجلات القديمة",
                    request=request
                )
            else:
                info_message = f"لا توجد سجلات أقدم من أو تساوي {cleanup_date} للحذف. إجمالي السجلات الحالية: {total_records}"
                
                # للطلبات AJAX
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False, 
                        'message': info_message,
                        'deleted_count': 0,
                        'remaining_count': total_records
                    })
                
                messages.info(request, info_message)
                
        except ValueError:
            error_message = "تاريخ غير صحيح"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_message})
            messages.error(request, error_message)
        except Exception as e:
            error_message = f"خطأ في تنظيف السجل: {str(e)}"
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': error_message})
            messages.error(request, error_message)
    
    return redirect("financial:audit_trail_list")


@login_required
def payment_sync_operations(request):
    """
    عمليات تزامن المدفوعات
    """
    try:
        from .models.payment_sync import PaymentSyncOperation
        
        operations = PaymentSyncOperation.objects.select_related(
            'created_by'
        ).order_by('-created_at')[:100]
        
        context = {
            "operations": operations,
            "page_title": "عمليات تزامن المدفوعات",
            "page_icon": "fas fa-sync-alt",
        }
        return render(request, "financial/banking/payment_sync_operations.html", context)
    except ImportError:
        messages.warning(request, "نماذج تزامن المدفوعات غير متاحة حالياً.")
        return render(request, "financial/banking/payment_sync_operations.html", {"operations": []})


@login_required
def payment_sync_logs(request):
    """
    سجلات تزامن المدفوعات
    """
    try:
        from .models.payment_sync import PaymentSyncError
        
        logs = PaymentSyncError.objects.order_by('-occurred_at')[:100]
        
        context = {
            "logs": logs,
            "page_title": "سجلات تزامن المدفوعات",
            "page_icon": "fas fa-list-alt",
        }
        return render(request, "financial/banking/payment_sync_logs.html", context)
    except ImportError:
        messages.warning(request, "نماذج سجلات التزامن غير متاحة حالياً.")
        return render(request, "financial/banking/payment_sync_logs.html", {"logs": []})


@login_required
def journal_entry_summary_api(request, journal_entry_id):
    """
    API لجلب ملخص القيد المحاسبي
    """
    try:
        journal_entry = get_object_or_404(JournalEntry, id=journal_entry_id)
        
        # جلب بيانات القيد
        data = {
            'id': journal_entry.id,
            'number': journal_entry.number,
            'reference': journal_entry.reference,
            'date': journal_entry.date.strftime('%Y-%m-%d') if journal_entry.date else '',
            'description': journal_entry.description,
            'status': journal_entry.status,
            'created_by': journal_entry.created_by.get_full_name() if journal_entry.created_by else 'غير محدد',
            'lines': []
        }
        
        # جلب بنود القيد
        for line in journal_entry.lines.all():
            data['lines'].append({
                'account_name': line.account.name,
                'account_code': line.account.code,
                'debit': float(line.debit) if line.debit else 0,
                'credit': float(line.credit) if line.credit else 0,
                'description': line.description or ''
            })
        
        return JsonResponse(data)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"خطأ في جلب ملخص القيد {journal_entry_id}: {str(e)}", exc_info=True)
        return JsonResponse({
            'error': 'حدث خطأ أثناء جلب تفاصيل القيد'
        }, status=500)


@login_required
def api_get_exchange_rate(request):
    """
    API لاستعلام سعر الصرف اللحظي والعملة الوظيفية الأساسية
    """
    try:
        currency_id = request.GET.get('currency_id')
        currency_code = request.GET.get('code')
        date_str = request.GET.get('date')

        from financial.models.currency import Currency
        from financial.services.exchange_rate_service import ExchangeRateService

        target_curr = None
        if currency_id:
            target_curr = Currency.objects.filter(pk=currency_id, is_active=True).first()
        elif currency_code:
            target_curr = Currency.objects.filter(code=currency_code, is_active=True).first()

        func_curr = ExchangeRateService.get_functional_currency()
        if not func_curr:
            return JsonResponse({'success': False, 'error': 'لم يتم تعيين العملة الأساسية للمؤسسة.'}, status=400)

        if not target_curr:
            target_curr = func_curr

        as_of_date = None
        if date_str:
            try:
                as_of_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        rate = Decimal('1.000000')
        if target_curr.code != func_curr.code:
            try:
                rate = ExchangeRateService.get_rate(target_curr.code, func_curr.code, as_of_date)
            except Exception as e:
                logger.warning(f"Exchange rate lookup notice: {e}")
                rate = Decimal('1.000000')

        return JsonResponse({
            'success': True,
            'rate': str(rate),
            'currency_id': target_curr.id,
            'currency_code': target_curr.code,
            'currency_symbol': target_curr.symbol or target_curr.code,
            'is_functional': target_curr.is_functional,
            'functional_code': func_curr.code,
            'functional_symbol': func_curr.symbol or func_curr.code,
        })
    except Exception as e:
        logger.error(f"Error in api_get_exchange_rate: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============== اكتمل ملف api_views.py بالكامل ==============
# تم نقل جميع دوال APIs والتصدير والتقارير بنجاح
