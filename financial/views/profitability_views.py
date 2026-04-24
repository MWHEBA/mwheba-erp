"""
Profitability Report Views
واجهات تقارير الربحية حسب التصنيفات المالية
"""
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import TemplateView
from django.urls import reverse_lazy
from django.http import HttpResponse
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..services.category_profitability_service import CategoryProfitabilityService
from ..models.categories import FinancialCategory


class ProfitabilityDashboardView(LoginRequiredMixin, TemplateView):
    """لوحة تقارير الربحية"""
    template_name = 'financial/reports/profitability_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # الحصول على فلاتر التاريخ
        date_from_str = self.request.GET.get('date_from')
        date_to_str = self.request.GET.get('date_to')
        
        # تحويل التواريخ
        date_from = None
        date_to = None
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # إذا لم يتم تحديد تواريخ، استخدم الشهر الحالي
        if not date_from and not date_to:
            today = datetime.now().date()
            date_from = today.replace(day=1)
            # آخر يوم في الشهر
            if today.month == 12:
                date_to = today.replace(day=31)
            else:
                date_to = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
        
        # جلب التقرير
        summary = CategoryProfitabilityService.get_all_summary(date_from, date_to)
        
        # جلب أفضل 3 تصنيفات
        top_categories = CategoryProfitabilityService.get_top_profitable_categories(
            limit=3, date_from=date_from, date_to=date_to
        )
        
        # جلب التصنيفات الخاسرة
        loss_categories = CategoryProfitabilityService.get_loss_making_categories(
            date_from=date_from, date_to=date_to
        )
        
        context.update({
            'active_menu': 'financial',
            'title': 'تقرير الربحية حسب التصنيفات',
            'breadcrumb_items': [
                {'title': 'الرئيسية', 'url': reverse_lazy('core:dashboard'), 'icon': 'fas fa-home'},
                {'title': 'المالية', 'url': reverse_lazy('financial:dashboard'), 'icon': 'fas fa-coins'},
                {'title': 'تقرير الربحية', 'active': True}
            ],
            'date_from': date_from,
            'date_to': date_to,
            'summary': summary,
            'top_categories': top_categories,
            'loss_categories': loss_categories,
        })
        
        return context


class CategoryDetailReportView(LoginRequiredMixin, TemplateView):
    """تقرير تفصيلي لتصنيف واحد"""
    template_name = 'financial/reports/category_detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        category_code = kwargs.get('code')
        
        # الحصول على فلاتر التاريخ
        date_from_str = self.request.GET.get('date_from')
        date_to_str = self.request.GET.get('date_to')
        
        date_from = None
        date_to = None
        
        if date_from_str:
            try:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        if date_to_str:
            try:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        # جلب التقرير التفصيلي
        report = CategoryProfitabilityService.get_category_report(
            category_code, date_from, date_to
        )
        
        # بناء URL التصدير
        export_url = reverse_lazy('financial:category_detail_export_excel', kwargs={'code': category_code})
        if date_from or date_to:
            params = []
            if date_from:
                params.append(f'date_from={date_from}')
            if date_to:
                params.append(f'date_to={date_to}')
            export_url = f"{export_url}?{'&'.join(params)}"
        
        context.update({
            'active_menu': 'financial',
            'title': f'تقرير تفصيلي: {report.get("category", {}).get("name", category_code)}',
            'breadcrumb_items': [
                {'title': 'الرئيسية', 'url': reverse_lazy('core:dashboard'), 'icon': 'fas fa-home'},
                {'title': 'المالية', 'url': reverse_lazy('financial:dashboard'), 'icon': 'fas fa-coins'},
                {'title': 'تقرير الربحية', 'url': reverse_lazy('financial:profitability_dashboard')},
                {'title': 'تقرير تفصيلي', 'active': True}
            ],
            'header_buttons': [
                {
                    'url': reverse_lazy('financial:profitability_dashboard'),
                    'icon': 'fa-arrow-right',
                    'text': 'العودة',
                    'class': 'btn-outline-secondary'
                },
                {
                    'url': export_url,
                    'icon': 'fa-file-excel',
                    'text': 'تصدير Excel',
                    'class': 'btn-success'
                }
            ],
            'date_from': date_from,
            'date_to': date_to,
            'report': report,
            'category_code': category_code,
        })
        
        return context


def export_profitability_excel(request):
    """تصدير تقرير الربحية إلى Excel"""
    # الحصول على فلاتر التاريخ
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    
    date_from = None
    date_to = None
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # جلب البيانات
    summary = CategoryProfitabilityService.get_all_summary(date_from, date_to)
    
    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير الربحية"
    
    # تنسيق العنوان
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان
    ws.merge_cells('A1:F1')
    ws['A1'] = 'تقرير الربحية حسب التصنيفات المالية'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # الفترة
    period_text = 'الفترة: '
    if date_from and date_to:
        period_text += f'من {date_from} إلى {date_to}'
    elif date_from:
        period_text += f'من {date_from}'
    elif date_to:
        period_text += f'حتى {date_to}'
    else:
        period_text += 'جميع الفترات'
    
    ws.merge_cells('A2:F2')
    ws['A2'] = period_text
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # رؤوس الأعمدة
    headers = ['التصنيف', 'الإيرادات', 'المصروفات', 'الربح/الخسارة', 'الهامش %', 'الحالة']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # البيانات
    row_num = 5
    category_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    subcategory_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    for category in summary.get('categories', []):
        # التصنيف الرئيسي
        category_font = Font(name='Arial', size=10, bold=True)
        ws.cell(row=row_num, column=1, value=category['name']).border = border
        ws.cell(row=row_num, column=1).font = category_font
        ws.cell(row=row_num, column=1).fill = category_fill
        
        ws.cell(row=row_num, column=2, value=float(category['revenues'])).border = border
        ws.cell(row=row_num, column=2).number_format = '#,##0.00'
        ws.cell(row=row_num, column=2).font = category_font
        ws.cell(row=row_num, column=2).fill = category_fill
        
        ws.cell(row=row_num, column=3, value=float(category['expenses'])).border = border
        ws.cell(row=row_num, column=3).number_format = '#,##0.00'
        ws.cell(row=row_num, column=3).font = category_font
        ws.cell(row=row_num, column=3).fill = category_fill
        
        ws.cell(row=row_num, column=4, value=float(category['profit'])).border = border
        ws.cell(row=row_num, column=4).number_format = '#,##0.00'
        ws.cell(row=row_num, column=4).font = category_font
        ws.cell(row=row_num, column=4).fill = category_fill
        
        ws.cell(row=row_num, column=5, value=float(category['margin'])).border = border
        ws.cell(row=row_num, column=5).number_format = '0.00'
        ws.cell(row=row_num, column=5).font = category_font
        ws.cell(row=row_num, column=5).fill = category_fill
        
        status_text = 'ربح' if category['status'] == 'profit' else 'خسارة' if category['status'] == 'loss' else 'تعادل'
        ws.cell(row=row_num, column=6, value=status_text).border = border
        ws.cell(row=row_num, column=6).font = category_font
        ws.cell(row=row_num, column=6).fill = category_fill
        
        row_num += 1
        
        # التصنيفات الفرعية
        for subcategory in category.get('subcategories', []):
            subcategory_font = Font(name='Arial', size=10)
            ws.cell(row=row_num, column=1, value=f"  ↳ {subcategory['name']}").border = border
            ws.cell(row=row_num, column=1).font = subcategory_font
            ws.cell(row=row_num, column=1).fill = subcategory_fill
            
            ws.cell(row=row_num, column=2, value=float(subcategory['revenues'])).border = border
            ws.cell(row=row_num, column=2).number_format = '#,##0.00'
            ws.cell(row=row_num, column=2).fill = subcategory_fill
            
            ws.cell(row=row_num, column=3, value=float(subcategory['expenses'])).border = border
            ws.cell(row=row_num, column=3).number_format = '#,##0.00'
            ws.cell(row=row_num, column=3).fill = subcategory_fill
            
            ws.cell(row=row_num, column=4, value=float(subcategory['profit'])).border = border
            ws.cell(row=row_num, column=4).number_format = '#,##0.00'
            ws.cell(row=row_num, column=4).fill = subcategory_fill
            
            ws.cell(row=row_num, column=5, value=float(subcategory['margin'])).border = border
            ws.cell(row=row_num, column=5).number_format = '0.00'
            ws.cell(row=row_num, column=5).fill = subcategory_fill
            
            sub_status_text = 'ربح' if subcategory['status'] == 'profit' else 'خسارة' if subcategory['status'] == 'loss' else 'تعادل'
            ws.cell(row=row_num, column=6, value=sub_status_text).border = border
            ws.cell(row=row_num, column=6).fill = subcategory_fill
            
            row_num += 1
    
    # الإجماليات
    row_num += 1
    totals = summary.get('totals', {})
    ws.cell(row=row_num, column=1, value='الإجمالي').font = Font(bold=True)
    ws.cell(row=row_num, column=1).border = border
    ws.cell(row=row_num, column=2, value=float(totals.get('revenues', 0))).font = Font(bold=True)
    ws.cell(row=row_num, column=2).number_format = '#,##0.00'
    ws.cell(row=row_num, column=2).border = border
    ws.cell(row=row_num, column=3, value=float(totals.get('expenses', 0))).font = Font(bold=True)
    ws.cell(row=row_num, column=3).number_format = '#,##0.00'
    ws.cell(row=row_num, column=3).border = border
    ws.cell(row=row_num, column=4, value=float(totals.get('profit', 0))).font = Font(bold=True)
    ws.cell(row=row_num, column=4).number_format = '#,##0.00'
    ws.cell(row=row_num, column=4).border = border
    ws.cell(row=row_num, column=5, value=float(totals.get('margin', 0))).font = Font(bold=True)
    ws.cell(row=row_num, column=5).number_format = '0.00'
    ws.cell(row=row_num, column=5).border = border
    ws.cell(row=row_num, column=6).border = border
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # إنشاء الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'profitability_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


def export_category_detail_excel(request, code):
    """تصدير التقرير التفصيلي لتصنيف إلى Excel"""
    # الحصول على فلاتر التاريخ
    date_from_str = request.GET.get('date_from')
    date_to_str = request.GET.get('date_to')
    
    date_from = None
    date_to = None
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    
    # جلب البيانات
    report = CategoryProfitabilityService.get_category_report(code, date_from, date_to)
    
    if not report.get('success'):
        return HttpResponse('خطأ في جلب البيانات', status=400)
    
    # إنشاء ملف Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير تفصيلي"
    
    # تنسيق العنوان
    title_font = Font(name='Arial', size=14, bold=True)
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان
    ws.merge_cells('A1:E1')
    ws['A1'] = f'تقرير تفصيلي: {report["category"]["name"]}'
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # الفترة
    period_text = 'الفترة: '
    if date_from and date_to:
        period_text += f'من {date_from.strftime("%d/%m/%Y")} إلى {date_to.strftime("%d/%m/%Y")}'
    elif date_from:
        period_text += f'من {date_from.strftime("%d/%m/%Y")}'
    elif date_to:
        period_text += f'حتى {date_to.strftime("%d/%m/%Y")}'
    else:
        period_text += 'جميع الفترات'
    
    ws.merge_cells('A2:E2')
    ws['A2'] = period_text
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # الملخص المالي
    row_num = 4
    ws.merge_cells(f'A{row_num}:E{row_num}')
    ws[f'A{row_num}'] = 'الملخص المالي'
    ws[f'A{row_num}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'A{row_num}'].fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    ws[f'A{row_num}'].alignment = Alignment(horizontal='center')
    
    row_num += 1
    summary_data = [
        ('إجمالي الإيرادات', float(report['summary']['gross_revenues'])),
        ('الاستردادات', float(report['summary']['refunds'])),
        ('صافي الإيرادات', float(report['summary']['revenues'])),
        ('المصروفات', float(report['summary']['expenses'])),
        ('الربح/الخسارة', float(report['summary']['profit'])),
        ('هامش الربح %', float(report['summary']['margin'])),
    ]
    
    for label, value in summary_data:
        ws[f'A{row_num}'] = label
        ws[f'A{row_num}'].font = Font(bold=True)
        ws[f'A{row_num}'].border = border
        ws[f'B{row_num}'] = value
        ws[f'B{row_num}'].number_format = '#,##0.00' if label != 'هامش الربح %' else '0.00'
        ws[f'B{row_num}'].border = border
        row_num += 1
    
    # قائمة الإيرادات
    row_num += 2
    ws.merge_cells(f'A{row_num}:E{row_num}')
    ws[f'A{row_num}'] = f'الإيرادات ({len(report["revenue_entries"])})'
    ws[f'A{row_num}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'A{row_num}'].fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
    ws[f'A{row_num}'].alignment = Alignment(horizontal='center')
    
    row_num += 1
    headers = ['التاريخ', 'رقم القيد', 'البيان', 'المرجع', 'المبلغ']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row_num += 1
    for entry in report['revenue_entries']:
        ws.cell(row=row_num, column=1, value=entry['date'].strftime('%d/%m/%Y')).border = border
        ws.cell(row=row_num, column=2, value=entry['number']).border = border
        ws.cell(row=row_num, column=3, value=entry['description']).border = border
        ws.cell(row=row_num, column=4, value=entry['reference']).border = border
        ws.cell(row=row_num, column=5, value=float(entry['amount'])).border = border
        ws.cell(row=row_num, column=5).number_format = '#,##0.00'
        row_num += 1
    
    # إجمالي الإيرادات
    ws.cell(row=row_num, column=4, value='الإجمالي:').font = Font(bold=True)
    ws.cell(row=row_num, column=4).border = border
    ws.cell(row=row_num, column=5, value=float(report['summary']['gross_revenues'])).font = Font(bold=True)
    ws.cell(row=row_num, column=5).number_format = '#,##0.00'
    ws.cell(row=row_num, column=5).border = border
    
    # قائمة الاستردادات (إذا وجدت)
    if report['refund_entries']:
        row_num += 2
        ws.merge_cells(f'A{row_num}:E{row_num}')
        ws[f'A{row_num}'] = f'الاستردادات ({len(report["refund_entries"])})'
        ws[f'A{row_num}'].font = Font(name='Arial', size=12, bold=True)
        ws[f'A{row_num}'].fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
        ws[f'A{row_num}'].alignment = Alignment(horizontal='center')
        
        row_num += 1
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row_num += 1
        for entry in report['refund_entries']:
            ws.cell(row=row_num, column=1, value=entry['date'].strftime('%d/%m/%Y')).border = border
            ws.cell(row=row_num, column=2, value=entry['number']).border = border
            ws.cell(row=row_num, column=3, value=entry['description']).border = border
            ws.cell(row=row_num, column=4, value=entry['reference']).border = border
            ws.cell(row=row_num, column=5, value=float(entry['amount'])).border = border
            ws.cell(row=row_num, column=5).number_format = '#,##0.00'
            row_num += 1
        
        # إجمالي الاستردادات
        ws.cell(row=row_num, column=4, value='الإجمالي:').font = Font(bold=True)
        ws.cell(row=row_num, column=4).border = border
        ws.cell(row=row_num, column=5, value=float(report['summary']['refunds'])).font = Font(bold=True)
        ws.cell(row=row_num, column=5).number_format = '#,##0.00'
        ws.cell(row=row_num, column=5).border = border
    
    # قائمة المصروفات
    row_num += 2
    ws.merge_cells(f'A{row_num}:E{row_num}')
    ws[f'A{row_num}'] = f'المصروفات ({len(report["expense_entries"])})'
    ws[f'A{row_num}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'A{row_num}'].fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
    ws[f'A{row_num}'].alignment = Alignment(horizontal='center')
    
    row_num += 1
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row_num += 1
    for entry in report['expense_entries']:
        ws.cell(row=row_num, column=1, value=entry['date'].strftime('%d/%m/%Y')).border = border
        ws.cell(row=row_num, column=2, value=entry['number']).border = border
        ws.cell(row=row_num, column=3, value=entry['description']).border = border
        ws.cell(row=row_num, column=4, value=entry['reference']).border = border
        ws.cell(row=row_num, column=5, value=float(entry['amount'])).border = border
        ws.cell(row=row_num, column=5).number_format = '#,##0.00'
        row_num += 1
    
    # إجمالي المصروفات
    ws.cell(row=row_num, column=4, value='الإجمالي:').font = Font(bold=True)
    ws.cell(row=row_num, column=4).border = border
    ws.cell(row=row_num, column=5, value=float(report['summary']['expenses'])).font = Font(bold=True)
    ws.cell(row=row_num, column=5).number_format = '#,##0.00'
    ws.cell(row=row_num, column=5).border = border
    
    # ضبط عرض الأعمدة
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    
    # إنشاء الاستجابة
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    # استخدام الكود بدل الاسم العربي
    date_str = ''
    if date_from and date_to:
        date_str = f"_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}"
    elif date_from:
        date_str = f"_from_{date_from.strftime('%Y%m%d')}"
    elif date_to:
        date_str = f"_to_{date_to.strftime('%Y%m%d')}"
    
    filename = f'profitability_{code}{date_str}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response
