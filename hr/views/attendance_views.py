"""
Views إدارة الحضور
"""
from .base_imports import *
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import permission_required
from django.db.models import Count
from decimal import Decimal, ROUND_HALF_UP
from ..models import Employee, Department, Shift, BiometricLog, AttendanceSummary, Attendance, RamadanSettings, AttendancePenalty
from ..services.attendance_service import AttendanceService
from ..services.attendance_summary_service import AttendanceSummaryService
from ..forms.attendance_forms import RamadanSettingsForm, AttendancePenaltyForm

__all__ = [
    'attendance_list',
    'attendance_export_excel',
    'attendance_check_in',
    'attendance_check_out',
    'attendance_summary_list',
    'attendance_summary_detail',  # Re-added with Smart Auto-recalculate
    'approve_attendance_summary',
    'recalculate_attendance_summary',
    'calculate_attendance_summaries',  # Moved from integrated_payroll_views
    'calculate_exempt_summaries',      # Calculate summaries for attendance-exempt employees
    'ramadan_settings_list',
    'ramadan_settings_create',
    'ramadan_settings_update',
    'ramadan_settings_delete',
    'fetch_ramadan_dates',
    'penalty_list',
    'penalty_create',
    'penalty_update',
    'penalty_delete',
    'penalty_toggle_active',
]


@login_required
def attendance_list(request):
    """قائمة الحضور - من سجلات Attendance المعالجة"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    # الفلاتر
    # التاريخ - افتراضي الشهر الحالي
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from or not date_to:
        # بداية ونهاية الدورة الحالية (تدعم الدورة المرنة)
        from hr.utils.payroll_helpers import get_payroll_period, get_payroll_month_for_date
        current_payroll_month = get_payroll_month_for_date(date.today())
        period_start, period_end, _ = get_payroll_period(current_payroll_month)
        if not date_from:
            date_from = period_start.strftime('%Y-%m-%d')
        if not date_to:
            date_to = min(period_end, date.today()).strftime('%Y-%m-%d')
    
    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # Generate missing attendance records for the selected period
    try:
        # Don't generate for future dates
        max_date = min(date_to_obj, date.today())
        if date_from_obj <= max_date:
            AttendanceService.generate_missing_attendances(date_from_obj, max_date)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error generating missing attendances: {str(e)}")
    
    # فلتر القسم
    department_id = request.GET.get('department')
    
    # فلتر الموظف
    employee_id = request.GET.get('employee')

    # فلتر الحالة
    status_filter = request.GET.get('status')

    # جلب أيام الإجازة الأسبوعية من الإعدادات
    import json as _json
    from core.models import SystemSetting
    weekly_off_days = SystemSetting.get_setting('hr_weekly_off_days', [4])
    if isinstance(weekly_off_days, str):
        try:
            weekly_off_days = _json.loads(weekly_off_days)
        except Exception:
            weekly_off_days = [4]

    # بناء قائمة التواريخ المسموح بها (بدون أيام الإجازة الأسبوعية والإجازات الرسمية)
    current = date_from_obj
    allowed_dates = []
    while current <= date_to_obj:
        if current.weekday() not in weekly_off_days:
            allowed_dates.append(current)
        current += timedelta(days=1)

    # استثناء الإجازات الرسمية
    official_holiday_dates = AttendanceService.get_official_holiday_dates(date_from_obj, date_to_obj)
    if official_holiday_dates:
        allowed_dates = [d for d in allowed_dates if d not in official_holiday_dates]

    # جلب سجلات الحضور المعالجة مع الأذونات المعتمدة
    from django.db.models import Prefetch
    from hr.models import PermissionRequest
    
    attendances = Attendance.objects.filter(
        date__in=allowed_dates,
        employee__attendance_exempt=False,  # استثناء المعفيين من البصمة
        employee__status='active'  # استثناء الموظفين المنهية خدمتهم
    ).select_related(
        'employee', 
        'employee__department', 
        'shift'
    ).prefetch_related(
        Prefetch(
            'employee__permissions',
            queryset=PermissionRequest.objects.filter(
                date__gte=date_from_obj,
                date__lte=date_to_obj,
                status='approved'
            ),
            to_attr='approved_permissions_in_period'
        )
    )
    
    # تطبيق فلتر القسم
    if department_id:
        attendances = attendances.filter(employee__department_id=department_id)
    
    # تطبيق فلتر الموظف
    if employee_id:
        attendances = attendances.filter(employee_id=employee_id)

    # تطبيق فلتر الحالة
    if status_filter:
        if status_filter == 'permission':
            # الأذونات محسوبة بعد الجلب، نفلتر بعدين
            pass
        else:
            attendances = attendances.filter(status=status_filter)
    
    attendances = attendances.order_by('-date', 'employee__name')
    
    # Pagination - 50 سجل في الصفحة
    paginator = Paginator(attendances, 50)
    page = request.GET.get('page', 1)
    
    try:
        attendances_page = paginator.page(page)
    except PageNotAnInteger:
        attendances_page = paginator.page(1)
    except EmptyPage:
        attendances_page = paginator.page(paginator.num_pages)
    
    # تحضير البيانات للعرض - فقط للصفحة الحالية
    attendance_data = []
    
    for attendance in attendances_page:
        # حساب الانصراف المبكر
        early_leave_minutes = attendance.early_leave_minutes if attendance.early_leave_minutes else 0
        
        # جلب حركات البصمة المرتبطة
        biometric_logs = attendance.biometric_logs.all().order_by('timestamp')
        movements = [log.timestamp for log in biometric_logs]
        
        # الحصول على الملخص الشهري (باستخدام الشهر الصحيح للدورة المرنة)
        from hr.utils.payroll_helpers import get_payroll_month_for_date
        payroll_month = get_payroll_month_for_date(attendance.date)
        summary, _ = AttendanceSummary.objects.get_or_create(
            employee=attendance.employee,
            month=payroll_month
        )
        
        # Calculate status manually to handle permissions
        status = attendance.status
        permission_info = None
        
        if status not in ['on_leave', 'permission']:
            # ✅ OPTIMIZED: استخدام الـ prefetched data بدلاً من query جديد
            approved_perms = getattr(attendance.employee, 'approved_permissions_in_period', [])
            matching_perm = next((p for p in approved_perms if p.date == attendance.date), None)
            
            if matching_perm:
                status = 'permission'
                permission_info = {
                    'type': matching_perm.permission_type.name_ar,
                    'is_extra': matching_perm.is_extra,
                    'duration': matching_perm.duration_hours,
                    'deduction_hours': matching_perm.deduction_hours if matching_perm.is_extra else None
                }
                
        attendance_data.append({
            'employee': attendance.employee,
            'date': attendance.date,
            'check_in': attendance.check_in,
            'check_out': attendance.check_out,
            'work_hours': float(attendance.work_hours) if attendance.work_hours else 0,
            'late_minutes': attendance.late_minutes,
            'early_leave_minutes': early_leave_minutes,
            'status': status,
            'permission_info': permission_info,
            'total_movements': len(movements),
            'movements': movements,
            'summary_id': summary.id,
            'overtime_hours': float(attendance.overtime_hours) if attendance.overtime_hours else 0,
        })
    
    # فلتر الحالة بعد البناء (لأن permission محسوبة بعد الجلب)
    if status_filter == 'permission':
        attendance_data = [a for a in attendance_data if a['status'] == 'permission']

    # حساب الإحصائيات
    present_count = sum(1 for a in attendance_data if a['status'] == 'present')
    late_count = sum(1 for a in attendance_data if a['status'] == 'late')
    absent_count = 0  # سيتم حسابه لاحقاً من قائمة الموظفين
    on_leave_count = 0  # من نظام الإجازات
    
    # جلب قوائم الفلاتر
    departments = Department.objects.filter(is_active=True).order_by('name_ar')
    employees = Employee.objects.filter(status='active', is_insurance_only=False).order_by('name')
    
    # إعداد headers للجدول الموحد
    headers = [
        {'key': 'date', 'label': 'التاريخ', 'sortable': True, 'width': '9%', 'template': 'hr/attendance/cells/date.html'},
        {'key': 'employee', 'label': 'الموظف', 'sortable': True, 'width': '13%', 'template': 'hr/attendance/cells/employee.html'},
        {'key': 'department', 'label': 'القسم', 'sortable': True, 'width': '10%', 'template': 'hr/attendance/cells/department.html'},
        {'key': 'check_in', 'label': 'الحضور', 'sortable': True, 'width': '9%', 'class': 'text-center', 'template': 'hr/attendance/cells/check_in.html'},
        {'key': 'check_out', 'label': 'الانصراف', 'sortable': True, 'width': '9%', 'class': 'text-center', 'template': 'hr/attendance/cells/check_out.html'},
        {'key': 'work_hours', 'label': 'ساعات العمل', 'sortable': True, 'width': '9%', 'class': 'text-center', 'template': 'hr/attendance/cells/work_hours.html'},
        {'key': 'late_minutes', 'label': 'التأخير', 'sortable': True, 'width': '7%', 'class': 'text-center', 'template': 'hr/attendance/cells/late_minutes.html'},
        {'key': 'early_leave_minutes', 'label': 'انصراف مبكر', 'sortable': True, 'width': '8%', 'class': 'text-center', 'template': 'hr/attendance/cells/early_leave.html'},
        {'key': 'total_movements', 'label': 'الحركات', 'sortable': True, 'width': '7%', 'class': 'text-center', 'template': 'hr/attendance/cells/movements.html'},
        {'key': 'status', 'label': 'الحالة', 'sortable': True, 'width': '8%', 'class': 'text-center', 'template': 'hr/attendance/cells/status.html'},
        {'key': 'actions', 'label': 'التفاصيل', 'sortable': False, 'width': '8%', 'class': 'text-center', 'template': 'hr/attendance/cells/actions.html'},
    ]
    
    context = {
        'attendance_data': attendance_data,
        'headers': headers,
        'date_from': date_from,
        'date_to': date_to,
        'date_from_display': date_from_obj,
        'date_to_display': date_to_obj,
        'today': date.today(),
        'present_count': present_count,
        'late_count': late_count,
        'absent_count': absent_count,
        'on_leave_count': on_leave_count,
        'departments': departments,
        'employees': employees,
        'status_filter': status_filter or '',
        'paginator': paginator,
        'page_obj': attendances_page,
        
        # بيانات الهيدر الموحد
        'page_title': 'سجل الحضور',
        'page_subtitle': f'متابعة حضور وانصراف الموظفين من نظام البصمة ({date_from_obj.strftime("%d/%m/%Y")} - {date_to_obj.strftime("%d/%m/%Y")})',
        'page_icon': 'fas fa-clock',
        'header_buttons': [
            {
                'onclick': 'processBiometricLogs()',
                'icon': 'fa-cogs',
                'text': 'معالجة البصمات',
                'class': 'btn-success',
                'id': 'btn-process-biometric',
            },
            {
                'url': reverse('hr:biometric_log_list'),
                'icon': 'fa-fingerprint',
                'text': 'سجلات البصمة الخام',
                'class': 'btn-info',
            },
        ],
        'breadcrumb_items': [
            {'title': 'الرئيسية', 'url': reverse('core:dashboard'), 'icon': 'fas fa-home'},
            {'title': 'الموارد البشرية', 'url': reverse('hr:employee_list'), 'icon': 'fas fa-users-cog'},
            {'title': 'لوحة تحكم البصمة', 'url': reverse('hr:biometric_dashboard'), 'icon': 'fas fa-fingerprint'},
            {'title': 'سجل الحضور', 'active': True},
        ],
    }
    return render(request, 'hr/attendance/list.html', context)


@login_required
def attendance_export_excel(request):
    """تصدير سجل الحضور كاملاً بدون pagination"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    import json as _json
    from core.models import SystemSetting

    # نفس منطق الفلترة في attendance_list
    date_from = request.GET.get('date_from', date.today().replace(day=1).strftime('%Y-%m-%d'))
    date_to   = request.GET.get('date_to',   date.today().strftime('%Y-%m-%d'))
    try:
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        date_to_obj   = datetime.strptime(date_to,   '%Y-%m-%d').date()
    except ValueError:
        date_from_obj = date.today().replace(day=1)
        date_to_obj   = date.today()

    department_id = request.GET.get('department')
    employee_id   = request.GET.get('employee')
    status_filter = request.GET.get('status')

    # أيام الإجازة الأسبوعية والرسمية
    weekly_off_days = SystemSetting.get_setting('hr_weekly_off_days', [4])
    if isinstance(weekly_off_days, str):
        try:
            weekly_off_days = _json.loads(weekly_off_days)
        except Exception:
            weekly_off_days = [4]

    current = date_from_obj
    allowed_dates = []
    while current <= date_to_obj:
        if current.weekday() not in weekly_off_days:
            allowed_dates.append(current)
        current += timedelta(days=1)

    official_holiday_dates = AttendanceService.get_official_holiday_dates(date_from_obj, date_to_obj)
    if official_holiday_dates:
        allowed_dates = [d for d in allowed_dates if d not in official_holiday_dates]

    # الـ queryset الكامل بدون pagination
    qs = Attendance.objects.filter(
        date__in=allowed_dates,
        employee__attendance_exempt=False,  # استثناء المعفيين من البصمة
        employee__status='active'           # استثناء الموظفين المنهية خدمتهم
    ).select_related('employee', 'employee__department', 'shift').order_by('date', 'employee__name')

    if department_id:
        qs = qs.filter(employee__department_id=department_id)
    if employee_id:
        qs = qs.filter(employee_id=employee_id)
    # فلتر الحالة - permission تُعالج بعد جلب الأذونات
    if status_filter and status_filter != 'permission':
        qs = qs.filter(status=status_filter)

    STATUS_MAP = {
        'present': '', 'absent': 'غائب', 'late': '',
        'half_day': 'نصف يوم', 'on_leave': 'في إجازة', 'permission': 'في إذن',
    }

    DAYS_AR = ['الاثنين', 'الثلاثاء', 'الأربعاء', 'الخميس', 'الجمعة', 'السبت', 'الأحد']

    # جلب الإجازات والأذونات المعتمدة في النطاق مرة واحدة
    from ..models import Leave, PermissionRequest

    # employee_id → leave_type name للأيام في النطاق
    leaves_qs = Leave.objects.filter(
        status='approved',
        start_date__lte=date_to_obj,
        end_date__gte=date_from_obj
    ).select_related('leave_type')
    # مفتاح: (employee_id, date) → اسم نوع الإجازة
    leave_map = {}
    for lv in leaves_qs:
        cur = max(lv.start_date, date_from_obj)
        end = min(lv.end_date, date_to_obj)
        while cur <= end:
            leave_map[(lv.employee_id, cur)] = lv.leave_type.name_ar
            cur += timedelta(days=1)

    # مفتاح: (employee_id, date) → "نوع الإذن (X س)"
    perms_qs = PermissionRequest.objects.filter(
        status='approved',
        date__gte=date_from_obj,
        date__lte=date_to_obj
    ).select_related('permission_type')
    perm_map = {}
    for p in perms_qs:
        key = (p.employee_id, p.date)
        hours = float(p.duration_hours)
        hours_str = f'{hours:.1f}'.rstrip('0').rstrip('.')
        entry = f'{p.permission_type.name_ar} ({hours_str} س)'
        if key in perm_map:
            perm_map[key] += f' / {entry}'
        else:
            perm_map[key] = entry

    # توقيت النظام (القاهرة) ديناميكياً من الإعدادات
    from django.utils import timezone as dj_timezone
    import pytz
    try:
        from django.conf import settings as dj_settings
        tz_name = getattr(dj_settings, 'TIME_ZONE', 'Africa/Cairo')
        local_tz = pytz.timezone(tz_name)
    except Exception:
        local_tz = pytz.timezone('Africa/Cairo')

    def to_local_time(dt):
        """تحويل datetime إلى توقيت النظام المحلي"""
        if dt is None:
            return None
        if dj_timezone.is_aware(dt):
            return dt.astimezone(local_tz)
        return local_tz.localize(dt)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'سجل الحضور'
    ws.sheet_view.rightToLeft = True

    # العنوان
    ws.merge_cells('A1:K1')
    ws['A1'] = f'سجل الحضور — {date_from_obj.strftime("%Y-%m-%d")} إلى {date_to_obj.strftime("%Y-%m-%d")}'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A1'].alignment = Alignment(horizontal='center')

    # الهيدر
    headers = ['التاريخ', 'اليوم', 'الموظف', 'الحضور', 'الانصراف', 'ساعات العمل', 'التأخير (د)', 'الانصراف المبكر (د)', 'الحالة', 'الإجازة', 'الأذونات (س)']
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # البيانات
    for row_idx, att in enumerate(qs, 3):
        check_in_local  = to_local_time(att.check_in)
        check_out_local = to_local_time(att.check_out)
        check_in_str  = check_in_local.strftime('%H:%M')  if check_in_local  else '—'
        check_out_str = check_out_local.strftime('%H:%M') if check_out_local else '—'
        leave_name  = leave_map.get((att.employee_id, att.date), '')
        perm_str    = perm_map.get((att.employee_id, att.date), '')

        # تحديد الحالة الفعلية (permission تُحسب من perm_map)
        actual_status = att.status
        if actual_status not in ('on_leave', 'permission') and perm_str:
            actual_status = 'permission'

        # فلتر الحالة لـ permission بعد الحساب
        if status_filter == 'permission' and actual_status != 'permission':
            continue

        ws.cell(row=row_idx, column=1,  value=att.date.strftime('%Y-%m-%d')).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=2,  value=DAYS_AR[att.date.weekday()]).alignment   = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=3,  value=att.employee.get_full_name_ar())
        ws.cell(row=row_idx, column=4,  value=check_in_str).alignment  = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=5,  value=check_out_str).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=6,  value=float(att.work_hours) if att.work_hours else 0).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=7,  value=att.late_minutes or 0).alignment         = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=8,  value=att.early_leave_minutes or 0).alignment  = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=9,  value=STATUS_MAP.get(actual_status, actual_status)).alignment = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=10, value=leave_name).alignment  = Alignment(horizontal='center')
        ws.cell(row=row_idx, column=11, value=perm_str).alignment    = Alignment(horizontal='center')

    # عرض الأعمدة
    for col, width in zip('ABCDEFGHIJK', [14, 12, 22, 10, 10, 14, 14, 18, 12, 18, 14]):
        ws.column_dimensions[col].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


@login_required
def attendance_check_in(request):
    """تسجيل الحضور"""
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        shift_id = request.POST.get('shift_id')
        
        if employee_id:
            try:
                employee = Employee.objects.get(pk=employee_id)
                shift = Shift.objects.get(pk=shift_id) if shift_id else None
                
                # Debug logging
                import logging
                logger = logging.getLogger(__name__)
                
                attendance = AttendanceService.record_check_in(employee, shift=shift)
                
                messages.success(request, f'تم تسجيل حضور {employee.get_full_name_ar()} بنجاح')
                return redirect('hr:attendance_list')
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Check-in failed: {str(e)}", exc_info=True)
                messages.error(request, f'خطأ: {str(e)}')
        else:
            messages.error(request, 'يرجى اختيار الموظف')
    
    context = {
        'employees': Employee.objects.filter(status='active', is_insurance_only=False),
        'shifts': Shift.objects.filter(is_active=True)
    }
    return render(request, 'hr/attendance/check_in.html', context)


@login_required
def attendance_summary_list(request):
    """قائمة ملخصات الحضور الشهرية"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Q
    from ..models import Leave
    
    # الفلاتر
    month_str = request.GET.get('month')
    department_id = request.GET.get('department')
    employee_id = request.GET.get('employee')
    status = request.GET.get('status')  # approved, pending
    exempt_type = request.GET.get('exempt_type', '')  # '' = الكل, 'exempt' = معفيين, 'non_exempt' = غير معفيين
    
    # تحديد الشهر الافتراضي (الشهر الحالي)
    if month_str:
        try:
            month_date = datetime.strptime(month_str, '%Y-%m').date()
        except ValueError:
            month_date = date.today().replace(day=1)
    else:
        month_date = date.today().replace(day=1)
    
    # جلب الملخصات - الموظفين النشطين فقط (استثناء التأمين فقط)
    summaries = AttendanceSummary.objects.filter(
        month=month_date,
        employee__status='active',
        employee__is_insurance_only=False
    ).select_related('employee', 'employee__department', 'employee__job_title')
    
    # تطبيق فلتر نوع الإعفاء
    if exempt_type == 'exempt':
        summaries = summaries.filter(employee__attendance_exempt=True)
    elif exempt_type == 'non_exempt':
        summaries = summaries.filter(employee__attendance_exempt=False)
    
    # تطبيق الفلاتر
    if department_id:
        summaries = summaries.filter(employee__department_id=department_id)
    
    if employee_id:
        summaries = summaries.filter(employee_id=employee_id)
    
    if status == 'approved':
        summaries = summaries.filter(is_approved=True)
    elif status == 'pending':
        summaries = summaries.filter(is_approved=False)
    
    summaries = summaries.order_by('-is_approved', 'employee__name')
    
    # Pagination
    paginator = Paginator(summaries, 50)
    page = request.GET.get('page', 1)
    
    try:
        summaries_page = paginator.page(page)
    except PageNotAnInteger:
        summaries_page = paginator.page(1)
    except EmptyPage:
        summaries_page = paginator.page(paginator.num_pages)
    
    # جلب الإجازات المعتمدة للشهر دفعة واحدة لتجنب N+1
    from hr.utils.payroll_helpers import get_payroll_period
    period_start, period_end, _ = get_payroll_period(month_date)
    
    employee_ids_in_page = [s.employee_id for s in summaries_page]
    approved_leaves_qs = Leave.objects.filter(
        employee_id__in=employee_ids_in_page,
        status='approved',
        start_date__lte=period_end,
        end_date__gte=period_start
    ).select_related('leave_type').order_by('start_date')
    
    # تجميع الإجازات حسب الموظف
    leaves_by_employee = {}
    for lv in approved_leaves_qs:
        leaves_by_employee.setdefault(lv.employee_id, []).append(lv)
    
    # تحضير البيانات للجدول
    summary_data = []
    for summary in summaries_page:
        # تحديد إذا كان يحتاج تحديث (للعرض فقط)
        needs_update = False
        if not summary.is_approved:
            if summary.updated_at:
                from django.utils import timezone
                age = timezone.now() - summary.updated_at
                if age > timedelta(minutes=5):
                    needs_update = True
        
        # present_days بالفعل يشمل (present + late + half_day)
        total_present = summary.present_days
        
        # استخدام net_penalizable_minutes من الملخص (بعد خصم السماح الشهري)
        adjusted_late_minutes = summary.net_penalizable_minutes
        
        # تجميع الإجازات للموظف في هذا الشهر
        emp_leaves = leaves_by_employee.get(summary.employee_id, [])
        leave_summary_parts = []
        total_leave_days = 0
        for lv in emp_leaves:
            # حساب الأيام الفعلية داخل الدورة
            lv_start = max(lv.start_date, period_start)
            lv_end = min(lv.end_date, period_end)
            days_in_period = (lv_end - lv_start).days + 1
            total_leave_days += days_in_period
            leave_summary_parts.append(f'{lv.leave_type.name_ar} ({days_in_period} أيام)')
        
        summary_data.append({
            'id': summary.id,
            'employee': summary.employee,
            'is_exempt': summary.employee.attendance_exempt,
            'department': summary.employee.department.name_ar if summary.employee.department else 'غير محدد',
            'total_working_days': summary.total_working_days,
            'present_days': total_present,
            'absent_days': summary.absent_days,
            'late_days': summary.late_days,
            'total_work_hours': float(summary.total_work_hours) if summary.total_work_hours else 0,
            'total_late_minutes': adjusted_late_minutes,
            'paid_leave_days': summary.paid_leave_days,
            'unpaid_leave_days': summary.unpaid_leave_days,
            'total_leave_days': total_leave_days,
            'leave_details': ' | '.join(leave_summary_parts) if leave_summary_parts else '',
            'is_approved': summary.is_approved,
            'approved_by': summary.approved_by,
            'approved_at': summary.approved_at,
            'needs_update': needs_update,
            'updated_at': summary.updated_at,
        })
    
    # إحصائيات عامة
    total_summaries = summaries.count()
    approved_count = summaries.filter(is_approved=True).count()
    pending_count = summaries.filter(is_approved=False).count()
    exempt_count = summaries.filter(employee__attendance_exempt=True).count()
    
    # قوائم الفلاتر
    departments = Department.objects.filter(is_active=True).order_by('name_ar')
    employees_qs = Employee.objects.filter(status='active', is_insurance_only=False).order_by('name')
    if exempt_type == 'exempt':
        employees_qs = employees_qs.filter(attendance_exempt=True)
    elif exempt_type == 'non_exempt':
        employees_qs = employees_qs.filter(attendance_exempt=False)
    
    context = {
        'summary_data': summary_data,
        'month_str': month_date.strftime('%Y-%m'),
        'month_display': month_date.strftime('%B %Y'),
        'month_date': month_date,
        'total_summaries': total_summaries,
        'approved_count': approved_count,
        'pending_count': pending_count,
        'exempt_count': exempt_count,
        'departments': departments,
        'employees': employees_qs,
        'paginator': paginator,
        'page_obj': summaries_page,
        'exempt_type': exempt_type,
        
        # بيانات الهيدر الموحد
        'page_title': 'ملخصات الحضور الشهرية',
        'page_subtitle': f'عرض وإدارة ملخصات حضور الموظفين لشهر {month_date.strftime("%B %Y")}',
        'page_icon': 'fas fa-calendar-check',
        'header_buttons': [
            {
                'url': reverse('hr:attendance_list'),
                'icon': 'fa-clock',
                'text': 'سجل الحضور اليومي',
                'class': 'btn-info',
            },
            {
                'onclick': f"calcExemptSummaries('{month_date.strftime('%Y-%m')}')",
                'icon': 'fa-calculator',
                'text': 'حساب ملخصات المعفيين',
                'class': 'btn-outline-secondary',
            },
        ],
        'breadcrumb_items': [
            {'title': 'الرئيسية', 'url': reverse('core:dashboard'), 'icon': 'fas fa-home'},
            {'title': 'الموارد البشرية', 'url': reverse('hr:employee_list'), 'icon': 'fas fa-users-cog'},
            {'title': 'لوحة تحكم البصمة', 'url': reverse('hr:biometric_dashboard'), 'icon': 'fas fa-fingerprint'},
            {'title': 'ملخصات الحضور', 'active': True},
        ],
    }
    return render(request, 'hr/attendance/summary_list.html', context)


# ============================================
# attendance_summary_detail with Smart Auto-recalculate
# ============================================

def _sync_leave_attendance_status(summary):
    """
    يتأكد إن كل أيام الإجازات المعتمدة في الشهر
    حالتها on_leave وليس absent — يُستدعى قبل أي recalculate
    """
    from hr.models import Leave, Attendance
    from hr.utils.payroll_helpers import get_payroll_period
    from datetime import timedelta

    start_date, end_date, _ = get_payroll_period(summary.month)

    approved_leaves = Leave.objects.filter(
        employee=summary.employee,
        status='approved',
        start_date__lte=end_date,
        end_date__gte=start_date
    )

    leave_dates = set()
    for lv in approved_leaves:
        cur = max(lv.start_date, start_date)
        while cur <= min(lv.end_date, end_date):
            leave_dates.add(cur)
            cur += timedelta(days=1)

    if leave_dates:
        Attendance.objects.filter(
            employee=summary.employee,
            date__in=leave_dates,
            status='absent'
        ).update(status='on_leave', notes='تم التحديث تلقائياً عند إعادة الحساب')


@login_required
def attendance_summary_detail(request, pk):
    """
    عرض تفاصيل ملخص الحضور مع Smart Auto-recalculate
    - يعيد الحساب تلقائياً لو البيانات قديمة (> 5 دقائق)
    - يستخدم locking لمنع concurrent calculations
    """
    from django.core.cache import cache
    from django.utils import timezone
    from hr.models import PermissionRequest
    from utils.helpers import arabic_date_format
    import time

    summary = get_object_or_404(AttendanceSummary, pk=pk)
    
    # Smart Auto-recalculate: Only if not approved and data is stale
    recalculated = False
    if not summary.is_approved:
        # Skip auto-recalculate if payroll already processed for this month
        payroll_locked = summary.employee.payrolls.filter(
            month=summary.month,
            status__in=['calculated', 'approved', 'paid']
        ).exists()

        should_recalculate = False

        if not payroll_locked:
            # Check if stale (older than 5 minutes)
            if summary.updated_at:
                age = timezone.now() - summary.updated_at
                if age > timedelta(minutes=5):
                    should_recalculate = True
            else:
                should_recalculate = True

            # Check if has new attendance data
            if not should_recalculate:
                from hr.utils.payroll_helpers import get_payroll_period as _gpp
                _s, _e, _ = _gpp(summary.month)
                latest_attendance = Attendance.objects.filter(
                    employee=summary.employee,
                    date__gte=_s,
                    date__lte=_e,
                ).order_by('-updated_at').first()

                if latest_attendance and latest_attendance.updated_at > summary.updated_at:
                    should_recalculate = True

            # Check if has new/changed extra permissions (is_deduction_exempt may have changed)
            if not should_recalculate:
                from hr.utils.payroll_helpers import get_payroll_period as _gpp2
                _s2, _e2, _ = _gpp2(summary.month)
                latest_extra_perm = PermissionRequest.objects.filter(
                    employee=summary.employee,
                    is_extra=True,
                    date__gte=_s2,
                    date__lte=_e2,
                ).order_by('-updated_at').first()

                if latest_extra_perm and latest_extra_perm.updated_at > summary.updated_at:
                    should_recalculate = True

        # Recalculate with lock to prevent concurrent calculations
        if should_recalculate:
            lock_key = f'summary_calc_{summary.id}'

            # Try to acquire lock (timeout 60 seconds)
            if cache.add(lock_key, 'locked', timeout=60):
                try:
                    # ✅ تأكد أولاً إن كل أيام الإجازات المعتمدة حالتها on_leave
                    _sync_leave_attendance_status(summary)
                    summary.calculate()
                    # ✅ أعد حساب LeaveSummary كمان عشان يشمل أي إجازات جديدة
                    from hr.models import LeaveSummary as _LS
                    _ls, _ = _LS.objects.get_or_create(
                        employee=summary.employee,
                        month=summary.month
                    )
                    _ls.calculate()
                    recalculated = True
                finally:
                    cache.delete(lock_key)
            else:
                # Another process is calculating, wait and refresh
                time.sleep(0.5)
                summary.refresh_from_db()

    # تحديد بداية ونهاية الدورة (تدعم الدورة المرنة)
    from hr.utils.payroll_helpers import get_payroll_period
    start_date, end_date, _ = get_payroll_period(summary.month)

    # السجلات اليومية مع الأذونات المعتمدة
    from django.db.models import Prefetch

    # استثناء أيام الإجازات الرسمية وأيام الإجازة الأسبوعية من الجدول اليومي
    official_holidays_in_month = AttendanceService.get_official_holiday_dates(start_date, end_date)

    # جلب أيام الإجازة الأسبوعية من الإعدادات
    from core.models import SystemSetting
    import json as _json
    weekly_off_days = SystemSetting.get_setting('hr_weekly_off_days', [4])
    if isinstance(weekly_off_days, str):
        weekly_off_days = _json.loads(weekly_off_days)

    # بناء قائمة تواريخ الإجازة الأسبوعية في الفترة
    weekly_off_dates = set()
    _cur = start_date
    while _cur <= end_date:
        if _cur.weekday() in weekly_off_days:
            weekly_off_dates.add(_cur)
        _cur += timedelta(days=1)

    excluded_dates = set(official_holidays_in_month) | weekly_off_dates

    daily_records_qs = summary.employee.attendances.filter(
        date__gte=start_date,
        date__lte=end_date,
    ).select_related('shift').prefetch_related(
        Prefetch(
            'permissions',
            queryset=PermissionRequest.objects.filter(status='approved').select_related('permission_type'),
            to_attr='approved_permissions'
        )
    ).order_by('date')

    if excluded_dates:
        daily_records_qs = daily_records_qs.exclude(date__in=excluded_dates)

    daily_records = daily_records_qs

    # إجازات الموظف المعتمدة في هذا الشهر
    from hr.models import Leave
    month_leaves = Leave.objects.filter(
        employee=summary.employee,
        status='approved',
        start_date__lte=end_date,
        end_date__gte=start_date,
    ).select_related('leave_type').order_by('start_date')

    # أذونات الشهر مجمعة بالتاريخ (تشمل الأذونات غير المربوطة بسجل حضور)
    month_permissions = PermissionRequest.objects.filter(
        employee=summary.employee,
        date__gte=start_date,
        date__lte=end_date,
        status='approved'
    ).select_related('permission_type')
    permissions_by_date = {}
    for perm in month_permissions:
        permissions_by_date.setdefault(perm.date, []).append(perm)

    # إجازات الشهر مجمعة بالتاريخ (لعرضها في السجل اليومي)
    from hr.models import Leave
    from datetime import timedelta as _td
    leaves_by_date = {}
    for leave in month_leaves:
        cur = max(leave.start_date, start_date)
        while cur <= min(leave.end_date, end_date):
            leaves_by_date.setdefault(cur, []).append(leave)
            cur += _td(days=1)

    # إحصائيات الأذونات
    all_permissions = PermissionRequest.objects.filter(
        employee=summary.employee,
        date__gte=start_date,
        date__lte=end_date,
        status='approved'
    )
    permissions_count = all_permissions.count()
    extra_permissions = all_permissions.filter(is_extra=True)
    extra_permissions_count = extra_permissions.count()
    # حساب ساعات الخصم: فقط الأذونات غير المعفاة
    extra_permissions_hours = sum(
        float(p.deduction_hours or 0) 
        for p in extra_permissions 
        if not p.is_deduction_exempt
    )

    # وصف دقائق التأخير الصافية
    net_min = summary.net_penalizable_minutes
    total_late = summary.total_late_minutes
    total_early = summary.total_early_leave_minutes
    if net_min == 0:
        late_minutes_description = 'لا توجد دقائق قابلة للخصم'
    else:
        monthly_grace = 0
        try:
            from core.models import SystemSetting
            monthly_grace = int(SystemSetting.get_setting('hr_monthly_grace_minutes', 0))
        except Exception:
            pass
        net_before_monthly = net_min + monthly_grace
        late_minutes_description = f'من إجمالي ({net_before_monthly} دقيقة) قبل خصم السماح'

    stats = {
        'total_working_days': summary.total_working_days,
        'present_days': summary.present_days,
        'late_days': summary.late_days,
        'half_days': summary.half_days,
        'absent_days': summary.absent_days,
        'paid_leave_days': summary.paid_leave_days,
        'unpaid_leave_days': summary.unpaid_leave_days,
        'permissions_count': permissions_count,
        'extra_permissions_subtitle': f'{extra_permissions_count} إضافي ({extra_permissions_hours:.2g} ساعة مخصومة)' if extra_permissions_count > 0 else 'لا توجد أذونات إضافية',
        'total_work_hours': summary.total_work_hours,
        'total_late_minutes': summary.total_late_minutes,
        'total_early_leave_minutes': summary.total_early_leave_minutes,
        'net_penalizable_minutes': net_min,
        'late_minutes_description': late_minutes_description,
        'present_subtitle': f'من أصل {summary.total_working_days} يوم عمل',
        'absent_subtitle': f'من أصل {summary.total_working_days} يوم عمل',
        'late_subtitle': f'إجمالي {summary.total_late_minutes} دقيقة',
    }

    # عنوان فرعي
    month_ar = arabic_date_format(summary.month)
    try:
        month_suffix = month_ar.split(' ', 1)[1]
    except Exception:
        month_suffix = summary.month.strftime('%Y-%m')
    
    # إضافة timestamp آخر تحديث للعنوان الفرعي
    subtitle = f'تقرير حضور: {month_suffix}'
    if summary.updated_at:
        subtitle += f' - آخر تحديث: {summary.updated_at.strftime("%Y-%m-%d %H:%M")}'
    
    # إعداد أزرار الهيدر
    is_hr_only = hasattr(request.user, 'role') and request.user.role and request.user.role.name == 'hr'
    header_buttons = [
        {
            'url': reverse('hr:calculate_single_payroll', args=[summary.employee.pk]) + f'?month={summary.month.strftime("%Y-%m")}',
            'icon': 'fa-calculator',
            'text': 'معالجة الراتب',
            'class': 'btn-outline-primary',
        },
    ]
    if not is_hr_only:
        header_buttons.append({
            'url': reverse('hr:payroll_list'),
            'icon': 'fa-money-bill-wave',
            'text': 'قائمة الرواتب',
            'class': 'btn-outline-secondary',
        })
    
    # إضافة زر إعادة الحساب دائماً (ما لم يكن معتمداً)
    if not summary.is_approved:
        header_buttons.insert(0, {
            'form_id': 'recalculate-form',
            'icon': 'fa-sync-alt',
            'text': 'إعادة الحساب',
            'class': 'btn-outline-warning',
        })

    # إضافة زر الاعتماد إذا لم يكن معتمداً
    if not summary.is_approved and request.user.has_perm('hr.change_attendancesummary'):
        header_buttons.insert(0, {
            'onclick': "openApproveModal()",
            'icon': 'fa-check-circle',
            'text': 'اعتماد الملخص',
            'class': 'btn-success',
        })

    # حساب تفاصيل جزاء التأخير
    penalty_details = None
    if summary.net_penalizable_minutes > 0 and summary.late_deduction_amount > 0:
        from hr.models import AttendancePenalty
        
        # البحث عن الجزاء المطبق
        penalty = AttendancePenalty.objects.filter(
            is_active=True,
            max_minutes__gte=summary.net_penalizable_minutes
        ).order_by('max_minutes').first()
        
        # fallback: النطاق المفتوح
        if not penalty:
            penalty = AttendancePenalty.objects.filter(
                is_active=True,
                max_minutes=0
            ).first()
        
        if penalty:
            penalty_details = {
                'name': penalty.name,
                'penalty_days': penalty.penalty_days,
            }

    # حساب الراتب اليومي — نفس query المستخدمة في LeaveSummary.calculate()
    daily_salary = None
    from django.db.models import Q as _Q
    contract = summary.employee.contracts.filter(
        status='active',
        start_date__lte=end_date
    ).filter(
        _Q(end_date__isnull=True) | _Q(end_date__gte=summary.month)
    ).order_by('-start_date').first()
    if contract:
        daily_salary = (contract.basic_salary / Decimal('30')).quantize(
            Decimal('0.01'),
            rounding=ROUND_HALF_UP
        )

    # حساب مبلغ خصم الأذونات الإضافية من الأذونات الفعلية (بدون الاعتماد على summary المحفوظ)
    extra_permissions_deduction_amount = Decimal('0')
    if extra_permissions_hours > 0 and contract:
        from hr.models import RamadanSettings
        non_exempt_perms = all_permissions.filter(is_extra=True, is_deduction_exempt=False).select_related('employee__shift')
        for perm in non_exempt_perms:
            is_ramadan_day = RamadanSettings.objects.filter(
                start_date__lte=perm.date, end_date__gte=perm.date
            ).exists()
            shift = perm.employee.shift
            if shift:
                if is_ramadan_day and shift.ramadan_start_time and shift.ramadan_end_time:
                    shift_hours = Decimal(str(shift.calculate_ramadan_work_hours()))
                else:
                    shift_hours = Decimal(str(shift.calculate_work_hours()))
                if shift_hours <= Decimal('0'):
                    shift_hours = Decimal('8')
            else:
                shift_hours = Decimal('8')
            day_salary = (contract.basic_salary / Decimal('30')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            day_hourly_rate = (day_salary / shift_hours).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            perm_hours = Decimal(str(perm.deduction_hours or perm.duration_hours))
            extra_permissions_deduction_amount += (perm_hours * day_hourly_rate)
        extra_permissions_deduction_amount = extra_permissions_deduction_amount.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # حساب معدل الساعة للأذونات الإضافية
    # المعدل يُحسب بناءً على الأيام الفعلية للأذونات (رمضان أو عادي)
    hourly_rate = None
    if summary.extra_permissions_hours and summary.extra_permissions_hours > 0:
        from hr.models import RamadanSettings, PermissionRequest
        
        extra_permissions = PermissionRequest.objects.filter(
            employee=summary.employee,
            status='approved',
            is_extra=True,
            is_deduction_exempt=False,  # فقط الأذونات غير المعفاة
            date__gte=start_date,
            date__lte=end_date,
        ).select_related('employee__shift')
        
        # حساب إجمالي الساعات والخصم لكل يوم
        total_hours = Decimal('0')
        total_deduction = Decimal('0')
        
        for perm in extra_permissions:
            # التحقق من رمضان لهذا اليوم المحدد
            is_ramadan_day = RamadanSettings.objects.filter(
                start_date__lte=perm.date,
                end_date__gte=perm.date
            ).exists()
            
            shift = perm.employee.shift
            if shift:
                if is_ramadan_day and shift.ramadan_start_time and shift.ramadan_end_time:
                    shift_hours = Decimal(str(shift.calculate_ramadan_work_hours()))
                else:
                    shift_hours = Decimal(str(shift.calculate_work_hours()))
                
                if shift_hours <= Decimal('0'):
                    shift_hours = Decimal('8')
            else:
                shift_hours = Decimal('8')
            
            # حساب معدل الساعة لهذا اليوم
            if contract:
                day_salary = (contract.basic_salary / Decimal('30')).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                day_hourly_rate = (day_salary / shift_hours).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                
                perm_hours = Decimal(str(perm.deduction_hours or perm.duration_hours))
                total_hours += perm_hours
                total_deduction += (perm_hours * day_hourly_rate)
        
        # حساب المتوسط المرجح
        if total_hours > 0:
            hourly_rate = (total_deduction / total_hours).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

    # حساب توزيع أيام الغياب حسب المعامل (مع استثناء الإجازات المعتمدة والإجازات الأسبوعية والرسمية)
    absence_breakdown = {}
    if summary.absent_days > 0:
        # استثناء أيام الإجازات المعتمدة
        from hr.models import Leave as _LeaveForBreakdown
        _approved_leave_dates_bd = set()
        for lv in _LeaveForBreakdown.objects.filter(
            employee=summary.employee,
            status='approved',
            start_date__lte=end_date,
            end_date__gte=start_date
        ):
            _cur = max(lv.start_date, start_date)
            while _cur <= min(lv.end_date, end_date):
                _approved_leave_dates_bd.add(_cur)
                _cur += timedelta(days=1)

        absent_records_bd = Attendance.objects.filter(
            employee=summary.employee,
            date__gte=start_date,
            date__lte=end_date,
            status='absent'
        ).exclude(
            date__in=excluded_dates  # استثناء الإجازات الأسبوعية والرسمية
        )
        if _approved_leave_dates_bd:
            absent_records_bd = absent_records_bd.exclude(date__in=_approved_leave_dates_bd)

        absent_records_bd = absent_records_bd.values('absence_multiplier').annotate(count=Count('id'))

        for record in absent_records_bd:
            multiplier = float(record['absence_multiplier'])
            count = record['count']
            absence_breakdown[multiplier] = count

    # جلب LeaveSummary للشهر لعرض خصم الإجازات غير المدفوعة
    # دايماً نعيد الحساب لو الملخص مش معتمد — عشان نضمن إن أي إجازة معتمدة تاريخها في الشهر ده تظهر
    from hr.models import LeaveSummary, Leave as _LeaveModel
    try:
        leave_summary, _ = LeaveSummary.objects.get_or_create(
            employee=summary.employee,
            month=summary.month
        )
        if not summary.is_approved:
            leave_summary.calculate()
    except Exception:
        leave_summary = None

    # حساب إجمالي الخصومات الكامل
    # نستخدم live_extra_permissions_deduction_amount (الأدق - يراعي رمضان) بدل القيمة المحفوظة في الـ DB
    leave_deduction = (leave_summary.deduction_amount if leave_summary and leave_summary.deduction_amount else Decimal('0'))
    total_deductions_full = (
        summary.absence_deduction_amount
        + summary.late_deduction_amount
        + extra_permissions_deduction_amount  # live value (Ramadan-aware)
        + leave_deduction
    )

    context = {
        'summary': summary,
        'leave_summary': leave_summary,
        'total_deductions_full': total_deductions_full,
        'employee': summary.employee,
        'daily_records': daily_records,
        'month_leaves': month_leaves,
        'leaves_by_date': leaves_by_date,
        'permissions_by_date': permissions_by_date,
        'stats': stats,
        'penalty_details': penalty_details,
        'daily_salary': daily_salary,
        'hourly_rate': hourly_rate,
        'absence_breakdown': absence_breakdown,
        'recalculated': recalculated,
        'period_start': start_date,
        'period_end': end_date,
        # ساعات ومبلغ الأذونات الإضافية محسوبة مباشرة (تستثني المعفيين من الخصم)
        'live_extra_permissions_hours': Decimal(str(extra_permissions_hours)),
        'live_extra_permissions_deduction_amount': extra_permissions_deduction_amount,
        'month_payroll': summary.employee.payrolls.filter(
            month=summary.month,
            status__in=['calculated', 'approved', 'paid']
        ).first(),
        'page_title': f'{summary.employee.get_full_name_ar()}',
        'page_subtitle': subtitle,
        'page_icon': 'fas fa-calendar-check',
        'header_buttons': header_buttons,
        'breadcrumb_items': [
            {'title': 'الرئيسية', 'url': reverse('core:dashboard'), 'icon': 'fas fa-home'},
            {'title': 'الموارد البشرية', 'url': reverse('hr:employee_list'), 'icon': 'fas fa-users-cog'},
            {'title': summary.employee.get_full_name_ar(), 'url': reverse('hr:employee_detail', args=[summary.employee.pk]), 'icon': 'fas fa-user'},
            {'title': 'لوحة تحكم البصمة', 'url': reverse('hr:biometric_dashboard'), 'icon': 'fas fa-fingerprint'},
            {'title': 'ملخصات الحضور', 'url': reverse('hr:attendance_summary_list'), 'icon': 'fas fa-calendar-check'},
            {'title': 'ملخص الحضور', 'active': True},
        ],
    }
    return render(request, 'hr/attendance/attendance_summary_detail.html', context)


@login_required
def approve_attendance_summary(request, pk):
    """اعتماد ملخص الحضور (AJAX Modal)"""
    from hr.forms import AttendanceSummaryApprovalForm
    from decimal import Decimal
    from django.http import JsonResponse
    
    summary = get_object_or_404(AttendanceSummary, pk=pk)
    
    if request.method == 'POST':
        # الاعتماد مباشرة بدون معامل غياب
        try:
            AttendanceSummaryService.approve_summary(summary, request.user)
            messages.success(request, 'تم اعتماد ملخص الحضور بنجاح')
            return JsonResponse({
                'success': True,
                'message': 'تم اعتماد ملخص الحضور بنجاح'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'حدث خطأ أثناء اعتماد الملخص: {e}'
            }, status=400)
    
    # GET request - not needed for modal
    return redirect('hr:attendance_summary_detail', pk=pk)


@login_required
@require_POST
def recalculate_attendance_summary(request, pk):
    """إعادة حساب ملخص الحضور."""
    summary = get_object_or_404(AttendanceSummary, pk=pk)
    try:
        AttendanceSummaryService.recalculate_summary(summary)
        # Warn the user that approval was reset and must be re-done before payroll
        messages.warning(
            request,
            'تمت إعادة حساب ملخص الحضور بنجاح — '
            'تم إلغاء الاعتماد السابق، يجب اعتماد الملخص مجدداً قبل حساب الراتب.'
        )
    except Exception as e:
        messages.error(request, f'حدث خطأ أثناء إعادة الحساب: {e}')
    return redirect('hr:attendance_summary_detail', pk=pk)


@login_required
@require_POST
def update_absence_multiplier(request, pk):
    """تحديث معامل الغياب ليوم معين"""
    import json
    from decimal import Decimal
    from django.http import JsonResponse
    
    try:
        attendance = get_object_or_404(Attendance, pk=pk)
        
        # التحقق من أن اليوم غياب
        if attendance.status != 'absent':
            return JsonResponse({
                'success': False,
                'message': 'لا يمكن تعديل معامل الغياب إلا لأيام الغياب'
            }, status=400)
        
        # التحقق من أن الملخص غير معتمد
        try:
            from hr.utils.payroll_helpers import get_payroll_month_for_date
            payroll_month = get_payroll_month_for_date(attendance.date)
            summary = AttendanceSummary.objects.get(
                employee=attendance.employee,
                month=payroll_month
            )
            if summary.is_approved:
                return JsonResponse({
                    'success': False,
                    'message': 'لا يمكن تعديل معامل الغياب بعد اعتماد الملخص'
                }, status=400)
        except AttendanceSummary.DoesNotExist:
            pass
        
        # قراءة البيانات
        data = json.loads(request.body)
        new_multiplier = Decimal(str(data.get('absence_multiplier', '1.0')))
        
        # التحقق من القيمة
        if new_multiplier not in [Decimal('1.0'), Decimal('2.0'), Decimal('3.0')]:
            return JsonResponse({
                'success': False,
                'message': 'قيمة المعامل غير صحيحة'
            }, status=400)
        
        # حفظ المعامل الجديد
        attendance.absence_multiplier = new_multiplier
        attendance.save(update_fields=['absence_multiplier'])
        
        # إعادة حساب الملخص إذا كان موجود
        try:
            from hr.utils.payroll_helpers import get_payroll_month_for_date
            payroll_month = get_payroll_month_for_date(attendance.date)
            summary = AttendanceSummary.objects.get(
                employee=attendance.employee,
                month=payroll_month
            )
            summary._calculate_financial_amounts()
            summary.save()
        except AttendanceSummary.DoesNotExist:
            pass
        
        return JsonResponse({
            'success': True,
            'message': 'تم تحديث معامل الغياب بنجاح'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'حدث خطأ: {str(e)}'
        }, status=500)
    return redirect('hr:attendance_summary_detail', pk=pk)


@login_required
def attendance_check_out(request):
    """تسجيل الانصراف"""
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        
        if employee_id:
            try:
                employee = Employee.objects.get(pk=employee_id)
                attendance = AttendanceService.record_check_out(employee)
                messages.success(request, f'تم تسجيل انصراف {employee.get_full_name_ar()} بنجاح')
                return redirect('hr:attendance_list')
            except Exception as e:
                messages.error(request, f'خطأ: {str(e)}')
        else:
            messages.error(request, 'يرجى اختيار الموظف')
    
    context = {
        'employees': Employee.objects.filter(status='active', is_insurance_only=False)
    }
    return render(request, 'hr/attendance/check_out.html', context)


# ============================================
# calculate_attendance_summaries (moved from integrated_payroll_views)
# ============================================

@login_required
@permission_required('hr.can_process_payroll', raise_exception=True)
@require_POST
def calculate_attendance_summaries(request):
    """حساب ملخصات الحضور لجميع الموظفين"""
    import logging
    logger = logging.getLogger(__name__)
    
    month_str = request.POST.get('month')
    if not month_str:
        month_date = date.today().replace(day=1)
    else:
        try:
            month_date = date.fromisoformat(month_str + '-01')
        except (ValueError, TypeError):
            messages.error(request, 'تاريخ غير صحيح')
            return redirect('hr:integrated_payroll_dashboard')
    
    # حساب الملخصات
    results = AttendanceSummaryService.calculate_all_summaries_for_month(month_date)
    
    messages.success(
        request,
        f'تم حساب {len(results["success"])} ملخص حضور بنجاح. '
        f'فشل {len(results["failed"])} ملخص.'
    )
    
    if results['failed']:
        for item in results['failed'][:5]:
            messages.warning(
                request,
                f'{item["employee"].get_full_name_ar()}: {item["error"]}'
            )
    
    url = reverse('hr:integrated_payroll_dashboard') + f'?month={month_date.strftime("%Y-%m")}'
    return redirect(url)


@login_required
@require_POST
def calculate_exempt_summaries(request):
    """حساب ملخصات الحضور للموظفين المعفيين من البصمة"""
    import logging
    logger = logging.getLogger(__name__)

    month_str = request.POST.get('month')
    if not month_str:
        month_date = date.today().replace(day=1)
    else:
        try:
            month_date = date.fromisoformat(month_str + '-01')
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'message': 'تاريخ غير صحيح'}, status=400)

    # جلب الموظفين المعفيين النشطين فقط
    exempt_employees = Employee.objects.filter(
        status='active',
        attendance_exempt=True,
        is_insurance_only=False
    )

    if not exempt_employees.exists():
        return JsonResponse({'success': False, 'message': 'لا يوجد موظفون معفيون من البصمة'}, status=404)

    results = AttendanceSummaryService.calculate_all_summaries_for_month(month_date, employees=exempt_employees)

    success_count = len(results['success'])
    failed_count = len(results['failed'])

    failed_names = [item['employee'].get_full_name_ar() for item in results['failed'][:5]]

    return JsonResponse({
        'success': True,
        'message': f'تم حساب {success_count} ملخص بنجاح' + (f'، فشل {failed_count}' if failed_count else ''),
        'success_count': success_count,
        'failed_count': failed_count,
        'failed_names': failed_names,
    })


# ============================================
# RamadanSettings Views
# ============================================

@login_required
def ramadan_settings_list(request):
    """قائمة إعدادات رمضان"""
    settings = RamadanSettings.objects.all()
    context = {
        'settings': settings,
        'page_title': 'إعدادات رمضان',
        'page_subtitle': 'إدارة تواريخ شهر رمضان لكل سنة هجرية',
        'page_icon': 'fas fa-moon',
        'header_buttons': [
            {
                'url': reverse('hr:ramadan_settings_create'),
                'icon': 'fa-plus',
                'text': 'إضافة سنة جديدة',
                'class': 'btn-primary',
            },
        ],
        'breadcrumb_items': [
            {'title': 'الرئيسية', 'url': reverse('core:dashboard'), 'icon': 'fas fa-home'},
            {'title': 'الموارد البشرية', 'url': reverse('hr:employee_list'), 'icon': 'fas fa-users-cog'},
            {'title': 'الإعدادات', 'url': reverse('hr:hr_settings'), 'icon': 'fas fa-cog'},
            {'title': 'إعدادات رمضان', 'active': True},
        ],
    }
    return render(request, 'hr/ramadan/list.html', context)


@login_required
def ramadan_settings_create(request):
    """إضافة إعداد رمضان جديد"""
    if request.method == 'POST':
        form = RamadanSettingsForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم إضافة إعدادات رمضان بنجاح')
            return redirect('hr:ramadan_settings_list')
    else:
        form = RamadanSettingsForm()

    context = {
        'form': form,
        'is_edit': False,
        'page_title': 'إضافة إعدادات رمضان',
        'page_icon': 'fas fa-moon',
        'header_buttons': [
            {'url': reverse('hr:ramadan_settings_list'), 'icon': 'fa-arrow-right', 'text': 'رجوع', 'class': 'btn-secondary'},
        ],
        'breadcrumb_items': [
            {'title': 'الرئيسية', 'url': reverse('core:dashboard'), 'icon': 'fas fa-home'},
            {'title': 'الإعدادات', 'url': reverse('hr:hr_settings'), 'icon': 'fas fa-cog'},
            {'title': 'إعدادات رمضان', 'url': reverse('hr:ramadan_settings_list'), 'icon': 'fas fa-moon'},
            {'title': 'إضافة', 'active': True},
        ],
    }
    return render(request, 'hr/ramadan/form.html', context)


@login_required
def ramadan_settings_update(request, pk):
    """تعديل إعداد رمضان"""
    setting = get_object_or_404(RamadanSettings, pk=pk)

    if request.method == 'POST':
        form = RamadanSettingsForm(request.POST, instance=setting)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تعديل إعدادات رمضان بنجاح')
            return redirect('hr:ramadan_settings_list')
    else:
        form = RamadanSettingsForm(instance=setting)

    context = {
        'form': form,
        'setting': setting,
        'is_edit': True,
        'page_title': f'تعديل رمضان {setting.hijri_year} هـ',
        'page_icon': 'fas fa-moon',
        'header_buttons': [
            {'url': reverse('hr:ramadan_settings_list'), 'icon': 'fa-arrow-right', 'text': 'رجوع', 'class': 'btn-secondary'},
        ],
        'breadcrumb_items': [
            {'title': 'الرئيسية', 'url': reverse('core:dashboard'), 'icon': 'fas fa-home'},
            {'title': 'الإعدادات', 'url': reverse('hr:hr_settings'), 'icon': 'fas fa-cog'},
            {'title': 'إعدادات رمضان', 'url': reverse('hr:ramadan_settings_list'), 'icon': 'fas fa-moon'},
            {'title': 'تعديل', 'active': True},
        ],
    }
    return render(request, 'hr/ramadan/form.html', context)


@login_required
@require_POST
def ramadan_settings_delete(request, pk):
    """حذف إعداد رمضان"""
    setting = get_object_or_404(RamadanSettings, pk=pk)
    year = setting.hijri_year
    setting.delete()
    messages.success(request, f'تم حذف إعدادات رمضان {year} هـ بنجاح')
    return redirect('hr:ramadan_settings_list')


@login_required
def fetch_ramadan_dates(request):
    """
    Fetch Ramadan start and end dates for a given Hijri year using AlAdhan API.
    Ramadan is always month 9 in the Hijri calendar.
    Returns JSON with start_date and end_date in YYYY-MM-DD format.
    """
    import urllib.request
    import json as json_lib

    hijri_year = request.GET.get('hijri_year', '').strip()

    if not hijri_year or not hijri_year.isdigit():
        return JsonResponse({'success': False, 'message': 'السنة الهجرية غير صحيحة'})

    hijri_year = int(hijri_year)
    if hijri_year < 1400 or hijri_year > 1600:
        return JsonResponse({'success': False, 'message': 'السنة الهجرية يجب أن تكون بين 1400 و 1600'})

    try:
        # Convert 1st of Ramadan (month 9) to Gregorian
        start_hijri = f"01-09-{hijri_year}"
        start_url = f"https://api.aladhan.com/v1/hToG/{start_hijri}"

        with urllib.request.urlopen(start_url, timeout=10) as resp:
            start_data = json_lib.loads(resp.read().decode())

        if start_data.get('code') != 200:
            return JsonResponse({'success': False, 'message': 'فشل في جلب تاريخ بداية رمضان من AlAdhan API'})

        greg_start = start_data['data']['gregorian']
        start_date = f"{greg_start['year']}-{greg_start['month']['number']:02d}-{int(greg_start['day']):02d}"

        # Convert 1st of Shawwal (month 10) to Gregorian — that's the day after Ramadan ends
        end_hijri = f"01-10-{hijri_year}"
        end_url = f"https://api.aladhan.com/v1/hToG/{end_hijri}"

        with urllib.request.urlopen(end_url, timeout=10) as resp:
            end_data = json_lib.loads(resp.read().decode())

        if end_data.get('code') != 200:
            return JsonResponse({'success': False, 'message': 'فشل في جلب تاريخ نهاية رمضان من AlAdhan API'})

        greg_end = end_data['data']['gregorian']
        # end_date is the last day of Ramadan = 1st Shawwal minus 1 day
        from datetime import date, timedelta
        end_date_obj = date(
            int(greg_end['year']),
            int(greg_end['month']['number']),
            int(greg_end['day'])
        ) - timedelta(days=1)
        end_date = end_date_obj.strftime('%Y-%m-%d')

        return JsonResponse({
            'success': True,
            'start_date': start_date,
            'end_date': end_date,
            'hijri_year': hijri_year,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': f'خطأ في الاتصال بـ AlAdhan API: {str(e)}'})


# ============================================
# AttendancePenalty Views
# ============================================

@login_required
def penalty_list(request):
    """قائمة جدول الجزاءات"""
    penalties = AttendancePenalty.objects.all()
    context = {
        'penalties': penalties,
        'page_title': 'جدول الجزاءات',
        'page_subtitle': 'إدارة نطاقات دقائق التأخير وقيمة الجزاء',
        'page_icon': 'fas fa-gavel',
        'header_buttons': [
            {
                'url': reverse('hr:penalty_create'),
                'icon': 'fa-plus',
                'text': 'إضافة نطاق جديد',
                'class': 'btn-primary',
            },
        ],
        'breadcrumb_items': [
            {'title': 'الرئيسية', 'url': reverse('core:dashboard'), 'icon': 'fas fa-home'},
            {'title': 'الموارد البشرية', 'url': reverse('hr:employee_list'), 'icon': 'fas fa-users-cog'},
            {'title': 'الإعدادات', 'url': reverse('hr:hr_settings'), 'icon': 'fas fa-cog'},
            {'title': 'جدول الجزاءات', 'active': True},
        ],
    }
    return render(request, 'hr/penalty/list.html', context)


@login_required
def penalty_create(request):
    """إضافة نطاق جزاء جديد"""
    if request.method == 'POST':
        form = AttendancePenaltyForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم إضافة نطاق الجزاء بنجاح')
            return redirect('hr:penalty_list')
    else:
        form = AttendancePenaltyForm()

    context = {
        'form': form,
        'is_edit': False,
        'page_title': 'إضافة نطاق جزاء',
        'page_icon': 'fas fa-gavel',
        'header_buttons': [
            {'url': reverse('hr:penalty_list'), 'icon': 'fa-arrow-right', 'text': 'رجوع', 'class': 'btn-secondary'},
        ],
        'breadcrumb_items': [
            {'title': 'الرئيسية', 'url': reverse('core:dashboard'), 'icon': 'fas fa-home'},
            {'title': 'الإعدادات', 'url': reverse('hr:hr_settings'), 'icon': 'fas fa-cog'},
            {'title': 'جدول الجزاءات', 'url': reverse('hr:penalty_list'), 'icon': 'fas fa-gavel'},
            {'title': 'إضافة', 'active': True},
        ],
    }
    return render(request, 'hr/penalty/form.html', context)


@login_required
def penalty_update(request, pk):
    """تعديل نطاق جزاء"""
    penalty = get_object_or_404(AttendancePenalty, pk=pk)

    if request.method == 'POST':
        form = AttendancePenaltyForm(request.POST, instance=penalty)
        if form.is_valid():
            form.save()
            messages.success(request, 'تم تعديل نطاق الجزاء بنجاح')
            return redirect('hr:penalty_list')
    else:
        form = AttendancePenaltyForm(instance=penalty)

    context = {
        'form': form,
        'penalty': penalty,
        'is_edit': True,
        'page_title': f'تعديل: {penalty.name}',
        'page_icon': 'fas fa-gavel',
        'header_buttons': [
            {'url': reverse('hr:penalty_list'), 'icon': 'fa-arrow-right', 'text': 'رجوع', 'class': 'btn-secondary'},
        ],
        'breadcrumb_items': [
            {'title': 'الرئيسية', 'url': reverse('core:dashboard'), 'icon': 'fas fa-home'},
            {'title': 'الإعدادات', 'url': reverse('hr:hr_settings'), 'icon': 'fas fa-cog'},
            {'title': 'جدول الجزاءات', 'url': reverse('hr:penalty_list'), 'icon': 'fas fa-gavel'},
            {'title': 'تعديل', 'active': True},
        ],
    }
    return render(request, 'hr/penalty/form.html', context)


@login_required
@require_POST
def penalty_delete(request, pk):
    """حذف نطاق جزاء"""
    penalty = get_object_or_404(AttendancePenalty, pk=pk)
    name = penalty.name
    penalty.delete()
    messages.success(request, f'تم حذف نطاق الجزاء "{name}" بنجاح')
    return redirect('hr:penalty_list')


@login_required
@require_POST
def penalty_toggle_active(request, pk):
    """تفعيل/تعطيل نطاق جزاء"""
    penalty = get_object_or_404(AttendancePenalty, pk=pk)
    penalty.is_active = not penalty.is_active
    penalty.save()
    status = 'تفعيل' if penalty.is_active else 'تعطيل'
    messages.success(request, f'تم {status} نطاق الجزاء "{penalty.name}" بنجاح')
    return redirect('hr:penalty_list')
