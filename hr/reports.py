"""
تقارير وحدة الموارد البشرية
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from datetime import datetime, date
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .models import Employee, Attendance, Leave, Payroll


@login_required
def reports_home(request):
    """الصفحة الرئيسية للتقارير"""
    return render(request, 'hr/reports/home.html')


@login_required
def attendance_report(request):
    """تقرير الحضور الشهري"""
    month_str = request.GET.get('month', date.today().strftime('%Y-%m'))
    department_id = request.GET.get('department')
    
    try:
        year, month = map(int, month_str.split('-'))
        month_date = date(year, month, 1)
    except:
        month_date = date.today().replace(day=1)
    
    # جلب الحضور
    attendances = Attendance.objects.filter(
        date__year=month_date.year,
        date__month=month_date.month
    ).select_related('employee', 'shift')
    
    if department_id:
        attendances = attendances.filter(employee__department_id=department_id)
    
    # حساب الإحصائيات
    stats = {
        'total_days': attendances.count(),
        'present': attendances.filter(status='present').count(),
        'late': attendances.filter(status='late').count(),
        'absent': attendances.filter(status='absent').count(),
        'total_work_hours': sum(float(a.work_hours) for a in attendances),
        'total_overtime': sum(float(a.overtime_hours) for a in attendances),
    }
    
    context = {
        'attendances': attendances,
        'stats': stats,
        'month': month_date,
        'month_str': month_str,
    }
    
    # تصدير Excel
    if request.GET.get('export') == 'excel':
        return export_attendance_excel(attendances, month_date, stats)
    
    return render(request, 'hr/reports/attendance.html', context)


@login_required
def leave_report(request):
    """تقرير الإجازات"""
    year = request.GET.get('year', date.today().year)
    department_id = request.GET.get('department')
    leave_type_id = request.GET.get('leave_type')
    
    # جلب الإجازات
    leaves = Leave.objects.filter(
        start_date__year=year
    ).select_related('employee', 'leave_type', 'approved_by')
    
    if department_id:
        leaves = leaves.filter(employee__department_id=department_id)
    
    if leave_type_id:
        leaves = leaves.filter(leave_type_id=leave_type_id)
    
    # حساب الإحصائيات
    stats = {
        'total_leaves': leaves.count(),
        'approved': leaves.filter(status='approved').count(),
        'pending': leaves.filter(status='pending').count(),
        'rejected': leaves.filter(status='rejected').count(),
        'total_days': sum(l.days_count for l in leaves),
    }
    
    context = {
        'leaves': leaves,
        'stats': stats,
        'year': year,
    }
    
    # تصدير Excel
    if request.GET.get('export') == 'excel':
        return export_leave_excel(leaves, year, stats)
    
    return render(request, 'hr/reports/leave.html', context)


@login_required
def payroll_report(request):
    """تقرير الرواتب الشهري"""
    month_str = request.GET.get('month', date.today().strftime('%Y-%m'))
    department_id = request.GET.get('department')
    
    try:
        year, month = map(int, month_str.split('-'))
        month_date = date(year, month, 1)
    except:
        month_date = date.today().replace(day=1)
    
    # جلب قسائم الرواتب
    payrolls = Payroll.objects.filter(
        month=month_date
    ).select_related('employee', 'contract', 'processed_by')
    
    if department_id:
        payrolls = payrolls.filter(employee__department_id=department_id)
    
    # حساب الإحصائيات
    stats = {
        'total_employees': payrolls.count(),
        'total_gross': sum(p.gross_salary for p in payrolls),
        'total_deductions': sum(p.total_deductions for p in payrolls),
        'total_net': sum(p.net_salary for p in payrolls),
        'approved': payrolls.filter(status='approved').count(),
        'pending': payrolls.filter(status='calculated').count(),
    }
    
    context = {
        'payrolls': payrolls,
        'stats': stats,
        'month': month_date,
        'month_str': month_str,
    }
    
    # تصدير Excel
    if request.GET.get('export') == 'excel':
        return export_payroll_excel(payrolls, month_date, stats)
    
    return render(request, 'hr/reports/payroll.html', context)


@login_required
def employee_report(request):
    """تقرير الموظفين"""
    department_id = request.GET.get('department')
    status = request.GET.get('status', 'active')
    
    # جلب الموظفين
    employees = Employee.objects.filter(
        status=status
    ).select_related('department', 'job_title')
    
    if department_id:
        employees = employees.filter(department_id=department_id)
    
    # حساب الإحصائيات
    stats = {
        'total_employees': employees.count(),
        'by_gender': {
            'male': employees.filter(gender='male').count(),
            'female': employees.filter(gender='female').count(),
        },
        'by_employment_type': {
            'full_time': employees.filter(employment_type='full_time').count(),
            'part_time': employees.filter(employment_type='part_time').count(),
            'contract': employees.filter(employment_type='contract').count(),
        },
    }
    
    context = {
        'employees': employees,
        'stats': stats,
        'status': status,
    }
    
    # تصدير Excel
    if request.GET.get('export') == 'excel':
        return export_employee_excel(employees, stats)
    
    return render(request, 'hr/reports/employee.html', context)


# دوال تصدير Excel

def export_attendance_excel(attendances, month, stats):
    """تصدير تقرير الحضور إلى Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير الحضور"
    
    # العنوان
    ws['A1'] = f"تقرير الحضور - {month.strftime('%Y-%m')}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:H1')
    
    # الإحصائيات
    ws['A3'] = "إجمالي الأيام:"
    ws['B3'] = stats['total_days']
    ws['D3'] = "الحضور:"
    ws['E3'] = stats['present']
    ws['F3'] = "التأخير:"
    ws['G3'] = stats['late']
    
    # العناوين
    headers = ['الموظف', 'التاريخ', 'الوردية', 'الحضور', 'الانصراف', 'ساعات العمل', 'التأخير', 'الحالة']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # البيانات
    for row, attendance in enumerate(attendances, 6):
        ws.cell(row=row, column=1, value=attendance.employee.get_full_name_ar())
        ws.cell(row=row, column=2, value=attendance.date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=attendance.shift.name)
        ws.cell(row=row, column=4, value=attendance.check_in.strftime('%H:%M') if attendance.check_in else '')
        ws.cell(row=row, column=5, value=attendance.check_out.strftime('%H:%M') if attendance.check_out else '')
        ws.cell(row=row, column=6, value=str(attendance.work_hours))
        ws.cell(row=row, column=7, value=f"{attendance.late_minutes} دقيقة" if attendance.late_minutes > 0 else '')
        ws.cell(row=row, column=8, value=attendance.get_status_display())
    
    # إعداد الاستجابة
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{month.strftime("%Y-%m")}.xlsx"'
    wb.save(response)
    return response


def export_leave_excel(leaves, year, stats):
    """تصدير تقرير الإجازات إلى Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير الإجازات"
    
    # العنوان
    ws['A1'] = f"تقرير الإجازات - {year}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:G1')
    
    # العناوين
    headers = ['الموظف', 'نوع الإجازة', 'من', 'إلى', 'عدد الأيام', 'الحالة', 'المعتمد']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # البيانات
    for row, leave in enumerate(leaves, 4):
        ws.cell(row=row, column=1, value=leave.employee.get_full_name_ar())
        ws.cell(row=row, column=2, value=leave.leave_type.name_ar)
        ws.cell(row=row, column=3, value=leave.start_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=4, value=leave.end_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=leave.days_count)
        ws.cell(row=row, column=6, value=leave.get_status_display())
        ws.cell(row=row, column=7, value=leave.approved_by.get_full_name() if leave.approved_by else '')
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="leave_report_{year}.xlsx"'
    wb.save(response)
    return response


def export_payroll_excel(payrolls, month, stats):
    """تصدير تقرير الرواتب إلى Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير الرواتب"
    
    # العنوان
    ws['A1'] = f"تقرير الرواتب - {month.strftime('%Y-%m')}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:H1')
    
    # العناوين
    headers = ['الموظف', 'الراتب الأساسي', 'البدلات', 'الإضافات', 'الخصومات', 'الإجمالي', 'الصافي', 'الحالة']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # البيانات
    for row, payroll in enumerate(payrolls, 4):
        ws.cell(row=row, column=1, value=payroll.employee.get_full_name_ar())
        ws.cell(row=row, column=2, value=float(payroll.basic_salary))
        ws.cell(row=row, column=3, value=float(payroll.allowances))
        ws.cell(row=row, column=4, value=float(payroll.total_additions))
        ws.cell(row=row, column=5, value=float(payroll.total_deductions))
        ws.cell(row=row, column=6, value=float(payroll.gross_salary))
        ws.cell(row=row, column=7, value=float(payroll.net_salary))
        ws.cell(row=row, column=8, value=payroll.get_status_display())
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="payroll_report_{month.strftime("%Y-%m")}.xlsx"'
    wb.save(response)
    return response


def export_employee_excel(employees, stats):
    """تصدير تقرير الموظفين إلى Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "تقرير الموظفين"
    
    # العنوان
    ws['A1'] = "تقرير الموظفين"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:H1')
    
    # العناوين
    headers = ['رقم الموظف', 'الاسم', 'القسم', 'الوظيفة', 'تاريخ التعيين', 'نوع التوظيف', 'الحالة', 'البريد']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # البيانات
    for row, employee in enumerate(employees, 4):
        ws.cell(row=row, column=1, value=employee.employee_number)
        ws.cell(row=row, column=2, value=employee.get_full_name_ar())
        ws.cell(row=row, column=3, value=employee.department.name_ar)
        ws.cell(row=row, column=4, value=employee.job_title.title_ar)
        ws.cell(row=row, column=5, value=employee.hire_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=6, value=employee.get_employment_type_display())
        ws.cell(row=row, column=7, value=employee.get_status_display())
        ws.cell(row=row, column=8, value=employee.work_email)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="employee_report.xlsx"'
    wb.save(response)
    return response
