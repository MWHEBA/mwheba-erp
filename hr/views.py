"""
Views لوحدة الموارد البشرية
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import models
from django.db.models import Count, Sum, Q
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import date, timedelta, datetime
import logging
import hmac
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Django REST Framework imports
from rest_framework.decorators import api_view, permission_classes, authentication_classes, throttle_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.throttling import AnonRateThrottle

# Local imports
from .models import Employee, Department, JobTitle, Attendance, Leave, LeaveBalance, LeaveType, Payroll, Shift, Advance, Contract, BiometricDevice, BiometricLog, BiometricSyncLog
from .forms.employee_forms import EmployeeForm, DepartmentForm
from .forms.attendance_forms import AttendanceForm
from .forms.leave_forms import LeaveRequestForm
from .forms.payroll_forms import PayrollProcessForm
from .services.attendance_service import AttendanceService
from .services.leave_service import LeaveService
from .services.payroll_service import PayrollService
from .reports import (
    reports_home, attendance_report, leave_report,
    payroll_report, employee_report
)

# Logger
logger = logging.getLogger(__name__)


@login_required
def dashboard(request):
    """Dashboard الموارد البشرية"""
    context = {
        'total_employees': Employee.objects.filter(status='active').count(),
        'total_departments': Department.objects.filter(is_active=True).count(),
        'pending_leaves': Leave.objects.filter(status='pending').count(),
        'today_attendance': Attendance.objects.filter(date=date.today()).count(),
    }
    return render(request, 'hr/dashboard.html', context)


@login_required
def employee_list(request):
    """قائمة الموظفين"""
    # جلب جميع الموظفين مع العلاقات
    employees = Employee.objects.select_related('department', 'job_title').all()
    
    # الفلترة حسب البحث
    search = request.GET.get('search', '')
    if search:
        employees = employees.filter(
            Q(first_name_ar__icontains=search) |
            Q(last_name_ar__icontains=search) |
            Q(employee_number__icontains=search)
        )
    
    # الفلترة حسب القسم
    department = request.GET.get('department', '')
    if department:
        employees = employees.filter(department_id=department)
    
    # الفلترة حسب الحالة
    status = request.GET.get('status', '')
    if status:
        employees = employees.filter(status=status)
    
    # الفلترة حسب المسمى الوظيفي
    job_title = request.GET.get('job_title', '')
    if job_title:
        employees = employees.filter(job_title_id=job_title)
    
    # جلب الأقسام والمسميات الوظيفية للفلترة
    departments = Department.objects.filter(is_active=True)
    job_titles = JobTitle.objects.filter(is_active=True)
    
    # الإحصائيات
    total_employees = Employee.objects.count()
    active_employees = Employee.objects.filter(status='active').count()
    total_departments = Department.objects.filter(is_active=True).count()
    total_job_titles = JobTitle.objects.filter(is_active=True).count()
    
    # تعريف رؤوس الجدول
    table_headers = [
        {'key': 'employee_number', 'label': 'رقم الموظف', 'sortable': True, 'class': 'text-center'},
        {'key': 'photo_display', 'label': 'الصورة', 'format': 'html', 'class': 'text-center'},
        {'key': 'full_name_display', 'label': 'الاسم', 'sortable': True, 'class': 'text-center fw-bold'},
        {'key': 'department_name', 'label': 'القسم', 'sortable': True, 'class': 'text-center'},
        {'key': 'job_title_name', 'label': 'المسمى الوظيفي', 'sortable': True, 'class': 'text-center'},
        {'key': 'hire_date', 'label': 'تاريخ التعيين', 'sortable': True, 'format': 'date', 'class': 'text-center'},
        {'key': 'years_of_service_display', 'label': 'سنوات الخدمة', 'sortable': True, 'class': 'text-center'},
        {'key': 'status_display', 'label': 'الحالة', 'sortable': True, 'format': 'html', 'class': 'text-center'},
    ]
    
    # تعريف أزرار الإجراءات
    table_actions = [
        {'url': 'hr:employee_detail', 'icon': 'fa-eye', 'label': 'عرض', 'class': 'action-view'},
        {'url': 'hr:employee_form_edit', 'icon': 'fa-edit', 'label': 'تعديل', 'class': 'action-edit'},
        {'url': 'hr:employee_delete', 'icon': 'fa-trash', 'label': 'حذف', 'class': 'action-delete'},
    ]
    
    # إضافة بيانات إضافية للعرض في الجدول
    for employee in employees:
        # عرض الصورة في عمود منفصل
        if employee.photo:
            employee.photo_display = f'<img src="{employee.photo.url}" class="rounded-circle" width="40" height="40">'
        else:
            first_letter = employee.first_name_ar[0] if employee.first_name_ar else 'م'
            employee.photo_display = f'<div class="avatar-placeholder rounded-circle bg-primary text-white d-inline-flex align-items-center justify-content-center" style="width: 40px; height: 40px; font-size: 1.1rem; font-weight: 600;">{first_letter}</div>'
        
        # عرض الاسم الكامل
        employee.full_name_display = employee.get_full_name_ar()
        
        # اسم القسم
        employee.department_name = employee.department.name_ar if employee.department else '-'
        
        # اسم المسمى الوظيفي
        employee.job_title_name = employee.job_title.title_ar if employee.job_title else '-'
        
        # سنوات الخدمة
        years = employee.years_of_service
        if years == 0:
            employee.years_of_service_display = 'أقل من سنة'
        elif years == 1:
            employee.years_of_service_display = 'سنة واحدة'
        elif years == 2:
            employee.years_of_service_display = 'سنتان'
        elif years <= 10:
            employee.years_of_service_display = f'{years} سنوات'
        else:
            employee.years_of_service_display = f'{years} سنة'
        
        # عرض الحالة
        status_badges = {
            'active': '<span class="badge bg-success">نشط</span>',
            'on_leave': '<span class="badge bg-warning text-dark">في إجازة</span>',
            'suspended': '<span class="badge bg-secondary">موقوف</span>',
            'terminated': '<span class="badge bg-danger">منتهي الخدمة</span>',
        }
        employee.status_display = status_badges.get(employee.status, '<span class="badge bg-secondary">غير محدد</span>')
    
    # معالجة التصدير
    export_format = request.GET.get('export', '')
    if export_format in ['csv', 'xlsx']:
        return export_employees(employees, export_format)
    
    context = {
        'employees': employees,
        'departments': departments,
        'job_titles': job_titles,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'total_departments': total_departments,
        'total_job_titles': total_job_titles,
        'show_stats': True,
        'table_headers': table_headers,
        'table_actions': table_actions,
    }
    return render(request, 'hr/employee/list.html', context)


def export_employees(employees, format_type):
    """تصدير قائمة الموظفين إلى CSV أو XLSX"""
    
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="employees_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        response.write('\ufeff')  # BOM for Excel UTF-8 support
        
        writer = csv.writer(response)
        # كتابة الهيدر
        writer.writerow([
            'رقم الموظف',
            'الاسم الأول',
            'اسم العائلة',
            'الرقم القومي',
            'تاريخ الميلاد',
            'الجنس',
            'الحالة الاجتماعية',
            'القسم',
            'المسمى الوظيفي',
            'تاريخ التعيين',
            'الحالة',
            'البريد الإلكتروني',
            'رقم الهاتف',
            'العنوان',
            'المدينة',
        ])
        
        # كتابة البيانات
        for emp in employees:
            status_map = {
                'active': 'نشط',
                'on_leave': 'في إجازة',
                'suspended': 'موقوف',
                'terminated': 'منتهي الخدمة',
            }
            gender_map = {'male': 'ذكر', 'female': 'أنثى'}
            marital_map = {'single': 'أعزب', 'married': 'متزوج', 'divorced': 'مطلق', 'widowed': 'أرمل'}
            
            writer.writerow([
                emp.employee_number or '',
                emp.first_name_ar or '',
                emp.last_name_ar or '',
                emp.national_id or '',
                emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else '',
                gender_map.get(emp.gender, emp.gender) if emp.gender else '',
                marital_map.get(emp.marital_status, emp.marital_status) if emp.marital_status else '',
                emp.department.name_ar if emp.department else '',
                emp.job_title.title_ar if emp.job_title else '',
                emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '',
                status_map.get(emp.status, emp.status),
                emp.work_email or '',
                emp.mobile_phone or '',
                emp.address or '',
                emp.city or '',
            ])
        
        return response
    
    elif format_type == 'xlsx':
        # إنشاء ملف Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'الموظفين'
        
        # تنسيق الرؤوس
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # كتابة الرؤوس
        headers = [
            'رقم الموظف',
            'الاسم الأول',
            'اسم العائلة',
            'الرقم القومي',
            'تاريخ الميلاد',
            'الجنس',
            'الحالة الاجتماعية',
            'القسم',
            'المسمى الوظيفي',
            'تاريخ التعيين',
            'الحالة',
            'البريد الإلكتروني',
            'رقم الهاتف',
            'العنوان',
            'المدينة',
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # كتابة البيانات
        status_map = {
            'active': 'نشط',
            'on_leave': 'في إجازة',
            'suspended': 'موقوف',
            'terminated': 'منتهي الخدمة',
        }
        
        gender_map = {'male': 'ذكر', 'female': 'أنثى'}
        marital_map = {'single': 'أعزب', 'married': 'متزوج', 'divorced': 'مطلق', 'widowed': 'أرمل'}
        
        for row_num, emp in enumerate(employees, 2):
            ws.cell(row=row_num, column=1, value=emp.employee_number or '')
            ws.cell(row=row_num, column=2, value=emp.first_name_ar or '')
            ws.cell(row=row_num, column=3, value=emp.last_name_ar or '')
            ws.cell(row=row_num, column=4, value=emp.national_id or '')
            ws.cell(row=row_num, column=5, value=emp.birth_date.strftime('%Y-%m-%d') if emp.birth_date else '')
            ws.cell(row=row_num, column=6, value=gender_map.get(emp.gender, emp.gender) if emp.gender else '')
            ws.cell(row=row_num, column=7, value=marital_map.get(emp.marital_status, emp.marital_status) if emp.marital_status else '')
            ws.cell(row=row_num, column=8, value=emp.department.name_ar if emp.department else '')
            ws.cell(row=row_num, column=9, value=emp.job_title.title_ar if emp.job_title else '')
            ws.cell(row=row_num, column=10, value=emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '')
            ws.cell(row=row_num, column=11, value=status_map.get(emp.status, emp.status))
            ws.cell(row=row_num, column=12, value=emp.work_email or '')
            ws.cell(row=row_num, column=13, value=emp.mobile_phone or '')
            ws.cell(row=row_num, column=14, value=emp.address or '')
            ws.cell(row=row_num, column=15, value=emp.city or '')
            
            # تنسيق الخلايا
            for col in range(1, 16):
                cell = ws.cell(row=row_num, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # ضبط عرض الأعمدة
        column_widths = [15, 20, 20, 18, 15, 12, 18, 20, 25, 15, 15, 30, 15, 30, 15]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # حفظ الملف
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="employees_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        
        return response


@login_required
def employee_import(request):
    """استيراد الموظفين من CSV أو Excel"""
    if request.method == 'POST':
        import_file = request.FILES.get('import_file')
        
        if not import_file:
            messages.error(request, 'الرجاء اختيار ملف للاستيراد')
            return redirect('hr:employee_list')
        
        file_extension = import_file.name.split('.')[-1].lower()
        
        try:
            if file_extension == 'csv':
                result = import_employees_from_csv(import_file, request.user)
            elif file_extension in ['xlsx', 'xls']:
                result = import_employees_from_excel(import_file, request.user)
            else:
                messages.error(request, 'نوع الملف غير مدعوم. الرجاء استخدام CSV أو Excel')
                return redirect('hr:employee_list')
            
            if result['success']:
                # رسالة النجاح الأساسية
                success_msg = f'تم استيراد {result["created"]} موظف جديد وتحديث {result["updated"]} موظف'
                
                # إضافة معلومات عن البيانات المتخطاة
                if result.get('skipped', 0) > 0:
                    success_msg += f' - تم تخطي {result["skipped"]} صف'
                
                messages.success(request, success_msg)
                
                # عرض تفاصيل الأخطاء إذا وجدت
                if result.get('errors'):
                    error_list = '<br>'.join(result['errors'])
                    messages.warning(
                        request,
                        f'تفاصيل البيانات المتخطاة:<br>{error_list}'
                    )
            else:
                messages.error(request, f'حدث خطأ: {result["error"]}')
                
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء الاستيراد: {str(e)}')
            logger.error(f'Import error: {str(e)}')
        
        return redirect('hr:employee_list')
    
    return redirect('hr:employee_list')


def import_employees_from_csv(csv_file, user):
    """استيراد الموظفين من ملف CSV"""
    import csv
    import io
    
    try:
        # قراءة الملف
        decoded_file = csv_file.read().decode('utf-8-sig')
        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)
        
        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_details = []
        
        for row_num, row in enumerate(reader, start=2):  # start=2 لأن الصف 1 هو الهيدر
            employee_number = row.get('رقم الموظف', '').strip()
            
            if not employee_number:
                skipped_count += 1
                error_details.append(f"الصف {row_num}: رقم الموظف مفقود")
                continue
            
            # تحويل الحالة من العربية للإنجليزية
            status_map_reverse = {
                'نشط': 'active',
                'في إجازة': 'on_leave',
                'موقوف': 'suspended',
                'منتهي الخدمة': 'terminated',
            }
            status = status_map_reverse.get(row.get('الحالة', '').strip(), 'active')
            
            # البحث عن القسم
            department = None
            dept_name = row.get('القسم', '').strip()
            if dept_name:
                department = Department.objects.filter(name_ar=dept_name).first()
            
            # البحث عن المسمى الوظيفي
            job_title = None
            job_name = row.get('المسمى الوظيفي', '').strip()
            if job_name:
                job_title = JobTitle.objects.filter(title_ar=job_name).first()
            
            # تحويل تاريخ التعيين
            hire_date = None
            hire_date_str = row.get('تاريخ التعيين', '').strip()
            if hire_date_str:
                try:
                    hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()
                except:
                    pass
            
            # التحقق من الحقول المطلوبة
            if not row.get('الاسم الأول', '').strip() or not row.get('اسم العائلة', '').strip():
                skipped_count += 1
                error_details.append(f"الصف {row_num}: الاسم الأول أو اسم العائلة مفقود")
                continue
            
            # تحويل تاريخ الميلاد
            birth_date = None
            birth_date_str = row.get('تاريخ الميلاد', '').strip()
            if birth_date_str:
                try:
                    birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                except:
                    pass
            
            # تحويل الجنس
            gender_map_reverse = {'ذكر': 'male', 'أنثى': 'female'}
            gender = gender_map_reverse.get(row.get('الجنس', '').strip(), '')
            
            # تحويل الحالة الاجتماعية
            marital_map_reverse = {'أعزب': 'single', 'متزوج': 'married', 'مطلق': 'divorced', 'أرمل': 'widowed'}
            marital_status = marital_map_reverse.get(row.get('الحالة الاجتماعية', '').strip(), '')
            
            # التحقق من وجود الموظف
            try:
                employee = Employee.objects.get(employee_number=employee_number)
                # تحديث الموظف الموجود
                employee.first_name_ar = row.get('الاسم الأول', '').strip()
                employee.last_name_ar = row.get('اسم العائلة', '').strip()
                employee.status = status
                
                # تحديث الرقم القومي فقط إذا كان مختلف
                new_national_id = row.get('الرقم القومي', '').strip()
                if new_national_id and new_national_id != employee.national_id:
                    # التحقق من عدم وجود الرقم القومي الجديد عند موظف آخر
                    if not Employee.objects.filter(national_id=new_national_id).exclude(id=employee.id).exists():
                        employee.national_id = new_national_id
                
                if birth_date:
                    employee.birth_date = birth_date
                if gender:
                    employee.gender = gender
                if marital_status:
                    employee.marital_status = marital_status
                if department:
                    employee.department = department
                if job_title:
                    employee.job_title = job_title
                if hire_date:
                    employee.hire_date = hire_date
                if row.get('البريد الإلكتروني', '').strip():
                    employee.work_email = row.get('البريد الإلكتروني', '').strip()
                if row.get('رقم الهاتف', '').strip():
                    employee.mobile_phone = row.get('رقم الهاتف', '').strip()
                if row.get('العنوان', '').strip():
                    employee.address = row.get('العنوان', '').strip()
                if row.get('المدينة', '').strip():
                    employee.city = row.get('المدينة', '').strip()
                
                employee.save()
                updated_count += 1
                
            except Employee.DoesNotExist:
                # إنشاء موظف جديد - التحقق من جميع الحقول المطلوبة
                national_id = row.get('الرقم القومي', '').strip()
                work_email = row.get('البريد الإلكتروني', '').strip()
                mobile_phone = row.get('رقم الهاتف', '').strip()
                address = row.get('العنوان', '').strip()
                city = row.get('المدينة', '').strip()
                
                # التحقق من جميع الحقول المطلوبة
                if not all([national_id, birth_date, gender, marital_status, 
                           department, job_title, work_email, mobile_phone, 
                           address, city]):
                    # تخطي إذا كانت أي من الحقول المطلوبة ناقصة
                    skipped_count += 1
                    missing_fields = []
                    if not national_id: missing_fields.append('الرقم القومي')
                    if not birth_date: missing_fields.append('تاريخ الميلاد')
                    if not gender: missing_fields.append('الجنس')
                    if not marital_status: missing_fields.append('الحالة الاجتماعية')
                    if not department: missing_fields.append('القسم')
                    if not job_title: missing_fields.append('المسمى الوظيفي')
                    if not work_email: missing_fields.append('البريد الإلكتروني')
                    if not mobile_phone: missing_fields.append('رقم الهاتف')
                    if not address: missing_fields.append('العنوان')
                    if not city: missing_fields.append('المدينة')
                    error_details.append(f"الصف {row_num}: حقول مفقودة ({', '.join(missing_fields)})")
                    continue
                
                # التحقق من عدم وجود الرقم القومي مسبقاً
                if Employee.objects.filter(national_id=national_id).exists():
                    # تخطي إذا كان الرقم القومي موجود عند موظف آخر
                    skipped_count += 1
                    error_details.append(f"الصف {row_num}: الرقم القومي {national_id} موجود مسبقاً")
                    continue
                
                # إنشاء الموظف الجديد
                employee = Employee.objects.create(
                    employee_number=employee_number,
                    first_name_ar=row.get('الاسم الأول', '').strip(),
                    last_name_ar=row.get('اسم العائلة', '').strip(),
                    national_id=national_id,
                    birth_date=birth_date,
                    gender=gender,
                    marital_status=marital_status,
                    department=department,
                    job_title=job_title,
                    hire_date=hire_date or timezone.now().date(),
                    status=status,
                    work_email=work_email,
                    mobile_phone=mobile_phone,
                    address=address,
                    city=city,
                    created_by=user,
                )
                created_count += 1
        
        return {
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'errors': error_details[:10]  # أول 10 أخطاء فقط
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


def import_employees_from_excel(excel_file, user):
    """استيراد الموظفين من ملف Excel"""
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        created_count = 0
        updated_count = 0
        skipped_count = 0
        error_details = []
        
        # قراءة الرؤوس من الصف الأول
        headers = {}
        for col in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col).value
            if header_value:
                headers[col] = header_value
        
        # قراءة البيانات
        for row_num in range(2, ws.max_row + 1):
            # جلب القيم حسب الرؤوس
            row_data = {}
            for col, header in headers.items():
                cell_value = ws.cell(row=row_num, column=col).value
                row_data[header] = str(cell_value).strip() if cell_value else ''
            
            employee_number = row_data.get('رقم الموظف', '').strip()
            
            if not employee_number:
                skipped_count += 1
                error_details.append(f"الصف {row_num}: رقم الموظف مفقود")
                continue
            
            # تحويل الحالة من العربية للإنجليزية
            status_map_reverse = {
                'نشط': 'active',
                'في إجازة': 'on_leave',
                'موقوف': 'suspended',
                'منتهي الخدمة': 'terminated',
            }
            status = status_map_reverse.get(row_data.get('الحالة', '').strip(), 'active')
            
            # البحث عن القسم
            department = None
            dept_name = row_data.get('القسم', '').strip()
            if dept_name:
                department = Department.objects.filter(name_ar=dept_name).first()
            
            # البحث عن المسمى الوظيفي
            job_title = None
            job_name = row_data.get('المسمى الوظيفي', '').strip()
            if job_name:
                job_title = JobTitle.objects.filter(title_ar=job_name).first()
            
            # تحويل تاريخ التعيين
            hire_date = None
            hire_date_str = row_data.get('تاريخ التعيين', '').strip()
            if hire_date_str:
                try:
                    hire_date = datetime.strptime(hire_date_str, '%Y-%m-%d').date()
                except:
                    pass
            
            # التحقق من الحقول المطلوبة
            if not row_data.get('الاسم الأول', '').strip() or not row_data.get('اسم العائلة', '').strip():
                skipped_count += 1
                error_details.append(f"الصف {row_num}: الاسم الأول أو اسم العائلة مفقود")
                continue
            
            # تحويل تاريخ الميلاد
            birth_date = None
            birth_date_str = row_data.get('تاريخ الميلاد', '').strip()
            if birth_date_str:
                try:
                    birth_date = datetime.strptime(birth_date_str, '%Y-%m-%d').date()
                except:
                    pass
            
            # تحويل الجنس
            gender_map_reverse = {'ذكر': 'male', 'أنثى': 'female'}
            gender = gender_map_reverse.get(row_data.get('الجنس', '').strip(), '')
            
            # تحويل الحالة الاجتماعية
            marital_map_reverse = {'أعزب': 'single', 'متزوج': 'married', 'مطلق': 'divorced', 'أرمل': 'widowed'}
            marital_status = marital_map_reverse.get(row_data.get('الحالة الاجتماعية', '').strip(), '')
            
            # التحقق من وجود الموظف
            try:
                employee = Employee.objects.get(employee_number=employee_number)
                # تحديث الموظف الموجود
                employee.first_name_ar = row_data.get('الاسم الأول', '').strip()
                employee.last_name_ar = row_data.get('اسم العائلة', '').strip()
                employee.status = status
                
                # تحديث الرقم القومي فقط إذا كان مختلف
                new_national_id = row_data.get('الرقم القومي', '').strip()
                if new_national_id and new_national_id != employee.national_id:
                    # التحقق من عدم وجود الرقم القومي الجديد عند موظف آخر
                    if not Employee.objects.filter(national_id=new_national_id).exclude(id=employee.id).exists():
                        employee.national_id = new_national_id
                
                if birth_date:
                    employee.birth_date = birth_date
                if gender:
                    employee.gender = gender
                if marital_status:
                    employee.marital_status = marital_status
                if department:
                    employee.department = department
                if job_title:
                    employee.job_title = job_title
                if hire_date:
                    employee.hire_date = hire_date
                if row_data.get('البريد الإلكتروني', '').strip():
                    employee.work_email = row_data.get('البريد الإلكتروني', '').strip()
                if row_data.get('رقم الهاتف', '').strip():
                    employee.mobile_phone = row_data.get('رقم الهاتف', '').strip()
                if row_data.get('العنوان', '').strip():
                    employee.address = row_data.get('العنوان', '').strip()
                if row_data.get('المدينة', '').strip():
                    employee.city = row_data.get('المدينة', '').strip()
                
                employee.save()
                updated_count += 1
                
            except Employee.DoesNotExist:
                # إنشاء موظف جديد - التحقق من جميع الحقول المطلوبة
                national_id = row_data.get('الرقم القومي', '').strip()
                work_email = row_data.get('البريد الإلكتروني', '').strip()
                mobile_phone = row_data.get('رقم الهاتف', '').strip()
                address = row_data.get('العنوان', '').strip()
                city = row_data.get('المدينة', '').strip()
                
                # التحقق من جميع الحقول المطلوبة
                if not all([national_id, birth_date, gender, marital_status, 
                           department, job_title, work_email, mobile_phone, 
                           address, city]):
                    # تخطي إذا كانت أي من الحقول المطلوبة ناقصة
                    skipped_count += 1
                    missing_fields = []
                    if not national_id: missing_fields.append('الرقم القومي')
                    if not birth_date: missing_fields.append('تاريخ الميلاد')
                    if not gender: missing_fields.append('الجنس')
                    if not marital_status: missing_fields.append('الحالة الاجتماعية')
                    if not department: missing_fields.append('القسم')
                    if not job_title: missing_fields.append('المسمى الوظيفي')
                    if not work_email: missing_fields.append('البريد الإلكتروني')
                    if not mobile_phone: missing_fields.append('رقم الهاتف')
                    if not address: missing_fields.append('العنوان')
                    if not city: missing_fields.append('المدينة')
                    error_details.append(f"الصف {row_num}: حقول مفقودة ({', '.join(missing_fields)})")
                    continue
                
                # التحقق من عدم وجود الرقم القومي مسبقاً
                if Employee.objects.filter(national_id=national_id).exists():
                    # تخطي إذا كان الرقم القومي موجود عند موظف آخر
                    skipped_count += 1
                    error_details.append(f"الصف {row_num}: الرقم القومي {national_id} موجود مسبقاً")
                    continue
                
                # إنشاء الموظف الجديد
                employee = Employee.objects.create(
                    employee_number=employee_number,
                    first_name_ar=row_data.get('الاسم الأول', '').strip(),
                    last_name_ar=row_data.get('اسم العائلة', '').strip(),
                    national_id=national_id,
                    birth_date=birth_date,
                    gender=gender,
                    marital_status=marital_status,
                    department=department,
                    job_title=job_title,
                    hire_date=hire_date or timezone.now().date(),
                    status=status,
                    work_email=work_email,
                    mobile_phone=mobile_phone,
                    address=address,
                    city=city,
                    created_by=user,
                )
                created_count += 1
        
        return {
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'skipped': skipped_count,
            'errors': error_details[:10]  # أول 10 أخطاء فقط
        }
        
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


@login_required
def employee_detail(request, pk):
    """تفاصيل الموظف"""
    from .services.user_employee_service import UserEmployeeService
    from django.db.models import Count
    from .models import Contract
    
    employee = get_object_or_404(Employee, pk=pk)
    
    # جلب المستخدمين غير المرتبطين للربط
    unlinked_users = UserEmployeeService.get_unlinked_users()
    
    # إحصائيات البصمة
    biometric_logs_count = BiometricLog.objects.filter(employee=employee).count()
    biometric_logs_last_30_days = BiometricLog.objects.filter(
        employee=employee,
        timestamp__gte=timezone.now() - timedelta(days=30)
    ).count()
    
    # آخر 5 سجلات بصمة
    recent_biometric_logs = BiometricLog.objects.filter(
        employee=employee
    ).select_related('device').order_by('-timestamp')[:5]
    
    # الحصول على ربط البصمة
    from hr.models import BiometricUserMapping
    biometric_mappings = BiometricUserMapping.objects.filter(
        employee=employee,
        is_active=True
    ).select_related('device')
    
    # جلب العقود الخاصة بالموظف
    contracts = Contract.objects.filter(
        employee=employee
    ).order_by('-start_date')
    
    # تحديد العقد النشط
    active_contract = contracts.filter(status='active').first()
    
    context = {
        'employee': employee,
        'unlinked_users': unlinked_users,
        'biometric_logs_count': biometric_logs_count,
        'biometric_logs_last_30_days': biometric_logs_last_30_days,
        'recent_biometric_logs': recent_biometric_logs,
        'biometric_mappings': biometric_mappings,
        'contracts': contracts,
        'active_contract': active_contract,
    }
    return render(request, 'hr/employee/detail.html', context)


@login_required
def employee_delete(request, pk):
    """حذف موظف"""
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.status = 'terminated'
        employee.save()
        messages.success(request, 'تم إنهاء خدمة الموظف')
        return redirect('hr:employee_list')
    
    # تجهيز البيانات للمودال الموحد
    item_fields = [
        {'label': 'الاسم', 'value': employee.get_full_name_ar()},
        {'label': 'رقم الموظف', 'value': employee.employee_number},
        {'label': 'القسم', 'value': employee.department.name_ar},
        {'label': 'الوظيفة', 'value': employee.job_title.title_ar},
    ]
    
    context = {
        'employee': employee,
        'item_fields': item_fields,
    }
    return render(request, 'hr/employee/delete_modal.html', context)


@login_required
def employee_form(request, pk=None):
    """نموذج موحد لإضافة/تعديل موظف"""
    employee = get_object_or_404(Employee, pk=pk) if pk else None
    
    if request.method == 'POST':
        form = EmployeeForm(request.POST, request.FILES, instance=employee)
        if form.is_valid():
            employee_obj = form.save(commit=False)
            
            # توليد رقم الموظف تلقائياً إذا لم يتم إدخاله (للإضافة فقط)
            if not pk and not employee_obj.employee_number:
                employee_obj.employee_number = EmployeeForm.generate_employee_number()
            
            # تعيين المستخدم الذي أنشأ السجل (للإضافة فقط)
            if not pk:
                employee_obj.created_by = request.user
            
            employee_obj.save()
            
            if pk:
                messages.success(request, 'تم تحديث بيانات الموظف بنجاح')
            else:
                messages.success(request, f'تم إضافة الموظف بنجاح - الرقم: {employee_obj.employee_number}')
            
            return redirect('hr:employee_detail', pk=employee_obj.pk)
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = EmployeeForm(instance=employee)
    
    # جلب الأقسام والمسميات الوظيفية والورديات
    departments = Department.objects.filter(is_active=True)
    job_titles = JobTitle.objects.filter(is_active=True)
    shifts = Shift.objects.filter(is_active=True)
    
    # توليد رقم الموظف المقترح (للإضافة فقط)
    next_employee_number = EmployeeForm.generate_employee_number() if not pk else None
    
    context = {
        'form': form,
        'employee': employee,
        'departments': departments,
        'job_titles': job_titles,
        'shifts': shifts,
        'next_employee_number': next_employee_number,
    }
    return render(request, 'hr/employee/form.html', context)


@login_required
def department_list(request):
    """قائمة الأقسام"""
    departments = Department.objects.all().select_related('manager', 'parent')
    context = {
        'departments': departments,
        'total_departments': departments.count(),
        'active_departments': departments.filter(is_active=True).count(),
        'total_employees': Employee.objects.filter(status='active').count(),
    }
    return render(request, 'hr/department/list.html', context)


@login_required
def department_delete(request, pk):
    """حذف قسم"""
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        dept_name = department.name_ar
        department.delete()
        messages.success(request, f'تم حذف القسم "{dept_name}" بنجاح')
        return redirect('hr:department_list')
    
    # تجهيز البيانات للمودال الموحد
    employees_count = department.employees.count()
    sub_departments_count = department.sub_departments.count()
    total_related = employees_count + sub_departments_count
    
    item_fields = [
        {'label': 'الاسم (عربي)', 'value': department.name_ar},
        {'label': 'الاسم (إنجليزي)', 'value': department.name_en or '-'},
        {'label': 'الكود', 'value': department.code},
        {'label': 'القسم الرئيسي', 'value': department.parent or '-'},
        {'label': 'المدير', 'value': department.manager.get_full_name_ar() if department.manager else '-'},
        {'label': 'الحالة', 'value': 'نشط' if department.is_active else 'غير نشط'},
    ]
    
    # بناء رسالة التحذير
    warning_parts = []
    if employees_count > 0:
        warning_parts.append(f'يوجد {employees_count} موظف مرتبط بهذا القسم')
    if sub_departments_count > 0:
        warning_parts.append(f'يوجد {sub_departments_count} قسم فرعي تابع لهذا القسم')
    
    warning_message = '<ul class="mb-0">' + ''.join([f'<li>{part}</li>' for part in warning_parts]) + '</ul>'
    
    context = {
        'department': department,
        'item_fields': item_fields,
        'total_related': total_related,
        'warning_message': warning_message,
        'show_final_warning': total_related > 0,
    }
    return render(request, 'hr/department/delete_modal.html', context)


@login_required
def department_form(request, pk=None):
    """نموذج موحد لإضافة/تعديل قسم"""
    department = get_object_or_404(Department, pk=pk) if pk else None
    
    # توليد الكود التلقائي (للإضافة فقط)
    next_code = None
    if not pk:
        last_dept = Department.objects.order_by('-id').first()
        if last_dept and last_dept.code.startswith('DEPT'):
            try:
                last_num = int(last_dept.code.replace('DEPT', ''))
                next_code = f'DEPT{str(last_num + 1).zfill(3)}'
            except:
                next_code = 'DEPT001'
        else:
            next_code = 'DEPT001'
    
    if request.method == 'POST':
        if not pk:
            # للإضافة: التأكد من عدم تكرار الكود
            code = request.POST.get('code', next_code)
            if Department.objects.filter(code=code).exists():
                last_dept = Department.objects.order_by('-id').first()
                if last_dept and last_dept.code.startswith('DEPT'):
                    try:
                        last_num = int(last_dept.code.replace('DEPT', ''))
                        code = f'DEPT{str(last_num + 1).zfill(3)}'
                    except:
                        code = f'DEPT{str(Department.objects.count() + 1).zfill(3)}'
                else:
                    code = f'DEPT{str(Department.objects.count() + 1).zfill(3)}'
            
            post_data = request.POST.copy()
            post_data['code'] = code
            form = DepartmentForm(post_data)
        else:
            # للتعديل
            form = DepartmentForm(request.POST, instance=department)
        
        if form.is_valid():
            dept = form.save()
            if pk:
                messages.success(request, 'تم تحديث القسم بنجاح')
            else:
                messages.success(request, f'تم إضافة القسم بنجاح - الكود: {dept.code}')
            return redirect('hr:department_list')
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = DepartmentForm(instance=department)
    
    # استبعاد القسم الحالي من قائمة الأقسام الرئيسية (للتعديل)
    if pk:
        departments = Department.objects.filter(is_active=True).exclude(pk=pk)
    else:
        departments = Department.objects.filter(is_active=True)
    
    context = {
        'form': form,
        'department': department,
        'next_code': next_code,
        'departments': departments,
        'employees': Employee.objects.filter(status='active'),
    }
    return render(request, 'hr/department/form.html', context)


@login_required
def attendance_list(request):
    """قائمة الحضور - محلل ذكي من BiometricLog"""
    from datetime import datetime, timedelta
    from collections import defaultdict
    from django.utils import timezone
    
    # الفلاتر
    # التاريخ - افتراضي الشهر الحالي
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if not date_from:
        # أول يوم في الشهر الحالي
        date_from = date.today().replace(day=1).strftime('%Y-%m-%d')
    if not date_to:
        # اليوم الحالي
        date_to = date.today().strftime('%Y-%m-%d')
    
    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
    
    # فلتر القسم
    department_id = request.GET.get('department')
    
    # فلتر الموظف
    employee_id = request.GET.get('employee')
    
    # جلب سجلات البصمة للفترة المحددة
    logs = BiometricLog.objects.filter(
        timestamp__date__gte=date_from_obj,
        timestamp__date__lte=date_to_obj,
        employee__isnull=False
    ).select_related('employee', 'employee__department')
    
    # تطبيق فلتر القسم
    if department_id:
        logs = logs.filter(employee__department_id=department_id)
    
    # تطبيق فلتر الموظف
    if employee_id:
        logs = logs.filter(employee_id=employee_id)
    
    logs = logs.order_by('employee', 'timestamp')
    
    # تجميع السجلات حسب الموظف واليوم
    employee_daily_logs = defaultdict(lambda: defaultdict(list))
    for log in logs:
        log_date = log.timestamp.date()
        employee_daily_logs[log.employee.id][log_date].append(log)
    
    # تحليل الحضور لكل موظف في كل يوم
    attendance_data = []
    
    for emp_id, daily_logs in employee_daily_logs.items():
        # معالجة كل يوم على حدة
        for log_date, day_logs in sorted(daily_logs.items()):
            if not day_logs:
                continue
            
            # ترتيب السجلات حسب الوقت
            day_logs.sort(key=lambda x: x.timestamp)
            
            employee = day_logs[0].employee
            total_movements = len(day_logs)
            movements = [log.timestamp for log in day_logs]
            
            # أول حركة = حضور
            check_in = movements[0]
            check_out = None
            
            # منطق ذكي للانصراف
            if total_movements == 1:
                # بصمة واحدة فقط - نعتبرها حضور بدون انصراف
                check_out = None
            elif total_movements == 2:
                # بصمتين - الثانية انصراف
                check_out = movements[1]
            else:
                # أكثر من بصمتين - نحلل
                # نبحث عن أكبر فجوة زمنية (تدل على خروج ثم عودة)
                max_gap = timedelta(0)
                potential_checkout = None
                
                for i in range(len(movements) - 1):
                    gap = movements[i + 1] - movements[i]
                    # لو الفجوة أكبر من ساعتين، نعتبرها خروج
                    if gap > timedelta(hours=2) and gap > max_gap:
                        max_gap = gap
                        potential_checkout = movements[i]
                
                # لو لقينا فجوة كبيرة، نستخدمها كانصراف
                # لو مافيش، نستخدم آخر بصمة
                if potential_checkout and max_gap > timedelta(hours=2):
                    check_out = potential_checkout
                else:
                    check_out = movements[-1]
            
            # حساب ساعات العمل
            work_hours = 0
            late_minutes = 0
            early_leave_minutes = 0
            status = 'present'
            
            if check_in and check_out:
                work_duration = check_out - check_in
                work_hours = round(work_duration.total_seconds() / 3600, 2)
                
                # الحصول على الوردية المربوطة بالموظف
                employee_shift = employee.shift if hasattr(employee, 'shift') and employee.shift else None
                
                if employee_shift:
                    # استخدام أوقات الوردية
                    work_start = datetime.combine(log_date, employee_shift.start_time)
                    work_start = timezone.make_aware(work_start)
                    
                    work_end = datetime.combine(log_date, employee_shift.end_time)
                    work_end = timezone.make_aware(work_end)
                else:
                    # افتراضي: 9 صباحاً - 5 مساءً
                    work_start = datetime.combine(log_date, datetime.strptime('09:00', '%H:%M').time())
                    work_start = timezone.make_aware(work_start)
                    
                    work_end = datetime.combine(log_date, datetime.strptime('17:00', '%H:%M').time())
                    work_end = timezone.make_aware(work_end)
                
                # حساب التأخير
                if check_in > work_start:
                    late_duration = check_in - work_start
                    late_minutes = int(late_duration.total_seconds() / 60)
                    if late_minutes > 15:
                        status = 'late'
                
                # حساب الانصراف المبكر
                if check_out < work_end:
                    early_duration = work_end - check_out
                    early_leave_minutes = int(early_duration.total_seconds() / 60)
                    if early_leave_minutes > 15:
                        if status == 'late':
                            status = 'late'  # يبقى متأخر
                        else:
                            status = 'early_leave'
            
            attendance_data.append({
                'employee': employee,
                'date': log_date,
                'check_in': check_in,
                'check_out': check_out,
                'work_hours': work_hours,
                'late_minutes': late_minutes,
                'early_leave_minutes': early_leave_minutes,
                'status': status,
                'total_movements': total_movements,
                'movements': movements
            })
    
    # ترتيب حسب التاريخ ثم اسم الموظف
    attendance_data.sort(key=lambda x: (x['date'], x['employee'].get_full_name_ar()), reverse=True)
    
    # حساب الإحصائيات
    present_count = sum(1 for a in attendance_data if a['status'] == 'present')
    late_count = sum(1 for a in attendance_data if a['status'] == 'late')
    absent_count = 0  # سيتم حسابه لاحقاً من قائمة الموظفين
    on_leave_count = 0  # من نظام الإجازات
    
    # جلب قوائم الفلاتر
    from hr.models import Department
    departments = Department.objects.filter(is_active=True).order_by('name_ar')
    employees = Employee.objects.filter(status='active').order_by('first_name_ar')
    
    # إعداد headers للجدول الموحد
    headers = [
        {'key': 'date', 'label': 'التاريخ', 'sortable': True, 'width': '9%', 'template': 'hr/attendance/cells/date.html'},
        {'key': 'employee', 'label': 'الموظف', 'sortable': True, 'width': '13%', 'template': 'hr/attendance/cells/employee.html'},
        {'key': 'department', 'label': 'القسم', 'sortable': True, 'width': '10%', 'template': 'hr/attendance/cells/department.html'},
        {'key': 'check_in', 'label': 'الحضور', 'sortable': True, 'width': '9%', 'class': 'text-center', 'template': 'hr/attendance/cells/check_in.html'},
        {'key': 'check_out', 'label': 'الانصراف', 'sortable': True, 'width': '9%', 'class': 'text-center', 'template': 'hr/attendance/cells/check_out.html'},
        {'key': 'work_hours', 'label': 'ساعات العمل', 'sortable': True, 'width': '9%', 'class': 'text-center', 'template': 'hr/attendance/cells/work_hours.html'},
        {'key': 'late_minutes', 'label': 'التأخير', 'sortable': True, 'width': '7%', 'class': 'text-center', 'template': 'hr/attendance/cells/late_minutes.html'},
        {'key': 'early_leave_minutes', 'label': 'انصراف مبكر', 'sortable': True, 'width': '8%', 'class': 'text-center', 'template': 'hr/attendance/cells/early_leave.html'},
        {'key': 'total_movements', 'label': 'الحركات', 'sortable': True, 'width': '7%', 'class': 'text-center', 'template': 'hr/attendance/cells/movements.html'},
        {'key': 'status', 'label': 'الحالة', 'sortable': True, 'width': '8%', 'class': 'text-center', 'template': 'hr/attendance/cells/status.html'},
        {'key': 'actions', 'label': 'التفاصيل', 'sortable': False, 'width': '8%', 'class': 'text-center', 'template': 'hr/attendance/cells/actions.html'},
    ]
    
    context = {
        'attendance_data': attendance_data,
        'headers': headers,
        'date_from': date_from,
        'date_to': date_to,
        'date_from_display': date_from_obj,
        'date_to_display': date_to_obj,
        'today': date.today(),
        'present_count': present_count,
        'late_count': late_count,
        'absent_count': absent_count,
        'on_leave_count': on_leave_count,
        'departments': departments,
        'employees': employees,
    }
    return render(request, 'hr/attendance/list.html', context)


@login_required
def attendance_check_in(request):
    """تسجيل الحضور"""
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        shift_id = request.POST.get('shift_id')
        
        if employee_id:
            try:
                employee = Employee.objects.get(pk=employee_id)
                shift = Shift.objects.get(pk=shift_id) if shift_id else None
                
                attendance = AttendanceService.check_in(employee, shift)
                messages.success(request, f'تم تسجيل حضور {employee.get_full_name_ar()} بنجاح')
                return redirect('hr:attendance_list')
            except Exception as e:
                messages.error(request, f'خطأ: {str(e)}')
        else:
            messages.error(request, 'يرجى اختيار الموظف')
    
    context = {
        'employees': Employee.objects.filter(status='active'),
        'shifts': Shift.objects.filter(is_active=True)
    }
    return render(request, 'hr/attendance/check_in.html', context)


@login_required
def attendance_check_out(request):
    """تسجيل الانصراف"""
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        
        if employee_id:
            try:
                employee = Employee.objects.get(pk=employee_id)
                attendance = AttendanceService.check_out(employee)
                messages.success(request, f'تم تسجيل انصراف {employee.get_full_name_ar()} بنجاح')
                return redirect('hr:attendance_list')
            except Exception as e:
                messages.error(request, f'خطأ: {str(e)}')
        else:
            messages.error(request, 'يرجى اختيار الموظف')
    
    context = {
        'employees': Employee.objects.filter(status='active')
    }
    return render(request, 'hr/attendance/check_out.html', context)


# ==================== الورديات ====================

@login_required
def shift_list(request):
    """قائمة الورديات"""
    shifts = Shift.objects.all().order_by('start_time')
    
    # إحصائيات
    stats = {
        'total_shifts': shifts.count(),
        'active_shifts': shifts.filter(is_active=True).count(),
        'morning_shifts': shifts.filter(shift_type='morning').count(),
        'evening_shifts': shifts.filter(shift_type='evening').count(),
        'night_shifts': shifts.filter(shift_type='night').count(),
    }
    
    context = {
        'shifts': shifts,
        'stats': stats,
    }
    return render(request, 'hr/shift/list.html', context)


@login_required
def shift_delete(request, pk):
    """حذف وردية"""
    shift = get_object_or_404(Shift, pk=pk)
    
    if request.method == 'POST':
        shift.is_active = False
        shift.save()
        messages.success(request, 'تم إلغاء تفعيل الوردية')
        return redirect('hr:shift_list')
    
    # عدد الموظفين المرتبطين
    employees_count = Attendance.objects.filter(shift=shift).values('employee').distinct().count()
    
    # تجهيز البيانات للمودال الموحد
    item_fields = [
        {'label': 'اسم الوردية', 'value': shift.name},
        {'label': 'النوع', 'value': shift.get_shift_type_display()},
        {'label': 'الوقت', 'value': f"{shift.start_time.strftime('%H:%M')} - {shift.end_time.strftime('%H:%M')}"},
        {'label': 'عدد الموظفين المرتبطين', 'value': employees_count},
    ]
    
    warning_message = f'تحذير: يوجد {employees_count} موظف مرتبط بهذه الوردية. سيتم إلغاء تفعيل الوردية فقط وليس حذفها نهائياً.'
    
    context = {
        'shift': shift,
        'item_fields': item_fields,
        'employees_count': employees_count,
        'warning_message': warning_message,
    }
    return render(request, 'hr/shift/delete_modal.html', context)


@login_required
def shift_assign_employees(request, pk):
    """تعيين موظفين للوردية"""
    shift = get_object_or_404(Shift, pk=pk)
    
    if request.method == 'POST':
        employee_ids = request.POST.getlist('employees')
        # هنا يمكن إضافة جدول لربط الموظفين بالورديات
        messages.success(request, f'تم تعيين {len(employee_ids)} موظف للوردية')
        return redirect('hr:shift_list')
    
    employees = Employee.objects.filter(status='active')
    context = {
        'shift': shift,
        'employees': employees,
    }
    return render(request, 'hr/shift/assign.html', context)


@login_required
def shift_form(request, pk=None):
    """نموذج موحد لإضافة/تعديل وردية"""
    from .forms.attendance_forms import ShiftForm
    
    shift = get_object_or_404(Shift, pk=pk) if pk else None
    
    if request.method == 'POST':
        form = ShiftForm(request.POST, instance=shift)
        if form.is_valid():
            form.save()
            if pk:
                messages.success(request, 'تم تحديث الوردية بنجاح')
            else:
                messages.success(request, 'تم إضافة الوردية بنجاح')
            return redirect('hr:shift_list')
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = ShiftForm(instance=shift)
    
    return render(request, 'hr/shift/form.html', {'form': form, 'shift': shift})


# ==================== ماكينات البصمة ====================

@login_required
def biometric_device_list(request):
    """قائمة ماكينات البصمة - جدول موحد"""
    devices = BiometricDevice.objects.select_related('department', 'created_by').all()
    
    # إحصائيات
    stats = {
        'total_devices': devices.count(),
        'active_devices': devices.filter(is_active=True, status='active').count(),
        'offline_devices': devices.filter(status='error').count(),
        'total_users': devices.aggregate(Sum('total_users'))['total_users__sum'] or 0,
        'total_records': devices.aggregate(Sum('total_records'))['total_records__sum'] or 0,
    }
    
    # Headers للجدول الموحد
    headers = [
        {'key': 'device_name', 'label': 'اسم الماكينة', 'width': '20%', 'template': 'hr/biometric/cells/device_name.html'},
        {'key': 'device_type', 'label': 'النوع', 'width': '10%', 'template': 'hr/biometric/cells/device_type.html'},
        {'key': 'location', 'label': 'الموقع', 'width': '15%', 'template': 'hr/biometric/cells/location.html'},
        {'key': 'ip_address', 'label': 'IP Address', 'width': '12%', 'template': 'hr/biometric/cells/ip_address.html'},
        {'key': 'status', 'label': 'الحالة', 'width': '12%', 'class': 'text-center', 'template': 'hr/biometric/cells/status.html'},
        {'key': 'last_connection', 'label': 'آخر اتصال', 'width': '13%', 'template': 'hr/biometric/cells/last_connection.html'},
        {'key': 'total_users', 'label': 'المستخدمين', 'width': '8%', 'template': 'hr/biometric/cells/total_users.html'},
    ]
    
    # أزرار الإجراءات
    action_buttons = [
        {
            'url': 'hr:biometric_device_detail',
            'icon': 'fa-eye',
            'class': 'btn-info',
            'label': 'عرض'
        },
        {
            'url': 'hr:biometric_device_form_edit',
            'icon': 'fa-edit',
            'class': 'btn-warning',
            'label': 'تعديل'
        }
    ]
    
    context = {
        'devices': devices,
        'stats': stats,
        'headers': headers,
        'action_buttons': action_buttons,
    }
    return render(request, 'hr/biometric/device_list.html', context)

@login_required
def biometric_device_detail(request, pk):
    """تفاصيل ماكينة البصمة - محسّن"""
    device = get_object_or_404(BiometricDevice, pk=pk)
    
    # العدد الافتراضي للسجلات
    default_logs_count = 10
    
    # آخر السجلات
    recent_logs = BiometricLog.objects.filter(device=device).select_related('employee')[:default_logs_count]
    
    # آخر عمليات المزامنة
    sync_logs = BiometricSyncLog.objects.filter(device=device).order_by('-started_at')[:10]
    
    # Headers للجدول الموحد - سجلات البصمة
    log_headers = [
        {'key': 'timestamp', 'label': 'الوقت', 'format': 'datetime_12h', 'width': '20%'},
        {'key': 'user_id', 'label': 'معرف المستخدم', 'width': '15%'},
        {'key': 'employee', 'label': 'الموظف', 'width': '25%', 'template': 'hr/biometric/cells/log_employee.html'},
        {'key': 'log_type', 'label': 'النوع', 'width': '20%', 'template': 'hr/biometric/cells/log_type.html'},
        {'key': 'is_processed', 'label': 'حالة المعالجة', 'width': '20%', 'template': 'hr/biometric/cells/log_status.html'},
    ]
    
    context = {
        'device': device,
        'recent_logs': recent_logs,
        'sync_logs': sync_logs,
        'log_headers': log_headers,
        'logs_count': default_logs_count,  # إضافة العدد للـ context
    }
    return render(request, 'hr/biometric/device_detail.html', context)


@login_required
def biometric_device_logs_ajax(request, pk):
    """جلب سجلات البصمة ديناميكياً حسب العدد المطلوب"""
    device = get_object_or_404(BiometricDevice, pk=pk)
    
    # الحصول على العدد المطلوب من الـ query parameter
    limit = int(request.GET.get('limit', 10))
    
    # جلب السجلات
    recent_logs = BiometricLog.objects.filter(device=device).select_related('employee')[:limit]
    
    # Headers للجدول
    log_headers = [
        {'key': 'timestamp', 'label': 'الوقت', 'format': 'datetime_12h', 'width': '20%'},
        {'key': 'user_id', 'label': 'معرف المستخدم', 'width': '15%'},
        {'key': 'employee', 'label': 'الموظف', 'width': '25%', 'template': 'hr/biometric/cells/log_employee.html'},
        {'key': 'log_type', 'label': 'النوع', 'width': '20%', 'template': 'hr/biometric/cells/log_type.html'},
        {'key': 'is_processed', 'label': 'حالة المعالجة', 'width': '20%', 'template': 'hr/biometric/cells/log_status.html'},
    ]
    
    context = {
        'data': recent_logs,
        'headers': log_headers,
        'empty_message': 'لا توجد سجلات بصمة',
        'empty_icon': 'fingerprint',
        'primary_key': 'id',
        'table_id': 'biometric-logs-table',
        'length_options': '5,10,20,50',
        'default_length': str(limit),  # القيمة المختارة حالياً
    }
    
    return render(request, 'components/data_table.html', context)


@login_required
def biometric_device_form(request, pk=None):
    """نموذج موحد لإضافة/تعديل ماكينة بصمة"""
    from .forms.biometric_forms import BiometricDeviceForm
    
    device = get_object_or_404(BiometricDevice, pk=pk) if pk else None
    
    if request.method == 'POST':
        form = BiometricDeviceForm(request.POST, instance=device)
        if form.is_valid():
            device_obj = form.save(commit=False)
            if not pk:
                device_obj.created_by = request.user
            device_obj.save()
            
            if pk:
                messages.success(request, 'تم تحديث الماكينة بنجاح')
                return redirect('hr:biometric_device_detail', pk=pk)
            else:
                messages.success(request, f'تم إضافة الماكينة بنجاح - الكود: {device_obj.device_code}')
                return redirect('hr:biometric_device_detail', pk=device_obj.pk)
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = BiometricDeviceForm(instance=device)
    
    return render(request, 'hr/biometric/device_form.html', {'form': form, 'device': device})


@login_required
def biometric_device_delete(request, pk):
    """حذف ماكينة البصمة"""
    device = get_object_or_404(BiometricDevice, pk=pk)
    
    if request.method == 'POST':
        device.is_active = False
        device.status = 'inactive'
        device.save()
        messages.success(request, 'تم إلغاء تفعيل الماكينة')
        return redirect('hr:biometric_device_list')
    
    logs_count = BiometricLog.objects.filter(device=device).count()
    
    # تجهيز البيانات للمودال الموحد
    item_fields = [
        {'label': 'اسم الماكينة', 'value': device.device_name},
        {'label': 'الموقع', 'value': device.location},
        {'label': 'IP Address', 'value': device.connection_string},
        {'label': 'عدد السجلات', 'value': logs_count},
    ]
    
    warning_message = 'سيتم إلغاء تفعيل الماكينة فقط. السجلات ستبقى محفوظة.' if logs_count > 0 else ''
    
    context = {
        'device': device,
        'item_fields': item_fields,
        'logs_count': logs_count,
        'warning_message': warning_message,
    }
    return render(request, 'hr/biometric/device_delete_modal.html', context)




@login_required
def biometric_device_test(request, pk):
    """اختبار الاتصال بالماكينة - AJAX Modal"""
    from .bridge_agent_utils import test_device_connection
    
    device = get_object_or_404(BiometricDevice, pk=pk)
    
    if request.method == 'GET':
        # عرض المودال
        context = {'device': device}
        return render(request, 'hr/biometric/device_test_modal.html', context)
    
    # POST - تنفيذ الاختبار
    if request.method == 'POST':
        success, message, details = test_device_connection(device)
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            response_data = {
                'success': success,
                'message': message
            }
            if details:
                response_data['details'] = details
            return JsonResponse(response_data)
        
        if success:
            messages.success(request, message)
        else:
            messages.error(request, message)
    
    return redirect('hr:biometric_device_detail', pk=pk)


@login_required
def biometric_agent_setup(request, pk):
    """صفحة إرشادات إعداد Bridge Agent"""
    device = get_object_or_404(BiometricDevice, pk=pk)
    
    context = {
        'device': device,
        'page_title': 'إعداد Bridge Agent',
    }
    
    return render(request, 'hr/biometric_agent_setup.html', context)


@login_required
def biometric_device_download_agent(request, pk):
    """تحميل ملف إعدادات Bridge Agent"""
    from .bridge_agent_utils import generate_agent_config, create_agent_zip_package, create_download_response
    
    device = get_object_or_404(BiometricDevice, pk=pk)
    
    # توليد الإعدادات
    config, agent_secret = generate_agent_config(device, request)
    
    # إنشاء ملف ZIP
    zip_buffer = create_agent_zip_package(device, config)
    
    # عرض رسالة للمستخدم
    messages.success(
        request,
        f'تم تحميل Bridge Agent. لا تنسى إضافة المفتاح السري في settings.py:\n'
        f'BRIDGE_AGENTS = {{\'{device.device_code}\': \'{agent_secret}\'}}'
    )
    
    # إرجاع الملف
    return create_download_response(device, zip_buffer)


@login_required
def biometric_log_list(request):
    """قائمة سجلات البصمات"""
    logs = BiometricLog.objects.select_related(
        'device', 'employee', 'attendance'
    )
    
    # فلترة
    device_id = request.GET.get('device')
    if device_id:
        logs = logs.filter(device_id=device_id)
    
    employee_id = request.GET.get('employee')
    if employee_id:
        logs = logs.filter(employee_id=employee_id)
    
    # فلتر حالة الربط (مربوط/غير مربوط)
    linked = request.GET.get('linked')
    if linked == 'true':
        logs = logs.filter(employee__isnull=False)
    elif linked == 'false':
        logs = logs.filter(employee__isnull=True)
    
    # الترتيب (بدون limit - DataTables هيعمل pagination)
    logs = logs.order_by('-timestamp')
    
    devices = BiometricDevice.objects.filter(is_active=True)
    
    # إعداد headers للجدول الموحد
    headers = [
        {'key': 'timestamp', 'label': 'التاريخ والوقت', 'sortable': True, 'width': '15%', 'template': 'components/cells/biometric_log_timestamp.html'},
        {'key': 'device', 'label': 'الماكينة', 'sortable': True, 'width': '15%', 'template': 'components/cells/biometric_log_device.html'},
        {'key': 'user_id', 'label': 'معرف المستخدم', 'sortable': False, 'width': '12%', 'class': 'text-center', 'template': 'components/cells/biometric_log_user_id.html'},
        {'key': 'employee', 'label': 'الموظف', 'sortable': True, 'width': '20%', 'class': 'text-center', 'template': 'components/cells/biometric_log_employee.html'},
        {'key': 'log_type', 'label': 'النوع', 'sortable': True, 'width': '12%', 'class': 'text-center', 'template': 'components/cells/biometric_log_punch_type.html'},
        {'key': 'employee', 'label': 'حالة الربط', 'sortable': True, 'width': '12%', 'class': 'text-center', 'template': 'components/cells/biometric_log_is_processed.html'},
        {'key': 'actions', 'label': 'الإجراءات', 'sortable': False, 'width': '14%', 'class': 'text-center', 'template': 'components/cells/biometric_log_actions.html'},
    ]
    
    context = {
        'logs': logs,
        'devices': devices,
        'headers': headers,
    }
    return render(request, 'hr/biometric/log_list.html', context)


# ==================== Bridge Agent API ====================

class BridgeSyncThrottle(AnonRateThrottle):
    """Rate limiting for Bridge Agent API"""
    rate = '100/hour'  # 100 requests per hour

@csrf_exempt
@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
@throttle_classes([BridgeSyncThrottle])
def biometric_bridge_sync(request):
    """
    API لاستقبال البيانات من Bridge Agent
    """
    # Logging بدلاً من print
    logger.info("=" * 60)
    logger.info("Bridge Sync Request Received!")
    logger.info(f"Headers: {dict(request.headers)}")
    logger.info(f"Data: {request.data}")
    logger.info("=" * 60)
    
    # التحقق من Agent Code و Secret
    auth_header = request.headers.get('Authorization', '')
    agent_code = request.data.get('agent_code')
    
    if not auth_header.startswith('Bearer '):
        return Response({'error': 'Invalid authorization header'}, status=401)
    
    agent_secret = auth_header.replace('Bearer ', '')
    
    # التحقق من Agent (يمكن تخزين الـ Agents في قاعدة البيانات)
    # هنا نستخدم إعدادات بسيطة
    from django.conf import settings
    valid_agents = getattr(settings, 'BRIDGE_AGENTS', {})
    
    # Debug logging
    logger.info(f"Agent Code: {agent_code}")
    logger.info(f"Agent Secret Received: {agent_secret}")
    logger.info(f"Valid Agents: {valid_agents}")
    
    if agent_code not in valid_agents:
        logger.error(f"Agent code '{agent_code}' not found in valid agents")
        return Response({'error': 'Invalid agent credentials'}, status=401)
    
    expected_secret = valid_agents[agent_code]
    # استخدام hmac.compare_digest لمنع timing attacks
    if not hmac.compare_digest(expected_secret, agent_secret):
        logger.error(f"Secret mismatch!")
        logger.error(f"  Agent Code: {agent_code}")
        return Response({'error': 'Invalid agent credentials'}, status=401)
    
    # جلب السجلات
    records = request.data.get('records', [])
    
    # البحث عن الماكينة المرتبطة بالـ Agent
    try:
        device = BiometricDevice.objects.get(device_code=agent_code)
    except BiometricDevice.DoesNotExist:
        return Response({'error': 'Device not found for this agent'}, status=404)
    
    # تحديث last_connection حتى لو مافيش سجلات (heartbeat)
    from django.utils import timezone
    device.last_connection = timezone.now()
    device.status = 'active'
    
    # تنظيف السجلات القديمة (أقدم من 30 يوم) - مرة واحدة كل 100 مزامنة
    import random
    if random.randint(1, 100) == 1:
        try:
            deleted = BiometricSyncLog.cleanup_old_logs(days=30)
            if deleted > 0:
                logger.info(f"Cleaned up {deleted} old sync logs")
        except Exception as e:
            logger.warning(f"Failed to cleanup old logs: {e}")
    
    # إنشاء سجل المزامنة
    sync_log = BiometricSyncLog.objects.create(
        device=device,
        started_at=timezone.now(),
        status='success',  # سيتم تحديثه لاحقاً
        records_fetched=len(records)
    )
    
    # لو مافيش سجلات، نرجع heartbeat response
    if not records:
        device.save()
        sync_log.completed_at = timezone.now()
        sync_log.status = 'success'
        sync_log.save()
        logger.info(f"Heartbeat received from {agent_code} - No records")
        return Response({
            'success': True,
            'message': 'Heartbeat received - No new records',
            'processed': 0,
            'skipped': 0,
            'total': 0
        })
    
    # معالجة السجلات
    processed = 0
    skipped = 0
    failed = 0
    
    for record in records:
        try:
            user_id = record.get('user_id')
            timestamp_str = record.get('timestamp')
            
            # تحويل التاريخ
            from dateutil import parser
            timestamp = parser.parse(timestamp_str)
            
            # تحديد نوع البصمة من بيانات الجهاز
            punch = record.get('punch')
            status_val = record.get('status')
            
            # تحديد log_type بناءً على punch أو status
            log_type = 'check_in'  # افتراضي
            if punch is not None:
                punch_map = {
                    0: 'check_in',
                    1: 'check_out',
                    2: 'break_start',
                    3: 'break_end'
                }
                log_type = punch_map.get(punch, 'check_in')
            elif status_val is not None:
                status_map = {
                    0: 'check_in',
                    1: 'check_out',
                    2: 'break_start',
                    3: 'break_end'
                }
                log_type = status_map.get(status_val, 'check_in')
            
            # محاولة ربط بالموظف
            employee = None
            try:
                employee = Employee.objects.get(employee_number=user_id)
            except Employee.DoesNotExist:
                pass
            
            # استخدام get_or_create لمنع race condition
            log, created = BiometricLog.objects.get_or_create(
                device=device,
                user_id=user_id,
                timestamp=timestamp,
                defaults={
                    'employee': employee,
                    'log_type': log_type,  # ✅ استخدام النوع المحدد من الجهاز
                    'is_processed': False,
                    'raw_data': record
                }
            )
            
            if created:
                processed += 1
            else:
                skipped += 1
                
        except Exception as e:
            logger.error(f"Error processing record: {e}")
            failed += 1
            continue
    
    # تحديث إحصائيات الماكينة
    device.total_records = BiometricLog.objects.filter(device=device).count()
    device.last_sync = timezone.now()
    device.save()
    
    # تحديث سجل المزامنة
    sync_log.completed_at = timezone.now()
    sync_log.records_processed = processed
    sync_log.records_failed = failed
    
    # تحديد حالة المزامنة
    if failed > 0 and processed > 0:
        sync_log.status = 'partial'
    elif failed > 0 and processed == 0:
        sync_log.status = 'failed'
        sync_log.error_message = f'فشلت معالجة {failed} سجل'
    else:
        sync_log.status = 'success'
    
    sync_log.save()
    
    return Response({
        'success': True,
        'message': f'Processed {processed} records',
        'processed': processed,
        'skipped': skipped,
        'total': len(records)
    })


@login_required
def leave_list(request):
    """قائمة الإجازات"""
    leaves = Leave.objects.select_related('employee', 'leave_type').all()
    return render(request, 'hr/leave/list.html', {'leaves': leaves})


@login_required
def leave_request(request):
    """طلب إجازة جديد"""
    from .services.leave_service import LeaveService
    
    # جلب الموظف الحالي
    try:
        employee = Employee.objects.get(user=request.user)
    except Employee.DoesNotExist:
        messages.error(request, 'لا يوجد حساب موظف مرتبط بحسابك')
        return redirect('hr:dashboard')
    
    # جلب أرصدة الإجازات للموظف
    current_year = date.today().year
    balances = LeaveBalance.objects.filter(
        employee=employee,
        year=current_year
    ).select_related('leave_type')
    
    # تحديث الأرصدة المستحقة
    for balance in balances:
        balance.update_accrued_days()
    
    if request.method == 'POST':
        form = LeaveRequestForm(request.POST, request.FILES)
        if form.is_valid():
            leave = form.save(commit=False)
            leave.employee = employee
            leave.requested_by = request.user
            
            # حساب عدد الأيام
            days_count = (leave.end_date - leave.start_date).days + 1
            leave.days_count = days_count
            
            # التحقق من الرصيد المتاح
            balance = balances.filter(leave_type=leave.leave_type).first()
            if balance and balance.remaining_days < days_count:
                messages.error(request, f'رصيدك غير كافٍ. المتبقي: {balance.remaining_days} يوم، المطلوب: {days_count} يوم')
            else:
                leave.save()
                messages.success(request, 'تم تقديم طلب الإجازة بنجاح')
                return redirect('hr:leave_detail', pk=leave.pk)
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = LeaveRequestForm()
    
    context = {
        'form': form,
        'employee': employee,
        'balances': balances,
        'current_year': current_year,
    }
    return render(request, 'hr/leave/request.html', context)


@login_required
def leave_detail(request, pk):
    """تفاصيل الإجازة"""
    leave = get_object_or_404(Leave, pk=pk)
    return render(request, 'hr/leave/detail.html', {'leave': leave})


@login_required
def leave_approve(request, pk):
    """اعتماد الإجازة"""
    leave = get_object_or_404(Leave, pk=pk)
    if request.method == 'POST':
        review_notes = request.POST.get('review_notes', '')
        try:
            LeaveService.approve_leave(leave, request.user, review_notes)
            messages.success(request, 'تم اعتماد الإجازة بنجاح')
            return redirect('hr:leave_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    return render(request, 'hr/leave/approve.html', {'leave': leave})


@login_required
def leave_reject(request, pk):
    """رفض الإجازة"""
    leave = get_object_or_404(Leave, pk=pk)
    if request.method == 'POST':
        review_notes = request.POST.get('review_notes', '')
        try:
            LeaveService.reject_leave(leave, request.user, review_notes)
            messages.success(request, 'تم رفض الإجازة')
            return redirect('hr:leave_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    return render(request, 'hr/leave/reject.html', {'leave': leave})


# ==================== أرصدة الإجازات ====================

@login_required
def leave_balance_list(request):
    """قائمة أرصدة الإجازات"""
    from core.models import SystemSetting
    from collections import defaultdict
    
    current_year = date.today().year
    year = request.GET.get('year', current_year)
    
    balances = LeaveBalance.objects.filter(year=year).select_related(
        'employee', 'employee__department', 'employee__job_title', 'leave_type'
    ).order_by('employee__employee_number')
    
    # تجميع الأرصدة حسب الموظف
    employees_balances = defaultdict(list)
    for balance in balances:
        employees_balances[balance.employee].append(balance)
    
    # تحويل إلى قائمة مرتبة
    grouped_balances = [
        {
            'employee': employee,
            'balances': employee_balances,
            'total_remaining': sum(b.remaining_days for b in employee_balances),
            'total_used': sum(b.used_days for b in employee_balances),
        }
        for employee, employee_balances in employees_balances.items()
    ]
    
    # ترتيب حسب رقم الموظف
    grouped_balances.sort(key=lambda x: x['employee'].employee_number)
    
    # إحصائيات
    from django.db.models import Sum, Avg
    stats = balances.aggregate(
        total_employees=Count('employee', distinct=True),
        total_days=Sum('total_days'),
        total_used=Sum('used_days'),
        total_remaining=Sum('remaining_days'),
        avg_remaining=Avg('remaining_days')
    )
    
    # جلب إعدادات الترحيل
    rollover_enabled = SystemSetting.get_setting('leave_rollover_enabled', False)
    
    context = {
        'balances': balances,
        'grouped_balances': grouped_balances,
        'stats': stats,
        'current_year': current_year,
        'selected_year': int(year),
        'years': range(current_year - 2, current_year + 2),
        'rollover_enabled': rollover_enabled,
    }
    return render(request, 'hr/leave_balance/list.html', context)


@login_required
def leave_balance_employee(request, employee_id):
    """أرصدة إجازات موظف محدد"""
    employee = get_object_or_404(Employee, pk=employee_id)
    current_year = date.today().year
    
    balances = LeaveBalance.objects.filter(
        employee=employee,
        year=current_year
    ).select_related('leave_type')
    
    # سجل الإجازات
    leaves = Leave.objects.filter(
        employee=employee,
        start_date__year=current_year
    ).select_related('leave_type').order_by('-start_date')
    
    context = {
        'employee': employee,
        'balances': balances,
        'leaves': leaves,
        'current_year': current_year,
    }
    return render(request, 'hr/leave_balance/employee.html', context)


@login_required
def leave_balance_update(request):
    """تحديث أرصدة الإجازات"""
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        leave_type_id = request.POST.get('leave_type_id')
        year = request.POST.get('year', date.today().year)
        total_days = request.POST.get('total_days')
        
        try:
            balance, created = LeaveBalance.objects.get_or_create(
                employee_id=employee_id,
                leave_type_id=leave_type_id,
                year=year,
                defaults={'total_days': total_days, 'remaining_days': total_days}
            )
            
            if not created:
                balance.total_days = int(total_days)
                balance.update_balance()
            
            messages.success(request, 'تم تحديث الرصيد بنجاح')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
        
        return redirect('hr:leave_balance_list')
    
    employees = Employee.objects.filter(status='active')
    leave_types = LeaveType.objects.filter(is_active=True)
    
    context = {
        'employees': employees,
        'leave_types': leave_types,
        'current_year': date.today().year,
    }
    return render(request, 'hr/leave_balance/update.html', context)


@login_required
def leave_balance_rollover(request):
    """ترحيل أرصدة الإجازات للسنة الجديدة"""
    from core.models import SystemSetting
    
    # التحقق من تفعيل الترحيل في الإعدادات
    rollover_enabled = SystemSetting.get_setting('leave_rollover_enabled', False)
    max_rollover_days = SystemSetting.get_setting('leave_rollover_max_days', 7)
    
    # جلب نوع الإجازة السنوية للأمثلة (أو أي نوع نشط)
    annual_leave = LeaveType.objects.filter(code='ANNUAL', is_active=True).first()
    if not annual_leave:
        # إذا لم توجد إجازة سنوية، خذ أول نوع نشط
        annual_leave = LeaveType.objects.filter(is_active=True).first()
    annual_days = annual_leave.max_days_per_year if annual_leave else 0
    
    if request.method == 'POST':
        if not rollover_enabled:
            messages.warning(request, 'ترحيل الإجازات غير مفعل في الإعدادات. يرجى تفعيله من إعدادات الموارد البشرية.')
            return redirect('hr:leave_balance_list')
        
        from_year = int(request.POST.get('from_year'))
        to_year = int(request.POST.get('to_year'))
        rollover_unused = request.POST.get('rollover_unused') == 'on'
        
        try:
            employees = Employee.objects.filter(status='active')
            leave_types = LeaveType.objects.filter(is_active=True)
            created_count = 0
            total_rollover_days = 0
            
            for employee in employees:
                for leave_type in leave_types:
                    # الحصول على الرصيد القديم
                    old_balance = LeaveBalance.objects.filter(
                        employee=employee,
                        leave_type=leave_type,
                        year=from_year
                    ).first()
                    
                    # حساب الرصيد الجديد
                    total_days = leave_type.max_days_per_year
                    if rollover_unused and old_balance and old_balance.remaining_days > 0:
                        # تطبيق الحد الأقصى للأيام المرحلة
                        rollover_days = min(old_balance.remaining_days, max_rollover_days)
                        total_days += rollover_days
                        total_rollover_days += rollover_days
                    
                    # إنشاء الرصيد الجديد
                    LeaveBalance.objects.get_or_create(
                        employee=employee,
                        leave_type=leave_type,
                        year=to_year,
                        defaults={
                            'total_days': total_days,
                            'remaining_days': total_days
                        }
                    )
                    created_count += 1
            
            if total_rollover_days > 0:
                messages.success(request, f'تم ترحيل {created_count} رصيد بنجاح (تم ترحيل {total_rollover_days} يوم بحد أقصى {max_rollover_days} يوم لكل رصيد)')
            else:
                messages.success(request, f'تم إنشاء {created_count} رصيد جديد بنجاح')
            return redirect('hr:leave_balance_list')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    current_year = date.today().year
    context = {
        'current_year': current_year,
        'years': range(current_year - 2, current_year + 3),
        'rollover_enabled': rollover_enabled,
        'max_rollover_days': max_rollover_days,
        'annual_days': annual_days,
    }
    return render(request, 'hr/leave_balance/rollover.html', context)


@login_required
def leave_balance_accrual_status(request, employee_id):
    """عرض حالة استحقاق إجازات موظف محدد"""
    from .services.leave_accrual_service import LeaveAccrualService
    from core.models import SystemSetting
    
    employee = get_object_or_404(Employee, pk=employee_id)
    status = LeaveAccrualService.get_employee_accrual_status(employee)
    
    # جلب الإعدادات للعرض في الـ template
    leave_settings = {
        'probation_months': SystemSetting.get_setting('leave_accrual_probation_months', 3),
        'partial_percentage': SystemSetting.get_setting('leave_accrual_partial_percentage', 25),
        'full_months': SystemSetting.get_setting('leave_accrual_full_months', 6),
    }
    
    context = {
        'employee': employee,
        'status': status,
        'leave_settings': leave_settings,
    }
    return render(request, 'hr/leave_balance/accrual_status.html', context)


@login_required
def payroll_list(request):
    """قائمة كشوف الرواتب"""
    payrolls = Payroll.objects.select_related('employee', 'salary').all()
    return render(request, 'hr/payroll/list.html', {'payrolls': payrolls})


@login_required
def payroll_run_list(request):
    """قائمة مسيرات الرواتب"""
    # جمع الرواتب حسب الشهر
    from django.db.models import Count, Sum
    payroll_runs = Payroll.objects.values('month').annotate(
        total_employees=Count('id'),
        total_amount=Sum('net_salary'),
        paid_count=Count('id', filter=models.Q(status='paid'))
    ).order_by('-month')
    
    return render(request, 'hr/payroll/run_list.html', {'payroll_runs': payroll_runs})


@login_required
def payroll_run_process(request):
    """معالجة مسيرة رواتب جديدة"""
    if request.method == 'POST':
        form = PayrollProcessForm(request.POST)
        if form.is_valid():
            try:
                month_str = form.cleaned_data['month']
                department = form.cleaned_data.get('department')
                
                payrolls = PayrollService.process_monthly_payroll(
                    month_str,
                    department,
                    request.user
                )
                messages.success(request, f'تم معالجة {len(payrolls)} كشف راتب بنجاح')
                return redirect('hr:payroll_run_detail', month=month_str)
            except Exception as e:
                messages.error(request, f'خطأ: {str(e)}')
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = PayrollProcessForm()
    
    return render(request, 'hr/payroll/run_process.html', {'form': form})


@login_required
def payroll_run_detail(request, month):
    """تفاصيل مسيرة رواتب شهر محدد"""
    from django.db.models import Sum
    
    payrolls = Payroll.objects.filter(month=month).select_related('employee', 'salary')
    
    stats = payrolls.aggregate(
        total_employees=Count('id'),
        total_gross=Sum('gross_salary'),
        total_deductions=Sum('total_deductions'),
        total_net=Sum('net_salary'),
        paid_count=Count('id', filter=models.Q(status='paid'))
    )
    
    context = {
        'month': month,
        'payrolls': payrolls,
        'stats': stats,
    }
    return render(request, 'hr/payroll/run_detail.html', context)


@login_required
def payroll_detail(request, pk):
    """تفاصيل كشف الراتب"""
    payroll = get_object_or_404(Payroll, pk=pk)
    return render(request, 'hr/payroll/detail.html', {'payroll': payroll})


# ==================== السلف ====================

@login_required
def advance_list(request):
    """قائمة السلف"""
    advances = Advance.objects.select_related('employee').all()
    return render(request, 'hr/advance/list.html', {'advances': advances})


@login_required
def advance_request(request):
    """طلب سلفة جديدة"""
    if request.method == 'POST':
        # سيتم إضافة النموذج لاحقاً
        messages.success(request, 'تم تقديم طلب السلفة بنجاح')
        return redirect('hr:advance_list')
    return render(request, 'hr/advance/request.html')


@login_required
def advance_detail(request, pk):
    """تفاصيل السلفة"""
    advance = get_object_or_404(Advance, pk=pk)
    return render(request, 'hr/advance/detail.html', {'advance': advance})


@login_required
def advance_approve(request, pk):
    """اعتماد السلفة"""
    advance = get_object_or_404(Advance, pk=pk)
    if request.method == 'POST':
        advance.status = 'approved'
        advance.approved_by = request.user
        advance.approved_at = date.today()
        advance.save()
        messages.success(request, 'تم اعتماد السلفة بنجاح')
        return redirect('hr:advance_detail', pk=pk)
    return render(request, 'hr/advance/approve.html', {'advance': advance})


@login_required
def advance_reject(request, pk):
    """رفض السلفة"""
    advance = get_object_or_404(Advance, pk=pk)
    if request.method == 'POST':
        advance.status = 'rejected'
        advance.save()
        messages.success(request, 'تم رفض السلفة')
        return redirect('hr:advance_detail', pk=pk)
    return render(request, 'hr/advance/reject.html', {'advance': advance})


@login_required
def salary_settings(request):
    """إعدادات الرواتب"""
    from .models import Salary
    salaries = Salary.objects.all()
    return render(request, 'hr/salary/settings.html', {'salaries': salaries})


# ==================== العقود ====================

@login_required
def contract_list(request):
    """قائمة العقود"""
    contracts = Contract.objects.select_related('employee', 'employee__department').all()
    
    # إحصائيات
    stats = {
        'total': contracts.count(),
        'active': contracts.filter(status='active').count(),
        'draft': contracts.filter(status='draft').count(),
        'suspended': contracts.filter(status='suspended').count(),
        'expiring_soon': contracts.filter(
            status='active',
            end_date__isnull=False,
            end_date__lte=date.today() + timedelta(days=30)
        ).count(),
    }
    
    # تعريف رؤوس الجدول
    headers = [
        {'key': 'contract_number', 'label': 'رقم العقد', 'sortable': True, 'class': 'text-center'},
        {'key': 'employee', 'label': 'الموظف', 'sortable': True, 'template': 'hr/contract/cells/employee.html'},
        {'key': 'contract_type', 'label': 'نوع العقد', 'sortable': True, 'template': 'hr/contract/cells/contract_type.html', 'class': 'text-center'},
        {'key': 'start_date', 'label': 'تاريخ البداية', 'sortable': True, 'format': 'date', 'class': 'text-center'},
        {'key': 'end_date', 'label': 'تاريخ النهاية', 'sortable': True, 'template': 'hr/contract/cells/end_date.html', 'class': 'text-center'},
        {'key': 'basic_salary', 'label': 'الراتب الأساسي', 'sortable': True, 'format': 'currency', 'class': 'text-center'},
        {'key': 'status', 'label': 'الحالة', 'sortable': True, 'template': 'hr/contract/cells/status.html', 'class': 'text-center'},
    ]
    
    # أزرار الإجراءات
    action_buttons = [
        {'url': 'hr:contract_detail', 'icon': 'fa-eye', 'label': 'عرض', 'class': 'action-view'},
        {'url': 'hr:contract_form_edit', 'icon': 'fa-edit', 'label': 'تعديل', 'class': 'action-edit'},
    ]
    
    context = {
        'headers': headers,
        'data': contracts,
        'action_buttons': action_buttons,
        'table_id': 'contracts-table',
        'primary_key': 'pk',
        'empty_message': 'لا توجد عقود',
        'currency_symbol': 'جنيه',
        'clickable_rows': True,
        'row_click_url': '/hr/contracts/0/',
        **{f'{k}_contracts': v for k, v in stats.items()},
        'show_stats': True,
    }
    
    return render(request, 'hr/contract/list.html', context)


@login_required
def contract_detail(request, pk):
    """تفاصيل العقد"""
    from core.models import SystemSetting
    from .models import ContractIncrease
    
    contract = get_object_or_404(Contract, pk=pk)
    
    # جلب رمز العملة
    try:
        currency_symbol = SystemSetting.get_currency_symbol()
    except:
        currency_symbol = 'جنيه'
    
    # جلب الزيادات المجدولة
    scheduled_increases = contract.scheduled_increases.all().order_by('increase_number')
    pending_increases = scheduled_increases.filter(status='pending')
    applied_increases = scheduled_increases.filter(status='applied')
    
    # تعريف رؤوس جدول المرفقات
    documents_headers = [
        {'key': 'file', 'label': '', 'template': 'hr/contract/cells/document_icon.html', 'class': 'text-center', 'width': '40', 'searchable': False},
        {'key': 'document_type', 'label': 'النوع', 'template': 'hr/contract/cells/document_type.html', 'searchable': True},
        {'key': 'title', 'label': 'العنوان', 'template': 'hr/contract/cells/document_title.html', 'searchable': True},
        {'key': 'file_size_mb', 'label': 'الحجم', 'template': 'hr/contract/cells/document_size.html', 'class': 'text-center', 'searchable': False},
        {'key': 'uploaded_at', 'label': 'تاريخ الرفع', 'template': 'hr/contract/cells/document_date.html', 'class': 'text-center', 'searchable': True},
        {'key': 'uploaded_by', 'label': 'رفع بواسطة', 'template': 'hr/contract/cells/document_uploader.html', 'searchable': True},
        {'key': 'actions', 'label': 'إجراءات', 'template': 'hr/contract/cells/document_actions.html', 'class': 'text-center', 'width': '100', 'searchable': False},
    ]
    
    context = {
        'contract': contract,
        'system_settings': {'currency_symbol': currency_symbol},
        'documents_headers': documents_headers,
        'primary_key': 'id',
        'scheduled_increases': scheduled_increases,
        'pending_increases': pending_increases,
        'applied_increases': applied_increases,
    }
    
    return render(request, 'hr/contract/detail.html', context)


@login_required
def contract_activate(request, pk):
    """تفعيل العقد (draft → active)"""
    contract = get_object_or_404(Contract, pk=pk)
    
    if contract.status == 'draft':
        contract.status = 'active'
        contract.save()
        messages.success(request, f'تم تفعيل العقد {contract.contract_number} بنجاح')
    else:
        messages.warning(request, 'لا يمكن تفعيل هذا العقد')
    
    return redirect('hr:contract_detail', pk=pk)


@login_required
def contract_suspend(request, pk):
    """إيقاف العقد مؤقتاً (active → suspended)"""
    contract = get_object_or_404(Contract, pk=pk)
    
    if contract.status == 'active':
        contract.status = 'suspended'
        contract.save()
        messages.warning(request, f'تم إيقاف العقد {contract.contract_number} مؤقتاً')
    else:
        messages.warning(request, 'لا يمكن إيقاف هذا العقد')
    
    return redirect('hr:contract_detail', pk=pk)


@login_required
def contract_reactivate(request, pk):
    """إعادة تفعيل العقد (suspended → active)"""
    contract = get_object_or_404(Contract, pk=pk)
    
    if contract.status == 'suspended':
        contract.status = 'active'
        contract.save()
        messages.success(request, f'تم إعادة تفعيل العقد {contract.contract_number} بنجاح')
    else:
        messages.warning(request, 'لا يمكن إعادة تفعيل هذا العقد')
    
    return redirect('hr:contract_detail', pk=pk)


@login_required
def contract_document_upload(request, pk):
    """رفع مرفق للعقد"""
    from .models.contract import ContractDocument
    
    contract = get_object_or_404(Contract, pk=pk)
    
    if request.method == 'POST':
        document_type = request.POST.get('document_type')
        title = request.POST.get('title')
        description = request.POST.get('description', '')
        file = request.FILES.get('file')
        
        if not all([document_type, title, file]):
            messages.error(request, 'يرجى ملء جميع الحقول المطلوبة')
            return redirect('hr:contract_detail', pk=pk)
        
        try:
            document = ContractDocument.objects.create(
                contract=contract,
                document_type=document_type,
                title=title,
                description=description,
                file=file,
                uploaded_by=request.user
            )
            messages.success(request, f'تم رفع المرفق "{title}" بنجاح')
        except Exception as e:
            messages.error(request, f'خطأ في رفع المرفق: {str(e)}')
    
    return redirect('hr:contract_detail', pk=pk)


@login_required
def contract_document_delete(request, pk, doc_id):
    """حذف مرفق من العقد"""
    from .models.contract import ContractDocument
    
    contract = get_object_or_404(Contract, pk=pk)
    document = get_object_or_404(ContractDocument, pk=doc_id, contract=contract)
    
    if request.method == 'POST':
        title = document.title
        document.delete()
        messages.success(request, f'تم حذف المرفق "{title}" بنجاح')
    
    return redirect('hr:contract_detail', pk=pk)


@login_required
def contract_amendment_create(request, pk):
    """إضافة تعديل على العقد"""
    from .models.contract import ContractAmendment
    
    contract = get_object_or_404(Contract, pk=pk)
    
    if request.method == 'POST':
        try:
            # توليد رقم تعديل تلقائي
            last_amendment = ContractAmendment.objects.filter(
                contract=contract
            ).order_by('-amendment_number').first()
            
            if last_amendment:
                # استخراج الرقم من آخر تعديل
                last_num = int(last_amendment.amendment_number.split('-')[-1])
                amendment_number = f"{contract.contract_number}-AMD-{last_num + 1:03d}"
            else:
                amendment_number = f"{contract.contract_number}-AMD-001"
            
            # إنشاء التعديل
            amendment = ContractAmendment.objects.create(
                contract=contract,
                amendment_number=amendment_number,
                amendment_type=request.POST.get('amendment_type'),
                effective_date=request.POST.get('effective_date'),
                description=request.POST.get('description'),
                old_value=request.POST.get('old_value', ''),
                new_value=request.POST.get('new_value', ''),
                created_by=request.user
            )
            
            messages.success(request, f'تم إضافة التعديل "{amendment.get_amendment_type_display()}" بنجاح')
            
        except Exception as e:
            messages.error(request, f'حدث خطأ أثناء إضافة التعديل: {str(e)}')
    
    return redirect('hr:contract_detail', pk=pk)


@login_required
def contract_form(request, pk=None):
    """نموذج موحد لإضافة/تعديل عقد"""
    from .forms.contract_forms import ContractForm
    from .models import SalaryComponent
    
    contract = get_object_or_404(Contract, pk=pk) if pk else None
    
    if request.method == 'POST':
        form = ContractForm(request.POST, instance=contract)
        if form.is_valid():
            contract_obj = form.save(commit=False)
            
            # توليد رقم العقد تلقائياً إذا لم يتم إدخاله (للإضافة فقط)
            if not pk and not contract_obj.contract_number:
                contract_obj.contract_number = ContractForm.generate_contract_number()
            
            # تعيين المستخدم الذي أنشأ السجل (للإضافة فقط)
            if not pk:
                contract_obj.created_by = request.user
                # تعيين نوع العقد الافتراضي (عقد محدد المدة)
                if not contract_obj.contract_type:
                    contract_obj.contract_type = 'contract'
            
            contract_obj.save()
            
            # حفظ مكونات الراتب بطريقة ذكية (update or create)
            from decimal import Decimal
            import logging
            logger = logging.getLogger(__name__)
            
            # جمع البيانات الجديدة من الـ POST مع IDs
            new_earnings = []
            new_deductions = []
            
            # طباعة POST data للتحليل
            logger.info("=" * 80)
            logger.info("POST DATA للمستحقات:")
            for key in sorted(request.POST.keys()):
                if 'earning' in key:
                    logger.info(f"{key} = {request.POST.get(key)}")
            logger.info("=" * 80)
            
            # جمع المستحقات - نرتبهم حسب counter
            earning_counters = sorted([
                key.split('_')[-1] for key in request.POST.keys() 
                if key.startswith('earning_name_')
            ], key=lambda x: int(x) if x.isdigit() else 0)
            
            for counter in earning_counters:
                name = request.POST.get(f'earning_name_{counter}')
                formula = request.POST.get(f'earning_formula_{counter}', '')
                amount = request.POST.get(f'earning_amount_{counter}')
                order = request.POST.get(f'earning_order_{counter}', 0)
                component_id = request.POST.get(f'earning_id_{counter}', '')
                
                if name and amount:
                    # تنظيف ID - تحويل string فاضي لـ None
                    clean_id = None
                    if component_id and component_id.strip():
                        try:
                            clean_id = int(component_id)
                        except (ValueError, TypeError):
                            clean_id = None
                    
                    earning_item = {
                        'id': clean_id,
                        'name': name,
                        'formula': formula,
                        'amount': Decimal(amount),
                        'order': int(order) if order else 0
                    }
                    new_earnings.append(earning_item)
                    logger.info(f"Counter {counter}: {earning_item}")
            
            # جمع الاستقطاعات - نرتبهم حسب counter
            deduction_counters = sorted([
                key.split('_')[-1] for key in request.POST.keys() 
                if key.startswith('deduction_name_')
            ], key=lambda x: int(x) if x.isdigit() else 0)
            
            for counter in deduction_counters:
                name = request.POST.get(f'deduction_name_{counter}')
                formula = request.POST.get(f'deduction_formula_{counter}', '')
                amount = request.POST.get(f'deduction_amount_{counter}')
                order = request.POST.get(f'deduction_order_{counter}', 0)
                component_id = request.POST.get(f'deduction_id_{counter}', '')
                
                if name and amount:
                    # تنظيف ID - تحويل string فاضي لـ None
                    clean_id = None
                    if component_id and component_id.strip():
                        try:
                            clean_id = int(component_id)
                        except (ValueError, TypeError):
                            clean_id = None
                    
                    new_deductions.append({
                        'id': clean_id,
                        'name': name,
                        'formula': formula,
                        'amount': Decimal(amount),
                        'order': int(order) if order else 0
                    })
            
            if pk:
                # للتعديل: استخدام ID tracking
                logger.info("=" * 80)
                logger.info(f"معالجة التعديل - العقد: {contract_obj.contract_number}")
                logger.info(f"عدد المستحقات في POST: {len(new_earnings)}")
                logger.info(f"عدد الاستقطاعات في POST: {len(new_deductions)}")
                
                # فصل البنود الموجودة عن الجديدة
                existing_earning_ids = []
                new_earnings_data = []
                
                for idx, item in enumerate(new_earnings):
                    logger.info(f"معالجة مستحق {idx+1}: ID={item['id']}, Name={item['name']}")
                    if item['id']:
                        existing_earning_ids.append(item['id'])
                        # تحديث الموجود مباشرة
                        try:
                            obj = SalaryComponent.objects.get(
                                id=item['id'],
                                contract=contract_obj,
                                component_type='earning'
                            )
                            obj.name = item['name']
                            obj.formula = item['formula']
                            obj.amount = item['amount']
                            obj.order = item['order']
                            obj.save()
                            logger.info(f"  → تم تحديث ID={item['id']}")
                        except SalaryComponent.DoesNotExist:
                            logger.error(f"  → خطأ: ID={item['id']} غير موجود!")
                    else:
                        new_earnings_data.append(item)
                        logger.info(f"  → سيتم إنشاء جديد: {item['name']}")
                
                # إنشاء المستحقات الجديدة
                logger.info(f"إنشاء {len(new_earnings_data)} مستحق جديد...")
                created_earning_ids = []
                for data in new_earnings_data:
                    new_obj = SalaryComponent.objects.create(
                        contract=contract_obj,
                        component_type='earning',
                        name=data['name'],
                        formula=data['formula'],
                        amount=data['amount'],
                        order=data['order']
                    )
                    created_earning_ids.append(new_obj.id)
                    logger.info(f"  → تم إنشاء: {data['name']} - ID={new_obj.id}")
                
                # حذف المستحقات المحذوفة (اللي مش موجودة في الـ POST)
                all_earning_ids = existing_earning_ids + created_earning_ids
                logger.info(f"IDs المحفوظة: {all_earning_ids}")
                deleted_earnings = contract_obj.salary_components.filter(
                    component_type='earning',
                    is_basic=False
                ).exclude(id__in=all_earning_ids)
                if deleted_earnings.exists():
                    logger.info(f"حذف {deleted_earnings.count()} مستحق...")
                    for item in deleted_earnings:
                        logger.info(f"  → حذف: {item.name} - ID={item.id}")
                    deleted_earnings.delete()
                logger.info("=" * 80)
                
                # معالجة الاستقطاعات
                existing_deduction_ids = []
                new_deductions_data = []
                
                for item in new_deductions:
                    if item['id']:
                        existing_deduction_ids.append(item['id'])
                        # تحديث الموجود مباشرة
                        try:
                            obj = SalaryComponent.objects.get(
                                id=item['id'],
                                contract=contract_obj,
                                component_type='deduction'
                            )
                            obj.name = item['name']
                            obj.formula = item['formula']
                            obj.amount = item['amount']
                            obj.order = item['order']
                            obj.save()
                        except SalaryComponent.DoesNotExist:
                            pass  # تجاهل لو الـ ID مش موجود
                    else:
                        new_deductions_data.append(item)
                
                # إنشاء الاستقطاعات الجديدة
                created_deduction_ids = []
                for data in new_deductions_data:
                    new_obj = SalaryComponent.objects.create(
                        contract=contract_obj,
                        component_type='deduction',
                        name=data['name'],
                        formula=data['formula'],
                        amount=data['amount'],
                        order=data['order']
                    )
                    created_deduction_ids.append(new_obj.id)
                
                # حذف الاستقطاعات المحذوفة (اللي مش موجودة في الـ POST)
                all_deduction_ids = existing_deduction_ids + created_deduction_ids
                contract_obj.salary_components.filter(
                    component_type='deduction'
                ).exclude(id__in=all_deduction_ids).delete()
            else:
                # للإضافة: إنشاء جديد
                for data in new_earnings:
                    data.pop('id', None)  # إزالة ID لو موجود
                    SalaryComponent.objects.create(
                        contract=contract_obj,
                        component_type='earning',
                        **data
                    )
                
                for data in new_deductions:
                    data.pop('id', None)  # إزالة ID لو موجود
                    SalaryComponent.objects.create(
                        contract=contract_obj,
                        component_type='deduction',
                        **data
                    )
            
            if pk:
                messages.success(request, 'تم تحديث العقد بنجاح')
            else:
                messages.success(request, f'تم إضافة العقد بنجاح - الرقم: {contract_obj.contract_number}')
            return redirect('hr:contract_list')
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = ContractForm(instance=contract)
    
    # توليد رقم العقد المقترح (للإضافة فقط)
    next_contract_number = ContractForm.generate_contract_number() if not pk else None
    
    # جلب مكونات الراتب الموجودة (للتعديل)
    earnings = []
    deductions = []
    if contract:
        earnings = contract.salary_components.filter(component_type='earning', is_basic=False)
        deductions = contract.salary_components.filter(component_type='deduction')
    
    # جلب إعدادات النظام
    from core.utils import get_default_currency
    try:
        currency_symbol = get_default_currency()
        system_settings = {
            'currency_symbol': currency_symbol
        }
    except:
        system_settings = {'currency_symbol': 'ج.م'}
    
    context = {
        'form': form,
        'contract': contract,
        'next_contract_number': next_contract_number,
        'earnings': earnings,
        'deductions': deductions,
        'system_settings': system_settings,
    }
    return render(request, 'hr/contract/form.html', context)


@login_required
def get_salary_component_templates(request):
    """API لجلب قوالب مكونات الراتب"""
    from django.http import JsonResponse
    from .models import SalaryComponentTemplate
    
    component_type = request.GET.get('type', 'earning')
    
    templates = SalaryComponentTemplate.objects.filter(
        component_type=component_type,
        is_active=True
    ).order_by('order', 'name')
    
    data = [{
        'id': t.id,
        'name': t.name,
        'formula': t.formula,
        'default_amount': str(t.default_amount),
        'description': t.description
    } for t in templates]
    
    return JsonResponse({'success': True, 'templates': data})


@login_required
def contract_renew(request, pk):
    """تجديد العقد - يفتح form كامل للتعديل"""
    from datetime import datetime, timedelta, date
    from .forms.contract_forms import ContractForm
    old_contract = get_object_or_404(Contract, pk=pk)
    
    if request.method == 'POST':
        form = ContractForm(request.POST)
        
        # إضافة الموظف يدوياً لأن الحقل disabled
        if form.data.get('employee') is None:
            mutable_post = request.POST.copy()
            mutable_post['employee'] = old_contract.employee.id
            form = ContractForm(mutable_post)
        
        if form.is_valid():
            new_contract = form.save(commit=False)
            new_contract.employee = old_contract.employee  # التأكد من الموظف
            new_contract.created_by = request.user
            new_contract.save()
            
            # تحديث العقد القديم
            old_contract.status = 'renewed'
            old_contract.renewed_to = new_contract
            old_contract.save()
            
            messages.success(request, f'تم تجديد العقد بنجاح. رقم العقد الجديد: {new_contract.contract_number}')
            return redirect('hr:contract_detail', pk=new_contract.pk)
    else:
        # تحديد تاريخ بداية العقد الجديد
        if old_contract.end_date:
            new_start_date = old_contract.end_date + timedelta(days=1)
        else:
            new_start_date = date.today()
        
        # جلب رقم البصمة من BiometricUserMapping
        from .models import BiometricUserMapping
        biometric_mapping = BiometricUserMapping.objects.filter(
            employee=old_contract.employee,
            is_active=True
        ).first()
        biometric_id = biometric_mapping.biometric_user_id if biometric_mapping else None
        
        # إنشاء form مع بيانات العقد القديم
        initial_data = {
            'employee': old_contract.employee,
            'contract_type': old_contract.contract_type,
            'job_title': old_contract.job_title,
            'department': old_contract.department,
            'biometric_user_id': biometric_id,
            'start_date': new_start_date,
            'basic_salary': old_contract.basic_salary,
            'probation_period_months': 0,  # بدون فترة تجربة في التجديد
            'terms_and_conditions': old_contract.terms_and_conditions,
            'auto_renew': old_contract.auto_renew,
            'renewal_notice_days': old_contract.renewal_notice_days,
            'status': 'active',  # العقد الجديد يكون ساري مباشرة
        }
        form = ContractForm(initial=initial_data)
        
        # تعطيل حقل الموظف (لا يمكن تغييره)
        form.fields['employee'].disabled = True
        form.fields['employee'].widget.attrs['readonly'] = 'readonly'
        form.fields['employee'].widget.attrs['style'] = 'background-color: #e9ecef; pointer-events: none;'
    
    context = {
        'form': form,
        'old_contract': old_contract,
        'is_renewal': True,
        'page_title': f'تجديد العقد: {old_contract.contract_number}',
    }
    return render(request, 'hr/contract/form.html', context)


@login_required
def contract_terminate(request, pk):
    """إنهاء العقد"""
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        termination_date = request.POST.get('termination_date')
        contract.terminate(termination_date)
        messages.success(request, 'تم إنهاء العقد بنجاح')
        return redirect('hr:contract_detail', pk=pk)
    return render(request, 'hr/contract/terminate.html', {'contract': contract})


@login_required
def contract_create_increase_schedule(request, pk):
    """إنشاء جدول زيادات مجدولة للعقد"""
    contract = get_object_or_404(Contract, pk=pk)
    
    if request.method == 'POST':
        try:
            annual_percentage = float(request.POST.get('annual_percentage'))
            installments = int(request.POST.get('installments'))
            interval_months = int(request.POST.get('interval_months'))
            
            # إنشاء الجدول
            increases = contract.create_increase_schedule(
                annual_percentage=annual_percentage,
                installments=installments,
                interval_months=interval_months,
                created_by=request.user
            )
            
            messages.success(request, f'تم إنشاء جدول الزيادات بنجاح ({len(increases)} زيادة مجدولة)')
        except Exception as e:
            messages.error(request, f'حدث خطأ: {str(e)}')
    
    return redirect('hr:contract_detail', pk=pk)


@login_required
@require_http_methods(["POST"])
def contract_increase_action(request, increase_id, action):
    """دالة موحدة لإجراءات الزيادات (تطبيق/إلغاء)"""
    from .models import ContractIncrease
    from django.http import JsonResponse
    
    increase = get_object_or_404(ContractIncrease, pk=increase_id)
    
    if action == 'apply':
        success, message = increase.apply_increase(applied_by=request.user)
    elif action == 'cancel':
        success, message = increase.cancel_increase()
    else:
        return JsonResponse({
            'success': False,
            'message': 'إجراء غير صحيح'
        }, status=400)
    
    return JsonResponse({
        'success': success,
        'message': message
    })


# الدوال القديمة للتوافق مع الروابط الموجودة (سيتم إزالتها لاحقاً)
@login_required
@require_http_methods(["POST"])
def contract_increase_apply(request, increase_id):
    """تطبيق زيادة مجدولة - استخدم contract_increase_action بدلاً منها"""
    return contract_increase_action(request, increase_id, 'apply')


@login_required
@require_http_methods(["POST"])
def contract_increase_cancel(request, increase_id):
    """إلغاء زيادة مجدولة - استخدم contract_increase_action بدلاً منها"""
    return contract_increase_action(request, increase_id, 'cancel')


@login_required
def contract_expiring(request):
    """العقود قرب الانتهاء"""
    # العقود التي ستنتهي خلال 60 يوم
    expiry_date = date.today() + timedelta(days=60)
    contracts = Contract.objects.filter(
        status='active',
        end_date__lte=expiry_date,
        end_date__gte=date.today()
    ).select_related('employee').order_by('end_date')
    
    return render(request, 'hr/contract/expiring.html', {'contracts': contracts})


# ==================== المسميات الوظيفية ====================

@login_required
def job_title_list(request):
    """قائمة المسميات الوظيفية"""
    job_titles = JobTitle.objects.select_related('department').filter(is_active=True)
    return render(request, 'hr/job_title/list.html', {'job_titles': job_titles})


@login_required
def job_title_delete(request, pk):
    """حذف مسمى وظيفي"""
    job_title = get_object_or_404(JobTitle, pk=pk)
    if request.method == 'POST':
        # التحقق من عدم وجود موظفين مرتبطين بهذه الوظيفة
        if job_title.employees.exists():
            messages.error(request, 'لا يمكن حذف هذه الوظيفة لأن هناك موظفين مرتبطين بها')
            return redirect('hr:job_title_list')
        
        # حذف نهائي (Hard Delete)
        job_title.delete()
        messages.success(request, 'تم حذف المسمى الوظيفي نهائياً')
        return redirect('hr:job_title_list')
    
    # تجهيز البيانات للمودال الموحد
    employees_count = job_title.employees.count()
    
    item_fields = [
        {'label': 'الكود', 'value': job_title.code},
        {'label': 'المسمى (عربي)', 'value': job_title.title_ar},
        {'label': 'المسمى (إنجليزي)', 'value': job_title.title_en or '-'},
        {'label': 'القسم', 'value': job_title.department.name_ar},
        {'label': 'عدد الموظفين', 'value': employees_count},
    ]
    
    warning_message = f'تحذير: يوجد {employees_count} موظف مرتبط بهذا المسمى الوظيفي. سيتم إلغاء تفعيل المسمى فقط وليس حذفه نهائياً.'
    
    context = {
        'job_title': job_title,
        'item_fields': item_fields,
        'employees_count': employees_count,
        'warning_message': warning_message,
    }
    return render(request, 'hr/job_title/delete_modal.html', context)


@login_required
def job_title_form(request, pk=None):
    """نموذج موحد لإضافة/تعديل مسمى وظيفي"""
    from .forms.employee_forms import JobTitleForm
    
    job_title = get_object_or_404(JobTitle, pk=pk) if pk else None
    
    if request.method == 'POST':
        form = JobTitleForm(request.POST, instance=job_title)
        if form.is_valid():
            form.save()
            if pk:
                messages.success(request, 'تم تحديث المسمى الوظيفي بنجاح')
            else:
                messages.success(request, 'تم إضافة المسمى الوظيفي بنجاح')
            return redirect('hr:job_title_list')
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = JobTitleForm(instance=job_title)
    
    return render(request, 'hr/job_title/form.html', {'form': form, 'job_title': job_title})


# ==================== الهيكل التنظيمي ====================

@login_required
def organization_chart(request):
    """عرض الهيكل التنظيمي"""
    # جلب الأقسام الرئيسية (التي ليس لها قسم أب)
    root_departments = Department.objects.filter(parent=None, is_active=True).prefetch_related('sub_departments')
    
    context = {
        'root_departments': root_departments,
        'total_departments': Department.objects.filter(is_active=True).count(),
        'total_employees': Employee.objects.filter(status='active').count(),
        'total_job_titles': JobTitle.objects.filter(is_active=True).count(),
    }
    return render(request, 'hr/organization/chart.html', context)


# ==================== إعدادات الموارد البشرية ====================

@login_required
def hr_settings(request):
    """صفحة إعدادات الموارد البشرية"""
    from core.models import SystemSetting
    
    if request.method == 'POST':
        # حفظ إعدادات الإجازات
        
        # الحقول النصية والرقمية
        numeric_settings = [
            'leave_accrual_probation_months',
            'leave_accrual_partial_percentage',
            'leave_accrual_full_months',
            'leave_rollover_max_days',
        ]
        
        for setting_key in numeric_settings:
            if setting_key in request.POST:
                value = request.POST.get(setting_key)
                SystemSetting.objects.filter(key=setting_key).update(value=value)
        
        # الـ checkboxes (تحتاج معالجة خاصة)
        checkbox_settings = [
            'leave_auto_create_balances',
            'leave_rollover_enabled',
        ]
        
        for setting_key in checkbox_settings:
            # checkbox يرسل 'on' إذا كان محدد، ولا يرسل شيء إذا لم يكن محدد
            value = 'true' if setting_key in request.POST else 'false'
            SystemSetting.objects.filter(key=setting_key).update(value=value)
        
        # حفظ عدد أيام أنواع الإجازات
        leave_types = LeaveType.objects.filter(is_active=True)
        for leave_type in leave_types:
            field_name = f'leave_type_{leave_type.id}_days'
            if field_name in request.POST:
                new_days = int(request.POST.get(field_name))
                if new_days != leave_type.max_days_per_year:
                    leave_type.max_days_per_year = new_days
                    leave_type.save()
        
        messages.success(request, 'تم حفظ إعدادات الإجازات بنجاح')
        return redirect('hr:hr_settings')
    
    # جلب إعدادات الإجازات
    leave_settings = {
        'probation_months': SystemSetting.get_setting('leave_accrual_probation_months', 3),
        'partial_percentage': SystemSetting.get_setting('leave_accrual_partial_percentage', 25),
        'full_months': SystemSetting.get_setting('leave_accrual_full_months', 6),
        'auto_create': SystemSetting.get_setting('leave_auto_create_balances', True),
        'rollover_enabled': SystemSetting.get_setting('leave_rollover_enabled', False),
        'rollover_max_days': SystemSetting.get_setting('leave_rollover_max_days', 7),
    }
    
    # جلب أنواع الإجازات
    leave_types = LeaveType.objects.filter(is_active=True).order_by('code')
    
    # إحصائيات قوالب مكونات الراتب
    from .models import SalaryComponentTemplate
    salary_templates_earnings = SalaryComponentTemplate.objects.filter(component_type='earning', is_active=True).count()
    salary_templates_deductions = SalaryComponentTemplate.objects.filter(component_type='deduction', is_active=True).count()
    
    context = {
        'active_menu': 'hr',
        'leave_settings': leave_settings,
        'leave_types': leave_types,
        
        # إحصائيات الأقسام
        'departments_count': Department.objects.filter(is_active=True).count(),
        'total_departments': Department.objects.count(),
        
        # إحصائيات المسميات الوظيفية
        'job_titles_count': JobTitle.objects.filter(is_active=True).count(),
        'total_job_titles': JobTitle.objects.count(),
        
        # إحصائيات الورديات
        'shifts_count': Shift.objects.filter(is_active=True).count(),
        'total_shifts': Shift.objects.count(),
        
        # إحصائيات ماكينات البصمة
        'biometric_devices_count': BiometricDevice.objects.filter(is_active=True).count(),
        'total_biometric_devices': BiometricDevice.objects.count(),
        
        # إحصائيات أرصدة الإجازات
        'leave_balances_count': LeaveBalance.objects.count(),
        'employees_with_balance': LeaveBalance.objects.values('employee').distinct().count(),
        
        # إحصائيات قوالب مكونات الراتب
        'salary_templates_earnings': salary_templates_earnings,
        'salary_templates_deductions': salary_templates_deductions,
    }
    return render(request, 'hr/settings.html', context)


# ==================== ربط الموظفين بالمستخدمين ====================

@login_required
def employee_create_user(request, pk):
    """إنشاء حساب مستخدم لموظف"""
    from .services.user_employee_service import UserEmployeeService
    
    employee = get_object_or_404(Employee, pk=pk)
    
    if employee.user:
        messages.warning(request, 'الموظف مرتبط بمستخدم بالفعل')
        return redirect('hr:employee_detail', pk=pk)
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        send_email = request.POST.get('send_email') == 'on'
        
        try:
            user, created_password = UserEmployeeService.create_user_for_employee(
                employee=employee,
                username=username,
                password=password,
                send_email=send_email
            )
            
            if send_email:
                messages.success(request, f'تم إنشاء الحساب بنجاح وإرسال البيانات للبريد: {employee.work_email}')
            else:
                messages.success(request, f'تم إنشاء الحساب بنجاح - اسم المستخدم: {username}')
            
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    return redirect('hr:employee_detail', pk=pk)


@login_required
def employee_link_user(request, pk):
    """ربط موظف بمستخدم موجود"""
    from .services.user_employee_service import UserEmployeeService
    from django.contrib.auth import get_user_model
    
    User = get_user_model()
    employee = get_object_or_404(Employee, pk=pk)
    
    if employee.user:
        messages.warning(request, 'الموظف مرتبط بمستخدم بالفعل')
        return redirect('hr:employee_detail', pk=pk)
    
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        
        try:
            user = User.objects.get(pk=user_id)
            UserEmployeeService.link_existing_user_to_employee(employee, user)
            messages.success(request, f'تم ربط الموظف بالمستخدم: {user.username}')
        except User.DoesNotExist:
            messages.error(request, 'المستخدم غير موجود')
        except Exception as e:
            messages.error(request, f'خطأ: {str(e)}')
    
    return redirect('hr:employee_detail', pk=pk)


@login_required
def employee_unlink_user(request, pk):
    """فك ربط موظف من مستخدم"""
    employee = get_object_or_404(Employee, pk=pk)
    
    if not employee.user:
        messages.warning(request, 'الموظف غير مرتبط بمستخدم')
        return redirect('hr:employee_detail', pk=pk)
    
    if request.method == 'POST':
        username = employee.user.username
        employee.user = None
        employee.save()
        messages.success(request, f'تم فك الربط من المستخدم: {username}')
    
    return redirect('hr:employee_detail', pk=pk)


# ==================== API للتحقق من التكرار ====================

@login_required
def check_employee_email(request):
    """التحقق من تكرار البريد الإلكتروني"""
    email = request.GET.get('email', '')
    employee_id = request.GET.get('employee_id', None)
    
    if not email:
        return JsonResponse({'available': True})
    
    # التحقق من وجود البريد
    query = Employee.objects.filter(work_email=email)
    if employee_id:
        query = query.exclude(pk=employee_id)
    
    exists = query.exists()
    
    return JsonResponse({
        'available': not exists,
        'message': 'هذا البريد الإلكتروني مستخدم بالفعل' if exists else 'البريد الإلكتروني متاح'
    })


@login_required
def check_employee_mobile(request):
    """التحقق من تكرار رقم الموبايل"""
    mobile = request.GET.get('mobile', '')
    employee_id = request.GET.get('employee_id', None)
    
    if not mobile:
        return JsonResponse({'available': True})
    
    # التحقق من وجود الرقم
    query = Employee.objects.filter(mobile_phone=mobile)
    if employee_id:
        query = query.exclude(pk=employee_id)
    
    exists = query.exists()
    
    return JsonResponse({
        'available': not exists,
        'message': 'رقم الموبايل مستخدم بالفعل' if exists else 'رقم الموبايل متاح'
    })


@login_required
def check_employee_national_id(request):
    """التحقق من تكرار الرقم القومي"""
    national_id = request.GET.get('national_id', '')
    employee_id = request.GET.get('employee_id', None)
    
    if not national_id:
        return JsonResponse({'available': True})
    
    # التحقق من وجود الرقم القومي
    query = Employee.objects.filter(national_id=national_id)
    if employee_id:
        query = query.exclude(pk=employee_id)
    
    exists = query.exists()
    
    return JsonResponse({
        'available': not exists,
        'message': 'هذا الرقم القومي مستخدم بالفعل' if exists else 'الرقم القومي متاح'
    })


# ==================== التقارير ====================

@login_required
def reports_home(request):
    """الصفحة الرئيسية للتقارير"""
    return render(request, 'hr/reports/home.html')


@login_required
def attendance_report(request):
    """تقرير الحضور"""
    return render(request, 'hr/reports/attendance.html')


@login_required
def leave_report(request):
    """تقرير الإجازات"""
    return render(request, 'hr/reports/leave.html')


@login_required
def payroll_report(request):
    """تقرير الرواتب"""
    return render(request, 'hr/reports/payroll.html')


@login_required
def employee_report(request):
    """تقرير الموظفين"""
    return render(request, 'hr/reports/employee.html')


# ==================== Salary Component Templates ====================

@login_required
def salary_component_templates_list(request):
    """قائمة قوالب مكونات الراتب"""
    from .models import SalaryComponentTemplate
    
    templates = SalaryComponentTemplate.objects.all().order_by('component_type', 'order', 'name')
    
    context = {
        'templates': templates,
        'earnings': templates.filter(component_type='earning'),
        'deductions': templates.filter(component_type='deduction'),
    }
    return render(request, 'hr/salary_component_templates/list.html', context)


@login_required
def salary_component_template_form(request, pk=None):
    """نموذج إضافة/تعديل قالب مكون الراتب"""
    from .models import SalaryComponentTemplate
    from .forms.salary_component_template_forms import SalaryComponentTemplateForm
    
    template = None
    if pk:
        template = get_object_or_404(SalaryComponentTemplate, pk=pk)
    
    if request.method == 'POST':
        form = SalaryComponentTemplateForm(request.POST, instance=template)
        if form.is_valid():
            template = form.save()
            messages.success(request, f'تم {"تعديل" if pk else "إضافة"} القالب بنجاح')
            return redirect('hr:salary_component_templates_list')
    else:
        form = SalaryComponentTemplateForm(instance=template)
    
    context = {
        'form': form,
        'template': template,
        'is_edit': pk is not None
    }
    return render(request, 'hr/salary_component_templates/form.html', context)


@login_required
def salary_component_template_delete(request, pk):
    """حذف قالب مكون الراتب"""
    from .models import SalaryComponentTemplate
    
    template = get_object_or_404(SalaryComponentTemplate, pk=pk)
    
    if request.method == 'POST':
        template_name = template.name
        template.delete()
        messages.success(request, f'تم حذف القالب "{template_name}" بنجاح')
        return redirect('hr:salary_component_templates_list')
    
    # تجهيز البيانات للمودال الموحد
    item_fields = [
        {'label': 'الاسم', 'value': template.name},
        {'label': 'النوع', 'value': template.get_component_type_display()},
    ]
    
    if template.formula:
        item_fields.append({'label': 'الصيغة', 'value': template.formula})
    if template.default_amount:
        item_fields.append({'label': 'المبلغ', 'value': template.default_amount})
    
    context = {
        'template': template,
        'item_fields': item_fields,
    }
    return render(request, 'hr/salary_component_templates/delete_modal.html', context)


# ==================== Biometric Dashboard ====================

@login_required
def biometric_dashboard(request):
    """لوحة تحكم شاملة لمراقبة سجلات البصمة"""
    from django.utils import timezone
    from datetime import timedelta
    from hr.models import BiometricLog, BiometricUserMapping, BiometricDevice, Attendance
    
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    # إحصائيات السجلات
    total_logs = BiometricLog.objects.count()
    linked_logs = BiometricLog.objects.filter(employee__isnull=False).count()
    unlinked_logs = total_logs - linked_logs
    # لا نحتاج processed_logs بعد الآن - كل السجلات المربوطة جاهزة للاستخدام
    
    # سجلات اليوم
    today_logs = BiometricLog.objects.filter(timestamp__date=today)
    today_total = today_logs.count()
    today_linked = today_logs.filter(employee__isnull=False).count()
    
    # سجلات الأسبوع
    week_logs = BiometricLog.objects.filter(timestamp__date__gte=week_ago)
    week_total = week_logs.count()
    week_linked = week_logs.filter(employee__isnull=False).count()
    
    # إحصائيات الربط
    total_mappings = BiometricUserMapping.objects.count()
    active_mappings = BiometricUserMapping.objects.filter(is_active=True).count()
    inactive_mappings = total_mappings - active_mappings
    
    # إحصائيات الأجهزة
    total_devices = BiometricDevice.objects.count()
    active_devices = BiometricDevice.objects.filter(is_active=True).count()
    online_devices = BiometricDevice.objects.filter(is_active=True, status='online').count()
    offline_devices = active_devices - online_devices
    
    # إحصائيات الحضور
    today_attendance = Attendance.objects.filter(date=today)
    today_attendance_total = today_attendance.count()
    today_present = today_attendance.filter(status='present').count()
    today_absent = today_attendance.filter(status='absent').count()
    today_late = today_attendance.filter(late_minutes__gt=0).count()
    
    # آخر السجلات غير المربوطة
    recent_unlinked = BiometricLog.objects.filter(
        employee__isnull=True
    ).select_related('device').order_by('-timestamp')[:10]
    
    # آخر السجلات المربوطة (جاهزة للاستخدام)
    recent_linked = BiometricLog.objects.filter(
        employee__isnull=False
    ).select_related('employee', 'device').order_by('-timestamp')[:10]
    
    # الأجهزة غير المتصلة
    offline_device_list = BiometricDevice.objects.filter(
        is_active=True,
        status='offline'
    )[:5]
    
    # إحصائيات يومية للأسبوع الماضي (للرسم البياني)
    daily_stats = []
    for i in range(7):
        day = today - timedelta(days=6-i)
        day_logs = BiometricLog.objects.filter(timestamp__date=day)
        daily_stats.append({
            'date': day.strftime('%Y-%m-%d'),
            'date_display': day.strftime('%d/%m'),
            'total': day_logs.count(),
            'linked': day_logs.filter(employee__isnull=False).count(),
        })
    
    # نسب الإنجاز
    link_percentage = (linked_logs / total_logs * 100) if total_logs > 0 else 0
    
    # نسب الأجهزة
    online_percentage = (online_devices / active_devices * 100) if active_devices > 0 else 0
    offline_percentage = (offline_devices / active_devices * 100) if active_devices > 0 else 0
    
    # نسب الأسبوع
    week_linked_percentage = (week_linked / week_total * 100) if week_total > 0 else 0
    
    context = {
        # إحصائيات عامة
        'total_logs': total_logs,
        'linked_logs': linked_logs,
        'unlinked_logs': unlinked_logs,
        'link_percentage': round(link_percentage, 1),
        
        # إحصائيات اليوم
        'today_total': today_total,
        'today_linked': today_linked,
        
        # إحصائيات الأسبوع
        'week_total': week_total,
        'week_linked': week_linked,
        'week_linked_percentage': round(week_linked_percentage, 1),
        
        # إحصائيات الربط
        'total_mappings': total_mappings,
        'active_mappings': active_mappings,
        'inactive_mappings': inactive_mappings,
        
        # إحصائيات الأجهزة
        'total_devices': total_devices,
        'active_devices': active_devices,
        'online_devices': online_devices,
        'offline_devices': offline_devices,
        'online_percentage': round(online_percentage, 1),
        'offline_percentage': round(offline_percentage, 1),
        
        # إحصائيات الحضور
        'today_attendance_total': today_attendance_total,
        'today_present': today_present,
        'today_absent': today_absent,
        'today_late': today_late,
        
        # قوائم
        'recent_unlinked': recent_unlinked,
        'recent_linked': recent_linked,
        'offline_device_list': offline_device_list,
        'daily_stats': daily_stats,
    }
    
    return render(request, 'hr/biometric/dashboard.html', context)


# ==================== BiometricUserMapping Management ====================

@login_required
def biometric_mapping_list(request):
    """قائمة ربط معرفات البصمة - جدول موحد"""
    from .models import BiometricUserMapping
    
    mappings = BiometricUserMapping.objects.select_related(
        'employee', 'device'
    ).all()
    
    # الفلترة حسب البحث
    search = request.GET.get('search', '')
    if search:
        mappings = mappings.filter(
            Q(employee__first_name_ar__icontains=search) |
            Q(employee__last_name_ar__icontains=search) |
            Q(employee__employee_number__icontains=search) |
            Q(biometric_user_id__icontains=search)
        )
    
    # الفلترة حسب الجهاز
    device_id = request.GET.get('device', '')
    if device_id:
        mappings = mappings.filter(device_id=device_id)
    
    # الفلترة حسب الحالة
    is_active = request.GET.get('is_active', '')
    if is_active:
        mappings = mappings.filter(is_active=(is_active == 'true'))
    
    # إحصائيات
    stats = {
        'total': BiometricUserMapping.objects.count(),
        'active': BiometricUserMapping.objects.filter(is_active=True).count(),
        'inactive': BiometricUserMapping.objects.filter(is_active=False).count(),
    }
    
    # Headers للجدول الموحد
    headers = [
        {'key': 'employee', 'label': 'الموظف', 'width': '25%', 'sortable': True},
        {'key': 'biometric_user_id', 'label': 'معرف البصمة', 'width': '15%', 'sortable': True},
        {'key': 'device', 'label': 'الماكينة', 'width': '20%', 'sortable': True},
        {'key': 'is_active', 'label': 'الحالة', 'width': '10%', 'class': 'text-center'},
        {'key': 'created_at', 'label': 'تاريخ الإنشاء', 'width': '15%', 'format': 'date'},
    ]
    
    # أزرار الإجراءات
    action_buttons = [
        {
            'url': 'hr:biometric_mapping_update',
            'icon': 'fa-edit',
            'class': 'btn-warning',
            'label': 'تعديل'
        },
        {
            'url': 'hr:biometric_mapping_delete',
            'icon': 'fa-trash',
            'class': 'btn-danger',
            'label': 'حذف'
        },
    ]
    
    # جلب الأجهزة للفلترة
    devices = BiometricDevice.objects.filter(is_active=True)
    
    context = {
        'mappings': mappings,
        'stats': stats,
        'headers': headers,
        'action_buttons': action_buttons,
        'devices': devices,
    }
    return render(request, 'hr/biometric/mapping_list.html', context)


@login_required
def biometric_mapping_create(request):
    """إنشاء ربط جديد"""
    from .forms.biometric_forms import BiometricUserMappingForm
    
    if request.method == 'POST':
        form = BiometricUserMappingForm(request.POST)
        if form.is_valid():
            mapping = form.save()
            messages.success(request, f'تم إنشاء الربط بنجاح: {mapping.employee.get_full_name_ar()}')
            return redirect('hr:biometric_mapping_list')
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = BiometricUserMappingForm()
    
    context = {
        'form': form,
        'is_edit': False,
    }
    return render(request, 'hr/biometric/mapping_form.html', context)


@login_required
def biometric_mapping_update(request, pk):
    """تعديل ربط موجود"""
    from .forms.biometric_forms import BiometricUserMappingForm
    from .models import BiometricUserMapping
    
    mapping = get_object_or_404(BiometricUserMapping, pk=pk)
    
    if request.method == 'POST':
        form = BiometricUserMappingForm(request.POST, instance=mapping)
        if form.is_valid():
            mapping = form.save()
            messages.success(request, f'تم تحديث الربط بنجاح: {mapping.employee.get_full_name_ar()}')
            return redirect('hr:biometric_mapping_list')
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = BiometricUserMappingForm(instance=mapping)
    
    context = {
        'form': form,
        'mapping': mapping,
        'is_edit': True,
    }
    return render(request, 'hr/biometric/mapping_form.html', context)


@login_required
def biometric_mapping_delete(request, pk):
    """حذف ربط"""
    from .models import BiometricUserMapping
    
    mapping = get_object_or_404(BiometricUserMapping, pk=pk)
    
    if request.method == 'POST':
        employee_name = mapping.employee.get_full_name_ar()
        biometric_id = mapping.biometric_user_id
        mapping.delete()
        messages.success(request, f'تم حذف الربط: {employee_name} ({biometric_id})')
        return redirect('hr:biometric_mapping_list')
    
    # تجهيز البيانات للمودال
    item_fields = [
        {'label': 'الموظف', 'value': mapping.employee.get_full_name_ar()},
        {'label': 'معرف البصمة', 'value': mapping.biometric_user_id},
        {'label': 'الماكينة', 'value': mapping.device.device_name if mapping.device else 'عام'},
        {'label': 'الحالة', 'value': 'نشط' if mapping.is_active else 'غير نشط'},
    ]
    
    context = {
        'mapping': mapping,
        'item_fields': item_fields,
    }
    return render(request, 'hr/biometric/mapping_delete.html', context)


@login_required
def biometric_mapping_bulk_import(request):
    """استيراد جماعي من CSV"""
    from .forms.biometric_forms import BulkMappingForm
    
    if request.method == 'POST':
        form = BulkMappingForm(request.POST, request.FILES)
        if form.is_valid():
            # معالجة الملف
            stats = form.process_csv()
            
            # عرض النتائج
            if stats['created'] > 0 or stats['updated'] > 0:
                messages.success(
                    request,
                    f'تم الاستيراد بنجاح: {stats["created"]} جديد، {stats["updated"]} محدث'
                )
            
            if stats['errors']:
                for error in stats['errors'][:5]:  # عرض أول 5 أخطاء
                    messages.warning(request, error)
                
                if len(stats['errors']) > 5:
                    messages.warning(request, f'... و {len(stats["errors"]) - 5} خطأ آخر')
            
            return redirect('hr:biometric_mapping_list')
        else:
            messages.error(request, 'يرجى تصحيح الأخطاء في النموذج')
    else:
        form = BulkMappingForm()
    
    context = {
        'form': form,
    }
    return render(request, 'hr/biometric/mapping_bulk_import.html', context)


# ==================== BiometricUserMapping APIs ====================

@api_view(['POST'])
@login_required
def api_link_single_log(request, log_id):
    """
    API لربط سجل واحد بموظف
    
    POST /hr/api/biometric/logs/<log_id>/link/
    Body: {"employee_id": 123} (اختياري)
    """
    from hr.utils.biometric_utils import link_single_log
    from hr.models import BiometricLog
    
    try:
        log = get_object_or_404(BiometricLog, pk=log_id)
        employee_id = request.data.get('employee_id')
        
        success, message = link_single_log(log, employee_id)
        
        return JsonResponse({
            'success': success,
            'message': message,
            'log': {
                'id': log.id,
                'user_id': log.user_id,
                'employee': log.employee.get_full_name_ar() if log.employee else None,
                'timestamp': log.timestamp.isoformat()
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@api_view(['POST'])
@login_required
def api_process_single_log(request, log_id):
    """
    API لمعالجة سجل واحد
    
    POST /hr/api/biometric/logs/<log_id>/process/
    """
    from hr.utils.biometric_utils import process_single_log
    from hr.models import BiometricLog
    
    try:
        log = get_object_or_404(BiometricLog, pk=log_id)
        
        success, message = process_single_log(log)
        
        return JsonResponse({
            'success': success,
            'message': message,
            'log': {
                'id': log.id,
                'is_processed': log.is_processed,
                'attendance_id': log.attendance.id if log.attendance else None
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@api_view(['GET'])
@login_required
def api_mapping_suggestions(request):
    """
    API للحصول على اقتراحات الربط التلقائي
    
    GET /hr/api/biometric/mapping/suggestions/
    """
    from hr.utils.biometric_utils import get_mapping_suggestions
    
    try:
        suggestions = get_mapping_suggestions()
        
        return JsonResponse({
            'success': True,
            'count': len(suggestions),
            'suggestions': suggestions
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@api_view(['POST'])
@login_required
def api_bulk_link_logs(request):
    """
    API للربط الجماعي
    
    POST /hr/api/biometric/logs/bulk-link/
    Body: {
        "device_id": 1 (optional),
        "unlinked_only": true,
        "limit": 100
    }
    """
    from hr.utils.biometric_utils import bulk_link_logs
    
    try:
        device_id = request.data.get('device_id')
        unlinked_only = request.data.get('unlinked_only', True)
        limit = request.data.get('limit')
        
        stats = bulk_link_logs(
            device_id=device_id,
            unlinked_only=unlinked_only,
            dry_run=False,
            limit=limit
        )
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@api_view(['POST'])
@login_required
def api_bulk_process_logs(request):
    """
    API للمعالجة الجماعية
    
    POST /hr/api/biometric/logs/bulk-process/
    Body: {
        "date": "2025-01-15" (optional),
        "employee_id": 1 (optional),
        "unprocessed_only": true
    }
    """
    from hr.utils.biometric_utils import bulk_process_logs
    from datetime import datetime
    
    try:
        date_str = request.data.get('date')
        target_date = None
        if date_str:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        employee_id = request.data.get('employee_id')
        unprocessed_only = request.data.get('unprocessed_only', True)
        
        stats = bulk_process_logs(
            date=target_date,
            employee_id=employee_id,
            unprocessed_only=unprocessed_only,
            dry_run=False
        )
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@api_view(['GET'])
@login_required
def api_biometric_stats(request):
    """
    API للحصول على إحصائيات البصمة
    
    GET /hr/api/biometric/stats/
    """
    from hr.models import BiometricLog, BiometricUserMapping, Attendance
    from django.utils import timezone
    
    try:
        today = timezone.now().date()
        
        # إحصائيات السجلات
        total_logs = BiometricLog.objects.count()
        linked_logs = BiometricLog.objects.filter(employee__isnull=False).count()
        today_logs = BiometricLog.objects.filter(timestamp__date=today).count()
        
        # إحصائيات الربط
        total_mappings = BiometricUserMapping.objects.count()
        active_mappings = BiometricUserMapping.objects.filter(is_active=True).count()
        
        return JsonResponse({
            'success': True,
            'stats': {
                'total': total_logs,
                'linked': linked_logs,
                'unlinked': total_logs - linked_logs,
                'today': today_logs
            },
            'mappings': {
                'total': total_mappings,
                'active': active_mappings,
                'inactive': total_mappings - active_mappings
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["GET"])
def employee_detail_api(request, pk):
    """
    API لجلب بيانات الموظف
    """
    import logging
    import traceback
    logger = logging.getLogger(__name__)
    
    try:
        print(f"=== API Call: محاولة جلب بيانات الموظف {pk} ===")
        logger.info(f"محاولة جلب بيانات الموظف {pk}")
        employee = Employee.objects.select_related('job_title', 'department').get(pk=pk)
        print(f"=== تم العثور على الموظف: {employee.first_name_ar} ===")
        logger.info(f"تم العثور على الموظف")
        
        # بناء البيانات خطوة بخطوة
        data = {}
        
        try:
            print("=== بناء البيانات الأساسية ===")
            data['id'] = employee.id
            data['employee_number'] = employee.employee_number
            data['name'] = f"{employee.first_name_ar} {employee.last_name_ar}"
            
            # جلب رقم البصمة من BiometricUserMapping
            from .models import BiometricUserMapping
            biometric_mapping = BiometricUserMapping.objects.filter(employee=employee, is_active=True).first()
            data['biometric_user_id'] = biometric_mapping.biometric_user_id if biometric_mapping else None
            
            print(f"=== البيانات الأساسية: {data} ===")
        except Exception as e:
            print(f"=== خطأ في البيانات الأساسية: {e} ===")
            logger.error(f"خطأ في البيانات الأساسية: {e}")
            raise
        
        # إضافة الوظيفة
        try:
            data['job_title'] = employee.job_title.id if employee.job_title else None
            data['job_title_name'] = str(employee.job_title) if employee.job_title else None
        except Exception as e:
            logger.error(f"خطأ في الوظيفة: {e}")
            data['job_title'] = None
            data['job_title_name'] = None
        
        # إضافة القسم
        try:
            data['department'] = employee.department.id if employee.department else None
            data['department_name'] = str(employee.department) if employee.department else None
        except Exception as e:
            logger.error(f"خطأ في القسم: {e}")
            data['department'] = None
            data['department_name'] = None
        
        logger.info(f"تم إرجاع البيانات بنجاح: {data}")
        return JsonResponse(data)
        
    except Employee.DoesNotExist:
        logger.error(f"الموظف {pk} غير موجود")
        return JsonResponse({
            'error': 'الموظف غير موجود'
        }, status=404)
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"=== خطأ في API ===")
        print(error_trace)
        logger.error(f"خطأ في API: {e}\n{error_trace}")
        return JsonResponse({
            'error': str(e),
            'traceback': error_trace
        }, status=500)


@login_required
@require_http_methods(["POST"])
def contract_check_overlap(request):
    """
    التحقق من تداخل العقود عبر AJAX
    """
    from datetime import datetime
    from django.http import JsonResponse
    
    try:
        employee_id = request.POST.get('employee_id')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        contract_id = request.POST.get('contract_id')
        
        if not employee_id or not start_date_str:
            return JsonResponse({
                'has_overlap': False,
                'message': ''
            })
        
        # تحويل التواريخ
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = None
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # البحث عن عقود متداخلة
        allowed_statuses = ['expired', 'terminated', 'suspended']
        overlapping_contracts = Contract.objects.filter(
            employee_id=employee_id
        ).exclude(
            status__in=allowed_statuses
        )
        
        # استثناء العقد الحالي عند التعديل
        if contract_id:
            overlapping_contracts = overlapping_contracts.exclude(pk=contract_id)
        
        # التحقق من التداخل
        for contract in overlapping_contracts:
            # حالة 1: العقد الموجود مفتوح
            if not contract.end_date:
                return JsonResponse({
                    'has_overlap': True,
                    'message': f'''
                        <div class="text-end">
                            <p class="mb-2">يوجد عقد ساري للموظف بدون تاريخ نهاية:</p>
                            <p class="mb-2"><strong>رقم العقد:</strong> {contract.contract_number}</p>
                            <p class="mb-3"><strong>تاريخ البداية:</strong> {contract.start_date.strftime("%d/%m/%Y")}</p>
                            <div class="alert alert-info mb-0">
                                <strong>الحل:</strong> يجب إنهاء أو إيقاف العقد الحالي أولاً
                            </div>
                        </div>
                    '''
                })
            
            # حالة 2: العقد الجديد مفتوح
            if not end_date:
                return JsonResponse({
                    'has_overlap': True,
                    'message': f'''
                        <div class="text-end">
                            <p class="mb-2">يوجد عقد ساري للموظف:</p>
                            <p class="mb-2"><strong>رقم العقد:</strong> {contract.contract_number}</p>
                            <p class="mb-2"><strong>الفترة:</strong> من {contract.start_date.strftime("%d/%m/%Y")} إلى {contract.end_date.strftime("%d/%m/%Y")}</p>
                            <div class="alert alert-info mb-0">
                                <strong>الحل:</strong> لا يمكن إنشاء عقد مفتوح. يجب تحديد تاريخ نهاية أو إنهاء العقد الحالي
                            </div>
                        </div>
                    '''
                })
            
            # حالة 3: تداخل في الفترات
            if start_date <= contract.end_date and end_date >= contract.start_date:
                return JsonResponse({
                    'has_overlap': True,
                    'message': f'''
                        <div class="text-end">
                            <p class="mb-2">يوجد تداخل مع عقد ساري:</p>
                            <p class="mb-2"><strong>رقم العقد:</strong> {contract.contract_number}</p>
                            <p class="mb-3"><strong>الفترة:</strong> من {contract.start_date.strftime("%d/%m/%Y")} إلى {contract.end_date.strftime("%d/%m/%Y")}</p>
                            <div class="alert alert-warning mb-2">
                                <strong>الحلول المتاحة:</strong>
                                <ol class="mb-0 mt-2">
                                    <li>إنهاء أو إيقاف العقد الحالي أولاً</li>
                                    <li>استخدام خاصية التجديد من صفحة العقد</li>
                                    <li>تعديل تاريخ البداية ليكون بعد {contract.end_date.strftime("%d/%m/%Y")}</li>
                                </ol>
                            </div>
                        </div>
                    '''
                })
        
        # لا يوجد تداخل
        return JsonResponse({
            'has_overlap': False,
            'message': ''
        })
    
    except Exception as e:
        return JsonResponse({
            'has_overlap': False,
            'message': '',
            'error': str(e)
        }, status=500)
