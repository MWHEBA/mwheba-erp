# -*- coding: utf-8 -*-
"""
Product Import Views
Flow: Upload → Preview (dry-run) → Confirm → Execute → Results
"""

import io
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.urls import reverse

from product.services.product_import_service import ProductImportService

SESSION_KEY = 'product_import_pending'


def _base_context(extra=None):
    from product.models.stock_management import Warehouse
    ctx = {
        'title': 'استيراد المنتجات',
        'page_subtitle': 'رفع ملف Excel أو CSV لاستيراد المنتجات دفعة واحدة',
        'page_icon': 'fas fa-file-import',
        'active_menu': 'products',
        'header_buttons': [
            {'url': reverse('product:product_import_template'), 'icon': 'fa-download',
             'text': 'تحميل القالب', 'class': 'btn-outline-info'},
            {'url': reverse('product:product_list'), 'icon': 'fa-arrow-right',
             'text': 'العودة للمنتجات', 'class': 'btn-outline-secondary'},
        ],
        'breadcrumb_items': [
            {'title': 'الرئيسية', 'url': reverse('core:dashboard'), 'icon': 'fas fa-home'},
            {'title': 'المنتجات', 'url': reverse('product:product_list'), 'icon': 'fas fa-boxes'},
            {'title': 'استيراد المنتجات', 'active': True},
        ],
        'warehouses': Warehouse.objects.filter(is_active=True).order_by('name'),
    }
    if extra:
        ctx.update(extra)
    return ctx


@login_required
def product_import(request):
    """Step 1: Upload form. Step 2: Preview. Step 3: Confirm → execute."""

    # ── CONFIRM ───────────────────────────────────────────────────────────────
    if request.method == 'POST' and request.POST.get('action') == 'confirm':
        pending = request.session.get(SESSION_KEY)
        if not pending:
            messages.error(request, 'انتهت صلاحية جلسة المعاينة. يرجى رفع الملف مجدداً.')
            return redirect('product:product_import')

        rows = pending['rows']
        update_existing = pending['update_existing']
        warehouse_id = pending.get('warehouse_id')

        service = ProductImportService(user=request.user)
        result = service._execute_from_preview(rows, update_existing, warehouse_id=warehouse_id)

        request.session.pop(SESSION_KEY, None)

        ctx = _base_context({'import_result': result, 'row_errors': result.get('errors', [])})

        if result['created'] > 0 or result['updated'] > 0:
            parts = []
            if result['created']:
                parts.append(f'{result["created"]} منتج جديد')
            if result['updated']:
                parts.append(f'{result["updated"]} منتج محدّث')
            messages.success(request, f'تم الاستيراد بنجاح: {" و ".join(parts)}')
        else:
            messages.warning(request, 'لم يتم إضافة أي منتجات.')

        return render(request, 'product/product_import.html', ctx)

    # ── UPLOAD ────────────────────────────────────────────────────────────────
    if request.method == 'POST' and request.POST.get('action') == 'upload':
        uploaded_file = request.FILES.get('import_file')
        update_existing = request.POST.get('update_existing') == 'on'
        warehouse_id = request.POST.get('warehouse_id') or None

        if not uploaded_file:
            ctx = _base_context({'import_error': 'يرجى اختيار ملف للاستيراد'})
            return render(request, 'product/product_import.html', ctx)

        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls', '.csv')):
            ctx = _base_context({'import_error': 'صيغة الملف غير مدعومة. يرجى استخدام Excel أو CSV'})
            return render(request, 'product/product_import.html', ctx)

        if uploaded_file.size > 5 * 1024 * 1024:
            ctx = _base_context({'import_error': 'حجم الملف كبير جداً. الحد الأقصى 5 ميجابايت'})
            return render(request, 'product/product_import.html', ctx)

        service = ProductImportService(user=request.user)
        result = service.import_from_file(uploaded_file, update_existing=update_existing, dry_run=True)

        if not result['success']:
            ctx = _base_context({'import_error': result.get('error'), 'row_errors': result.get('errors', [])})
            return render(request, 'product/product_import.html', ctx)

        request.session[SESSION_KEY] = {
            'rows': result['preview_rows'],
            'update_existing': update_existing,
            'filename': uploaded_file.name,
            'warehouse_id': warehouse_id,
        }

        from django.core.paginator import Paginator
        paginator = Paginator(result['preview_rows'], 50)
        page_obj = paginator.get_page(request.GET.get('page', 1))

        ctx = _base_context({
            'preview': result,
            'preview_rows_page': page_obj,
            'filename': uploaded_file.name,
            'update_existing': update_existing,
            'selected_warehouse_id': warehouse_id,
        })
        return render(request, 'product/product_import.html', ctx)

    # ── GET: pagination or upload form ────────────────────────────────────────
    pending = request.session.get(SESSION_KEY)
    if pending and request.GET.get('page'):
        from django.core.paginator import Paginator
        rows = pending['rows']
        paginator = Paginator(rows, 50)
        page_obj = paginator.get_page(request.GET.get('page', 1))

        valid_count  = sum(1 for r in rows if r['action'] == 'new')
        update_count = sum(1 for r in rows if r['action'] == 'update')
        error_count  = sum(1 for r in rows if r['action'] == 'error')
        skip_count   = sum(1 for r in rows if r['action'] == 'skip')

        preview = {
            'valid_count': valid_count,
            'update_count': update_count,
            'error_count': error_count,
            'skip_count': skip_count,
            'total_rows': len(rows),
        }
        ctx = _base_context({
            'preview': preview,
            'preview_rows_page': page_obj,
            'filename': pending.get('filename', ''),
            'update_existing': pending['update_existing'],
            'selected_warehouse_id': pending.get('warehouse_id'),
        })
        return render(request, 'product/product_import.html', ctx)

    request.session.pop(SESSION_KEY, None)
    return render(request, 'product/product_import.html', _base_context())


@login_required
def product_import_template(request):
    """Download the Excel import template."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'المنتجات'

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        req_fill    = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        opt_fill    = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
        center      = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right       = Alignment(horizontal='right',  vertical='center')
        thin        = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'),  bottom=Side(style='thin'))

        headers = [
            ('اسم المنتج *',          30, True),
            ('التصنيف *',             20, True),
            ('وحدة القياس *',         18, True),
            ('سعر التكلفة *',         16, True),
            ('سعر البيع *',           16, True),
            ('كود المنتج',            18, False),
            ('الباركود',              16, False),
            ('الوصف',                 30, False),
            ('الحد الأدنى للمخزون',   22, False),
            ('نشط (نعم/لا)',          18, False),
            ('خدمة (نعم/لا)',         18, False),
            ('نوع المنتج',            20, False),
            ('الكمية',                14, False),
        ]

        for ci, (h, w, req) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font, cell.fill, cell.alignment, cell.border = header_font, header_fill, center, thin
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 30

        samples = [
            ['قلم رصاص',       'قرطاسية',  'قطعة', 2,   5,   '',            '', '',             50, 'نعم', 'لا',  'stationery', 100],
            ['ورق A4',         'قرطاسية',  'رزمة', 25,  40,  '',            '', 'ورق طباعة',    20, 'نعم', 'لا',  'stationery', 50],
            ['خدمة الطباعة',   'خدمات',    'خدمة', 0,   50,  'SRV-PRINT',   '', 'طباعة وثائق',  0,  'نعم', 'نعم', 'general',    0],
            ['حبر طابعة',      'مستلزمات', 'قطعة', 80,  120, '',            '', '',             10, 'نعم', 'لا',  'general',    20],
        ]
        for ri, row in enumerate(samples, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = req_fill if headers[ci-1][2] else opt_fill
                cell.alignment, cell.border = right, thin

        ws2 = wb.create_sheet('تعليمات')
        notes = [
            ('تعليمات الاستيراد', True),
            ('', False),
            ('الأعمدة المطلوبة (*):', True),
            ('• اسم المنتج، التصنيف، وحدة القياس، سعر التكلفة، سعر البيع', False),
            ('', False),
            ('الأعمدة الاختيارية:', True),
            ('• كود المنتج: يُولَّد تلقائياً إن تُرك فارغاً', False),
            ('• نشط: نعم أو لا (افتراضي: نعم)', False),
            ('• خدمة: نعم أو لا (افتراضي: لا)', False),
            ('• نوع المنتج: educational, uniform, stationery, activity, kitchen, cleaning, medical, general', False),
            ('• الكمية: الكمية الافتتاحية في المخزن (افتراضي: 0)', False),
            ('', False),
            ('ملاحظات:', True),
            ('• تأكد من أن التصنيفات ووحدات القياس موجودة في النظام', False),
            ('• المنتجات المكررة تُتخطى ما لم تفعّل خيار التحديث', False),
            ('• يمكن استخدام الأعمدة بالعربي أو الإنجليزي', False),
        ]
        for ri, (txt, bold) in enumerate(notes, 1):
            c = ws2.cell(row=ri, column=1, value=txt)
            if bold:
                c.font = Font(bold=True, size=12)
        ws2.column_dimensions['A'].width = 80

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        resp = HttpResponse(out.getvalue(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="product_import_template.xlsx"'
        return resp

    except ImportError:
        import csv
        resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        resp['Content-Disposition'] = 'attachment; filename="product_import_template.csv"'
        writer = csv.writer(resp)
        writer.writerow(ProductImportService.get_template_arabic_headers())
        writer.writerow(['قلم رصاص', 'قرطاسية', 'قطعة', '2', '5', '', '', '', '50', 'نعم', 'لا', 'stationery', '100'])
        return resp
