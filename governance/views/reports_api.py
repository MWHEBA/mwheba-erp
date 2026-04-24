"""
Reports Builder API Views
"""

from django.http import JsonResponse
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils import timezone
import json
import logging

from governance.services.reports_builder_service import ReportsBuilderService
from governance.models import SavedReport, ReportSchedule, ReportExecution

logger = logging.getLogger(__name__)


class GetAvailableFieldsAPI(LoginRequiredMixin, View):
    """API للحصول على الحقول المتاحة لمصدر بيانات"""
    
    def get(self, request):
        data_source = request.GET.get('data_source')
        
        if not data_source:
            return JsonResponse({'success': False, 'error': 'مصدر البيانات مطلوب'}, status=400)
        
        fields = ReportsBuilderService.get_available_fields(data_source)
        
        return JsonResponse({
            'success': True,
            'fields': fields
        })


class GetAvailableFieldsAPI(LoginRequiredMixin, View):
    """API للحصول على الحقول المتاحة لمصدر بيانات"""
    
    def get(self, request):
        data_source = request.GET.get('data_source')
        
        if not data_source:
            return JsonResponse({'success': False, 'error': 'مصدر البيانات مطلوب'}, status=400)
        
        fields = ReportsBuilderService.get_available_fields(data_source)
        
        return JsonResponse({
            'success': True,
            'fields': fields
        })


class GenerateReportAPI(LoginRequiredMixin, View):
    """API لتوليد تقرير"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            
            # Validate required fields
            required_fields = ['data_source', 'selected_fields', 'report_type']
            for field in required_fields:
                if field not in data:
                    return JsonResponse({
                        'success': False,
                        'error': f'الحقل {field} مطلوب'
                    }, status=400)
            
            # Execute report
            result = ReportsBuilderService.execute_report(data, user=request.user)
            
            return JsonResponse(result)
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'بيانات JSON غير صالحة'
            }, status=400)
        except Exception as e:
            logger.error(f"Error generating report: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class SaveReportAPI(LoginRequiredMixin, View):
    """API لحفظ تقرير"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            
            # Validate required fields
            if not data.get('name'):
                return JsonResponse({
                    'success': False,
                    'error': 'اسم التقرير مطلوب'
                }, status=400)
            
            # Create saved report
            report = SavedReport.objects.create(
                name=data['name'],
                description=data.get('description', ''),
                report_type=data.get('report_type', 'table'),
                data_source=data.get('data_source'),
                selected_fields=data.get('selected_fields', []),
                filters=data.get('filters', {}),
                group_by=data.get('group_by', ''),
                sort_by=data.get('sort_by', ''),
                sort_order=data.get('sort_order', 'asc'),
                created_by=request.user,
                is_public=data.get('is_public', False)
            )
            
            return JsonResponse({
                'success': True,
                'report_id': report.id,
                'message': 'تم حفظ التقرير بنجاح'
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'بيانات JSON غير صالحة'
            }, status=400)
        except Exception as e:
            logger.error(f"Error saving report: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class RunSavedReportAPI(LoginRequiredMixin, View):
    """API لتشغيل تقرير محفوظ"""
    
    def post(self, request, report_id):
        try:
            # Get report
            report = SavedReport.objects.get(id=report_id)
            
            # Check access
            if not report.can_user_access(request.user):
                return JsonResponse({
                    'success': False,
                    'error': 'ليس لديك صلاحية الوصول لهذا التقرير'
                }, status=403)
            
            # Create execution record
            execution = ReportExecution.objects.create(
                report=report,
                status='RUNNING',
                triggered_by=request.user
            )
            
            # Execute report
            report_config = {
                'data_source': report.data_source,
                'selected_fields': report.selected_fields,
                'filters': report.filters,
                'group_by': report.group_by,
                'sort_by': report.sort_by,
                'sort_order': report.sort_order,
                'report_type': report.report_type,
            }
            
            result = ReportsBuilderService.execute_report(report_config, user=request.user)
            
            # Update execution record
            if result['success']:
                execution.mark_as_success(result['rows_count'], result['data'])
                report.increment_run_count()
            else:
                execution.mark_as_failed(result['error'])
            
            return JsonResponse(result)
            
        except SavedReport.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'التقرير غير موجود'
            }, status=404)
        except Exception as e:
            logger.error(f"Error running report: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class DeleteReportAPI(LoginRequiredMixin, View):
    """API لحذف تقرير"""
    
    def post(self, request, report_id):
        try:
            report = SavedReport.objects.get(id=report_id)
            
            # Check permission
            if report.created_by != request.user and not request.user.is_superuser:
                return JsonResponse({
                    'success': False,
                    'error': 'ليس لديك صلاحية حذف هذا التقرير'
                }, status=403)
            
            report.status = 'ARCHIVED'
            report.save()
            
            return JsonResponse({
                'success': True,
                'message': 'تم حذف التقرير بنجاح'
            })
            
        except SavedReport.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'التقرير غير موجود'
            }, status=404)
        except Exception as e:
            logger.error(f"Error deleting report: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class CreateScheduleAPI(LoginRequiredMixin, View):
    """API لإنشاء جدولة تقرير"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            
            # Validate required fields
            required_fields = ['report_id', 'frequency', 'schedule_time', 'email_recipients']
            for field in required_fields:
                if field not in data:
                    return JsonResponse({
                        'success': False,
                        'error': f'الحقل {field} مطلوب'
                    }, status=400)
            
            # Get report
            report = SavedReport.objects.get(id=data['report_id'])
            
            # Check access
            if not report.can_user_access(request.user):
                return JsonResponse({
                    'success': False,
                    'error': 'ليس لديك صلاحية جدولة هذا التقرير'
                }, status=403)
            
            # Parse time
            from datetime import datetime
            schedule_time = datetime.strptime(data['schedule_time'], '%H:%M').time()
            
            # Create schedule
            schedule = ReportSchedule.objects.create(
                report=report,
                frequency=data['frequency'],
                schedule_time=schedule_time,
                day_of_week=data.get('day_of_week'),
                day_of_month=data.get('day_of_month'),
                email_recipients=data['email_recipients'],
                created_by=request.user
            )
            
            # Calculate next run
            schedule.calculate_next_run()
            
            return JsonResponse({
                'success': True,
                'schedule_id': schedule.id,
                'next_run_at': schedule.next_run_at.isoformat() if schedule.next_run_at else None,
                'message': 'تم إنشاء الجدولة بنجاح'
            })
            
        except SavedReport.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'التقرير غير موجود'
            }, status=404)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'بيانات JSON غير صالحة'
            }, status=400)
        except Exception as e:
            logger.error(f"Error creating schedule: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class PauseScheduleAPI(LoginRequiredMixin, View):
    """API لإيقاف جدولة مؤقتاً"""
    
    def post(self, request, schedule_id):
        try:
            schedule = ReportSchedule.objects.get(id=schedule_id)
            
            # Check permission
            if schedule.created_by != request.user and not request.user.is_superuser:
                return JsonResponse({
                    'success': False,
                    'error': 'ليس لديك صلاحية تعديل هذه الجدولة'
                }, status=403)
            
            schedule.pause()
            
            return JsonResponse({
                'success': True,
                'message': 'تم إيقاف الجدولة مؤقتاً'
            })
            
        except ReportSchedule.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'الجدولة غير موجودة'
            }, status=404)
        except Exception as e:
            logger.error(f"Error pausing schedule: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class ResumeScheduleAPI(LoginRequiredMixin, View):
    """API لاستئناف جدولة"""
    
    def post(self, request, schedule_id):
        try:
            schedule = ReportSchedule.objects.get(id=schedule_id)
            
            # Check permission
            if schedule.created_by != request.user and not request.user.is_superuser:
                return JsonResponse({
                    'success': False,
                    'error': 'ليس لديك صلاحية تعديل هذه الجدولة'
                }, status=403)
            
            schedule.resume()
            
            return JsonResponse({
                'success': True,
                'message': 'تم استئناف الجدولة',
                'next_run_at': schedule.next_run_at.isoformat() if schedule.next_run_at else None
            })
            
        except ReportSchedule.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'الجدولة غير موجودة'
            }, status=404)
        except Exception as e:
            logger.error(f"Error resuming schedule: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)


class DownloadReportAPI(LoginRequiredMixin, View):
    """API لتحميل تقرير"""
    
    def get(self, request, report_id):
        try:
            from django.http import HttpResponse
            import io
            
            # Get report
            try:
                report = SavedReport.objects.get(id=report_id)
            except SavedReport.DoesNotExist:
                logger.warning(f"Report {report_id} not found")
                return JsonResponse({
                    'success': False,
                    'error': 'التقرير غير موجود'
                }, status=404)
            
            # Check access
            if not report.can_user_access(request.user):
                logger.warning(f"User {request.user.username} denied access to report {report_id}")
                return JsonResponse({
                    'success': False,
                    'error': 'ليس لديك صلاحية الوصول لهذا التقرير'
                }, status=403)
            
            # Get format
            format_type = request.GET.get('format', 'excel')
            logger.info(f"Downloading report {report_id} in {format_type} format")
            
            # Execute report to get data
            report_config = {
                'data_source': report.data_source,
                'selected_fields': report.selected_fields if report.selected_fields else [],
                'filters': report.filters if report.filters else {},
                'group_by': report.group_by or '',
                'sort_by': report.sort_by or '',
                'sort_order': report.sort_order or 'asc',
            }
            
            logger.debug(f"Report config: {report_config}")
            
            result = ReportsBuilderService.execute_report(report_config, user=request.user)
            
            if not result['success']:
                logger.error(f"Report execution failed: {result.get('error')}")
                return JsonResponse({
                    'success': False,
                    'error': result['error']
                }, status=500)
            
            data = result['data']
            logger.info(f"Report executed successfully with {len(data)} rows")
            
            # Generate file based on format
            if format_type == 'excel':
                return self._generate_excel(report, data)
            elif format_type == 'csv':
                return self._generate_csv(report, data)
            elif format_type == 'pdf':
                return self._generate_pdf(report, data)
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'صيغة غير مدعومة'
                }, status=400)
            
        except Exception as e:
            logger.error(f"Error downloading report {report_id}: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': f'حدث خطأ أثناء تحميل التقرير: {str(e)}'
            }, status=500)
    
    def _generate_excel(self, report, data):
        """توليد ملف Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse
            import io
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = report.name[:31]  # Excel sheet name limit
            
            # Add headers and data
            if data:
                headers = list(data[0].keys())
                num_cols = len(headers)
                
                # Add title (merge cells based on actual column count)
                if num_cols > 0:
                    last_col = get_column_letter(num_cols)
                    ws.merge_cells(f'A1:{last_col}1')
                    title_cell = ws['A1']
                    title_cell.value = report.name
                    title_cell.font = Font(size=16, bold=True)
                    title_cell.alignment = Alignment(horizontal='center')
                
                # Add headers
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=3, column=col_idx)
                    cell.value = header
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                # Add data
                for row_idx, row_data in enumerate(data, start=4):
                    for col_idx, header in enumerate(headers, start=1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = row_data.get(header, '')
                        # Handle different data types
                        try:
                            cell.value = value
                        except Exception:
                            cell.value = str(value) if value is not None else ''
                
                # Auto-adjust column widths
                for col_idx in range(1, num_cols + 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = 0
                    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            try:
                                if cell.value:
                                    max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            else:
                # No data - add title only
                ws['A1'] = report.name
                ws['A1'].font = Font(size=16, bold=True)
                ws['A3'] = 'لا توجد بيانات'
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create response
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{report.name}.xlsx"'
            
            return response
            
        except ImportError as e:
            logger.error(f"openpyxl not installed: {e}")
            return JsonResponse({
                'success': False,
                'error': 'مكتبة openpyxl غير مثبتة'
            }, status=500)
        except Exception as e:
            logger.error(f"Error generating Excel: {e}", exc_info=True)
            return JsonResponse({
                'success': False,
                'error': f'خطأ في توليد ملف Excel: {str(e)}'
            }, status=500)
    
    def _generate_csv(self, report, data):
        """توليد ملف CSV"""
        import csv
        from django.http import HttpResponse
        import io
        
        output = io.StringIO()
        
        if data:
            headers = list(data[0].keys())
            writer = csv.DictWriter(output, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
        
        response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="{report.name}.csv"'
        
        return response
    
    def _generate_pdf(self, report, data):
        """توليد ملف PDF"""
        from django.http import HttpResponse
        
        # For now, return a simple message
        # You can implement full PDF generation using reportlab or weasyprint
        return JsonResponse({
            'success': False,
            'error': 'تصدير PDF قيد التطوير. استخدم Excel أو CSV'
        }, status=501)

